from __future__ import annotations

import re
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

DROP_DIR = Path("drop-excel-files")
OUTPUT_DIR = Path("pdf-output")
COMPLAINT_NO_RE = re.compile(r"\b\d{3}-\d{7}\b")
EXCEL_SUFFIXES = {".xlsx", ".xlsm", ".xltx", ".xltm"}


def main() -> None:
    workbooks = list(find_workbooks(DROP_DIR))
    if not workbooks:
        raise SystemExit(f"No Excel files found in {DROP_DIR}")

    OUTPUT_DIR.mkdir(exist_ok=True)

    total = 0
    for workbook_path in workbooks:
        count = convert_workbook(workbook_path, OUTPUT_DIR)
        total += count
        print(f"{workbook_path.name}: created {count} PDF(s)")

    print(f"Done. Created {total} PDF(s) in {OUTPUT_DIR}")


def find_workbooks(directory: Path) -> Iterable[Path]:
    for path in sorted(directory.iterdir()):
        if path.name.startswith(("~$", ".~lock.")):
            continue
        if path.is_file() and path.suffix.lower() in EXCEL_SUFFIXES:
            yield path


def convert_workbook(workbook_path: Path, output_dir: Path) -> int:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    created = 0

    for worksheet in workbook.worksheets:
        header_row_index, headers = find_header_row(
            worksheet.iter_rows(values_only=True)
        )
        if header_row_index is None:
            continue

        complaint_index = find_complaint_column(headers)

        for row_index, row in enumerate(
            worksheet.iter_rows(min_row=header_row_index + 1, values_only=True),
            start=header_row_index + 1,
        ):
            if row_is_empty(row):
                continue

            row_data = build_row_data(headers, row)
            complaint_number = get_complaint_number(row_data, row, complaint_index)
            if not complaint_number:
                print(
                    f"Skipping {workbook_path.name} / {worksheet.title} row {row_index}: "
                    "no complaint number found"
                )
                continue

            write_pdf(
                output_path=output_dir / f"{complaint_number}.pdf",
                complaint_number=complaint_number,
                workbook_name=workbook_path.name,
                sheet_name=worksheet.title,
                row_index=row_index,
                row_data=row_data,
            )
            created += 1

    workbook.close()
    return created


def find_header_row(rows: Iterable[tuple[object, ...]]) -> tuple[int | None, list[str]]:
    for index, row in enumerate(rows, start=1):
        values = [clean_text(value) for value in row]
        non_empty = [value for value in values if value]
        if not non_empty:
            continue
        if any(COMPLAINT_NO_RE.fullmatch(value) for value in non_empty):
            continue
        normalized_values = {normalize_header(value) for value in non_empty}
        if normalized_values.intersection({"complaintno", "complaintnumber"}):
            return index, make_unique_headers(values)
    return None, []


def make_unique_headers(headers: list[str]) -> list[str]:
    result: list[str] = []
    seen: dict[str, int] = {}

    for column_number, header in enumerate(headers, start=1):
        base = header or f"Column {column_number}"
        count = seen.get(base, 0) + 1
        seen[base] = count
        result.append(base if count == 1 else f"{base} ({count})")

    return result


def find_complaint_column(headers: list[str]) -> int | None:
    for index, header in enumerate(headers):
        if normalize_header(header) in {"complaintno", "complaintnumber"}:
            return index
    return None


def normalize_header(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", value.lower())


def build_row_data(
    headers: list[str], row: tuple[object, ...]
) -> list[tuple[str, str]]:
    data: list[tuple[str, str]] = []
    width = max(len(headers), len(row))

    for index in range(width):
        header = headers[index] if index < len(headers) else f"Column {index + 1}"
        value = row[index] if index < len(row) else None
        data.append((header, clean_text(value)))

    return data


def get_complaint_number(
    row_data: list[tuple[str, str]],
    row: tuple[object, ...],
    complaint_index: int | None,
) -> str | None:
    if complaint_index is not None and complaint_index < len(row):
        match = COMPLAINT_NO_RE.search(clean_text(row[complaint_index]))
        if match:
            return match.group(0)

    for _, value in row_data:
        match = COMPLAINT_NO_RE.search(value)
        if match:
            return match.group(0)

    return None


def row_is_empty(row: tuple[object, ...]) -> bool:
    return not any(clean_text(value) for value in row)


def clean_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, Decimal):
        return format(value, "f")
    return str(value).strip()


def write_pdf(
    *,
    output_path: Path,
    complaint_number: str,
    workbook_name: str,
    sheet_name: str,
    row_index: int,
    row_data: list[tuple[str, str]],
) -> None:
    styles = getSampleStyleSheet()
    normal = ParagraphStyle(
        "RowValue",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=9,
        leading=12,
        alignment=TA_LEFT,
        wordWrap="CJK",
    )
    label = ParagraphStyle(
        "RowLabel",
        parent=normal,
        fontName="Helvetica-Bold",
        textColor=colors.HexColor("#1f2937"),
    )
    title = ParagraphStyle(
        "ComplaintTitle",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=16,
        leading=20,
        spaceAfter=8,
        textColor=colors.HexColor("#111827"),
    )
    meta = ParagraphStyle(
        "Meta",
        parent=normal,
        fontSize=8,
        leading=10,
        textColor=colors.HexColor("#4b5563"),
        spaceAfter=12,  # Added a bit of space below the meta text
    )

    # A slightly larger label style for items broken out of the table
    long_item_label = ParagraphStyle(
        "LongItemLabel",
        parent=label,
        fontSize=11,
        spaceBefore=12,
        spaceAfter=4,
    )

    document = SimpleDocTemplate(
        str(output_path),
        pagesize=A4,
        rightMargin=16 * mm,
        leftMargin=16 * mm,
        topMargin=14 * mm,
        bottomMargin=14 * mm,
        title=f"Complaint {complaint_number}",
        author="excel-row-to-pdf",
    )

    story = [
        Paragraph(f"Complaint {escape_text(complaint_number)}", title),
        Paragraph(
            escape_text(
                f"Source: {workbook_name} | Sheet: {sheet_name} | Row: {row_index}"
            ),
            meta,
        ),
    ]

    table_rows = []
    long_items = []

    for field, value in row_data:
        # Check if the text is too long or has too many newlines for a single table cell.
        # 1500 chars is a safe threshold that guarantees it won't exceed a page height.
        if value and (len(value) > 1500 or value.count("\n") > 25):
            long_items.append((field, value))
        else:
            table_rows.append(
                [
                    Paragraph(escape_text(field), label),
                    Paragraph(escape_text(value) if value else "-", normal),
                ]
            )

    # 1. Render the standard data inside the styled Table
    if table_rows:
        table = Table(table_rows, colWidths=[48 * mm, 130 * mm], repeatRows=0)
        table.setStyle(
            TableStyle(
                [
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#f3f4f6")),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#d1d5db")),
                    ("LEFTPADDING", (0, 0), (-1, -1), 5),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ]
            )
        )
        story.append(table)

    # 2. Render massive data blocks below the table as standard paginating flowables
    if long_items:
        story.append(Spacer(1, 6 * mm))
        for field, value in long_items:
            story.append(Paragraph(escape_text(field), long_item_label))
            story.append(Paragraph(escape_text(value), normal))
            story.append(Spacer(1, 4 * mm))

    document.build(story)


def escape_text(value: str) -> str:
    return (
        value.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace("\n", "<br/>")
    )


if __name__ == "__main__":
    main()
