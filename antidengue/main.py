import asyncio
import datetime
import hashlib
import json
import logging
import os
import re
import shutil
import sqlite3
import sys
import time
import uuid
from collections import Counter
from decimal import Decimal, InvalidOperation
from pathlib import Path

import nats  # Removed httpx
import pandas as pd
from dotenv import load_dotenv
from nats.js.api import RetentionPolicy, StorageType, StreamConfig
from nats.errors import NoRespondersError, TimeoutError as NatsTimeoutError
from nats.js.errors import NotFoundError
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

# Import your custom scraper task
from scraper import scrape_portal_reports
from hotspot_analysis import write_hotspot_review_report
from simple_activity_analysis import write_simple_activity_review_report

# ==========================================
# 1. CONFIGURATION & FOLDER PATHS
# ==========================================
BASE_DIR = Path(__file__).parent.resolve()
load_dotenv(BASE_DIR / ".env")


def _resolve_project_path(value: str | Path) -> Path:
    candidate = Path(value)
    return candidate if candidate.is_absolute() else (BASE_DIR / candidate).resolve()


def _get_logger():
    return logging.getLogger("antidengue")


def _configure_cli_logging() -> None:
    """Emit subprocess logs for Celery's streamed-command collector."""
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s | %(levelname)s | %(name)s | %(message)s",
        stream=sys.stdout,
    )


OUTPUT_DIR = BASE_DIR / "output-files"
ARCHIVE_DIR = BASE_DIR / "archived-files"
UNMAPPED_REPORT_DIR = BASE_DIR / "unmapped-officer-reports"
MANUAL_UNFILTERED_DIR = BASE_DIR / "drop-manually-unfilterd-files"
MANUAL_UNFILTERED_DIR_ALIASES = [
    MANUAL_UNFILTERED_DIR,
    BASE_DIR / "drop-manual-unfiltered-files",
]
MANUAL_SELECTED_DIR = BASE_DIR / "drop-manual-selected-files"
MASTER_FILE = _resolve_project_path(
    os.getenv("MASTER_FILE_PATH", "List-school-M-EE-09-04-2026.xlsx")
)
WHATSAPP_RECIPIENTS_FILE = _resolve_project_path(
    os.getenv("WHATSAPP_RECIPIENTS_FILE", "whatsapp_recipients.csv")
)
OFFICERS_LIST_FILE = _resolve_project_path(
    os.getenv("OFFICERS_LIST_FILE", "officers_list.csv")
)
POCKETBASE_DB_PATH = _resolve_project_path(
    os.getenv("POCKETBASE_DB_PATH", "../antidengue-pocketbase/pb_data/data.db")
)
RUNTIME_SNAPSHOT_PATH = (
    _resolve_project_path(os.environ["ANTIDENGUE_RUNTIME_SNAPSHOT"])
    if os.getenv("ANTIDENGUE_RUNTIME_SNAPSHOT")
    else None
)
RUNTIME_SNAPSHOT_SCHEMA_VERSION = 1
REPORT_SOURCE = os.getenv("REPORT_SOURCE", "unfiltered").strip().lower()
NATS_URL = os.getenv("NATS_URL", "nats://localhost:4222")
NATS_SUBJECT = os.getenv("NATS_SUBJECT", "whatsapp.pending")
NATS_STREAM = os.getenv("NATS_STREAM", "pending_stream")
NATS_HEALTH_SUBJECT = os.getenv("NATS_HEALTH_SUBJECT", "whatsapp.worker.health")
NATS_PUBLISH_ATTEMPTS = max(1, int(os.getenv("NATS_PUBLISH_ATTEMPTS", "3")))
NATS_PUBLISH_RETRY_DELAY_SECONDS = max(
    0.0, float(os.getenv("NATS_PUBLISH_RETRY_DELAY_SECONDS", "2"))
)
DEFAULT_SEND_DELAY_MS = int(os.getenv("WA_SEND_DELAY_MS", "1500"))
GENERATE_SCREENSHOT = os.getenv("GENERATE_SCREENSHOT", "false").strip().lower() in {
    "1",
    "true",
    "yes",
    "on",
}
DEFAULT_ATTACHMENT_TEXT_MODE = os.getenv("WA_ATTACHMENT_TEXT_MODE", "separate")
UNFILTERED_DORMANT_ACTIVITY_COLUMNS = [
    "Simple Activities",
    "Patient Activities",
    "Vector Surveillance Activities",
    "Larvae Case Response",
    "TPV Activities",
    "Total Activities",
]
TOTAL_ACTIVITY_COLUMN = "Total Activities"
MASTER_REPORT_COLUMNS = [
    "Sr. No.",
    "District",
    "Tehsil",
    "Markaz",
    "School Name",
    "School EMIS",
    "School Type",
    "DEOs Wise",
    "School Level",
]
DETAILED_ACTIVITY_COLUMNS = [
    column
    for column in UNFILTERED_DORMANT_ACTIVITY_COLUMNS
    if column != TOTAL_ACTIVITY_COLUMN
]
QUALITY_MAX_DORMANCY_RATIO = float(
    os.getenv(
        "QUALITY_MAX_DORMANCY_RATIO",
        os.getenv("QUALITY_MAX_MASTER_COVERAGE", "0.98"),
    )
)
# Compatibility for callers and old summaries that used the ambiguous name.
QUALITY_MAX_MASTER_COVERAGE = QUALITY_MAX_DORMANCY_RATIO
QUALITY_MAX_UNMAPPED_RATIO = float(os.getenv("QUALITY_MAX_UNMAPPED_RATIO", "0.10"))
QUALITY_MIN_EXPORT_COVERAGE = float(os.getenv("QUALITY_MIN_EXPORT_COVERAGE", "0.98"))
QUALITY_DORMANCY_WARNING_AFTER_HOUR = int(
    os.getenv("QUALITY_DORMANCY_WARNING_AFTER_HOUR", "8")
)
PORTAL_DUPLICATE_RAW_POLICY = (
    os.getenv("PORTAL_DUPLICATE_RAW_POLICY", "block_send")
    .strip()
    .lower()
    .replace("-", "_")
)
if PORTAL_DUPLICATE_RAW_POLICY not in {"block_send", "warn", "allow"}:
    PORTAL_DUPLICATE_RAW_POLICY = "block_send"


def _clean_optional_text(value) -> str | None:
    if value is None:
        return None

    text = str(value).strip()
    if not text or text.lower() == "nan":
        return None

    return text


def _parse_bool(value, default: bool = True) -> bool:
    if value is None:
        return default

    text = str(value).strip().lower()
    if not text:
        return default

    return text not in {"0", "false", "no", "off"}


INCLUDE_EXCEL_FOR_DDEO = _parse_bool(os.getenv("WA_INCLUDE_EXCEL_FOR_DDEO"), True)
INCLUDE_EXCEL_FOR_AEO = _parse_bool(os.getenv("WA_INCLUDE_EXCEL_FOR_AEO"), False)
WHATSAPP_REQUIRE_WORKER_READY = _parse_bool(os.getenv("WA_REQUIRE_WORKER_READY"), True)
WHATSAPP_WAIT_FOR_DELIVERY = _parse_bool(os.getenv("WA_WAIT_FOR_DELIVERY"), True)
WHATSAPP_HEALTH_TIMEOUT_SECONDS = float(os.getenv("WA_HEALTH_TIMEOUT_SECONDS", "5"))
WHATSAPP_HEALTH_WAIT_SECONDS = float(os.getenv("WA_HEALTH_WAIT_SECONDS", "120"))
WHATSAPP_DELIVERY_TIMEOUT_SECONDS = float(
    os.getenv("WA_DELIVERY_TIMEOUT_SECONDS", "900")
)
POCKETBASE_ENABLED = _parse_bool(os.getenv("POCKETBASE_ENABLED"), True)


def _read_runtime_snapshot() -> dict | None:
    if RUNTIME_SNAPSHOT_PATH is None:
        return None
    try:
        snapshot = json.loads(RUNTIME_SNAPSHOT_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise RuntimeError(
            f"Cannot read AntiDengue runtime snapshot: {RUNTIME_SNAPSHOT_PATH}"
        ) from exc
    if snapshot.get("schema_version") != RUNTIME_SNAPSHOT_SCHEMA_VERSION:
        raise RuntimeError(
            "Unsupported AntiDengue runtime snapshot schema: "
            f"{snapshot.get('schema_version')!r}"
        )
    if snapshot.get("source") != "automation-platform-postgresql":
        raise RuntimeError("AntiDengue runtime snapshot has an untrusted data source.")
    return snapshot


def _runtime_snapshot_dataframe(key: str, columns: list[str] | None = None) -> pd.DataFrame | None:
    snapshot = _read_runtime_snapshot()
    if snapshot is None:
        return None
    rows = snapshot.get(key)
    if not isinstance(rows, list):
        raise RuntimeError(f"Runtime snapshot field {key!r} must be a list.")
    return pd.DataFrame(rows, columns=columns)


def _parse_delay_ms(value) -> int | None:
    text = _clean_optional_text(value)
    if text is None:
        return None

    parsed = int(text)
    if parsed < 0:
        raise ValueError("delay_ms must be a non-negative integer")
    return parsed


def _normalize_attachment_text_mode(value, default: str = "separate") -> str:
    text = _clean_optional_text(value)
    normalized_default = (default or "separate").strip().lower()

    if normalized_default not in {"caption", "separate"}:
        normalized_default = "separate"

    if text is None:
        return normalized_default

    normalized = text.strip().lower()
    if normalized not in {"caption", "separate"}:
        raise ValueError("attachment_text_mode must be either 'caption' or 'separate'")

    return normalized


def _should_include_excel_for_roles(role_labels: list[str]) -> bool:
    normalized_roles = {role.strip().upper() for role in role_labels}

    if "DDEO" in normalized_roles:
        return INCLUDE_EXCEL_FOR_DDEO

    if "AEO" in normalized_roles:
        return INCLUDE_EXCEL_FOR_AEO

    return False


def _resolve_from_recipients_file(value: str | None) -> Path | None:
    text = _clean_optional_text(value)
    if text is None:
        return None

    candidate = Path(text)
    if not candidate.is_absolute():
        candidate = (WHATSAPP_RECIPIENTS_FILE.parent / candidate).resolve()

    return candidate


def _normalize_target(target: str, recipient_type: str | None) -> tuple[str, str]:
    cleaned_target = _clean_optional_text(target)
    if cleaned_target is None:
        raise ValueError("target is required")

    raw_type = _clean_optional_text(recipient_type)
    normalized_type = (
        raw_type.lower()
        if raw_type
        else ("group" if cleaned_target.endswith("@g.us") else "contact")
    )

    if normalized_type == "group":
        if not cleaned_target.endswith("@g.us"):
            raise ValueError(
                f"group targets must end with @g.us, received: {cleaned_target}"
            )
        return cleaned_target, normalized_type

    if normalized_type == "contact":
        if cleaned_target.endswith("@s.whatsapp.net"):
            return cleaned_target, normalized_type

        digits = re.sub(r"\D", "", cleaned_target)
        if not digits:
            raise ValueError(f"invalid contact target: {cleaned_target}")

        return f"{digits}@s.whatsapp.net", normalized_type

    raise ValueError(f"unsupported recipient type: {normalized_type}")


def _render_template(template: str | None, variables: dict[str, str]) -> str | None:
    text = _clean_optional_text(template)
    if text is None:
        return None

    return re.sub(
        r"\{\{\s*([a-zA-Z0-9_]+)\s*\}\}",
        lambda match: str(variables.get(match.group(1), "")),
        text,
    )


def _normalize_dataframe_columns(dataframe: pd.DataFrame) -> pd.DataFrame:
    dataframe = dataframe.copy()
    dataframe.columns = (
        dataframe.columns.astype(str)
        .str.strip()
        .str.lower()
        .str.replace(r"\s+", "_", regex=True)
    )
    return dataframe.loc[:, ~dataframe.columns.str.startswith("unnamed")]


def _read_dataframe_by_content(file_path: Path, *, dtype=str) -> pd.DataFrame:
    if not file_path.exists():
        raise FileNotFoundError(f"File not found: {file_path}")

    if file_path.stat().st_size == 0:
        raise ValueError(f"File is empty: {file_path}")

    with file_path.open("rb") as handle:
        signature = handle.read(8)

    if signature.startswith(b"PK"):
        return pd.read_excel(file_path, dtype=dtype, engine="openpyxl")

    if signature.startswith(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"):
        return pd.read_excel(file_path, dtype=dtype, engine="xlrd")

    csv_encodings = ("utf-8", "utf-8-sig", "cp1252", "latin1")
    last_error = None

    for encoding in csv_encodings:
        try:
            return pd.read_csv(file_path, dtype=dtype, encoding=encoding)
        except (UnicodeDecodeError, pd.errors.EmptyDataError) as exc:
            last_error = exc

    suffix = file_path.suffix.lower()
    if suffix == ".xlsx":
        return pd.read_excel(file_path, dtype=dtype, engine="openpyxl")
    if suffix == ".xls":
        return pd.read_excel(file_path, dtype=dtype, engine="xlrd")

    raise last_error or ValueError(f"Unable to read file: {file_path}")


def _read_table_dataframe(file_path: Path) -> pd.DataFrame:
    dataframe = _read_dataframe_by_content(file_path, dtype=str)
    return _normalize_dataframe_columns(dataframe)


def _file_sha256(file_path: Path) -> str:
    digest = hashlib.sha256()
    with file_path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _json_safe(value):
    if isinstance(value, Path):
        return str(value)
    if isinstance(value, datetime.datetime):
        return value.isoformat()
    if isinstance(value, (set, tuple)):
        return sorted(value)
    if isinstance(value, list):
        return [_json_safe(item) for item in value]
    if isinstance(value, dict):
        return {str(key): _json_safe(item) for key, item in value.items()}
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    if hasattr(value, "item"):
        try:
            return value.item()
        except (TypeError, ValueError):
            pass
    return value


def _write_json_file(output_path: Path, payload: dict) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        json.dumps(_json_safe(payload), indent=2, sort_keys=True),
        encoding="utf-8",
    )


def _pocketbase_record_id(prefix: str, *parts: object) -> str:
    text = ":".join(str(part) for part in parts if part is not None)
    return hashlib.sha1(f"{prefix}:{text}".encode("utf-8")).hexdigest()[:15]


def _pocketbase_json(value) -> str:
    return json.dumps(_json_safe(value), sort_keys=True, ensure_ascii=False)


def _pocketbase_table_names(conn: sqlite3.Connection) -> set[str]:
    return {
        str(row[0])
        for row in conn.execute(
            "select name from sqlite_master where type = 'table'"
        ).fetchall()
    }


def _pocketbase_table_columns(conn: sqlite3.Connection, table: str) -> set[str]:
    return {str(row[1]) for row in conn.execute(f"pragma table_info({table})").fetchall()}


def _connect_pocketbase(readonly: bool = False) -> sqlite3.Connection:
    if not POCKETBASE_DB_PATH.is_file():
        raise FileNotFoundError(f"PocketBase database not found: {POCKETBASE_DB_PATH}")

    if readonly:
        conn = sqlite3.connect(f"file:{POCKETBASE_DB_PATH}?mode=ro", uri=True)
    else:
        conn = sqlite3.connect(POCKETBASE_DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def _pocketbase_has_normalized_mapping_schema(conn: sqlite3.Connection) -> bool:
    tables = _pocketbase_table_names(conn)
    required_tables = {
        "schools",
        "ddeo_officers",
        "aeo_officers",
        "ddeo_jurisdictions",
        "aeo_jurisdictions",
        "school_ddeo_overrides",
        "school_aeo_overrides",
        "tehsils",
        "markazes",
    }
    if required_tables.difference(tables):
        return False

    required_columns = {
        "schools": {"id", "emis", "name", "wing_ref", "tehsil_ref", "markaz_ref", "active"},
        "ddeo_officers": {"id", "name", "normalized_mobile", "active"},
        "aeo_officers": {"id", "name", "normalized_mobile", "active"},
        "ddeo_jurisdictions": {"ddeo_ref", "wing_ref", "tehsil_ref", "active"},
        "aeo_jurisdictions": {"aeo_ref", "wing_ref", "markaz_ref", "active"},
        "school_ddeo_overrides": {"school_ref", "ddeo_ref", "active"},
        "school_aeo_overrides": {"school_ref", "aeo_ref", "active"},
    }
    return all(
        columns.issubset(_pocketbase_table_columns(conn, table))
        for table, columns in required_columns.items()
    )


def _pivot_pocketbase_mapping_rows(rows: list[sqlite3.Row]) -> pd.DataFrame:
    by_emis: dict[str, dict[str, str]] = {}
    for row in rows:
        emis = _clean_optional_text(row["emis"])
        if emis is None:
            continue

        item = by_emis.setdefault(
            emis,
            {
                "emis": emis,
                "school_name": _clean_optional_text(row["school_name"]) or "",
                "tehsil": _clean_optional_text(row["school_tehsil"]) or "",
                "markaz": _clean_optional_text(row["school_markaz"]) or "",
                "ddeo_name": "",
                "ddeo_cell_number": "",
                "aeo_name": "",
                "aeo_cell_number": "",
            },
        )

        role = (_clean_optional_text(row["officer_role"]) or "").strip().upper()
        officer_name = _clean_optional_text(row["officer_name"]) or ""
        officer_mobile = _clean_optional_text(row["officer_mobile"]) or ""
        if role == "DDEO":
            item["ddeo_name"] = officer_name
            item["ddeo_cell_number"] = officer_mobile
        elif role == "AEO":
            item["aeo_name"] = officer_name
            item["aeo_cell_number"] = officer_mobile

    if not by_emis:
        raise RuntimeError("PocketBase officer mapping is empty.")

    officers_df = pd.DataFrame(list(by_emis.values()))
    officers_df["emis_normalized"] = officers_df["emis"].map(_normalize_emis)
    officers_df["school_name_normalized"] = (
        officers_df["school_name"].astype(str).str.strip()
    )
    return officers_df


def _read_normalized_officers_dataframe_from_pocketbase(
    conn: sqlite3.Connection,
) -> pd.DataFrame:
    rows = conn.execute(
        """
        select
          s.emis,
          s.name as school_name,
          t.name as school_tehsil,
          m.name as school_markaz,
          'DDEO' as officer_role,
          d.name as officer_name,
          d.normalized_mobile as officer_mobile
        from schools s
        join school_ddeo_overrides o
          on o.school_ref = s.id
         and o.active = 1
        join ddeo_officers d
          on d.id = o.ddeo_ref
         and d.active = 1
        left join tehsils t
          on t.id = s.tehsil_ref
        left join markazes m
          on m.id = s.markaz_ref
        where s.active = 1

        union all

        select
          s.emis,
          s.name as school_name,
          t.name as school_tehsil,
          m.name as school_markaz,
          'DDEO' as officer_role,
          d.name as officer_name,
          d.normalized_mobile as officer_mobile
        from schools s
        join ddeo_jurisdictions j
          on j.wing_ref = s.wing_ref
         and j.tehsil_ref = s.tehsil_ref
         and j.active = 1
        join ddeo_officers d
          on d.id = j.ddeo_ref
         and d.active = 1
        left join tehsils t
          on t.id = s.tehsil_ref
        left join markazes m
          on m.id = s.markaz_ref
        where s.active = 1
          and not exists (
            select 1
            from school_ddeo_overrides o
            where o.school_ref = s.id
              and o.active = 1
          )

        union all

        select
          s.emis,
          s.name as school_name,
          t.name as school_tehsil,
          m.name as school_markaz,
          'AEO' as officer_role,
          a.name as officer_name,
          a.normalized_mobile as officer_mobile
        from schools s
        join school_aeo_overrides o
          on o.school_ref = s.id
         and o.active = 1
        join aeo_officers a
          on a.id = o.aeo_ref
         and a.active = 1
        left join tehsils t
          on t.id = s.tehsil_ref
        left join markazes m
          on m.id = s.markaz_ref
        where s.active = 1

        union all

        select
          s.emis,
          s.name as school_name,
          t.name as school_tehsil,
          m.name as school_markaz,
          'AEO' as officer_role,
          a.name as officer_name,
          a.normalized_mobile as officer_mobile
        from schools s
        join aeo_jurisdictions j
          on j.wing_ref = s.wing_ref
         and j.markaz_ref = s.markaz_ref
         and (j.tehsil_ref = '' or j.tehsil_ref = s.tehsil_ref)
         and j.active = 1
        join aeo_officers a
          on a.id = j.aeo_ref
         and a.active = 1
        left join tehsils t
          on t.id = s.tehsil_ref
        left join markazes m
          on m.id = s.markaz_ref
        where s.active = 1
          and not exists (
            select 1
            from school_aeo_overrides o
            where o.school_ref = s.id
              and o.active = 1
          )

        order by emis, officer_role
        """
    ).fetchall()
    return _pivot_pocketbase_mapping_rows(rows)


def _read_legacy_officers_dataframe_from_pocketbase(
    conn: sqlite3.Connection,
) -> pd.DataFrame:
    rows = conn.execute(
        """
        select
          s.emis,
          s.name as school_name,
          s.tehsil as school_tehsil,
          s.markaz as school_markaz,
          sa.officer_role,
          o.name as officer_name,
          o.normalized_mobile as officer_mobile
        from schools s
        left join school_assignments sa
          on sa.school_emis = s.emis and sa.active = 1
        left join officers o
          on o.normalized_mobile = sa.officer_mobile
         and o.role = sa.officer_role
         and o.active = 1
        where s.active = 1
        order by s.emis, sa.officer_role
        """
    ).fetchall()
    return _pivot_pocketbase_mapping_rows(rows)


def _read_officers_dataframe_from_pocketbase() -> pd.DataFrame:
    with _connect_pocketbase(readonly=True) as conn:
        required_tables = {"schools"}
        missing_tables = required_tables.difference(_pocketbase_table_names(conn))
        if missing_tables:
            raise RuntimeError(
                "PocketBase is missing required collections: "
                + ", ".join(sorted(missing_tables))
            )

        if _pocketbase_has_normalized_mapping_schema(conn):
            return _read_normalized_officers_dataframe_from_pocketbase(conn)
        return _read_legacy_officers_dataframe_from_pocketbase(conn)


def _school_level_from_name(name: str | None) -> str:
    normalized = (_clean_optional_text(name) or "").upper()
    prefix = normalized.split(" ", 1)[0] if normalized else ""
    if prefix == "GPS":
        return "Primary"
    if prefix == "GES":
        return "Middle"
    if prefix == "GMMS":
        return "sMosque"
    return ""


def _deos_wise_from_wing(wing_name: str | None) -> str:
    normalized = (_clean_optional_text(wing_name) or "").upper().replace(" ", "")
    if normalized in {"DEOMEE", "MEE"}:
        return "M-EE"
    return _clean_optional_text(wing_name) or ""


def _read_master_dataframe_from_pocketbase() -> pd.DataFrame:
    with _connect_pocketbase(readonly=True) as conn:
        tables = _pocketbase_table_names(conn)
        required_tables = {"schools", "districts", "tehsils", "markazes", "wings"}
        missing_tables = required_tables.difference(tables)
        if missing_tables:
            raise RuntimeError(
                "PocketBase is missing required master-data collections: "
                + ", ".join(sorted(missing_tables))
            )

        school_columns = _pocketbase_table_columns(conn, "schools")
        school_type_expr = "s.school_type" if "school_type" in school_columns else "''"
        deos_wise_expr = "s.deos_wise" if "deos_wise" in school_columns else "''"
        school_level_expr = (
            "s.school_level" if "school_level" in school_columns else "''"
        )

        rows = conn.execute(
            f"""
            select
              s.emis as school_emis,
              s.name as school_name,
              d.name as district,
              t.name as tehsil,
              m.name as markaz,
              w.name as wing,
              {school_type_expr} as school_type,
              {deos_wise_expr} as deos_wise,
              {school_level_expr} as school_level
            from schools s
            left join districts d on d.id = s.district_ref
            left join tehsils t on t.id = s.tehsil_ref
            left join markazes m on m.id = s.markaz_ref
            left join wings w on w.id = s.wing_ref
            where s.active = 1
            order by t.name, m.name, s.name, s.emis
            """
        ).fetchall()

    if not rows:
        raise RuntimeError("PocketBase master-school collection is empty.")

    records = []
    for index, row in enumerate(rows, start=1):
        school_name = _clean_optional_text(row["school_name"]) or ""
        records.append(
            {
                "Sr. No.": index,
                "District": _clean_optional_text(row["district"]) or "",
                "Tehsil": _clean_optional_text(row["tehsil"]) or "",
                "Markaz": _clean_optional_text(row["markaz"]) or "",
                "School Name": school_name,
                "School EMIS": _clean_optional_text(row["school_emis"]) or "",
                "School Type": _clean_optional_text(row["school_type"]) or "Male",
                "DEOs Wise": _clean_optional_text(row["deos_wise"])
                or _deos_wise_from_wing(row["wing"]),
                "School Level": _clean_optional_text(row["school_level"])
                or _school_level_from_name(school_name),
            }
        )

    return pd.DataFrame(records, columns=MASTER_REPORT_COLUMNS)


def _read_master_dataframe_from_file() -> pd.DataFrame:
    master_df = pd.read_excel(MASTER_FILE, dtype=str)
    master_df.columns = master_df.columns.str.strip()
    return master_df


def _read_master_dataframe() -> pd.DataFrame:
    logger = _get_logger()
    snapshot_df = _runtime_snapshot_dataframe("master_schools")
    if snapshot_df is not None:
        missing = set(MASTER_REPORT_COLUMNS).difference(snapshot_df.columns)
        if missing:
            raise RuntimeError(
                "Runtime snapshot master data is missing columns: "
                + ", ".join(sorted(missing))
            )
        logger.info(
            "Loaded master school list from platform PostgreSQL snapshot: "
            f"{len(snapshot_df)} schools."
        )
        return snapshot_df[MASTER_REPORT_COLUMNS].copy()
    return _read_master_dataframe_from_file()


def _master_school_count() -> int:
    return len(_read_master_dataframe())


def _read_officers_dataframe_from_csv() -> pd.DataFrame:
    officers_df = _read_table_dataframe(OFFICERS_LIST_FILE)
    required_columns = {
        "emis",
        "school_name",
        "ddeo_name",
        "ddeo_cell_number",
        "aeo_name",
        "aeo_cell_number",
    }
    missing_columns = sorted(required_columns.difference(officers_df.columns))

    if missing_columns:
        raise ValueError(
            f"Officers list is missing required columns: {', '.join(missing_columns)}"
        )

    officers_df["emis_normalized"] = officers_df["emis"].map(_normalize_emis)
    officers_df["school_name_normalized"] = (
        officers_df["school_name"].astype(str).str.strip()
    )

    duplicate_emis = officers_df[
        officers_df["emis_normalized"].notna()
        & officers_df["emis_normalized"].duplicated(keep=False)
    ]
    if not duplicate_emis.empty:
        sample = ", ".join(sorted(duplicate_emis["emis_normalized"].unique())[:10])
        raise ValueError(
            f"Officers list contains duplicate EMIS mappings. Sample: {sample}"
        )

    return officers_df


def _validate_runtime_config() -> None:
    snapshot = _read_runtime_snapshot()
    fallback_files_required = snapshot is None
    missing_files = []
    if fallback_files_required:
        missing_files = [
            file_path
            for file_path in [MASTER_FILE, OFFICERS_LIST_FILE, WHATSAPP_RECIPIENTS_FILE]
            if not file_path.exists()
        ]
    if missing_files:
        formatted = ", ".join(str(path) for path in missing_files)
        raise FileNotFoundError(f"Required configuration file(s) missing: {formatted}")

    if not 0 < QUALITY_MAX_DORMANCY_RATIO <= 1:
        raise ValueError("QUALITY_MAX_DORMANCY_RATIO must be > 0 and <= 1.")

    if not 0 <= QUALITY_MAX_UNMAPPED_RATIO <= 1:
        raise ValueError("QUALITY_MAX_UNMAPPED_RATIO must be between 0 and 1.")

    if not 0 < QUALITY_MIN_EXPORT_COVERAGE <= 1:
        raise ValueError("QUALITY_MIN_EXPORT_COVERAGE must be > 0 and <= 1.")

    if not 0 <= QUALITY_DORMANCY_WARNING_AFTER_HOUR <= 23:
        raise ValueError(
            "QUALITY_DORMANCY_WARNING_AFTER_HOUR must be between 0 and 23."
        )

    if REPORT_SOURCE not in {"unfiltered", "all", "filtered", "dormant"} and not (
        REPORT_SOURCE.startswith("http://") or REPORT_SOURCE.startswith("https://")
    ):
        raise ValueError(
            "REPORT_SOURCE must be 'unfiltered', 'filtered', or a full report URL."
        )


def _default_dispatch_settings() -> dict[str, bool]:
    return {
        "allow_individual": True,
        "allow_groups": True,
        "manual_only": False,
        "attach_excel": True,
        "send_failed_only": False,
        "require_preview": False,
    }


def _read_dispatch_settings_from_pocketbase() -> dict[str, bool]:
    settings = _default_dispatch_settings()
    if not POCKETBASE_ENABLED:
        return settings

    try:
        with _connect_pocketbase(readonly=True) as conn:
            if "dispatch_settings" not in _pocketbase_table_names(conn):
                return settings
            row = conn.execute(
                """
                select allow_individual, allow_groups, manual_only, attach_excel,
                       send_failed_only, require_preview
                from dispatch_settings
                where active = 1
                order by name
                limit 1
                """
            ).fetchone()
            if row is None:
                return settings
            for key in settings:
                settings[key] = _parse_bool(row[key], settings[key])
    except Exception as exc:
        _get_logger().warning(
            "Could not read dispatch settings from PocketBase; using safe defaults. "
            f"Error: {exc}"
        )
    return settings


def _read_dispatch_settings() -> dict[str, bool]:
    settings = _default_dispatch_settings()
    snapshot = _read_runtime_snapshot()
    if snapshot is None:
        return settings
    configured = snapshot.get("dispatch_settings") or {}
    if not isinstance(configured, dict):
        raise RuntimeError("Runtime snapshot dispatch_settings must be an object.")
    for key, default in settings.items():
        settings[key] = _parse_bool(configured.get(key), default)
    return settings


def _read_group_routes_dataframe_from_pocketbase(purpose: str) -> pd.DataFrame:
    columns = [
        "enabled",
        "name",
        "type",
        "target",
        "text",
        "image_path",
        "excel_path",
        "excel_filename",
        "attachment_text_mode",
        "delay_ms",
        "message_mode",
        "route_kind",
        "attach_excel",
        "manual_only",
        "tehsil_ref",
        "markaz_ref",
        "tehsil",
        "markaz",
    ]
    with _connect_pocketbase(readonly=True) as conn:
        if "whatsapp_groups" not in _pocketbase_table_names(conn):
            return pd.DataFrame(columns=columns)

        table_columns = _pocketbase_table_columns(conn, "whatsapp_groups")
        required = {"enabled", "name", "target", "route_kind", "message_mode"}
        missing = required.difference(table_columns)
        if missing:
            raise RuntimeError(
                "PocketBase whatsapp_groups is missing columns: "
                + ", ".join(sorted(missing))
            )

        purpose = purpose.strip().lower()
        if purpose == "fallback":
            purpose_clause = """
              and (
                wg.route_kind = 'fallback'
                or wg.message_mode = 'failed_fallback'
              )
            """
        else:
            purpose_clause = """
              and wg.route_kind != 'fallback'
              and wg.message_mode != 'failed_fallback'
            """

        rows = conn.execute(
            f"""
            select
              wg.enabled,
              wg.name,
              'group' as type,
              wg.target,
              '' as text,
              '' as image_path,
              '' as excel_path,
              '' as excel_filename,
              coalesce(wg.attachment_text_mode, '') as attachment_text_mode,
              coalesce(wg.delay_ms, 0) as delay_ms,
              coalesce(wg.message_mode, 'full_report') as message_mode,
              coalesce(wg.route_kind, 'district') as route_kind,
              coalesce(wg.attach_excel, 1) as attach_excel,
              coalesce(wg.manual_only, 0) as manual_only,
              coalesce(wg.tehsil_ref, '') as tehsil_ref,
              coalesce(wg.markaz_ref, '') as markaz_ref,
              coalesce(t.name, '') as tehsil,
              coalesce(m.name, '') as markaz
            from whatsapp_groups wg
            left join tehsils t
              on t.id = wg.tehsil_ref
            left join markazes m
              on m.id = wg.markaz_ref
            where wg.enabled = 1
              and coalesce(wg.manual_only, 0) = 0
              {purpose_clause}
            order by wg.route_kind, wg.name, wg.target
            """
        ).fetchall()

    return pd.DataFrame([dict(row) for row in rows], columns=columns)


def _pocketbase_has_group_routes_table() -> bool:
    if not POCKETBASE_ENABLED:
        return False
    try:
        with _connect_pocketbase(readonly=True) as conn:
            return "whatsapp_groups" in _pocketbase_table_names(conn)
    except Exception:
        return False


def _runtime_has_group_routes() -> bool:
    snapshot = _read_runtime_snapshot()
    return snapshot is not None and isinstance(snapshot.get("group_routes"), list)


def _read_group_routes_dataframe(purpose: str) -> pd.DataFrame:
    routes = _runtime_snapshot_dataframe("group_routes")
    if routes is None:
        return pd.DataFrame()
    if routes.empty:
        return routes
    fallback = (
        routes.get("route_kind", pd.Series("", index=routes.index)).eq("fallback")
        | routes.get("message_mode", pd.Series("", index=routes.index)).eq("failed_fallback")
    )
    return routes[fallback if purpose.strip().lower() == "fallback" else ~fallback].copy()


def _read_recipients_dataframe_from_pocketbase(purpose: str = "normal") -> pd.DataFrame:
    columns = [
        "enabled",
        "name",
        "type",
        "target",
        "text",
        "image_path",
        "excel_path",
        "excel_filename",
        "attachment_text_mode",
        "delay_ms",
    ]
    with _connect_pocketbase(readonly=True) as conn:
        has_group_routes = "whatsapp_groups" in _pocketbase_table_names(conn)

    if has_group_routes:
        group_routes = _read_group_routes_dataframe_from_pocketbase(purpose)
        return group_routes

    with _connect_pocketbase(readonly=True) as conn:
        if "whatsapp_recipients" not in _pocketbase_table_names(conn):
            raise RuntimeError("PocketBase is missing whatsapp_recipients collection.")

        table_columns = _pocketbase_table_columns(conn, "whatsapp_recipients")
        missing_columns = {"enabled", "name", "type", "target"}.difference(
            table_columns
        )
        if missing_columns:
            raise RuntimeError(
                "PocketBase whatsapp_recipients is missing columns: "
                + ", ".join(sorted(missing_columns))
            )

        selectable_columns = [
            column if column in table_columns else f"'' as {column}"
            for column in columns
        ]
        rows = conn.execute(
            f"""
            select {", ".join(selectable_columns)}
            from whatsapp_recipients
            order by enabled desc, name, target
            """
        ).fetchall()

    return pd.DataFrame([dict(row) for row in rows], columns=columns)


def _read_recipients_dataframe(purpose: str = "normal") -> pd.DataFrame:
    logger = _get_logger()
    routes_df = _runtime_snapshot_dataframe("group_routes")
    if routes_df is not None:
        recipients_df = routes_df
        logger.info(
            "Loaded WhatsApp routes from platform PostgreSQL snapshot: "
            f"{len(recipients_df)} target(s)."
        )
    else:
        recipients_df = _read_table_dataframe(WHATSAPP_RECIPIENTS_FILE)

    try:
        if "target" not in recipients_df.columns:
            raise KeyError("The recipients file must include a 'target' column.")
    except Exception as exc:
        raise ValueError(
            f"Invalid fixed WhatsApp recipients configuration: {exc}"
        ) from exc

    return recipients_df


def _normalize_emis(value) -> str | None:
    text = _clean_optional_text(value)
    if text is None:
        return None

    digits = re.sub(r"\D", "", text)
    return digits or None


def _normalize_pk_mobile(value) -> str:
    text = _clean_optional_text(value)
    if text is None:
        raise ValueError("mobile number is required")

    cleaned = text.replace("'", "").replace('"', "").replace(" ", "")
    cleaned = cleaned.replace("-", "").replace("(", "").replace(")", "")
    cleaned = cleaned.replace("+", "")

    if re.search(r"[eE]", cleaned):
        try:
            cleaned = str(Decimal(cleaned).quantize(Decimal("1")))
        except InvalidOperation as exc:
            raise ValueError(f"invalid scientific mobile number: {text}") from exc

    digits = re.sub(r"\D", "", cleaned)

    if digits.startswith("0092"):
        digits = digits[2:]
    elif digits.startswith("92") and len(digits) == 12:
        pass
    elif digits.startswith("03") and len(digits) == 11:
        digits = f"92{digits[1:]}"
    elif digits.startswith("3") and len(digits) == 10:
        digits = f"92{digits}"
    else:
        raise ValueError(f"unsupported Pakistani mobile format: {text}")

    if not re.fullmatch(r"92[0-9]{10}", digits):
        raise ValueError(f"invalid Pakistani mobile number: {text}")

    if digits[2] != "3":
        raise ValueError(
            f"Pakistani mobile number must start with 3 after country code: {text}"
        )

    return digits


def _create_job_payload(
    *,
    target: str,
    recipient_type: str,
    recipient_name: str | None,
    attachment_text_mode: str,
    text: str | None,
    excel_path: Path | None,
    image_path: Path | None,
    excel_filename: str | None,
    delay_ms: int,
    status_subject: str | None = None,
) -> dict:
    if excel_path and not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    if image_path and not image_path.exists():
        raise FileNotFoundError(f"Image file not found: {image_path}")

    if not text and excel_path is None and image_path is None:
        raise ValueError("Each recipient needs text, excel_path, or image_path.")

    return {
        "job_id": str(uuid.uuid4()),
        "target": target,
        "type": recipient_type,
        "recipient_name": recipient_name,
        "attachment_text_mode": attachment_text_mode,
        "text": text,
        "excel_path": str(excel_path) if excel_path else None,
        "excel_filename": excel_filename,
        "image_path": str(image_path) if image_path else None,
        "delay_ms": delay_ms,
        "status_subject": status_subject,
    }


def build_fixed_whatsapp_payloads(
    excel_path: Path,
    image_path: Path | None,
    message_body: str,
    *,
    purpose: str = "normal",
) -> list[dict]:
    recipients_df = _read_recipients_dataframe(purpose)
    payloads: list[dict] = []
    errors: list[str] = []

    generated_excel_path = excel_path.resolve()
    generated_image_path = (
        image_path.resolve() if image_path and image_path.exists() else None
    )

    for row_number, row in enumerate(recipients_df.to_dict(orient="records"), start=2):
        try:
            if not _parse_bool(row.get("enabled"), default=True):
                continue
            if _parse_bool(row.get("manual_only"), default=False):
                continue

            target, recipient_type = _normalize_target(
                row.get("target"), row.get("type")
            )
            recipient_name = _clean_optional_text(row.get("name"))
            variables = {
                "name": recipient_name or "",
                "target": target,
                "type": recipient_type,
            }

            text = _render_template(row.get("text"), variables) or message_body
            excel_override = _resolve_from_recipients_file(row.get("excel_path"))
            image_override = _resolve_from_recipients_file(row.get("image_path"))
            attach_excel = _parse_bool(row.get("attach_excel"), default=True)

            final_excel_path = (excel_override or generated_excel_path) if attach_excel else None
            final_image_path = image_override or generated_image_path

            if final_excel_path and not final_excel_path.exists():
                raise FileNotFoundError(f"Excel file not found: {final_excel_path}")

            if image_override and not image_override.exists():
                raise FileNotFoundError(f"Image file not found: {image_override}")

            excel_filename = (
                _clean_optional_text(row.get("excel_filename")) or final_excel_path.name
                if final_excel_path
                else None
            )
            delay_ms = _parse_delay_ms(row.get("delay_ms")) or DEFAULT_SEND_DELAY_MS
            attachment_text_mode = _normalize_attachment_text_mode(
                row.get("attachment_text_mode"),
                default=DEFAULT_ATTACHMENT_TEXT_MODE,
            )
            payloads.append(
                _create_job_payload(
                    target=target,
                    recipient_type=recipient_type,
                    recipient_name=recipient_name,
                    attachment_text_mode=attachment_text_mode,
                    text=text,
                    excel_path=final_excel_path,
                    image_path=final_image_path,
                    excel_filename=excel_filename,
                    delay_ms=delay_ms,
                )
            )
        except Exception as exc:
            errors.append(
                f"Row {row_number}: {row.get('target', '(missing target)')} -> {exc}"
            )

    if errors:
        raise ValueError("Recipient file validation failed:\n" + "\n".join(errors))

    return payloads


def _read_officers_dataframe() -> pd.DataFrame:
    logger = _get_logger()
    officers_df = _runtime_snapshot_dataframe("officer_mappings")
    if officers_df is not None:
        officers_df["emis_normalized"] = officers_df["emis"].map(_normalize_emis)
        officers_df["school_name_normalized"] = officers_df["school_name"].astype(str).str.strip()
        logger.info(
            "Loaded officer mapping from platform PostgreSQL snapshot: "
            f"{len(officers_df)} schools."
        )
        return officers_df
    return _read_officers_dataframe_from_csv()


def _build_officer_message(
    *,
    title_text: str,
    role_labels: list[str],
    recipient_name: str,
    schools: list[dict[str, str]],
) -> str:
    normalized_roles = {role.strip().upper() for role in role_labels}
    show_tehsil_summary = "DDEO" in normalized_roles
    tehsil_counts = Counter()
    markaz_counts = Counter()
    school_lines: list[str] = []

    for index, school in enumerate(schools, start=1):
        school_name = school["school_name"]
        tehsil = school.get("tehsil")
        markaz = school.get("markaz")

        if tehsil:
            tehsil_counts[tehsil] += 1

        if markaz:
            markaz_counts[markaz] += 1

        location_parts = []
        if tehsil:
            location_parts.append(f"Tehsil: {tehsil}")
        if markaz:
            location_parts.append(f"Markaz: {markaz}")

        location_suffix = f" ({', '.join(location_parts)})" if location_parts else ""
        school_lines.append(f"{index}. {school_name}{location_suffix}")

    intro_lines = [
        f"🚨 *{title_text}*",
        "",
        f"*For:* {recipient_name}",
        f"*Role:* {', '.join(role_labels)}",
        f"*Dormant schools mapped to you in this run:* {len(schools)}",
    ]

    if show_tehsil_summary and tehsil_counts:
        intro_lines.extend(
            [
                "",
                "📍 *Tehsil Summary*",
                *[
                    f"{index}. {tehsil}: {count}"
                    for index, (tehsil, count) in enumerate(
                        sorted(tehsil_counts.items()), start=1
                    )
                ],
            ]
        )

    if markaz_counts:
        intro_lines.extend(
            [
                "",
                "🏫 *Markaz Summary*",
                *[
                    f"{index}. {markaz}: {count}"
                    for index, (markaz, count) in enumerate(
                        sorted(markaz_counts.items()), start=1
                    )
                ],
            ]
        )

    intro_lines.extend(
        [
            "",
            "📋 *Schools*",
            *(school_lines or ["1. Not available"]),
        ]
    )

    return "\n".join(intro_lines)


def _dormant_dispatch_rows(final_df: pd.DataFrame) -> list[dict[str, str]]:
    school_emis_column = _find_column(final_df, "School EMIS")
    school_name_column = _find_column(final_df, "School Name")
    tehsil_column = _find_column(final_df, "Tehsil")
    markaz_column = _find_column(final_df, "Markaz")
    if school_emis_column is None or school_name_column is None:
        raise ValueError(
            "Processed report is missing 'School EMIS' or 'School Name' columns."
        )

    dormant_df = final_df.copy()
    dormant_df["emis_normalized"] = dormant_df[school_emis_column].map(_normalize_emis)
    officers_df = _read_officers_dataframe()
    merged_df = dormant_df.merge(
        officers_df,
        on="emis_normalized",
        how="left",
        suffixes=("_dormant", "_officer"),
    )

    rows: list[dict[str, str]] = []
    for row in merged_df.to_dict(orient="records"):
        rows.append(
            {
                "emis": _clean_optional_text(row.get(school_emis_column)) or "",
                "school_name": _clean_optional_text(row.get(school_name_column))
                or "Unknown School",
                "tehsil": (
                    _clean_optional_text(row.get(tehsil_column))
                    if tehsil_column
                    else None
                )
                or _clean_optional_text(row.get("tehsil"))
                or "",
                "markaz": (
                    _clean_optional_text(row.get(markaz_column))
                    if markaz_column
                    else None
                )
                or _clean_optional_text(row.get("markaz"))
                or "",
                "ddeo_name": _clean_optional_text(row.get("ddeo_name")) or "",
                "ddeo_mobile": _clean_optional_text(row.get("ddeo_cell_number")) or "",
                "aeo_name": _clean_optional_text(row.get("aeo_name")) or "",
                "aeo_mobile": _clean_optional_text(row.get("aeo_cell_number")) or "",
            }
        )
    return sorted(
        rows,
        key=lambda item: (item["tehsil"], item["markaz"], item["school_name"]),
    )


def _route_scoped_rows(route: dict, rows: list[dict[str, str]]) -> list[dict[str, str]]:
    route_kind = (_clean_optional_text(route.get("route_kind")) or "district").lower()
    tehsil = _clean_optional_text(route.get("tehsil"))
    markaz = _clean_optional_text(route.get("markaz"))
    if route_kind == "tehsil" and tehsil is None:
        return []
    if route_kind == "markaz" and markaz is None:
        return []

    scoped = rows
    if tehsil:
        scoped = [
            row
            for row in scoped
            if row.get("tehsil", "").strip().upper() == tehsil.strip().upper()
        ]
    if markaz:
        scoped = [
            row
            for row in scoped
            if row.get("markaz", "").strip().upper() == markaz.strip().upper()
        ]
    return scoped


def _route_scope_label(route: dict) -> str:
    scope_parts = [
        _clean_optional_text(route.get("tehsil")),
        _clean_optional_text(route.get("markaz")),
    ]
    return " / ".join(part for part in scope_parts if part) or str(
        route.get("route_kind") or "district"
    ).title()


def _safe_filename_part(value: object, *, fallback: str = "route") -> str:
    text = _clean_optional_text(value) or fallback
    text = re.sub(r"[^A-Za-z0-9._ -]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip(" ._-")
    return (text or fallback)[:90]


def _scoped_dataframe_for_dispatch_rows(
    final_df: pd.DataFrame,
    scoped_rows: list[dict[str, str]],
) -> pd.DataFrame:
    if not scoped_rows:
        return final_df.iloc[0:0].copy()

    school_emis_column = _find_column(final_df, "School EMIS")
    if school_emis_column is None:
        return final_df.iloc[0:0].copy()

    ordered_emis = [
        _normalize_emis(row.get("emis"))
        for row in scoped_rows
        if _normalize_emis(row.get("emis"))
    ]
    if not ordered_emis:
        return final_df.iloc[0:0].copy()

    order_map = {emis: index for index, emis in enumerate(ordered_emis)}
    scoped_df = final_df.copy()
    scoped_df["_dispatch_order"] = scoped_df[school_emis_column].map(
        lambda value: order_map.get(_normalize_emis(value))
    )
    scoped_df = scoped_df[scoped_df["_dispatch_order"].notna()].copy()
    scoped_df = scoped_df.sort_values("_dispatch_order").drop(
        columns=["_dispatch_order"]
    )
    scoped_df = scoped_df.reset_index(drop=True)

    sr_col = next(
        (
            column
            for column in scoped_df.columns
            if "sr" in column.strip().lower()
            or "serial" in column.strip().lower()
        ),
        None,
    )
    if sr_col:
        scoped_df[sr_col] = range(1, len(scoped_df) + 1)
    return scoped_df


def _build_route_excel_path(
    *,
    route: dict,
    route_index: int,
    base_excel_path: Path,
) -> Path:
    route_name = _safe_filename_part(route.get("name"), fallback="group")
    route_kind = _safe_filename_part(route.get("route_kind"), fallback="route")
    route_scope = _safe_filename_part(_route_scope_label(route), fallback="scope")
    filename = f"{route_index:02d} - {route_name} - {route_kind} - {route_scope}.xlsx"
    return base_excel_path.parent / "group-route-excels" / filename


def _build_route_excel_title(
    *,
    title_text: str,
    route: dict,
    row_count: int,
) -> str:
    return (
        f"{title_text} | {route.get('name') or 'Group'} | "
        f"{_route_scope_label(route)} | {_school_count_text(row_count)}"
    )


def _school_count_text(count: int) -> str:
    return f"{count} {'school' if count == 1 else 'schools'}"


def _dispatch_school_line(index: int, row: dict[str, str]) -> str:
    emis = row.get("emis", "").strip()
    emis_prefix = f"{emis} - " if emis else ""
    return f"{index}. {emis_prefix}{row.get('school_name') or 'Unknown School'}"


def _dispatch_tehsil_summary_lines(rows: list[dict[str, str]]) -> list[str]:
    tehsil_counts: Counter[str] = Counter(row.get("tehsil") or "Unmapped" for row in rows)
    if not tehsil_counts:
        return []
    return [
        "",
        "*TEHSIL SUMMARY*",
        *[
            f"{index}. {name}: {_school_count_text(count)}"
            for index, (name, count) in enumerate(sorted(tehsil_counts.items()), start=1)
        ],
    ]


def _dispatch_markaz_summary_lines(rows: list[dict[str, str]]) -> list[str]:
    if not rows:
        return []

    grouped_by_tehsil: dict[str, Counter[str]] = {}
    for row in rows:
        tehsil = row.get("tehsil") or "Unmapped"
        markaz = row.get("markaz") or "Unmapped"
        grouped_by_tehsil.setdefault(tehsil, Counter())[markaz] += 1

    lines = ["", "*MARKAZ SUMMARY*"]
    for tehsil, markaz_counts in sorted(grouped_by_tehsil.items()):
        tehsil_total = sum(markaz_counts.values())
        lines.extend(["", f"*{tehsil} - {_school_count_text(tehsil_total)}*"])
        lines.extend(
            f"{index}. {markaz}: {_school_count_text(count)}"
            for index, (markaz, count) in enumerate(
                sorted(markaz_counts.items()), start=1
            )
        )
    return lines


def _dispatch_hierarchy_sections(
    rows: list[dict[str, str]],
    *,
    include_heading: bool = True,
) -> list[str]:
    if not rows:
        return ["", "No dormant schools in this group's current scope."]

    grouped_by_tehsil: dict[str, list[dict[str, str]]] = {}
    for row in rows:
        grouped_by_tehsil.setdefault(row.get("tehsil") or "Unmapped", []).append(row)

    lines: list[str] = []
    if include_heading:
        lines.extend(["", "*DETAILS BY TEHSIL / MARKAZ*"])

    for tehsil, tehsil_rows in sorted(grouped_by_tehsil.items()):
        sorted_tehsil_rows = sorted(
            tehsil_rows,
            key=lambda item: (item.get("markaz", ""), item.get("school_name", "")),
        )
        lines.extend(["", f"*{tehsil} - {_school_count_text(len(sorted_tehsil_rows))}*"])

        grouped_by_markaz: dict[str, list[dict[str, str]]] = {}
        for row in sorted_tehsil_rows:
            grouped_by_markaz.setdefault(row.get("markaz") or "Unmapped", []).append(row)

        for markaz, markaz_rows in sorted(grouped_by_markaz.items()):
            sorted_markaz_rows = sorted(
                markaz_rows,
                key=lambda item: item.get("school_name", ""),
            )
            lines.extend(
                [
                    "",
                    f"*{markaz} - {_school_count_text(len(sorted_markaz_rows))}*",
                ]
            )
            lines.extend(
                _dispatch_school_line(index, school)
                for index, school in enumerate(sorted_markaz_rows, start=1)
            )

    return lines


def _dispatch_markaz_sections(rows: list[dict[str, str]]) -> list[str]:
    if not rows:
        return ["", "No dormant schools in this group's current scope."]

    grouped_by_markaz: dict[str, list[dict[str, str]]] = {}
    for row in rows:
        grouped_by_markaz.setdefault(row.get("markaz") or "Unmapped", []).append(row)

    lines: list[str] = []
    for markaz, markaz_rows in sorted(grouped_by_markaz.items()):
        sorted_markaz_rows = sorted(
            markaz_rows,
            key=lambda item: item.get("school_name", ""),
        )
        lines.extend(["", f"*{markaz} - {_school_count_text(len(sorted_markaz_rows))}*"])
        lines.extend(
            _dispatch_school_line(index, row)
            for index, row in enumerate(sorted_markaz_rows, start=1)
        )
    return lines


def _dispatch_grouped_sections(
    rows: list[dict[str, str]],
    *,
    name_key: str,
    mobile_key: str,
) -> list[str]:
    grouped: dict[tuple[str, str], list[dict[str, str]]] = {}
    for row in rows:
        key = (
            row.get(name_key) or "Unmapped",
            row.get(mobile_key) or "",
        )
        grouped.setdefault(key, []).append(row)

    lines: list[str] = []
    for (name, mobile), group_rows in sorted(grouped.items()):
        heading = name if not mobile else f"{name} ({mobile})"
        lines.extend(["", f"*{heading} - {_school_count_text(len(group_rows))}*"])
        lines.extend(_dispatch_hierarchy_sections(group_rows, include_heading=False))
    return lines


def _dispatch_summary_only_sections(
    rows: list[dict[str, str]],
    *,
    include_tehsil: bool,
    include_markaz: bool,
) -> list[str]:
    if not rows:
        return ["", "No dormant schools in this group's current scope."]

    lines: list[str] = []
    if include_tehsil:
        lines.extend(_dispatch_tehsil_summary_lines(rows))
    if include_markaz:
        lines.extend(_dispatch_markaz_summary_lines(rows))
    return lines


def _build_group_route_message(
    route: dict,
    rows: list[dict[str, str]],
    *,
    title_text: str,
    default_message: str,
) -> str:
    message_mode = str(route.get("message_mode") or "full_report")
    route_name = _clean_optional_text(route.get("name")) or _clean_optional_text(route.get("target")) or "Group"
    scope = _route_scope_label(route)

    if message_mode == "failed_fallback":
        return default_message

    lines = [
        f"*{title_text}*",
        f"*Group:* {route_name}",
        f"*Scope:* {scope}",
        f"*Total dormant:* {_school_count_text(len(rows))}",
    ]

    if message_mode == "ddeo_wise":
        lines.extend(_dispatch_grouped_sections(rows, name_key="ddeo_name", mobile_key="ddeo_mobile"))
    elif message_mode == "aeo_wise":
        lines.extend(_dispatch_grouped_sections(rows, name_key="aeo_name", mobile_key="aeo_mobile"))
    elif message_mode == "tehsil_summary":
        lines.extend(
            _dispatch_summary_only_sections(
                rows, include_tehsil=True, include_markaz=False
            )
        )
    elif message_mode == "markaz_summary":
        lines.extend(
            _dispatch_summary_only_sections(
                rows, include_tehsil=False, include_markaz=True
            )
        )
    elif message_mode == "tehsil_markaz_summary":
        lines.extend(
            _dispatch_summary_only_sections(
                rows, include_tehsil=True, include_markaz=True
            )
        )
    else:
        lines.extend(_dispatch_tehsil_summary_lines(rows))
        lines.extend(_dispatch_hierarchy_sections(rows))

    lines.extend(
        [
            "",
            "Please ensure activities are submitted/updated on the portal today.",
        ]
    )
    return "\n".join(lines)


def build_group_route_payloads(
    final_df: pd.DataFrame,
    excel_path: Path,
    image_path: Path | None,
    title_text: str,
    message_body: str,
    *,
    purpose: str = "normal",
) -> list[dict]:
    if not _runtime_has_group_routes():
        return build_fixed_whatsapp_payloads(
            excel_path,
            image_path,
            message_body,
            purpose=purpose,
        )

    routes_df = _read_group_routes_dataframe(purpose)
    if routes_df.empty:
        return []

    dormant_rows = _dormant_dispatch_rows(final_df)
    payloads: list[dict] = []
    errors: list[str] = []
    generated_excel_path = excel_path.resolve()
    generated_image_path = (
        image_path.resolve() if image_path and image_path.exists() else None
    )

    for row_number, route in enumerate(routes_df.to_dict(orient="records"), start=1):
        try:
            target, recipient_type = _normalize_target(
                route.get("target"),
                route.get("type") or "group",
            )
            scoped_rows = _route_scoped_rows(route, dormant_rows)
            if purpose != "fallback" and not scoped_rows:
                continue

            text = _render_template(route.get("text"), {"name": route.get("name") or ""})
            if not text:
                text = _build_group_route_message(
                    route,
                    scoped_rows,
                    title_text=title_text,
                    default_message=message_body,
                )

            attach_excel = _parse_bool(route.get("attach_excel"), default=True)
            final_excel_path = None
            excel_filename = None
            if attach_excel:
                if purpose == "fallback":
                    final_excel_path = generated_excel_path
                    excel_filename = generated_excel_path.name
                else:
                    scoped_df = _scoped_dataframe_for_dispatch_rows(
                        final_df, scoped_rows
                    )
                    final_excel_path = _build_route_excel_path(
                        route=route,
                        route_index=row_number,
                        base_excel_path=generated_excel_path,
                    )
                    load_excel_report(
                        scoped_df,
                        final_excel_path,
                        _build_route_excel_title(
                            title_text=title_text,
                            route=route,
                            row_count=len(scoped_df),
                        ),
                    )
                    excel_filename = final_excel_path.name
            delay_ms = _parse_delay_ms(route.get("delay_ms")) or DEFAULT_SEND_DELAY_MS
            attachment_text_mode = _normalize_attachment_text_mode(
                route.get("attachment_text_mode"),
                default=DEFAULT_ATTACHMENT_TEXT_MODE,
            )

            payload = _create_job_payload(
                target=target,
                recipient_type=recipient_type,
                recipient_name=_clean_optional_text(route.get("name")),
                attachment_text_mode=attachment_text_mode,
                text=text,
                excel_path=final_excel_path,
                image_path=generated_image_path,
                excel_filename=excel_filename,
                delay_ms=delay_ms,
            )
            payload["dispatch_route"] = {
                "route_kind": route.get("route_kind"),
                "message_mode": route.get("message_mode"),
                "tehsil": route.get("tehsil"),
                "markaz": route.get("markaz"),
                "row_count": len(scoped_rows),
            }
            payloads.append(payload)
        except Exception as exc:
            errors.append(
                f"Group route {row_number}: {route.get('name') or route.get('target') or '(missing target)'} -> {exc}"
            )

    if errors:
        raise ValueError("Group dispatch route validation failed:\n" + "\n".join(errors))
    return payloads


def _compact_audit_list(values: list[str]) -> str:
    return "\n".join(sorted({value for value in values if value}))


def build_dynamic_officer_payloads_with_audit(
    final_df: pd.DataFrame,
    excel_path: Path,
    image_path: Path | None,
    title_text: str,
) -> tuple[list[dict], list[dict]]:
    officers_df = _read_officers_dataframe()
    dormant_df = final_df.copy()

    school_emis_column = next(
        (
            column
            for column in dormant_df.columns
            if column.strip().lower() == "school emis"
        ),
        None,
    )
    school_name_column = next(
        (
            column
            for column in dormant_df.columns
            if column.strip().lower() == "school name"
        ),
        None,
    )

    if school_emis_column is None or school_name_column is None:
        raise ValueError(
            "Processed report is missing 'School EMIS' or 'School Name' columns."
        )

    dormant_df["emis_normalized"] = dormant_df[school_emis_column].map(_normalize_emis)
    merged_df = dormant_df.merge(
        officers_df,
        on="emis_normalized",
        how="left",
        suffixes=("_dormant", "_officer"),
    )

    contacts_by_phone: dict[str, dict] = {}
    unmatched_emis: set[str] = set()
    skipped_contact_rows: list[str] = []

    for row in merged_df.to_dict(orient="records"):
        emis_value = row.get("emis_normalized")
        school_name = (
            _clean_optional_text(row.get(school_name_column)) or "Unknown School"
        )

        if (
            _clean_optional_text(row.get("ddeo_name")) is None
            and _clean_optional_text(row.get("aeo_name")) is None
        ):
            if emis_value:
                unmatched_emis.add(emis_value)
            continue

        officer_specs = [
            ("DDEO", row.get("ddeo_name"), row.get("ddeo_cell_number")),
            ("AEO", row.get("aeo_name"), row.get("aeo_cell_number")),
        ]

        for role_label, officer_name, officer_mobile in officer_specs:
            if (
                _clean_optional_text(officer_name) is None
                or _clean_optional_text(officer_mobile) is None
            ):
                continue

            try:
                normalized_mobile = _normalize_pk_mobile(officer_mobile)
            except Exception as exc:
                skipped_contact_rows.append(
                    f"{role_label} {officer_name} / {officer_mobile} / EMIS {emis_value}: {exc}"
                )
                continue

            contact_entry = contacts_by_phone.setdefault(
                normalized_mobile,
                {
                    "names": set(),
                    "roles": set(),
                    "schools": [],
                    "school_keys_seen": set(),
                },
            )

            contact_entry["names"].add(_clean_optional_text(officer_name))
            contact_entry["roles"].add(role_label)

            school_key = (
                school_name,
                _clean_optional_text(row.get("tehsil")) or "",
                _clean_optional_text(row.get("markaz")) or "",
            )

            if school_key not in contact_entry["school_keys_seen"]:
                contact_entry["school_keys_seen"].add(school_key)
                contact_entry["schools"].append(
                    {
                        "school_name": school_name,
                        "tehsil": _clean_optional_text(row.get("tehsil")) or "",
                        "markaz": _clean_optional_text(row.get("markaz")) or "",
                    }
                )

    payloads: list[dict] = []
    audit_rows: list[dict] = []

    for normalized_mobile, details in sorted(contacts_by_phone.items()):
        recipient_name = ", ".join(sorted(name for name in details["names"] if name))
        role_labels = sorted(details["roles"])
        include_excel = _should_include_excel_for_roles(role_labels)
        sorted_schools = sorted(
            details["schools"],
            key=lambda school: (
                school["tehsil"],
                school["markaz"],
                school["school_name"],
            ),
        )
        officer_message = _build_officer_message(
            title_text=title_text,
            role_labels=role_labels,
            recipient_name=recipient_name or "Officer",
            schools=sorted_schools,
        )

        payload = _create_job_payload(
            target=normalized_mobile,
            recipient_type="contact",
            recipient_name=recipient_name or None,
            attachment_text_mode=_normalize_attachment_text_mode(
                DEFAULT_ATTACHMENT_TEXT_MODE
            ),
            text=officer_message,
            excel_path=excel_path.resolve() if include_excel else None,
            image_path=(
                image_path.resolve() if image_path and image_path.exists() else None
            ),
            excel_filename=excel_path.name,
            delay_ms=DEFAULT_SEND_DELAY_MS,
        )
        payloads.append(payload)
        audit_rows.append(
            {
                "Job ID": payload["job_id"],
                "Target": f"{normalized_mobile}@s.whatsapp.net",
                "Mobile Number": normalized_mobile,
                "Officer Name": recipient_name or "Officer",
                "Roles": ", ".join(role_labels),
                "School Count": len(sorted_schools),
                "Tehsils": _compact_audit_list(
                    [school["tehsil"] for school in sorted_schools]
                ),
                "Markazes": _compact_audit_list(
                    [school["markaz"] for school in sorted_schools]
                ),
                "Schools": _compact_audit_list(
                    [school["school_name"] for school in sorted_schools]
                ),
                "Includes Excel": include_excel,
                "Delivery Status": "not queued",
                "WhatsApp Message IDs": "",
                "WhatsApp Ack Statuses": "",
                "WhatsApp Remote JIDs": "",
                "WhatsApp Requested Target": "",
                "WhatsApp Actual Send Target": "",
                "WhatsApp Canonical Target": "",
                "WhatsApp Known LID": "",
                "WhatsApp Target Policy": "",
                "WhatsApp Token Preflight": "",
                "Delivery Error": "",
            }
        )

    if unmatched_emis:
        sample_unmatched = ", ".join(sorted(unmatched_emis)[:10])
        print(
            f"Warning: {len(unmatched_emis)} dormant school EMIS values had no officer mapping. Sample: {sample_unmatched}"
        )

    if skipped_contact_rows:
        sample_skips = " | ".join(skipped_contact_rows[:5])
        print(
            f"Warning: skipped {len(skipped_contact_rows)} officer contact rows due to invalid mobile data. Sample: {sample_skips}"
        )

    return payloads, audit_rows


def build_dynamic_officer_payloads(
    final_df: pd.DataFrame,
    excel_path: Path,
    image_path: Path | None,
    title_text: str,
) -> list[dict]:
    payloads, _ = build_dynamic_officer_payloads_with_audit(
        final_df, excel_path, image_path, title_text
    )
    return payloads


def _find_column(dataframe: pd.DataFrame, column_name: str) -> str | None:
    normalized_name = column_name.strip().lower()
    return next(
        (
            column
            for column in dataframe.columns
            if str(column).strip().lower() == normalized_name
        ),
        None,
    )


def _resolve_columns_by_name(
    dataframe: pd.DataFrame, expected_columns: list[str]
) -> dict[str, str | None]:
    columns_by_name = {
        str(column).strip().lower(): column for column in dataframe.columns
    }
    return {
        expected_column: columns_by_name.get(expected_column.strip().lower())
        for expected_column in expected_columns
    }


def _coerce_activity_values(
    dataframe: pd.DataFrame, columns: list[str]
) -> pd.DataFrame:
    return dataframe[columns].apply(
        lambda column: pd.to_numeric(
            column.astype(str).str.strip().replace("", pd.NA),
            errors="coerce",
        )
    )


def _filter_dormant_report_rows(
    user_df: pd.DataFrame, logger=None
) -> tuple[pd.DataFrame, dict]:
    resolved_columns = _resolve_columns_by_name(
        user_df, UNFILTERED_DORMANT_ACTIVITY_COLUMNS
    )
    total_column = resolved_columns[TOTAL_ACTIVITY_COLUMN]
    detailed_columns = [
        resolved_columns[column]
        for column in DETAILED_ACTIVITY_COLUMNS
        if resolved_columns[column] is not None
    ]
    missing_columns = [
        column for column, resolved in resolved_columns.items() if resolved is None
    ]
    diagnostics = {
        "input_rows": len(user_df),
        "output_rows": len(user_df),
        "filter_applied": False,
        "filter_strategy": "already_filtered",
        "available_activity_columns": [
            column
            for column, resolved in resolved_columns.items()
            if resolved is not None
        ],
        "missing_activity_columns": missing_columns,
    }

    if total_column is not None:
        total_values = pd.to_numeric(
            user_df[total_column].astype(str).str.strip().replace("", pd.NA),
            errors="coerce",
        )
        diagnostics["invalid_total_activity_rows"] = int(total_values.isna().sum())
        filtered_df = user_df[total_values.eq(0)].copy()
        diagnostics.update(
            {
                "output_rows": len(filtered_df),
                "filter_applied": True,
                "filter_strategy": TOTAL_ACTIVITY_COLUMN,
            }
        )

        if detailed_columns:
            detailed_values = _coerce_activity_values(user_df, detailed_columns)
            detailed_zero_count = int(detailed_values.eq(0).all(axis=1).sum())
            diagnostics["zero_rows_across_available_detail_columns"] = (
                detailed_zero_count
            )
            diagnostics["dormant_activity_conflict_rows"] = int(
                (total_values.eq(0) & detailed_values.gt(0).any(axis=1)).sum()
            )

        if logger and missing_columns:
            logger.info(
                "Optional activity detail columns are unavailable: "
                f"{', '.join(missing_columns)}. Used authoritative "
                f"'{TOTAL_ACTIVITY_COLUMN}' to filter dormant rows."
            )

        return filtered_df, diagnostics

    if detailed_columns:
        detailed_values = _coerce_activity_values(user_df, detailed_columns)
        filtered_df = user_df[detailed_values.eq(0).all(axis=1)].copy()
        diagnostics.update(
            {
                "output_rows": len(filtered_df),
                "filter_applied": True,
                "filter_strategy": "available_detail_activity_columns",
            }
        )

        if logger:
            logger.warning(
                f"'{TOTAL_ACTIVITY_COLUMN}' column not found. Filtered dormant rows "
                "using available detailed activity columns: "
                f"{', '.join(str(column) for column in detailed_columns)}. "
                f"Missing columns: {', '.join(missing_columns)}"
            )

        return filtered_df, diagnostics

    if logger:
        logger.warning(
            "No activity columns found; treating report as already filtered. "
            "Use REPORT_SOURCE=filtered only for this shape. Missing columns: "
            f"{', '.join(missing_columns)}"
        )

    return user_df.copy(), diagnostics


def _build_officer_mapping_audit(
    final_df: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    officers_df = _read_officers_dataframe()
    dormant_df = final_df.copy()

    school_emis_column = _find_column(dormant_df, "School EMIS")
    school_name_column = _find_column(dormant_df, "School Name")
    tehsil_column = _find_column(dormant_df, "Tehsil")
    markaz_column = _find_column(dormant_df, "Markaz")

    if school_emis_column is None or school_name_column is None:
        raise ValueError(
            "Processed report is missing 'School EMIS' or 'School Name' columns."
        )

    dormant_df["emis_normalized"] = dormant_df[school_emis_column].map(_normalize_emis)
    merged_df = dormant_df.merge(
        officers_df,
        on="emis_normalized",
        how="left",
        suffixes=("_dormant", "_officer"),
    )

    audit_rows = []
    invalid_contact_rows = []

    for row in merged_df.to_dict(orient="records"):
        emis_value = _clean_optional_text(row.get(school_emis_column)) or row.get(
            "emis_normalized"
        )
        school_name = (
            _clean_optional_text(row.get(school_name_column)) or "Unknown School"
        )
        ddeo_name = _clean_optional_text(row.get("ddeo_name"))
        aeo_name = _clean_optional_text(row.get("aeo_name"))
        ddeo_cell = _clean_optional_text(row.get("ddeo_cell_number"))
        aeo_cell = _clean_optional_text(row.get("aeo_cell_number"))

        status_parts = []
        if ddeo_name:
            status_parts.append("DDEO mapped")
        if aeo_name:
            status_parts.append("AEO mapped")
        if not status_parts:
            status_parts.append("No officer mapping")

        audit_rows.append(
            {
                "School EMIS": emis_value,
                "School Name": school_name,
                "Tehsil": _clean_optional_text(row.get(tehsil_column)) or "",
                "Markaz": _clean_optional_text(row.get(markaz_column)) or "",
                "DDEO Name": ddeo_name or "",
                "DDEO Cell Number": ddeo_cell or "",
                "AEO Name": aeo_name or "",
                "AEO Cell Number": aeo_cell or "",
                "Mapping Status": ", ".join(status_parts),
            }
        )

        for role_label, officer_name, officer_mobile in (
            ("DDEO", ddeo_name, ddeo_cell),
            ("AEO", aeo_name, aeo_cell),
        ):
            if officer_name is None or officer_mobile is None:
                continue

            try:
                _normalize_pk_mobile(officer_mobile)
            except Exception as exc:
                invalid_contact_rows.append(
                    {
                        "School EMIS": emis_value,
                        "School Name": school_name,
                        "Role": role_label,
                        "Officer Name": officer_name,
                        "Mobile Number": officer_mobile,
                        "Issue": str(exc),
                    }
                )

    audit_df = pd.DataFrame(audit_rows)
    unmatched_df = audit_df[audit_df["Mapping Status"] == "No officer mapping"].copy()
    invalid_contacts_df = pd.DataFrame(invalid_contact_rows)

    return unmatched_df, invalid_contacts_df, audit_df


def _build_activity_evidence(
    final_df: pd.DataFrame, filtered_user_df: pd.DataFrame
) -> pd.DataFrame:
    school_emis_column = _find_column(final_df, "School EMIS")
    school_name_column = _find_column(final_df, "School Name")
    tehsil_column = _find_column(final_df, "Tehsil")
    markaz_column = _find_column(final_df, "Markaz")

    if school_emis_column is None or school_name_column is None:
        raise ValueError("Processed report is missing 'School EMIS' or 'School Name'.")

    raw_columns_by_name = _resolve_columns_by_name(
        filtered_user_df, UNFILTERED_DORMANT_ACTIVITY_COLUMNS
    )
    available_activity_columns = [
        resolved for resolved in raw_columns_by_name.values() if resolved is not None
    ]
    evidence_source = filtered_user_df.copy()
    evidence_source["emis_normalized"] = evidence_source["Extracted_EMIS"].map(
        _normalize_emis
    )

    school_columns = [
        column
        for column in [
            school_emis_column,
            school_name_column,
            tehsil_column,
            markaz_column,
        ]
        if column is not None
    ]
    schools_df = final_df[school_columns].copy()
    schools_df["emis_normalized"] = schools_df[school_emis_column].map(_normalize_emis)

    raw_evidence_columns = ["emis_normalized", "Username", *available_activity_columns]
    evidence_df = schools_df.merge(
        evidence_source[raw_evidence_columns],
        on="emis_normalized",
        how="left",
    )
    evidence_df = evidence_df.drop(columns=["emis_normalized"])

    ordered_columns = [
        column
        for column in [
            school_emis_column,
            school_name_column,
            tehsil_column,
            markaz_column,
            "Username",
            *UNFILTERED_DORMANT_ACTIVITY_COLUMNS,
        ]
        if column in evidence_df.columns
    ]
    return evidence_df[ordered_columns].drop_duplicates()


# ==========================================
# 2. DATA ENGINEERING PIPELINE FUNCTIONS
# ==========================================
def _extract_raw_data_with_diagnostics(
    file_path: Path, logger=None
) -> tuple[pd.DataFrame, pd.DataFrame, dict]:
    logger = logger or _get_logger()
    user_df = _read_dataframe_by_content(file_path, dtype=str)
    logger.info(
        f"Loaded raw report from {file_path.name}: "
        f"{len(user_df)} rows, {len(user_df.columns)} columns."
    )

    user_df.columns = user_df.columns.str.strip()
    raw_columns = list(user_df.columns)
    if "Username" not in user_df.columns:
        raise KeyError("Could not find 'Username' column!")
    exported_emis_values = sorted(
        user_df["Username"]
        .astype(str)
        .str.extract(r"^(\d+)", expand=False)
        .dropna()
        .unique()
        .tolist()
    )
    filtered_user_df, filter_diagnostics = _filter_dormant_report_rows(user_df, logger)
    logger.info(
        "Dormant filter result: "
        f"{filter_diagnostics['output_rows']} of "
        f"{filter_diagnostics['input_rows']} rows kept "
        f"using {filter_diagnostics['filter_strategy']}."
    )

    filtered_user_df["Extracted_EMIS"] = (
        filtered_user_df["Username"].astype(str).str.extract(r"^(\d+)", expand=False)
    )
    dormant_emis_values = sorted(
        {
            normalized
            for normalized in filtered_user_df["Extracted_EMIS"].map(_normalize_emis)
            if normalized
        }
    )

    diagnostics = {
        "source_file": str(file_path),
        "source_file_name": file_path.name,
        "source_file_size_bytes": file_path.stat().st_size,
        "source_file_sha256": _file_sha256(file_path),
        "raw_columns": raw_columns,
        # Working sets for reconciliation. Only aggregate counts and bounded
        # samples are copied into run_summary.json by the quality gate.
        "exported_emis_values": exported_emis_values,
        "dormant_emis_values": dormant_emis_values,
        **filter_diagnostics,
    }
    return filtered_user_df[["Extracted_EMIS"]].copy(), filtered_user_df, diagnostics


def extract_raw_data(file_path: Path) -> pd.DataFrame:
    user_df_clean, _, _ = _extract_raw_data_with_diagnostics(file_path)
    return user_df_clean


def transform_data(user_df_clean: pd.DataFrame) -> pd.DataFrame:
    logger = _get_logger()
    master_df = _read_master_dataframe()
    master_df.columns = master_df.columns.str.strip()

    master_df["School EMIS"] = master_df["School EMIS"].astype(str).str.strip()
    user_df_clean["Extracted_EMIS"] = (
        user_df_clean["Extracted_EMIS"].astype(str).str.strip()
    )

    merged_df = pd.merge(
        user_df_clean,
        master_df,
        left_on="Extracted_EMIS",
        right_on="School EMIS",
        how="inner",
    )
    final_df = merged_df[master_df.columns].drop_duplicates()

    # Sort by Tehsil
    tehsil_col = next(
        (c for c in final_df.columns if c.strip().lower() == "tehsil"), None
    )
    if tehsil_col:
        final_df = final_df.sort_values(by=tehsil_col, ascending=True).reset_index(
            drop=True
        )

    # Reset Serial Numbers
    sr_col = next(
        (
            c
            for c in final_df.columns
            if "sr" in c.strip().lower() or "serial" in c.strip().lower()
        ),
        None,
    )
    if sr_col:
        final_df[sr_col] = range(1, len(final_df) + 1)

    logger.info(f"Processed {len(final_df)} unique records.")
    return final_df


def load_excel_report(final_df: pd.DataFrame, output_path: Path, title_text: str):
    logger = _get_logger()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        final_df.to_excel(writer, index=False, startrow=2, sheet_name="Report")

        workbook = writer.book
        worksheet = writer.sheets["Report"]
        num_columns = len(final_df.columns)

        # Title Row
        worksheet["A1"] = title_text
        worksheet["A1"].font = Font(bold=True, size=14, color="000080")
        if num_columns > 0:
            worksheet.merge_cells(
                start_row=1, start_column=1, end_row=1, end_column=num_columns
            )
            worksheet["A1"].alignment = Alignment(
                horizontal="center", vertical="center"
            )

        # Apply Thin Borders & Widths
        thin_border = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000"),
        )

        for col_idx in range(1, num_columns + 1):
            header_cell = worksheet.cell(row=3, column=col_idx)
            header_cell.font = Font(bold=True)
            header_cell.border = thin_border

            column_letter = get_column_letter(col_idx)
            max_length = 0

            for row_idx in range(3, worksheet.max_row + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))

            worksheet.column_dimensions[column_letter].width = max_length + 2


def load_screenshot(final_df: pd.DataFrame, image_path: Path, title_text: str):
    import dataframe_image as dfi

    logger = _get_logger()
    styled_df = (
        final_df.style.set_caption(title_text)
        .set_table_styles(
            [
                {
                    "selector": "caption",
                    "props": [
                        ("color", "#000080"),
                        ("font-size", "18px"),
                        ("font-weight", "bold"),
                        ("text-align", "center"),
                        ("padding-bottom", "15px"),
                    ],
                },
                {
                    "selector": "th",
                    "props": [
                        ("background-color", "#f8f9fa"),
                        ("font-weight", "bold"),
                        ("border", "1px solid black"),
                        ("text-align", "left"),
                    ],
                },
                {
                    "selector": "td",
                    "props": [("border", "1px solid black"), ("text-align", "left")],
                },
            ]
        )
        .hide(axis="index")
    )

    dfi.export(styled_df, str(image_path), max_rows=-1, dpi=300)


def load_officer_mapping_audit(
    final_df: pd.DataFrame, output_path: Path, title_text: str
):
    logger = _get_logger()
    unmatched_df, invalid_contacts_df, audit_df = _build_officer_mapping_audit(final_df)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    if invalid_contacts_df.empty:
        invalid_contacts_df = pd.DataFrame(
            columns=[
                "School EMIS",
                "School Name",
                "Role",
                "Officer Name",
                "Mobile Number",
                "Issue",
            ]
        )

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        unmatched_df.to_excel(writer, index=False, sheet_name="Unmapped Schools")
        invalid_contacts_df.to_excel(writer, index=False, sheet_name="Invalid Contacts")
        audit_df.to_excel(writer, index=False, sheet_name="All Dormant Mapping")

        workbook = writer.book
        for worksheet in workbook.worksheets:
            worksheet.freeze_panes = "A2"
            for cell in worksheet[1]:
                cell.font = Font(bold=True)

            for column_cells in worksheet.columns:
                column_letter = get_column_letter(column_cells[0].column)
                max_length = max(
                    len(str(cell.value)) if cell.value is not None else 0
                    for cell in column_cells
                )
                worksheet.column_dimensions[column_letter].width = min(
                    max(max_length + 2, 12), 45
                )

    logger.info(
        f"Officer mapping audit saved to {output_path}. "
        f"Unmapped schools: {len(unmatched_df)}, invalid contacts: {len(invalid_contacts_df)}."
    )


def load_officer_delivery_audit(audit_rows: list[dict], output_path: Path):
    logger = _get_logger()
    output_path.parent.mkdir(parents=True, exist_ok=True)

    audit_df = pd.DataFrame(audit_rows)
    if audit_df.empty:
        audit_df = pd.DataFrame(
            columns=[
                "Job ID",
                "Target",
                "Mobile Number",
                "Officer Name",
                "Roles",
                "School Count",
                "Tehsils",
                "Markazes",
                "Schools",
                "Includes Excel",
                "Delivery Status",
                "WhatsApp Message IDs",
                "WhatsApp Ack Statuses",
                "WhatsApp Remote JIDs",
                "WhatsApp Requested Target",
                "WhatsApp Actual Send Target",
                "WhatsApp Canonical Target",
                "WhatsApp Known LID",
                "WhatsApp Target Policy",
                "WhatsApp Token Preflight",
                "Delivery Error",
            ]
        )

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        audit_df.to_excel(writer, index=False, sheet_name="Officer Delivery")
        worksheet = writer.book["Officer Delivery"]
        worksheet.freeze_panes = "A2"
        for cell in worksheet[1]:
            cell.font = Font(bold=True)

        for column_cells in worksheet.columns:
            column_letter = get_column_letter(column_cells[0].column)
            max_length = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in column_cells
            )
            worksheet.column_dimensions[column_letter].width = min(
                max(max_length + 2, 12), 55
            )

        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    logger.info(f"Officer delivery audit saved to {output_path}.")


def load_activity_evidence_report(
    final_df: pd.DataFrame, filtered_user_df: pd.DataFrame, output_path: Path
):
    logger = _get_logger()
    evidence_df = _build_activity_evidence(final_df, filtered_user_df)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        evidence_df.to_excel(writer, index=False, sheet_name="Activity Evidence")

        worksheet = writer.sheets["Activity Evidence"]
        worksheet.freeze_panes = "A2"
        for cell in worksheet[1]:
            cell.font = Font(bold=True)

        for column_cells in worksheet.columns:
            column_letter = get_column_letter(column_cells[0].column)
            max_length = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in column_cells
            )
            worksheet.column_dimensions[column_letter].width = min(
                max(max_length + 2, 12), 45
            )

    logger.info(f"Activity evidence report saved to {output_path}.")
    return evidence_df


# --- NEW: Helper to create a nice text message ---
# --- UPDATED: Helper to create a Tehsil-wise text message ---
def create_whatsapp_summary(df: pd.DataFrame, title_text: str) -> str:
    total_users = len(df)
    msg = f"🚨 *{title_text}*\n\n"
    msg += f"📊 *Total Dormant Users:* {total_users}\n"

    # Find the Tehsil column dynamically (ignoring case/spaces)
    tehsil_col = next((c for c in df.columns if c.strip().lower() == "tehsil"), None)

    if tehsil_col:
        # Get counts per tehsil and sort alphabetically
        tehsil_counts = df[tehsil_col].value_counts().sort_index()
        for tehsil, count in tehsil_counts.items():
            # Clean up the tehsil name just in case there are weird spaces in the Excel
            clean_tehsil = str(tehsil).strip().title()
            msg += f"Tehsil {clean_tehsil} = {count}\n"
    else:
        msg += "*(Tehsil breakdown not available in this report)*\n"

    msg += "\nPlease review the attached Excel report for details."
    return msg


def _summarize_delivery_operation_result(status: dict) -> str:
    operation_result = status.get("operationResult") or []
    if not isinstance(operation_result, list):
        return ""

    parts: list[str] = []
    for item in operation_result:
        if not isinstance(item, dict):
            continue

        message_id = _clean_optional_text(item.get("id"))
        status_name = _clean_optional_text(item.get("statusName"))
        status_code = item.get("status")
        remote_jid = _clean_optional_text(item.get("remoteJid"))
        status_text = status_name or (
            f"status {status_code}" if status_code is not None else None
        )
        item_parts = [
            value
            for value in [message_id, status_text, remote_jid]
            if value is not None
        ]
        if item_parts:
            parts.append(" / ".join(item_parts))

    return "; ".join(parts)


def _build_delivery_failure_summary(
    failed_officer_rows: list[dict],
    delivery_statuses: dict[str, dict],
    title_text: str,
) -> str:
    lines = [
        f"*Delivery failure alert: {title_text}*",
        "",
        (
            "The dormant-user report was generated, but WhatsApp did not confirm "
            "delivery to these mapped officers. Please share the report manually "
            "or confirm their WhatsApp availability."
        ),
        "",
        "*Failed officers*",
    ]

    for index, row in enumerate(failed_officer_rows, start=1):
        status = delivery_statuses.get(row.get("Job ID"), {})
        error = (
            _clean_optional_text(row.get("Delivery Error"))
            or _clean_optional_text(status.get("error"))
            or "No delivery confirmation"
        )
        operation_summary = _summarize_delivery_operation_result(status)
        line = (
            f"{index}. {row.get('Officer Name')} - {row.get('Mobile Number')} "
            f"({row.get('Roles')}, schools: {row.get('School Count')}): {error}"
        )
        actual_target = (
            _clean_optional_text(row.get("WhatsApp Actual Send Target"))
            or _clean_optional_text(status.get("actualSendTarget"))
        )
        known_lid = (
            _clean_optional_text(row.get("WhatsApp Known LID"))
            or _clean_optional_text(status.get("knownLidTarget"))
        )
        if actual_target or known_lid:
            line += (
                f" [target: {actual_target or row.get('Target')}; "
                f"known LID: {known_lid or '-'}]"
            )
        if operation_summary:
            line += f" [{operation_summary}]"
        lines.append(line)

    lines.extend(
        [
            "",
            "The Excel report is attached again for escalation.",
        ]
    )
    return "\n".join(lines)


def _build_failed_delivery_fallback_payloads(
    *,
    failed_officer_rows: list[dict],
    delivery_statuses: dict[str, dict],
    excel_path: Path,
    image_path: Path | None,
    title_text: str,
) -> list[dict]:
    if not failed_officer_rows:
        return []

    summary_text = _build_delivery_failure_summary(
        failed_officer_rows,
        delivery_statuses,
        title_text,
    )
    fallback_payloads = build_fixed_whatsapp_payloads(
        excel_path,
        image_path,
        summary_text,
        purpose="fallback",
    )

    for payload in fallback_payloads:
        payload["recipient_name"] = (
            f"{payload.get('recipient_name') or 'Fixed recipient'} - delivery escalation"
        )

    return fallback_payloads


async def _request_whatsapp_worker_health(nc) -> dict:
    health_wait_seconds = max(
        WHATSAPP_HEALTH_TIMEOUT_SECONDS,
        WHATSAPP_HEALTH_WAIT_SECONDS,
    )
    deadline = time.monotonic() + health_wait_seconds
    attempt = 0
    last_error: Exception | None = None

    while True:
        attempt += 1
        try:
            response = await nc.request(
                NATS_HEALTH_SUBJECT,
                b"{}",
                timeout=WHATSAPP_HEALTH_TIMEOUT_SECONDS,
            )
            body = response.data.decode("utf-8") if response.data else "{}"
            data = json.loads(body)
            if not data.get("ready"):
                last_error = RuntimeError(
                    "WhatsApp worker responded but is not ready: "
                    + json.dumps(data, sort_keys=True)
                )
                remaining = deadline - time.monotonic()
                if remaining <= 0:
                    break
                await asyncio.sleep(min(2.0, max(0.1, remaining)))
                continue
            return data
        except (NoRespondersError, NatsTimeoutError) as exc:
            last_error = exc
            remaining = deadline - time.monotonic()
            if remaining <= 0:
                break
            await asyncio.sleep(min(2.0, max(0.1, remaining)))

    raise RuntimeError(
        "WhatsApp worker is not answering "
        f"{NATS_HEALTH_SUBJECT!r} after {health_wait_seconds:.0f}s "
        f"and {attempt} health request(s). Last error: {last_error}. "
        "Make sure the WhatsApp worker is connected and using the same "
        "NATS_HEALTH_SUBJECT."
    )


def _delivery_operation_result_has_message_id(operation_result: object) -> bool:
    return isinstance(operation_result, list) and any(
        isinstance(item, dict) and bool(item.get("id")) for item in operation_result
    )


def _delivery_operation_result_has_terminal_error(operation_result: object) -> bool:
    return isinstance(operation_result, list) and any(
        isinstance(item, dict) and str(item.get("statusName") or "").upper() == "ERROR"
        for item in operation_result
    )


def _delivery_status_is_pending_confirmation(status: dict) -> bool:
    status_name = str(status.get("status") or "").strip().lower()
    if status_name == "sent_pending_confirmation":
        return True

    error = str(status.get("error") or "")
    return (
        status_name == "failed"
        and str(status.get("type") or "").strip().lower() == "group"
        and "Timed out waiting for WhatsApp" in error
        and _delivery_operation_result_has_message_id(status.get("operationResult"))
        and not _delivery_operation_result_has_terminal_error(
            status.get("operationResult")
        )
    )


async def _collect_delivery_statuses(
    subscription,
    expected_job_ids: set[str],
) -> dict:
    statuses: dict[str, dict] = {}
    deadline = time.monotonic() + WHATSAPP_DELIVERY_TIMEOUT_SECONDS

    while expected_job_ids.difference(statuses):
        remaining = deadline - time.monotonic()
        if remaining <= 0:
            break

        try:
            message = await subscription.next_msg(timeout=min(remaining, 30))
        except NatsTimeoutError:
            continue

        try:
            payload = json.loads(message.data.decode("utf-8"))
        except Exception as exc:
            payload = {"status": "invalid_status_payload", "error": str(exc)}

        job_id = payload.get("jobId") or payload.get("job_id")
        if job_id:
            statuses[str(job_id)] = payload

    missing = sorted(expected_job_ids.difference(statuses))
    delivered = [
        status for status in statuses.values() if status.get("status") == "delivered"
    ]
    pending_confirmation = [
        status
        for status in statuses.values()
        if _delivery_status_is_pending_confirmation(status)
    ]
    failed = [
        status
        for status in statuses.values()
        if status.get("status") not in {"delivered", None}
        and not _delivery_status_is_pending_confirmation(status)
    ]
    return {
        "received": len(statuses),
        "delivered": len(delivered),
        "sent": len(delivered) + len(pending_confirmation),
        "pending_confirmation": len(pending_confirmation),
        "failed": len(failed),
        "missing": missing,
        "statuses": statuses,
        "timed_out": bool(missing),
    }


async def _publish_jetstream_payload(js, payload: dict) -> None:
    """Publish once logically, retrying safely with JetStream deduplication."""
    encoded_payload = json.dumps(payload).encode("utf-8")
    message_id = str(payload.get("job_id") or "").strip()
    if not message_id:
        message_id = hashlib.sha256(encoded_payload).hexdigest()
    for attempt in range(1, NATS_PUBLISH_ATTEMPTS + 1):
        try:
            await js.publish(
                NATS_SUBJECT,
                encoded_payload,
                headers={"Nats-Msg-Id": message_id},
            )
            return
        except Exception:
            if attempt >= NATS_PUBLISH_ATTEMPTS:
                raise
            _get_logger().warning(
                "JetStream publish failed for job %s on attempt %s/%s; "
                "retrying with the same idempotency key.",
                message_id,
                attempt,
                NATS_PUBLISH_ATTEMPTS,
            )
            await asyncio.sleep(NATS_PUBLISH_RETRY_DELAY_SECONDS)


async def _publish_to_nats(payloads: list[dict], *, wait_for_delivery: bool):
    """Async helper to connect and publish to JetStream."""
    nc = await nats.connect(NATS_URL)
    js = nc.jetstream()
    delivery_subscription = None
    status_subject = None
    try:
        health = None
        if WHATSAPP_REQUIRE_WORKER_READY:
            health = await _request_whatsapp_worker_health(nc)

        if wait_for_delivery:
            status_subject = f"whatsapp.delivery.antidengue.{uuid.uuid4().hex}"
            delivery_subscription = await nc.subscribe(status_subject)
            for payload in payloads:
                payload["status_subject"] = status_subject

        try:
            owning_stream = await js.find_stream_name_by_subject(NATS_SUBJECT)
        except NotFoundError:
            owning_stream = None

        stream_config = StreamConfig(
            name=owning_stream or NATS_STREAM,
            subjects=[NATS_SUBJECT],
            retention=RetentionPolicy.WORK_QUEUE,
            storage=StorageType.FILE,
        )

        if owning_stream:
            stream_info = await js.stream_info(owning_stream)
            existing_subjects = set(stream_info.config.subjects or [])
            if NATS_SUBJECT not in existing_subjects:
                stream_config.subjects = sorted([*existing_subjects, NATS_SUBJECT])
                await js.update_stream(stream_config)
        else:
            await js.add_stream(stream_config)

        for payload in payloads:
            await _publish_jetstream_payload(js, payload)

        await nc.flush()

        delivery_result = None
        if delivery_subscription is not None:
            delivery_result = await _collect_delivery_statuses(
                delivery_subscription,
                {payload["job_id"] for payload in payloads},
            )

        return {
            "health": health,
            "delivery": delivery_result,
            "status_subject": status_subject,
        }
    finally:
        if delivery_subscription is not None:
            try:
                await delivery_subscription.unsubscribe()
            except Exception:
                pass
        await nc.close()


def _format_privacy_token_preflight(operation_result: list) -> str:
    for item in operation_result:
        if not isinstance(item, dict):
            continue
        preflight = item.get("privacyTokenPreflight")
        if not isinstance(preflight, dict):
            continue

        before = preflight.get("before") or {}
        after = preflight.get("after") or {}
        issuance = preflight.get("issuance") or {}
        return (
            f"checked={preflight.get('checked')}; "
            f"recovered={preflight.get('recovered')}; "
            f"before_usable={before.get('usable')}; "
            f"before_expired={before.get('expired')}; "
            f"after_usable={after.get('usable')}; "
            f"token_timestamp={after.get('timestamp')}; "
            f"issue_target={issuance.get('issueTarget') or ''}"
        )
    return ""


def send_whatsapp_alert(
    final_df: pd.DataFrame,
    excel_path: Path,
    image_path: Path | None,
    title_text: str,
    message_body: str,
    dry_run: bool = False,
):
    logger = _get_logger()
    dispatch_settings = _read_dispatch_settings()
    automation_blocked = bool(
        dispatch_settings.get("manual_only")
        or dispatch_settings.get("require_preview")
    )
    officer_delivery_audit: list[dict] = []
    try:
        dynamic_payloads, officer_delivery_audit = (
            build_dynamic_officer_payloads_with_audit(
                final_df, excel_path, image_path, title_text
            )
        )
    except Exception as exc:
        raise RuntimeError(
            "Dynamic officer recipient build failed; refusing to send a partial "
            f"WhatsApp run. Error: {exc}"
        )

    if automation_blocked:
        skipped_status = (
            "manual pending"
            if dispatch_settings.get("manual_only")
            else "preview required"
        )
        skipped_reason = (
            "Dispatch Center manual-only mode is enabled."
            if dispatch_settings.get("manual_only")
            else "Dispatch Center preview is required before sending."
        )
    elif not dispatch_settings.get("allow_individual", True):
        skipped_status = "skipped"
        skipped_reason = "Individual direct delivery is disabled in Dispatch Center."
    else:
        skipped_status = ""
        skipped_reason = ""

    dynamic_payloads_to_queue = dynamic_payloads
    if skipped_status:
        dynamic_payloads_to_queue = []
        for audit_row in officer_delivery_audit:
            audit_row["Delivery Status"] = skipped_status
            audit_row["Delivery Error"] = skipped_reason

    if dispatch_settings.get("allow_groups", True) and not automation_blocked:
        fixed_payloads = build_group_route_payloads(
            final_df,
            excel_path,
            image_path,
            title_text,
            message_body,
            purpose="normal",
        )
    else:
        fixed_payloads = []

    payloads = [*dynamic_payloads_to_queue, *fixed_payloads]
    queued_dynamic_job_ids = {
        str(payload.get("job_id"))
        for payload in dynamic_payloads_to_queue
        if payload.get("job_id")
    }
    if not dispatch_settings.get("attach_excel", True):
        for payload in payloads:
            payload["excel_path"] = None
            payload["excel_filename"] = None

    if not payloads:
        skipped_parts = []
        if automation_blocked:
            skipped_parts.append(skipped_reason)
        if not dispatch_settings.get("allow_individual", True):
            skipped_parts.append("Individual direct delivery is disabled.")
        if not dispatch_settings.get("allow_groups", True):
            skipped_parts.append("Group delivery is disabled.")
        skipped_reason_text = " ".join(skipped_parts).strip()
        logger.warning(
            skipped_reason_text
            or "No enabled WhatsApp recipients found. Nothing was queued."
        )
        return {
            "dry_run": dry_run,
            "dynamic_payloads": len(dynamic_payloads),
            "fixed_payloads": len(fixed_payloads),
            "total_payloads": 0,
            "queued": False,
            "officer_delivery_audit": officer_delivery_audit,
            "dispatch_plan": _build_dispatch_plan_items(
                dynamic_payloads=dynamic_payloads,
                dynamic_payloads_to_queue=dynamic_payloads_to_queue,
                fixed_payloads=fixed_payloads,
                officer_delivery_audit=officer_delivery_audit,
                dry_run=dry_run,
                skipped_status=skipped_status or "skipped",
                skipped_reason=skipped_reason_text,
            ),
            "dispatch_settings": dispatch_settings,
            "skipped_reason": skipped_reason_text,
        }

    if dry_run:
        logger.info(
            "Dry-run enabled; WhatsApp payloads were built but not queued: "
            f"{len(dynamic_payloads_to_queue)} dynamic officer recipients, "
            f"{len(fixed_payloads)} fixed recipients, "
            f"{len(payloads)} total."
        )
        return {
            "dry_run": True,
            "dynamic_payloads": len(dynamic_payloads),
            "queued_dynamic_payloads": len(dynamic_payloads_to_queue),
            "fixed_payloads": len(fixed_payloads),
            "total_payloads": len(payloads),
            "queued": False,
            "officer_delivery_audit": officer_delivery_audit,
            "dispatch_plan": _build_dispatch_plan_items(
                dynamic_payloads=dynamic_payloads,
                dynamic_payloads_to_queue=dynamic_payloads_to_queue,
                fixed_payloads=fixed_payloads,
                officer_delivery_audit=officer_delivery_audit,
                dry_run=True,
            ),
            "dispatch_settings": dispatch_settings,
        }

    fallback_payloads: list[dict] = []
    fallback_delivery_statuses: dict[str, dict] = {}
    publish_result = asyncio.run(
        _publish_to_nats(payloads, wait_for_delivery=WHATSAPP_WAIT_FOR_DELIVERY)
    )
    delivery_result = publish_result.get("delivery") or {}
    delivery_statuses = delivery_result.get("statuses") or {}

    for audit_row in officer_delivery_audit:
        job_id = str(audit_row.get("Job ID") or "")
        status = delivery_statuses.get(job_id)
        if status:
            audit_row["Delivery Status"] = status.get("status") or "unknown"
            audit_row["Delivery Error"] = status.get("error") or ""
            identity = status.get("identity") or {}
            audit_row["WhatsApp Requested Target"] = (
                status.get("requestedTarget")
                or identity.get("requestedTarget")
                or audit_row.get("Target")
                or ""
            )
            audit_row["WhatsApp Actual Send Target"] = (
                status.get("actualSendTarget")
                or identity.get("actualSendTarget")
                or ""
            )
            audit_row["WhatsApp Canonical Target"] = (
                status.get("canonicalTarget")
                or identity.get("canonicalTarget")
                or audit_row.get("Target")
                or ""
            )
            audit_row["WhatsApp Known LID"] = (
                status.get("knownLidTarget")
                or identity.get("knownLidTarget")
                or ""
            )
            audit_row["WhatsApp Target Policy"] = (
                status.get("targetPolicy")
                or identity.get("policy")
                or ""
            )
            operation_result = status.get("operationResult") or []
            if isinstance(operation_result, list):
                audit_row["WhatsApp Message IDs"] = "\n".join(
                    str(item.get("id"))
                    for item in operation_result
                    if isinstance(item, dict) and item.get("id")
                )
                audit_row["WhatsApp Ack Statuses"] = "\n".join(
                    str(item.get("statusName") or item.get("status"))
                    for item in operation_result
                    if isinstance(item, dict)
                    and (
                        item.get("statusName") is not None
                        or item.get("status") is not None
                    )
                )
                audit_row["WhatsApp Remote JIDs"] = "\n".join(
                    str(item.get("remoteJid"))
                    for item in operation_result
                    if isinstance(item, dict) and item.get("remoteJid")
                )
                audit_row["WhatsApp Token Preflight"] = (
                    _format_privacy_token_preflight(operation_result)
                )
        elif WHATSAPP_WAIT_FOR_DELIVERY and job_id in queued_dynamic_job_ids:
            audit_row["Delivery Status"] = "missing status"
            audit_row["WhatsApp Requested Target"] = audit_row.get("Target") or ""
            audit_row["WhatsApp Canonical Target"] = audit_row.get("Target") or ""
            audit_row["Delivery Error"] = (
                "No delivery status received from WhatsApp worker"
            )
        elif job_id in queued_dynamic_job_ids:
            audit_row["Delivery Status"] = "queued"
            audit_row["WhatsApp Requested Target"] = audit_row.get("Target") or ""
            audit_row["WhatsApp Canonical Target"] = audit_row.get("Target") or ""
            audit_row["Delivery Error"] = ""

    if WHATSAPP_WAIT_FOR_DELIVERY:
        failed_officer_rows = [
            row
            for row in officer_delivery_audit
            if str(row.get("Job ID") or "") in queued_dynamic_job_ids
            and row.get("Delivery Status") != "delivered"
        ]
        fallback_result = None
        if failed_officer_rows:
            try:
                fallback_payloads = _build_failed_delivery_fallback_payloads(
                    failed_officer_rows=failed_officer_rows,
                    delivery_statuses=delivery_statuses,
                    excel_path=excel_path,
                    image_path=image_path,
                    title_text=title_text,
                )
                if fallback_payloads:
                    logger.error(
                        "Officer delivery failed for "
                        f"{len(failed_officer_rows)} recipient(s); queueing "
                        f"{len(fallback_payloads)} escalation payload(s)."
                    )
                    fallback_publish_result = asyncio.run(
                        _publish_to_nats(
                            fallback_payloads,
                            wait_for_delivery=WHATSAPP_WAIT_FOR_DELIVERY,
                        )
                    )
                    fallback_result = {
                        "payloads": len(fallback_payloads),
                        "delivery": fallback_publish_result.get("delivery"),
                        "worker_health": fallback_publish_result.get("health"),
                    }
                    fallback_delivery_statuses = (
                        (fallback_publish_result.get("delivery") or {}).get("statuses")
                        or {}
                    )
            except Exception as exc:
                fallback_result = {
                    "payloads": 0,
                    "delivery": None,
                    "error": str(exc),
                }
                logger.error(f"Failed to queue officer delivery escalation: {exc}")

        if delivery_result.get("timed_out") or delivery_result.get("failed"):
            error_message = (
                "WhatsApp delivery did not complete cleanly: "
                f"{delivery_result.get('delivered', 0)} delivered, "
                f"{delivery_result.get('failed', 0)} failed, "
                f"{len(delivery_result.get('missing', []))} missing status updates."
            )
            logger.error(error_message)
            return {
                "dry_run": False,
                "dynamic_payloads": len(dynamic_payloads),
                "queued_dynamic_payloads": len(dynamic_payloads_to_queue),
                "fixed_payloads": len(fixed_payloads),
                "total_payloads": len(payloads),
                "queued": False,
                "worker_health": publish_result.get("health"),
                "delivery": delivery_result or None,
                "fallback_delivery": fallback_result,
                "officer_delivery_audit": officer_delivery_audit,
                "dispatch_plan": _build_dispatch_plan_items(
                    dynamic_payloads=dynamic_payloads,
                    dynamic_payloads_to_queue=dynamic_payloads_to_queue,
                    fixed_payloads=fixed_payloads,
                    officer_delivery_audit=officer_delivery_audit,
                    delivery_statuses=delivery_statuses,
                    fallback_payloads=fallback_payloads,
                    fallback_delivery_statuses=fallback_delivery_statuses,
                ),
                "dispatch_settings": dispatch_settings,
                "error": error_message,
            }
        if failed_officer_rows:
            failed_names = ", ".join(
                f"{row.get('Officer Name')} ({row.get('Mobile Number')}: {row.get('Delivery Status')})"
                for row in failed_officer_rows[:5]
            )
            error_message = (
                "Officer WhatsApp delivery audit has non-delivered recipients: "
                f"{failed_names}"
            )
            logger.error(error_message)
            return {
                "dry_run": False,
                "dynamic_payloads": len(dynamic_payloads),
                "queued_dynamic_payloads": len(dynamic_payloads_to_queue),
                "fixed_payloads": len(fixed_payloads),
                "total_payloads": len(payloads),
                "queued": False,
                "worker_health": publish_result.get("health"),
                "delivery": delivery_result or None,
                "fallback_delivery": fallback_result,
                "officer_delivery_audit": officer_delivery_audit,
                "dispatch_plan": _build_dispatch_plan_items(
                    dynamic_payloads=dynamic_payloads,
                    dynamic_payloads_to_queue=dynamic_payloads_to_queue,
                    fixed_payloads=fixed_payloads,
                    officer_delivery_audit=officer_delivery_audit,
                    delivery_statuses=delivery_statuses,
                    fallback_payloads=fallback_payloads,
                    fallback_delivery_statuses=fallback_delivery_statuses,
                ),
                "dispatch_settings": dispatch_settings,
                "error": error_message,
            }

        pending_confirmation_count = int(delivery_result.get("pending_confirmation") or 0)
        if pending_confirmation_count:
            logger.warning(
                "WhatsApp payloads sent with pending group acknowledgements: "
                f"{delivery_result.get('sent', len(payloads))} sent, "
                f"{pending_confirmation_count} pending confirmation."
            )
        else:
            logger.info(
                "WhatsApp payloads delivered successfully: "
                f"{len(dynamic_payloads_to_queue)} dynamic officer recipients, "
                f"{len(fixed_payloads)} fixed recipients, "
                f"{delivery_result.get('delivered', len(payloads))} delivered."
            )
    else:
        logger.info(
            "Queued WhatsApp payloads successfully: "
            f"{len(dynamic_payloads_to_queue)} dynamic officer recipients, "
            f"{len(fixed_payloads)} fixed recipients, "
            f"{len(payloads)} total."
        )

    return {
        "dry_run": False,
        "dynamic_payloads": len(dynamic_payloads),
        "queued_dynamic_payloads": len(dynamic_payloads_to_queue),
        "fixed_payloads": len(fixed_payloads),
        "total_payloads": len(payloads),
        "queued": True,
        "worker_health": publish_result.get("health"),
        "delivery": delivery_result or None,
        "officer_delivery_audit": officer_delivery_audit,
        "dispatch_plan": _build_dispatch_plan_items(
            dynamic_payloads=dynamic_payloads,
            dynamic_payloads_to_queue=dynamic_payloads_to_queue,
            fixed_payloads=fixed_payloads,
            officer_delivery_audit=officer_delivery_audit,
            delivery_statuses=delivery_statuses,
            fallback_payloads=fallback_payloads,
            fallback_delivery_statuses=fallback_delivery_statuses,
        ),
        "dispatch_settings": dispatch_settings,
    }


def archive_raw_file(file_path: Path):
    ARCHIVE_DIR.mkdir(parents=True, exist_ok=True)
    archive_path = ARCHIVE_DIR / file_path.name
    if archive_path.exists():
        archive_path = (
            ARCHIVE_DIR / f"{file_path.stem}_{int(time.time())}{file_path.suffix}"
        )
    shutil.move(str(file_path), str(archive_path))


def _create_run_directories(current_time_obj: datetime.datetime) -> tuple[Path, Path]:
    timestamp_folder_name = current_time_obj.strftime("%Y-%m-%d_%H-%M-%S")
    run_output_dir = OUTPUT_DIR / timestamp_folder_name
    run_unmapped_report_dir = UNMAPPED_REPORT_DIR / timestamp_folder_name

    counter = 2
    while run_output_dir.exists() or run_unmapped_report_dir.exists():
        suffix = f"_{counter}"
        run_output_dir = OUTPUT_DIR / f"{timestamp_folder_name}{suffix}"
        run_unmapped_report_dir = (
            UNMAPPED_REPORT_DIR / f"{timestamp_folder_name}{suffix}"
        )
        counter += 1

    run_output_dir.mkdir(parents=True, exist_ok=False)
    run_unmapped_report_dir.mkdir(parents=True, exist_ok=False)
    return run_output_dir, run_unmapped_report_dir


def _quality_gate(
    *,
    final_df: pd.DataFrame,
    extraction_diagnostics: dict,
    unmatched_df: pd.DataFrame,
    invalid_contacts_df: pd.DataFrame,
    enforce_officer_delivery_quality: bool = True,
    observed_at: datetime.datetime | None = None,
) -> dict:
    master_df = _read_master_dataframe()
    master_emis_column = _find_column(master_df, "School EMIS")
    master_emis = {
        normalized
        for normalized in (
            master_df[master_emis_column].map(_normalize_emis)
            if master_emis_column is not None
            else []
        )
        if normalized
    }
    master_count = len(master_emis) if master_emis else len(master_df)
    final_count = len(final_df)
    unmapped_count = len(unmatched_df)
    invalid_contact_count = len(invalid_contacts_df)
    exported_emis = set(extraction_diagnostics.get("exported_emis_values") or [])
    dormant_emis = set(extraction_diagnostics.get("dormant_emis_values") or [])
    if not dormant_emis:
        final_emis_column = _find_column(final_df, "School EMIS")
        if final_emis_column is not None:
            dormant_emis = {
                normalized
                for normalized in final_df[final_emis_column].map(_normalize_emis)
                if normalized
            }
    if not exported_emis:
        # Backward compatibility for direct callers that predate source
        # reconciliation. Dormant rows are the only safe available proxy.
        exported_emis = set(dormant_emis)

    reconciled_export_emis = exported_emis & master_emis if master_emis else exported_emis
    unknown_export_emis = exported_emis - master_emis if master_emis else set()
    master_missing_from_export = master_emis - exported_emis if exported_emis else set()
    reconciled_dormant_emis = dormant_emis & master_emis if master_emis else dormant_emis
    export_coverage_ratio = (
        len(reconciled_export_emis) / master_count if master_count else 0
    )
    dormancy_ratio = (
        len(reconciled_dormant_emis) / len(reconciled_export_emis)
        if reconciled_export_emis
        else 0
    )
    coverage_ratio = final_count / master_count if master_count else 0
    unmapped_ratio = unmapped_count / final_count if final_count else 0
    invalid_contact_ratio = (
        invalid_contact_count / (final_count * 2) if final_count else 0
    )

    errors: list[str] = []
    warnings: list[str] = []
    issues: list[dict] = []

    def add_issue(code: str, severity: str, message: str, **context) -> None:
        (errors if severity == "blocked" else warnings).append(message)
        issues.append(
            {
                "code": code,
                "severity": severity,
                "message": message,
                **context,
            }
        )

    filter_applied = bool(extraction_diagnostics.get("filter_applied"))
    filter_strategy = extraction_diagnostics.get("filter_strategy")
    missing_activity_columns = (
        extraction_diagnostics.get("missing_activity_columns") or []
    )
    available_activity_columns = (
        extraction_diagnostics.get("available_activity_columns") or []
    )

    total_activity_available = TOTAL_ACTIVITY_COLUMN in available_activity_columns
    available_detail_columns = [
        column
        for column in DETAILED_ACTIVITY_COLUMNS
        if column in available_activity_columns
    ]
    schema_mode = (
        "total_activities"
        if total_activity_available
        else "complete_detail_columns"
        if len(available_detail_columns) == len(DETAILED_ACTIVITY_COLUMNS)
        else "partial_detail_columns"
        if available_detail_columns
        else "no_activity_columns"
    )
    invalid_total_activity_rows = int(
        extraction_diagnostics.get("invalid_total_activity_rows") or 0
    )
    dormant_activity_conflict_rows = int(
        extraction_diagnostics.get("dormant_activity_conflict_rows") or 0
    )

    # Detailed fields such as TPV are optional capabilities when the portal
    # provides its authoritative Total Activities value.
    if REPORT_SOURCE in {"unfiltered", "all"} and schema_mode == "partial_detail_columns":
        add_issue(
            "unreliable_activity_schema",
            "blocked",
            "The unfiltered portal export has no Total Activities column and only "
            "a partial set of detailed activity columns; dormant status cannot be "
            "determined safely.",
            missing_columns=missing_activity_columns,
        )

    if total_activity_available and invalid_total_activity_rows:
        add_issue(
            "invalid_total_activity_values",
            "blocked",
            f"Total Activities is blank or non-numeric for "
            f"{invalid_total_activity_rows} portal row(s); dormant status cannot be "
            "determined safely.",
            affected_rows=invalid_total_activity_rows,
        )

    if total_activity_available and dormant_activity_conflict_rows:
        add_issue(
            "contradictory_activity_totals",
            "blocked",
            f"{dormant_activity_conflict_rows} portal row(s) have Total Activities "
            "equal to zero but a positive detailed activity value.",
            affected_rows=dormant_activity_conflict_rows,
        )

    if REPORT_SOURCE in {"unfiltered", "all"} and not filter_applied:
        add_issue(
            "dormant_filter_not_applied",
            "blocked",
            "REPORT_SOURCE is unfiltered, but no local dormant filter was applied.",
        )

    if REPORT_SOURCE in {"unfiltered", "all"} and not available_activity_columns:
        add_issue(
            "activity_columns_missing",
            "blocked",
            "Unfiltered report has no activity columns; refusing to send WhatsApp alerts.",
        )

    if final_count == master_count and not filter_applied:
        add_issue(
            "full_master_without_filter",
            "blocked",
            "Inactive school count equals the full master list and no local filter was applied.",
        )
    elif (
        dormancy_ratio >= QUALITY_MAX_DORMANCY_RATIO
        and (observed_at or datetime.datetime.now()).hour
        >= QUALITY_DORMANCY_WARNING_AFTER_HOUR
    ):
        message = (
            f"Dormant rate is {dormancy_ratio:.1%} of reconciled portal schools "
            f"({len(reconciled_dormant_emis)}/{len(reconciled_export_emis)}); "
            f"portal export coverage is {export_coverage_ratio:.1%} "
            f"({len(reconciled_export_emis)}/{master_count})."
        )
        if filter_strategy == "already_filtered":
            add_issue("high_dormancy_rate", "blocked", message)
        else:
            add_issue("high_dormancy_rate", "warning", message)

    if exported_emis and export_coverage_ratio < QUALITY_MIN_EXPORT_COVERAGE:
        add_issue(
            "low_portal_export_coverage",
            "warning",
            f"Portal export covers only {export_coverage_ratio:.1%} of the master list "
            f"({len(reconciled_export_emis)}/{master_count}).",
            master_missing_sample=sorted(master_missing_from_export)[:20],
            unknown_export_sample=sorted(unknown_export_emis)[:20],
        )

    if (
        enforce_officer_delivery_quality
        and final_count
        and unmapped_ratio > QUALITY_MAX_UNMAPPED_RATIO
    ):
        message = (
            f"Officer mapping coverage is too low: {unmapped_count} unmapped of "
            f"{final_count} inactive schools ({unmapped_ratio:.1%})."
        )
        add_issue("officer_mapping_incomplete", "blocked", message)

    if (
        enforce_officer_delivery_quality
        and final_count
        and invalid_contact_ratio > QUALITY_MAX_UNMAPPED_RATIO
    ):
        message = (
            f"Too many invalid officer contacts: {invalid_contact_count} invalid "
            f"contact rows for {final_count} inactive schools."
        )
        add_issue("officer_contacts_invalid", "blocked", message)

    return {
        "passed": not errors,
        "errors": errors,
        "warnings": warnings,
        "issues": issues,
        "activity_schema_mode": schema_mode,
        "optional_missing_activity_columns": (
            missing_activity_columns if total_activity_available else []
        ),
        "invalid_total_activity_rows": invalid_total_activity_rows,
        "dormant_activity_conflict_rows": dormant_activity_conflict_rows,
        "master_count": master_count,
        "final_school_count": final_count,
        "master_coverage_ratio": coverage_ratio,
        "exported_school_count": len(exported_emis),
        "reconciled_export_school_count": len(reconciled_export_emis),
        "export_coverage_ratio": export_coverage_ratio,
        "dormant_reconciled_school_count": len(reconciled_dormant_emis),
        "dormancy_ratio": dormancy_ratio,
        "unknown_export_school_count": len(unknown_export_emis),
        "unknown_export_emis_sample": sorted(unknown_export_emis)[:20],
        "master_missing_from_export_count": len(master_missing_from_export),
        "master_missing_from_export_emis_sample": sorted(master_missing_from_export)[:20],
        "dormancy_rate_check_applicable": (
            (observed_at or datetime.datetime.now()).hour
            >= QUALITY_DORMANCY_WARNING_AFTER_HOUR
        ),
        "unmapped_school_count": unmapped_count,
        "unmapped_ratio": unmapped_ratio,
        "invalid_contact_count": invalid_contact_count,
        "invalid_contact_ratio": invalid_contact_ratio,
        "officer_delivery_quality_enforced": enforce_officer_delivery_quality,
        "officer_mapping_check_applicable": enforce_officer_delivery_quality,
        "max_master_coverage": QUALITY_MAX_DORMANCY_RATIO,
        "max_dormancy_ratio": QUALITY_MAX_DORMANCY_RATIO,
        "min_export_coverage": QUALITY_MIN_EXPORT_COVERAGE,
        "dormancy_warning_after_hour": QUALITY_DORMANCY_WARNING_AFTER_HOUR,
        "max_unmapped_ratio": QUALITY_MAX_UNMAPPED_RATIO,
    }


def _should_enforce_officer_delivery_quality(
    *,
    dry_run: bool,
    dispatch_settings: dict[str, bool],
) -> bool:
    """Only make officer mapping blocking when direct delivery can be queued."""
    return bool(
        not dry_run
        and dispatch_settings.get("allow_individual", True)
        and not dispatch_settings.get("manual_only", False)
        and not dispatch_settings.get("require_preview", False)
    )


def _build_run_summary(
    *,
    current_time_obj: datetime.datetime,
    file_path: Path,
    run_output_dir: Path,
    run_unmapped_report_dir: Path,
    excel_path: Path | None,
    mapping_audit_path: Path | None,
    officer_delivery_audit_path: Path | None,
    activity_evidence_path: Path | None,
    extraction_diagnostics: dict,
    quality_gate: dict,
    whatsapp_result: dict | None,
    dry_run: bool,
    lifecycle_events: list[dict] | None = None,
) -> dict:
    return {
        "run_started_at": current_time_obj.isoformat(),
        "dry_run": dry_run,
        "report_source": REPORT_SOURCE,
        "lifecycle": lifecycle_events or [],
        "input": {
            "path": str(file_path),
            "name": file_path.name,
            "source": extraction_diagnostics.get("input_source"),
            "size_bytes": extraction_diagnostics.get("source_file_size_bytes"),
            "sha256": extraction_diagnostics.get("source_file_sha256"),
            "portal_duplicate_raw_check": extraction_diagnostics.get(
                "portal_duplicate_raw_check"
            ),
        },
        "portal_acquisition": extraction_diagnostics.get("portal_acquisition"),
        "filter": {
            "strategy": extraction_diagnostics.get("filter_strategy"),
            "applied": extraction_diagnostics.get("filter_applied"),
            "raw_rows": extraction_diagnostics.get("input_rows"),
            "dormant_raw_rows": extraction_diagnostics.get("output_rows"),
            "available_activity_columns": extraction_diagnostics.get(
                "available_activity_columns"
            ),
            "missing_activity_columns": extraction_diagnostics.get(
                "missing_activity_columns"
            ),
        },
        "quality_gate": quality_gate,
        "outputs": {
            "output_dir": run_output_dir,
            "audit_dir": run_unmapped_report_dir,
            "excel_report": excel_path,
            "officer_mapping_audit": mapping_audit_path,
            "officer_delivery_audit": officer_delivery_audit_path,
            "activity_evidence": activity_evidence_path,
        },
        "whatsapp": whatsapp_result
        or {
            "queued": False,
            "dry_run": dry_run,
            "total_payloads": 0,
            "skipped_reason": "not attempted",
        },
    }


def _summary_run_key(summary: dict) -> str:
    started = str(summary.get("run_started_at") or uuid.uuid4())
    return re.sub(r"[^0-9A-Za-z_.:-]+", "_", started)[:150]


def _derive_pocketbase_run_status(summary: dict) -> str:
    quality_gate = summary.get("quality_gate") or {}
    whatsapp = summary.get("whatsapp") or {}

    if not quality_gate.get("passed", True):
        return "failed"

    if whatsapp.get("error"):
        return "partial"

    if whatsapp.get("skipped_reason") == "stale duplicate portal export":
        return "completed_no_change"

    if whatsapp.get("skipped_reason"):
        return "completed"

    delivery = whatsapp.get("delivery") or {}
    if delivery.get("failed") or delivery.get("timed_out") or delivery.get("missing"):
        return "partial"

    fallback = whatsapp.get("fallback_delivery") or {}
    fallback_delivery = fallback.get("delivery") or {}
    if fallback.get("error") or fallback_delivery.get("failed") or fallback_delivery.get("timed_out"):
        return "partial"

    return "completed"


def _raise_if_summary_not_completed(summary: dict, logger) -> None:
    status = _derive_pocketbase_run_status(summary)
    if status in {"completed", "completed_no_change"}:
        return

    whatsapp = summary.get("whatsapp") or {}
    quality_gate = summary.get("quality_gate") or {}
    message = (
        whatsapp.get("error")
        or whatsapp.get("skipped_reason")
        or "; ".join(str(item) for item in quality_gate.get("errors") or [])
        or f"Run finished with status: {status}"
    )
    logger.error(f"Pipeline finished with status {status}: {message}")
    raise RuntimeError(f"Pipeline finished with status {status}: {message}")


def _new_lifecycle_event(
    stage: str,
    status: str,
    message: str,
    payload: dict | None = None,
) -> dict:
    return {
        "stage": stage,
        "status": status,
        "message": message,
        "payload": payload or {},
        "created_at": datetime.datetime.now().isoformat(timespec="seconds"),
    }


def _append_lifecycle_event(
    lifecycle_events: list[dict],
    stage: str,
    status: str,
    message: str,
    payload: dict | None = None,
) -> None:
    lifecycle_events.append(_new_lifecycle_event(stage, status, message, payload))


def _normalize_plan_status(value: object) -> str:
    status = str(value or "planned").strip().lower().replace(" ", "_")
    aliases = {
        "pending": "planned",
        "not_queued": "skipped",
        "manual_pending": "manual_pending",
        "manual": "manual_pending",
        "preview_required": "preview_required",
        "missing_status": "missing_status",
    }
    status = aliases.get(status, status)
    allowed = {
        "planned",
        "queued",
        "delivered",
        "sent_pending_confirmation",
        "manual_pending",
        "preview_required",
        "skipped",
        "blocked",
        "failed",
        "missing_status",
    }
    return status if status in allowed else "failed"


def _dispatch_plan_item_from_payload(
    payload: dict,
    *,
    channel: str,
    status: str,
    cause: str = "",
    row_count: int | None = None,
    message_mode: str = "",
    route_kind: str = "",
) -> dict:
    route = payload.get("dispatch_route") or {}
    route_kind = str(route_kind or route.get("route_kind") or "").strip().lower()
    message_mode = str(message_mode or route.get("message_mode") or "").strip().lower()
    if row_count is None:
        try:
            row_count = int(route.get("row_count") or 0)
        except (TypeError, ValueError):
            row_count = 0

    return {
        "job_id": str(payload.get("job_id") or ""),
        "channel": channel,
        "recipient_name": str(payload.get("recipient_name") or ""),
        "target": str(payload.get("target") or ""),
        "route_kind": route_kind,
        "message_mode": message_mode,
        "row_count": int(row_count or 0),
        "excel_path": str(payload.get("excel_path") or ""),
        "status": _normalize_plan_status(status),
        "cause": str(cause or ""),
        "payload": payload,
        "created_at": datetime.datetime.now().isoformat(timespec="seconds"),
        "updated_at": datetime.datetime.now().isoformat(timespec="seconds"),
    }


def _build_dispatch_plan_items(
    *,
    dynamic_payloads: list[dict],
    dynamic_payloads_to_queue: list[dict],
    fixed_payloads: list[dict],
    officer_delivery_audit: list[dict],
    delivery_statuses: dict[str, dict] | None = None,
    dry_run: bool = False,
    skipped_status: str = "",
    skipped_reason: str = "",
    fallback_payloads: list[dict] | None = None,
    fallback_delivery_statuses: dict[str, dict] | None = None,
) -> list[dict]:
    delivery_statuses = delivery_statuses or {}
    fallback_payloads = fallback_payloads or []
    fallback_delivery_statuses = fallback_delivery_statuses or {}
    queued_dynamic_job_ids = {
        str(payload.get("job_id"))
        for payload in dynamic_payloads_to_queue
        if payload.get("job_id")
    }
    audit_by_job_id = {
        str(row.get("Job ID") or ""): row
        for row in officer_delivery_audit
        if isinstance(row, dict) and row.get("Job ID")
    }

    items: list[dict] = []
    for payload in dynamic_payloads:
        job_id = str(payload.get("job_id") or "")
        audit_row = audit_by_job_id.get(job_id, {})
        if skipped_status:
            status = skipped_status
            cause = skipped_reason
        elif dry_run:
            status = "planned"
            cause = "Dry run; not queued."
        elif job_id in queued_dynamic_job_ids:
            status = audit_row.get("Delivery Status") or "queued"
            cause = audit_row.get("Delivery Error") or ""
        else:
            status = "skipped"
            cause = "Individual direct delivery was not queued."

        items.append(
            _dispatch_plan_item_from_payload(
                payload,
                channel="individual",
                status=str(status),
                cause=str(cause),
                row_count=int(audit_row.get("School Count") or 0),
                message_mode="individual",
                route_kind="",
            )
        )

    for payload in fixed_payloads:
        job_id = str(payload.get("job_id") or "")
        status_payload = delivery_statuses.get(job_id) or {}
        if dry_run:
            status = "planned"
            cause = "Dry run; not queued."
        elif status_payload:
            status = status_payload.get("status") or "queued"
            cause = status_payload.get("error") or ""
        else:
            status = "queued"
            cause = ""
        items.append(
            _dispatch_plan_item_from_payload(
                payload,
                channel="group",
                status=str(status),
                cause=str(cause),
            )
        )

    for payload in fallback_payloads:
        job_id = str(payload.get("job_id") or "")
        status_payload = fallback_delivery_statuses.get(job_id) or {}
        if dry_run:
            status = "planned"
            cause = "Dry run; not queued."
        elif status_payload:
            status = status_payload.get("status") or "queued"
            cause = status_payload.get("error") or ""
        else:
            status = "queued"
            cause = ""
        items.append(
            _dispatch_plan_item_from_payload(
                payload,
                channel="fallback",
                status=str(status),
                cause=str(cause),
            )
        )

    return items


def _insert_delivery_event(
    conn: sqlite3.Connection,
    *,
    run_key: str,
    job_id: str,
    target: str,
    recipient_name: str,
    role: str,
    status: str,
    cause: str,
    payload: dict,
) -> None:
    conn.execute(
        """
        insert or replace into delivery_events
        (id, run_key, job_id, target, recipient_name, role, status, cause, payload)
        values (:id, :run_key, :job_id, :target, :recipient_name, :role, :status, :cause, :payload)
        """,
        {
            "id": _pocketbase_record_id("delivery", run_key, job_id, target, role),
            "run_key": run_key,
            "job_id": job_id,
            "target": target,
            "recipient_name": recipient_name,
            "role": role,
            "status": (status or "pending").lower(),
            "cause": cause or "",
            "payload": _pocketbase_json(payload),
        },
    )


def _persist_delivery_events(
    conn: sqlite3.Connection, run_key: str, whatsapp_result: dict
) -> None:
    conn.execute("delete from delivery_events where run_key = ?", (run_key,))
    officer_rows = whatsapp_result.get("officer_delivery_audit") or []
    seen_job_ids: set[str] = set()

    for row in officer_rows:
        if not isinstance(row, dict):
            continue
        job_id = str(row.get("Job ID") or "")
        if not job_id:
            continue
        seen_job_ids.add(job_id)
        _insert_delivery_event(
            conn,
            run_key=run_key,
            job_id=job_id,
            target=str(row.get("Target") or row.get("Mobile Number") or ""),
            recipient_name=str(row.get("Officer Name") or ""),
            role=str(row.get("Roles") or ""),
            status=str(row.get("Delivery Status") or "pending"),
            cause=str(row.get("Delivery Error") or ""),
            payload=row,
        )

    delivery = whatsapp_result.get("delivery") or {}
    for job_id, status_payload in (delivery.get("statuses") or {}).items():
        job_id = str(job_id)
        if job_id in seen_job_ids:
            continue
        if not isinstance(status_payload, dict):
            status_payload = {"status": str(status_payload)}
        _insert_delivery_event(
            conn,
            run_key=run_key,
            job_id=job_id,
            target=str(status_payload.get("target") or ""),
            recipient_name=str(status_payload.get("recipient_name") or "Fixed recipient"),
            role="FIXED",
            status=str(status_payload.get("status") or "pending"),
            cause=str(status_payload.get("error") or ""),
            payload=status_payload,
        )

    fallback = whatsapp_result.get("fallback_delivery") or {}
    fallback_delivery = fallback.get("delivery") or {}
    for job_id, status_payload in (fallback_delivery.get("statuses") or {}).items():
        job_id = f"fallback:{job_id}"
        if not isinstance(status_payload, dict):
            status_payload = {"status": str(status_payload)}
        _insert_delivery_event(
            conn,
            run_key=run_key,
            job_id=job_id,
            target=str(status_payload.get("target") or ""),
            recipient_name=str(status_payload.get("recipient_name") or "Fallback recipient"),
            role="FALLBACK",
            status=str(status_payload.get("status") or "pending"),
            cause=str(status_payload.get("error") or fallback.get("error") or ""),
            payload=status_payload,
        )


def _persist_quality_issues(
    conn: sqlite3.Connection, run_key: str, quality_gate: dict
) -> None:
    conn.execute("delete from data_quality_issues where run_key = ?", (run_key,))
    items: list[tuple[str, str]] = [
        ("warning", message) for message in quality_gate.get("warnings") or []
    ]
    items.extend(("error", message) for message in quality_gate.get("errors") or [])

    for index, (severity, message) in enumerate(items):
        conn.execute(
            """
            insert or replace into data_quality_issues
            (id, run_key, severity, category, message, context)
            values (:id, :run_key, :severity, :category, :message, :context)
            """,
            {
                "id": _pocketbase_record_id("quality", run_key, index, severity),
                "run_key": run_key,
                "severity": severity,
                "category": "quality_gate",
                "message": str(message),
                "context": _pocketbase_json(quality_gate),
            },
        )


def _persist_run_stage_events(
    conn: sqlite3.Connection,
    run_key: str,
    lifecycle_events: list[dict],
) -> None:
    if "run_stage_events" not in _pocketbase_table_names(conn):
        return

    conn.execute("delete from run_stage_events where run_key = ?", (run_key,))
    for index, event in enumerate(lifecycle_events or []):
        if not isinstance(event, dict):
            continue
        stage = str(event.get("stage") or "started").strip().lower()
        status = str(event.get("status") or "ok").strip().lower()
        if status not in {"started", "ok", "warning", "blocked", "failed"}:
            status = "warning"
        conn.execute(
            """
            insert or replace into run_stage_events
            (id, run_key, stage, status, message, payload, created_at)
            values (:id, :run_key, :stage, :status, :message, :payload, :created_at)
            """,
            {
                "id": _pocketbase_record_id("stage", run_key, index, stage),
                "run_key": run_key,
                "stage": stage,
                "status": status,
                "message": str(event.get("message") or ""),
                "payload": _pocketbase_json(event.get("payload") or {}),
                "created_at": str(
                    event.get("created_at")
                    or datetime.datetime.now().isoformat(timespec="seconds")
                ),
            },
        )


def _persist_dispatch_plan_items(
    conn: sqlite3.Connection,
    run_key: str,
    dispatch_plan: list[dict],
) -> None:
    if "dispatch_plan_items" not in _pocketbase_table_names(conn):
        return

    columns = _pocketbase_table_columns(conn, "dispatch_plan_items")
    conn.execute("delete from dispatch_plan_items where run_key = ?", (run_key,))
    for index, item in enumerate(dispatch_plan or []):
        if not isinstance(item, dict):
            continue
        record = {
            "id": _pocketbase_record_id(
                "dispatch_plan",
                run_key,
                index,
                item.get("job_id") or "",
                item.get("target") or "",
            ),
            "run_key": run_key,
            "job_id": str(item.get("job_id") or ""),
            "channel": str(item.get("channel") or "group"),
            "recipient_name": str(item.get("recipient_name") or ""),
            "target": str(item.get("target") or ""),
            "route_kind": str(item.get("route_kind") or ""),
            "message_mode": str(item.get("message_mode") or ""),
            "row_count": int(item.get("row_count") or 0),
            "excel_path": str(item.get("excel_path") or ""),
            "status": _normalize_plan_status(item.get("status")),
            "cause": str(item.get("cause") or ""),
            "payload": _pocketbase_json(item.get("payload") or {}),
            "created_at": str(
                item.get("created_at")
                or datetime.datetime.now().isoformat(timespec="seconds")
            ),
            "updated_at": str(
                item.get("updated_at")
                or datetime.datetime.now().isoformat(timespec="seconds")
            ),
        }
        writable = {
            key: value
            for key, value in record.items()
            if key in columns or key == "id"
        }
        conn.execute(
            "insert or replace into dispatch_plan_items "
            f"({', '.join(writable)}) values "
            f"({', '.join(':' + key for key in writable)})",
            writable,
        )


def _persist_dormant_records(
    conn: sqlite3.Connection, run_key: str, dormant_df: pd.DataFrame | None
) -> None:
    if "dormant_records" not in _pocketbase_table_names(conn):
        return

    conn.execute("delete from dormant_records where run_key = ?", (run_key,))
    if dormant_df is None or dormant_df.empty:
        return

    dormant_columns = _pocketbase_table_columns(conn, "dormant_records")
    school_emis_column = _find_column(dormant_df, "School EMIS")
    school_name_column = _find_column(dormant_df, "School Name")
    tehsil_column = _find_column(dormant_df, "Tehsil")
    markaz_column = _find_column(dormant_df, "Markaz")

    school_refs_by_emis: dict[str, str] = {}
    if "school_ref" in dormant_columns and "schools" in _pocketbase_table_names(conn):
        school_refs_by_emis = {
            _normalize_emis(row["emis"]) or "": str(row["id"] or "")
            for row in conn.execute("select id, emis from schools where active = 1").fetchall()
        }

    for index, row in enumerate(dormant_df.to_dict(orient="records")):
        school_emis = (
            _clean_optional_text(row.get(school_emis_column))
            if school_emis_column
            else ""
        )
        school_name = (
            _clean_optional_text(row.get(school_name_column))
            if school_name_column
            else None
        ) or "Unknown School"
        tehsil = (
            _clean_optional_text(row.get(tehsil_column)) if tehsil_column else ""
        ) or ""
        markaz = (
            _clean_optional_text(row.get(markaz_column)) if markaz_column else ""
        ) or ""
        record = {
            "id": _pocketbase_record_id("dormant", run_key, index, school_emis),
            "run_key": run_key,
            "school_ref": school_refs_by_emis.get(_normalize_emis(school_emis) or "", ""),
            "school_emis": school_emis or "",
            "school_name": school_name,
            "tehsil": tehsil,
            "markaz": markaz,
            "row_data": _pocketbase_json(row),
        }
        writable = {
            key: value
            for key, value in record.items()
            if key in dormant_columns or key in {"id"}
        }
        columns = ", ".join(writable)
        placeholders = ", ".join(":" + key for key in writable)
        conn.execute(
            f"insert or replace into dormant_records ({columns}) values ({placeholders})",
            writable,
        )


def _persist_run_summary_to_pocketbase(
    summary: dict,
    dormant_df: pd.DataFrame | None,
    logger=None,
) -> None:
    if not POCKETBASE_ENABLED:
        return

    logger = logger or _get_logger()
    try:
        with _connect_pocketbase(readonly=False) as conn:
            required_tables = {
                "report_runs",
                "delivery_events",
                "data_quality_issues",
            }
            missing_tables = required_tables.difference(_pocketbase_table_names(conn))
            if missing_tables:
                raise RuntimeError(
                    "PocketBase is missing required collections: "
                    + ", ".join(sorted(missing_tables))
                )

            run_key = _summary_run_key(summary)
            quality_gate = summary.get("quality_gate") or {}
            whatsapp_result = summary.get("whatsapp") or {}
            input_info = summary.get("input") or {}
            source = str(input_info.get("source") or "").strip().lower()
            if source not in {"portal", "manual"}:
                source = (
                    "manual"
                    if "drop-manual" in str(input_info.get("path") or "")
                    else "portal"
                )
            conn.execute(
                """
                insert or replace into report_runs
                (id, started_at, finished_at, source, status, raw_file_name, raw_file_sha256, summary)
                values (:id, :started_at, :finished_at, :source, :status, :raw_file_name, :raw_file_sha256, :summary)
                """,
                {
                    "id": _pocketbase_record_id("run", run_key),
                    "started_at": str(summary.get("run_started_at") or ""),
                    "finished_at": datetime.datetime.now().isoformat(),
                    "source": source,
                    "status": _derive_pocketbase_run_status(summary),
                    "raw_file_name": str(input_info.get("name") or ""),
                    "raw_file_sha256": str(input_info.get("sha256") or ""),
                    "summary": _pocketbase_json(summary),
                },
            )
            _persist_dormant_records(conn, run_key, dormant_df)
            _persist_quality_issues(conn, run_key, quality_gate)
            _persist_delivery_events(conn, run_key, whatsapp_result)
            _persist_run_stage_events(conn, run_key, summary.get("lifecycle") or [])
            _persist_dispatch_plan_items(
                conn,
                run_key,
                whatsapp_result.get("dispatch_plan") or [],
            )
        logger.info(f"PocketBase run history updated for {run_key}.")
    except Exception as exc:
        logger.warning(f"Could not persist AntiDengue run to PocketBase: {exc}")


def _save_run_summary(
    summary_path: Path,
    summary: dict,
    dormant_df: pd.DataFrame | None,
    logger=None,
) -> None:
    _write_json_file(summary_path, summary)


def _previous_portal_run_from_runtime_snapshot(
    raw_sha256: str,
    current_started_at: str,
) -> dict | None:
    if not raw_sha256:
        return None
    snapshot = _read_runtime_snapshot()
    if snapshot is None:
        return None
    runs = snapshot.get("previous_portal_runs") or []
    if not isinstance(runs, list):
        raise RuntimeError("Runtime snapshot previous_portal_runs must be a list.")
    successful_statuses = {
        "completed",
        "completed_no_change",
        "completed_with_delivery_errors",
    }
    matches = [
        row
        for row in runs
        if isinstance(row, dict)
        and str(row.get("raw_file_sha256") or "") == raw_sha256
        and str(row.get("started_at") or "") < current_started_at
        and str(row.get("status") or "").strip().lower() in successful_statuses
    ]
    return sorted(
        matches,
        key=lambda row: str(row.get("started_at") or ""),
        reverse=True,
    )[0] if matches else None


def _previous_portal_run_from_pocketbase(
    raw_sha256: str,
    current_started_at: str,
) -> dict | None:
    if not POCKETBASE_ENABLED or not raw_sha256:
        return None

    try:
        with _connect_pocketbase(readonly=True) as conn:
            table_names = _pocketbase_table_names(conn)
            if "report_runs" not in table_names:
                return None
            columns = _pocketbase_table_columns(conn, "report_runs")
            required_columns = {"started_at", "source", "status", "raw_file_sha256"}
            if not required_columns.issubset(columns):
                return None

            row = conn.execute(
                """
                select started_at, status, raw_file_name, raw_file_sha256
                  from report_runs
                 where source = 'portal'
                   and raw_file_sha256 = :sha
                   and started_at < :current_started_at
                 order by started_at desc
                 limit 1
                """,
                {
                    "sha": raw_sha256,
                    "current_started_at": current_started_at,
                },
            ).fetchone()
            return dict(row) if row else None
    except Exception:
        return None


def _portal_report_signature(portal_acquisition: dict | None) -> dict:
    """Build a stable signature for the complete requested portal report bundle."""

    captures = (portal_acquisition or {}).get("reports") or []
    reports: list[dict[str, str]] = []
    for capture in captures:
        if not isinstance(capture, dict):
            continue
        report_key = str(capture.get("report_key") or "").strip()
        if not report_key:
            continue
        reports.append(
            {
                "report_key": report_key,
                "status": str(capture.get("status") or "").strip().lower(),
                "sha256": str(capture.get("sha256") or "").strip().lower(),
            }
        )
    reports.sort(key=lambda item: item["report_key"])
    complete = bool(reports) and all(
        item["status"] == "completed" and item["sha256"] for item in reports
    )
    canonical = json.dumps(reports, sort_keys=True, separators=(",", ":"))
    return {
        "complete": complete,
        "fingerprint": hashlib.sha256(canonical.encode("utf-8")).hexdigest()
        if complete
        else "",
        "reports": reports,
        "report_keys": [item["report_key"] for item in reports],
    }


def _portal_report_changes(current: dict, previous: dict) -> dict:
    current_by_key = {item["report_key"]: item for item in current.get("reports") or []}
    previous_by_key = {item["report_key"]: item for item in previous.get("reports") or []}
    all_keys = sorted(set(current_by_key) | set(previous_by_key))
    unchanged = [key for key in all_keys if current_by_key.get(key) == previous_by_key.get(key)]
    changed = [key for key in all_keys if current_by_key.get(key) != previous_by_key.get(key)]
    return {"unchanged_report_keys": unchanged, "changed_report_keys": changed}


def _previous_portal_run_from_local_summaries(
    raw_sha256: str,
    current_started_at: str,
) -> dict | None:
    if not raw_sha256:
        return None

    matches: list[dict] = []
    for summary_path in OUTPUT_DIR.glob("*/run_summary.json"):
        try:
            summary = json.loads(summary_path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            continue

        started_at = str(summary.get("run_started_at") or "")
        if not started_at or started_at >= current_started_at:
            continue

        input_info = summary.get("input") or {}
        if str(input_info.get("sha256") or "") != raw_sha256:
            continue

        if str(summary.get("report_source") or "").strip().lower() not in {
            "unfiltered",
            "filtered",
            "dormant",
            "all",
        }:
            continue

        previous_status = _derive_pocketbase_run_status(summary)
        whatsapp = summary.get("whatsapp") or {}
        if previous_status not in {"completed", "completed_no_change"}:
            continue
        # Platform Test Run/preview source jobs intentionally execute the legacy
        # extractor with dry_run=True. A preview must never consume the future
        # real Send Now/scheduled delivery by being mistaken for a prior send.
        if bool(whatsapp.get("dry_run")) and previous_status != "completed_no_change":
            continue

        matches.append(
            {
                "started_at": started_at,
                "status": previous_status,
                "raw_file_name": str(input_info.get("name") or ""),
                "raw_file_sha256": raw_sha256,
                "summary_path": str(summary_path),
                "portal_acquisition": summary.get("portal_acquisition") or {},
                "portal_report_signature": _portal_report_signature(
                    summary.get("portal_acquisition") or {}
                ),
            }
        )

    return sorted(matches, key=lambda row: row["started_at"], reverse=True)[0] if matches else None


def _detect_duplicate_portal_export(
    *,
    extraction_diagnostics: dict,
    current_time_obj: datetime.datetime,
    input_source: str,
    portal_acquisition: dict | None = None,
) -> dict:
    raw_sha256 = str(extraction_diagnostics.get("source_file_sha256") or "")
    current_started_at = current_time_obj.isoformat()
    current_bundle = _portal_report_signature(portal_acquisition)
    result = {
        "duplicate": False,
        "primary_duplicate": False,
        "duplicate_scope": "none",
        "policy": PORTAL_DUPLICATE_RAW_POLICY,
        "current_sha256": raw_sha256,
        "current_report_bundle": current_bundle,
    }
    if input_source != "portal" or not raw_sha256:
        return result

    local_previous = _previous_portal_run_from_local_summaries(
        raw_sha256,
        current_started_at,
    )
    runtime_previous = _previous_portal_run_from_runtime_snapshot(
        raw_sha256,
        current_started_at,
    )
    candidates = [item for item in (local_previous, runtime_previous) if item]
    if not candidates:
        return result
    previous = max(candidates, key=lambda item: str(item.get("started_at") or ""))

    result.update(
        {
            "primary_duplicate": True,
            "previous_run_started_at": str(previous.get("started_at") or ""),
            "previous_status": str(previous.get("status") or ""),
            "previous_raw_file_name": str(previous.get("raw_file_name") or ""),
            "previous_summary_path": str(previous.get("summary_path") or ""),
        }
    )

    previous_bundle = dict(
        previous.get("portal_report_signature")
        or _portal_report_signature(previous.get("portal_acquisition") or {})
    )
    if current_bundle.get("complete") and previous_bundle.get("complete"):
        changes = _portal_report_changes(current_bundle, previous_bundle)
        result.update(changes)
        if current_bundle.get("fingerprint") == previous_bundle.get("fingerprint"):
            result.update(
                {
                    "duplicate": True,
                    "duplicate_scope": "complete_report_bundle",
                    "previous_report_bundle": previous_bundle,
                }
            )
        else:
            result["duplicate_scope"] = "primary_report_only"
        return result

    # An incomplete current report bundle must never be converted into a
    # successful no-change run. Let the normal acquisition/quality gate expose
    # the missing or failed component. The legacy primary-only behaviour is
    # retained only when no component-level capture metadata exists at all.
    if current_bundle.get("reports"):
        result["duplicate_scope"] = "incomplete_report_bundle"
    elif not current_bundle.get("complete"):
        result.update({"duplicate": True, "duplicate_scope": "legacy_primary_report"})
    else:
        result["duplicate_scope"] = "primary_report_only"
    return result


def _duplicate_portal_export_warning(stale_check: dict) -> str:
    previous = stale_check.get("previous_run_started_at") or "an earlier run"
    sha = stale_check.get("current_sha256") or "-"
    if stale_check.get("duplicate"):
        return (
            "Portal returned the exact same complete report bundle as "
            f"{previous} (primary sha256={sha}). "
            "The run is completed as no-change and duplicate preview/WhatsApp "
            "delivery is skipped safely."
        )
    changed = ", ".join(stale_check.get("changed_report_keys") or []) or "other reports"
    return (
        "The dormant-users export matched "
        f"{previous} (sha256={sha}), but {changed} changed. "
        "The changed distance/timing evidence will continue through consolidated "
        "preview generation instead of being incorrectly blocked."
    )


def _process_raw_report_file(
    file_path: Path,
    logger,
    dry_run: bool = False,
    *,
    input_source: str = "portal",
    portal_acquisition: dict | None = None,
) -> dict:
    # Step 1: Extract & Transform
    raw_df, filtered_user_df, extraction_diagnostics = (
        _extract_raw_data_with_diagnostics(file_path, logger)
    )
    clean_df = transform_data(raw_df)

    # Step 2: Setup dynamic Output Folders & Filenames
    current_time_obj = datetime.datetime.now()
    extraction_diagnostics["input_source"] = input_source
    extraction_diagnostics["portal_acquisition"] = portal_acquisition
    run_output_dir, run_unmapped_report_dir = _create_run_directories(current_time_obj)
    if portal_acquisition:
        for capture in portal_acquisition.get("reports") or []:
            if capture.get("status") != "completed":
                continue
            source_path = Path(str(capture.get("path") or ""))
            if not source_path.is_file():
                continue
            report_key = str(capture.get("report_key") or "")
            if report_key == "hotspot_distance":
                label = "Hotspot Distance Review"
                writer = write_hotspot_review_report
            elif report_key == "simple_activity_list":
                label = "Simple Activity Timing Review"
                writer = write_simple_activity_review_report
            else:
                continue
            review_path = run_output_dir / f"{label} - {current_time_obj.strftime('%d-%m-%Y - %I-%M %p')}.xlsx"
            review = writer(source_path, review_path)
            capture["review_report"] = review
            analysis = capture.setdefault("analysis", {})
            analysis["review_candidate_count"] = review["candidate_count"]
            analysis["review_report_path"] = review["path"]
            analysis["review_selection_mode"] = review["selection_mode"]
            logger.info(
                f"Generated {label} report with "
                f"{review['candidate_count']} classified activity row(s)."
            )
    lifecycle_events: list[dict] = []
    _append_lifecycle_event(
        lifecycle_events,
        "started",
        "started",
        "Run started.",
        {
            "input_source": input_source,
            "dry_run": dry_run,
            "file": str(file_path),
        },
    )
    _append_lifecycle_event(
        lifecycle_events,
        "downloaded",
        "ok",
        f"Raw report loaded with {extraction_diagnostics.get('input_rows', len(raw_df))} row(s).",
        {
            "raw_rows": extraction_diagnostics.get("input_rows"),
            "raw_file_sha256": extraction_diagnostics.get("source_file_sha256"),
            "raw_file_size_bytes": extraction_diagnostics.get("source_file_size_bytes"),
        },
    )
    _append_lifecycle_event(
        lifecycle_events,
        "processed",
        "ok",
        f"Dormant filter produced {len(clean_df)} row(s).",
        {
            "dormant_rows": len(clean_df),
            "filter_strategy": extraction_diagnostics.get("filter_strategy"),
            "filter_applied": extraction_diagnostics.get("filter_applied"),
        },
    )

    title_text = f"Anti-Dengue Dormant Users - {current_time_obj.strftime('%d-%b-%Y - %I:%M %p')}"

    date_str = current_time_obj.strftime("%d-%m-%Y")
    time_str = current_time_obj.strftime("%I-%M %p")
    excel_filename = f"Anti-Dengue App Dormant Users - {date_str} - {time_str}.xlsx"

    excel_path = run_output_dir / excel_filename
    image_path = run_output_dir / "Inactive_Schools_Report_Screenshot.png"
    summary_path = run_output_dir / "run_summary.json"
    activity_evidence_path = (
        run_output_dir / f"Activity Evidence - {date_str} - {time_str}.xlsx"
    )
    mapping_audit_path = (
        run_unmapped_report_dir
        / f"Officer Mapping Audit - {date_str} - {time_str}.xlsx"
    )
    officer_delivery_audit_path = (
        run_output_dir / f"Officer Delivery Audit - {date_str} - {time_str}.xlsx"
    )
    generated_image_path = None

    unmatched_df = pd.DataFrame()
    invalid_contacts_df = pd.DataFrame()

    # Step 3: Check if zero schools left (Graceful exit)
    if len(clean_df) == 0:
        logger.info("TARGET REACHED: ZERO INACTIVE SCHOOLS LEFT! Stopping this file.")
        # Preserve an explicit, schema-valid empty canonical workbook. Native
        # preview renderers must never infer dormant schools from another
        # report category merely because it also contains School EMIS fields.
        load_excel_report(clean_df, excel_path, title_text)
        _append_lifecycle_event(
            lifecycle_events,
            "validated",
            "ok",
            "No dormant schools found after filtering.",
            {"dormant_rows": 0, "excel_report": str(excel_path)},
        )
        _append_lifecycle_event(
            lifecycle_events,
            "dispatch_blocked",
            "ok",
            "Dispatch skipped because there are zero inactive schools.",
        )
        _append_lifecycle_event(
            lifecycle_events,
            "summary_saved",
            "ok",
            f"Run summary prepared at {summary_path}.",
        )
        _append_lifecycle_event(
            lifecycle_events,
            "completed",
            "ok",
            "Run completed with zero inactive schools.",
        )
        quality_gate = {
            "passed": True,
            "errors": [],
            "warnings": [],
            "master_count": _master_school_count(),
            "final_school_count": 0,
            "master_coverage_ratio": 0,
            "unmapped_school_count": 0,
            "unmapped_ratio": 0,
            "invalid_contact_count": 0,
            "invalid_contact_ratio": 0,
        }
        summary = _build_run_summary(
            current_time_obj=current_time_obj,
            file_path=file_path,
            run_output_dir=run_output_dir,
            run_unmapped_report_dir=run_unmapped_report_dir,
            excel_path=excel_path,
            mapping_audit_path=None,
            officer_delivery_audit_path=None,
            activity_evidence_path=None,
            extraction_diagnostics=extraction_diagnostics,
            quality_gate=quality_gate,
            whatsapp_result={
                "queued": False,
                "dry_run": dry_run,
                "total_payloads": 0,
                "skipped_reason": "zero inactive schools",
            },
            dry_run=dry_run,
            lifecycle_events=lifecycle_events,
        )
        _save_run_summary(summary_path, summary, clean_df, logger)
        if not dry_run:
            archive_raw_file(file_path)
        return summary

    # Step 4: Load outputs and audits
    load_excel_report(clean_df, excel_path, title_text)
    unmatched_df, invalid_contacts_df, _ = _build_officer_mapping_audit(clean_df)
    load_officer_mapping_audit(clean_df, mapping_audit_path, title_text)
    load_activity_evidence_report(clean_df, filtered_user_df, activity_evidence_path)
    _append_lifecycle_event(
        lifecycle_events,
        "processed",
        "ok",
        "Generated Excel report, officer mapping audit, and activity evidence.",
        {
            "excel_report": str(excel_path),
            "officer_mapping_audit": str(mapping_audit_path),
            "activity_evidence": str(activity_evidence_path),
            "unmapped_schools": len(unmatched_df),
            "invalid_contacts": len(invalid_contacts_df),
        },
    )

    if GENERATE_SCREENSHOT:
        try:
            load_screenshot(clean_df, image_path, title_text)
            generated_image_path = image_path
        except Exception as e:
            logger.error(f"Screenshot failed, but pipeline will continue. Error: {e}")

    # Step 5: Validate before sending
    dispatch_settings = _read_dispatch_settings()
    enforce_officer_delivery_quality = _should_enforce_officer_delivery_quality(
        dry_run=dry_run,
        dispatch_settings=dispatch_settings,
    )
    quality_gate = _quality_gate(
        final_df=clean_df,
        extraction_diagnostics=extraction_diagnostics,
        unmatched_df=unmatched_df,
        invalid_contacts_df=invalid_contacts_df,
        enforce_officer_delivery_quality=enforce_officer_delivery_quality,
        observed_at=current_time_obj,
    )
    for warning in quality_gate["warnings"]:
        logger.warning(f"Data quality warning: {warning}")

    stale_check = _detect_duplicate_portal_export(
        extraction_diagnostics=extraction_diagnostics,
        current_time_obj=current_time_obj,
        input_source=input_source,
        portal_acquisition=portal_acquisition,
    )
    extraction_diagnostics["portal_duplicate_raw_check"] = stale_check
    if stale_check.get("duplicate") or stale_check.get("primary_duplicate"):
        warning = _duplicate_portal_export_warning(stale_check)
        if warning not in quality_gate["warnings"]:
            quality_gate["warnings"].append(warning)
            quality_gate.setdefault("issues", []).append(
                {
                    "code": (
                        "duplicate_portal_report_bundle"
                        if stale_check.get("duplicate")
                        else "unchanged_dormant_export_with_changed_components"
                    ),
                    "severity": "warning",
                    "message": warning,
                }
            )
        logger.warning(warning)

    if not quality_gate["passed"]:
        _append_lifecycle_event(
            lifecycle_events,
            "validated",
            "failed",
            "Data quality gate failed.",
            {
                "errors": quality_gate.get("errors") or [],
                "warnings": quality_gate.get("warnings") or [],
            },
        )
        for error in quality_gate["errors"]:
            logger.error(f"Data quality error: {error}")
        whatsapp_result = {
            "queued": False,
            "dry_run": dry_run,
            "total_payloads": 0,
            "skipped_reason": "data quality gate failed",
        }
        _append_lifecycle_event(
            lifecycle_events,
            "dispatch_blocked",
            "blocked",
            "Dispatch blocked because the data quality gate failed.",
        )
        _append_lifecycle_event(
            lifecycle_events,
            "summary_saved",
            "ok",
            f"Run summary prepared at {summary_path}.",
        )
        _append_lifecycle_event(
            lifecycle_events,
            "failed",
            "failed",
            "Run ended before dispatch because validation failed.",
        )
        summary = _build_run_summary(
            current_time_obj=current_time_obj,
            file_path=file_path,
            run_output_dir=run_output_dir,
            run_unmapped_report_dir=run_unmapped_report_dir,
            excel_path=excel_path,
            mapping_audit_path=mapping_audit_path,
            officer_delivery_audit_path=None,
            activity_evidence_path=activity_evidence_path,
            extraction_diagnostics=extraction_diagnostics,
            quality_gate=quality_gate,
            whatsapp_result=whatsapp_result,
            dry_run=dry_run,
            lifecycle_events=lifecycle_events,
        )
        _save_run_summary(summary_path, summary, clean_df, logger)
        logger.error(f"WhatsApp sending blocked. Run summary saved to {summary_path}.")
        return summary

    _append_lifecycle_event(
        lifecycle_events,
        "validated",
        "warning" if quality_gate.get("warnings") else "ok",
        "Data quality gate passed.",
        {
            "warnings": quality_gate.get("warnings") or [],
            "dormant_rows": len(clean_df),
            "duplicate_portal_export": bool(stale_check.get("duplicate")),
        },
    )

    if (
        input_source == "portal"
        and stale_check.get("duplicate")
        and PORTAL_DUPLICATE_RAW_POLICY == "block_send"
    ):
        whatsapp_result = {
            "queued": False,
            "dry_run": dry_run,
            "total_payloads": 0,
            "skipped_reason": "stale duplicate portal export",
            "outcome": "no_change",
            "no_change": True,
            "stale_duplicate": stale_check,
        }
        _append_lifecycle_event(
            lifecycle_events,
            "dispatch_blocked",
            "blocked",
            "Dispatch blocked because the portal returned a duplicate raw export.",
            stale_check,
        )
        _append_lifecycle_event(
            lifecycle_events,
            "summary_saved",
            "ok",
            f"Run summary prepared at {summary_path}.",
        )
        _append_lifecycle_event(
            lifecycle_events,
            "completed_no_change",
            "ok",
            "Run completed safely with no dispatch because the portal export was unchanged.",
            {"status": "completed_no_change", "dispatch_prevented": True},
        )
        summary = _build_run_summary(
            current_time_obj=current_time_obj,
            file_path=file_path,
            run_output_dir=run_output_dir,
            run_unmapped_report_dir=run_unmapped_report_dir,
            excel_path=excel_path,
            mapping_audit_path=mapping_audit_path,
            officer_delivery_audit_path=None,
            activity_evidence_path=activity_evidence_path,
            extraction_diagnostics=extraction_diagnostics,
            quality_gate=quality_gate,
            whatsapp_result=whatsapp_result,
            dry_run=dry_run,
            lifecycle_events=lifecycle_events,
        )
        _save_run_summary(summary_path, summary, clean_df, logger)
        logger.info(
            "Portal report bundle was unchanged. Duplicate WhatsApp dispatch was prevented "
            f"and the run completed as no-change. Run summary saved to {summary_path}."
        )
        return summary

    # Step 6: Notify and cleanup
    body_text = create_whatsapp_summary(clean_df, title_text)
    whatsapp_result = None
    _append_lifecycle_event(
        lifecycle_events,
        "dispatch_started",
        "started",
        "WhatsApp dispatch started.",
        {"dormant_rows": len(clean_df), "dry_run": dry_run},
    )

    pre_dispatch_result = {
        "queued": False,
        "dry_run": dry_run,
        "total_payloads": 0,
        "dispatch_state": "started",
        "error": "Dispatch started but final WhatsApp status has not been recorded yet.",
    }
    pre_dispatch_summary = _build_run_summary(
        current_time_obj=current_time_obj,
        file_path=file_path,
        run_output_dir=run_output_dir,
        run_unmapped_report_dir=run_unmapped_report_dir,
        excel_path=excel_path,
        mapping_audit_path=mapping_audit_path,
        officer_delivery_audit_path=None,
        activity_evidence_path=activity_evidence_path,
        extraction_diagnostics=extraction_diagnostics,
        quality_gate=quality_gate,
        whatsapp_result=pre_dispatch_result,
        dry_run=dry_run,
        lifecycle_events=lifecycle_events,
    )
    _save_run_summary(summary_path, pre_dispatch_summary, clean_df, logger)

    try:
        whatsapp_result = send_whatsapp_alert(
            clean_df,
            excel_path,
            generated_image_path,
            title_text,
            body_text,
            dry_run,
        )
    except Exception as e:
        logger.error(f"CRITICAL: Failed to push to NATS. Error: {e}")
        whatsapp_result = {
            "queued": False,
            "dry_run": dry_run,
            "total_payloads": 0,
            "error": str(e),
        }
    dispatch_status = "ok"
    dispatch_message = "WhatsApp dispatch completed."
    if whatsapp_result and whatsapp_result.get("error"):
        dispatch_status = "failed"
        dispatch_message = str(whatsapp_result.get("error") or dispatch_message)
    elif whatsapp_result and whatsapp_result.get("skipped_reason"):
        dispatch_status = "blocked"
        dispatch_message = str(whatsapp_result.get("skipped_reason") or dispatch_message)
    elif not (whatsapp_result or {}).get("queued"):
        dispatch_status = "warning"
        dispatch_message = "WhatsApp dispatch finished without queueing messages."
    _append_lifecycle_event(
        lifecycle_events,
        "dispatch_finished",
        dispatch_status,
        dispatch_message,
        {
            "total_payloads": (whatsapp_result or {}).get("total_payloads"),
            "queued": (whatsapp_result or {}).get("queued"),
            "dynamic_payloads": (whatsapp_result or {}).get("dynamic_payloads"),
            "queued_dynamic_payloads": (whatsapp_result or {}).get("queued_dynamic_payloads"),
            "fixed_payloads": (whatsapp_result or {}).get("fixed_payloads"),
        },
    )

    officer_delivery_audit_rows = (
        whatsapp_result.get("officer_delivery_audit") if whatsapp_result else None
    )
    if officer_delivery_audit_rows is not None:
        load_officer_delivery_audit(
            officer_delivery_audit_rows,
            officer_delivery_audit_path,
        )

    projected_summary = _build_run_summary(
        current_time_obj=current_time_obj,
        file_path=file_path,
        run_output_dir=run_output_dir,
        run_unmapped_report_dir=run_unmapped_report_dir,
        excel_path=excel_path,
        mapping_audit_path=mapping_audit_path,
        officer_delivery_audit_path=(
            officer_delivery_audit_path
            if officer_delivery_audit_rows is not None
            else None
        ),
        activity_evidence_path=activity_evidence_path,
        extraction_diagnostics=extraction_diagnostics,
        quality_gate=quality_gate,
        whatsapp_result=whatsapp_result,
        dry_run=dry_run,
        lifecycle_events=lifecycle_events,
    )
    final_status = _derive_pocketbase_run_status(projected_summary)
    _append_lifecycle_event(
        lifecycle_events,
        "summary_saved",
        "ok",
        f"Run summary prepared at {summary_path}.",
    )
    successful_statuses = {"completed", "completed_no_change"}
    _append_lifecycle_event(
        lifecycle_events,
        final_status if final_status in successful_statuses else "failed",
        "ok" if final_status in successful_statuses else "warning",
        f"Run finished with status: {final_status}.",
        {"status": final_status},
    )
    summary = _build_run_summary(
        current_time_obj=current_time_obj,
        file_path=file_path,
        run_output_dir=run_output_dir,
        run_unmapped_report_dir=run_unmapped_report_dir,
        excel_path=excel_path,
        mapping_audit_path=mapping_audit_path,
        officer_delivery_audit_path=(
            officer_delivery_audit_path
            if officer_delivery_audit_rows is not None
            else None
        ),
        activity_evidence_path=activity_evidence_path,
        extraction_diagnostics=extraction_diagnostics,
        quality_gate=quality_gate,
        whatsapp_result=whatsapp_result,
        dry_run=dry_run,
        lifecycle_events=lifecycle_events,
    )
    _save_run_summary(summary_path, summary, clean_df, logger)

    if not dry_run and whatsapp_result and whatsapp_result.get("queued"):
        archive_raw_file(file_path)

    return summary


def _iter_manual_unfiltered_files() -> list[Path]:
    files: list[Path] = []
    for manual_dir in MANUAL_UNFILTERED_DIR_ALIASES:
        if not manual_dir.exists():
            continue

        files.extend(
            path
            for path in manual_dir.iterdir()
            if path.is_file()
            and not path.name.startswith(".")
            and not path.name.startswith("~$")
        )

    return sorted(files, key=lambda path: (path.stat().st_mtime, path.name))


def _stage_manual_selected_file(source_file: Path) -> Path:
    source_file = source_file.expanduser().resolve(strict=False)
    if not source_file.exists():
        raise FileNotFoundError(f"Manual report file not found: {source_file}")
    if not source_file.is_file():
        raise ValueError(f"Manual report path is not a file: {source_file}")
    if source_file.stat().st_size == 0:
        raise ValueError(f"Manual report file is empty: {source_file}")

    MANUAL_SELECTED_DIR.mkdir(parents=True, exist_ok=True)
    safe_stem = re.sub(r"[^A-Za-z0-9._-]+", "_", source_file.stem).strip("._-")
    if not safe_stem:
        safe_stem = "manual_report"
    suffix = source_file.suffix or ".xls"
    staged_path = (
        MANUAL_SELECTED_DIR
        / f"{safe_stem}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:8]}{suffix}"
    )
    shutil.copy2(source_file, staged_path)
    return staged_path


# ==========================================
# 3. PIPELINE ENTRY POINTS
# ==========================================
def process_file_flow(dry_run: bool = False):
    logger = _get_logger()
    logger.info(f"Starting Hourly Pipeline... dry_run={dry_run}")

    # Step 1: Scrape the data
    portal_bundle = scrape_portal_reports(logger=logger)
    summary = _process_raw_report_file(
        portal_bundle.primary_path,
        logger,
        dry_run=dry_run,
        portal_acquisition=portal_bundle.as_dict(),
    )
    _raise_if_summary_not_completed(summary, logger)
    logger.info("Pipeline completed successfully.")


def process_manual_unfiltered_flow(dry_run: bool = False):
    logger = _get_logger()
    logger.info(
        "Starting manual unfiltered pipeline from: "
        + ", ".join(str(path) for path in MANUAL_UNFILTERED_DIR_ALIASES)
        + f". dry_run={dry_run}"
    )

    manual_files = _iter_manual_unfiltered_files()
    if not manual_files:
        logger.info("No manual unfiltered files found to process.")
        return

    processed_count = 0
    failed_count = 0

    for file_path in manual_files:
        logger.info(f"Processing manual unfiltered file: {file_path.name}")
        try:
            summary = _process_raw_report_file(
                file_path,
                logger,
                dry_run=dry_run,
                input_source="manual",
            )
            _raise_if_summary_not_completed(summary, logger)
            processed_count += 1
        except Exception as exc:
            failed_count += 1
            logger.error(
                f"Failed to process manual file {file_path}. "
                "The file was left in place for review. Error: "
                f"{exc}"
            )

    logger.info(
        "Manual unfiltered pipeline completed: "
        f"{processed_count} processed, {failed_count} failed."
    )


def process_selected_manual_file_flow(source_file: str, dry_run: bool = False):
    logger = _get_logger()
    logger.info(
        f"Starting selected manual report pipeline from {source_file}. "
        f"dry_run={dry_run}"
    )
    staged_path = _stage_manual_selected_file(Path(source_file))
    logger.info(f"Manual report copied to staging file: {staged_path}")
    summary = _process_raw_report_file(
        staged_path,
        logger,
        dry_run=dry_run,
        input_source="manual",
    )
    _raise_if_summary_not_completed(summary, logger)
    logger.info("Selected manual report pipeline completed successfully.")


if __name__ == "__main__":
    _configure_cli_logging()
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    ARCHIVE_DIR.mkdir(parents=True, exist_ok=True)
    UNMAPPED_REPORT_DIR.mkdir(parents=True, exist_ok=True)
    MANUAL_SELECTED_DIR.mkdir(parents=True, exist_ok=True)
    for manual_dir in MANUAL_UNFILTERED_DIR_ALIASES:
        manual_dir.mkdir(parents=True, exist_ok=True)

    try:
        _validate_runtime_config()
    except Exception as exc:
        print(f"CRITICAL: {exc}")
        exit(1)

    raw_args = sys.argv[1:]
    normalized_args = [argument.strip().lower() for argument in raw_args]
    dry_run = "--dry-run" in normalized_args or "dry-run" in normalized_args
    args = [
        argument
        for argument in raw_args
        if argument.strip().lower() not in {"--dry-run", "dry-run"}
    ]
    command = args[0].strip().lower() if args else "portal"
    if command in {"portal", "run", "download"}:
        process_file_flow(dry_run=dry_run)
    elif command in {"manual-unfiltered", "manual-unfilterd", "manual"}:
        process_manual_unfiltered_flow(dry_run=dry_run)
    elif command in {"manual-file", "selected-file", "manual-selected"}:
        source_file = ""
        remaining_args = args[1:]
        if remaining_args and remaining_args[0] in {"--file", "-f"}:
            source_file = remaining_args[1] if len(remaining_args) > 1 else ""
        elif remaining_args:
            source_file = remaining_args[0]

        if not source_file:
            print("Missing manual report file. Use: python main.py manual-file --file <path>")
            exit(2)
        process_selected_manual_file_flow(source_file=source_file, dry_run=dry_run)
    else:
        print(
            "Unknown command. Use 'python main.py' for portal download or "
            "'python main.py manual-unfiltered [--dry-run]' for files in "
            f"{MANUAL_UNFILTERED_DIR_ALIASES[1].name}, or "
            "'python main.py manual-file --file <path> [--dry-run]'."
        )
        exit(2)
