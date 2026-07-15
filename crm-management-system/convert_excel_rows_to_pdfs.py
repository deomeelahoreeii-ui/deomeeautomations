from __future__ import annotations

import argparse
import csv
import logging
import re
import sys
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from main import configure_logging

LOGGER_NAME = "pmdu_automation"
DEFAULT_INPUT_DIR = "phase1-crm/unprocessed-crm/sheets"
DEFAULT_OUTPUT_DIR = "crm-main-complaints"
COMPLAINT_NO_RE = re.compile(r"\b\d{3}-\d{7}\b")
EXCEL_SUFFIXES = {".xlsx", ".xlsm", ".xltx", ".xltm"}
CSV_SUFFIXES = {".csv"}
SUPPORTED_SUFFIXES = EXCEL_SUFFIXES | CSV_SUFFIXES


def resolve_project_path(project_root: Path, value: str) -> Path:
    path = Path(value).expanduser()
    if not path.is_absolute():
        path = project_root / path
    return path.resolve(strict=False)


def find_input_files(directory: Path) -> Iterable[Path]:
    if directory.is_file() and directory.suffix.lower() in SUPPORTED_SUFFIXES:
        yield directory
        return

    for path in sorted(directory.iterdir()):
        if path.name.startswith(("~$", ".~lock.")):
            continue
        if path.is_file() and path.suffix.lower() in SUPPORTED_SUFFIXES:
            yield path


def convert_input_file(input_path: Path, output_dir: Path, overwrite: bool) -> int:
    if input_path.suffix.lower() in CSV_SUFFIXES:
        return convert_csv(input_path, output_dir, overwrite)
    return convert_workbook(input_path, output_dir, overwrite)


def convert_workbook(workbook_path: Path, output_dir: Path, overwrite: bool) -> int:
    logger = logging.getLogger(LOGGER_NAME)
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    created = 0
    skipped_existing = 0

    try:
        for worksheet in workbook.worksheets:
            header_row_index, headers = find_header_row(
                worksheet.iter_rows(values_only=True)
            )
            if header_row_index is None:
                continue

            complaint_index = find_complaint_column(headers)

            for row_index, row in enumerate(
                worksheet.iter_rows(min_row=header_row_index + 1, values_only=True),
                start=header_row_index + 1,
            ):
                if row_is_empty(row):
                    continue

                row_data = build_row_data(headers, row)
                complaint_number = get_complaint_number(row_data, row, complaint_index)
                if not complaint_number:
                    logger.warning(
                        "Skipping %s / %s row %s: no complaint number found.",
                        workbook_path.name,
                        worksheet.title,
                        row_index,
                    )
                    continue

                output_path = output_dir / f"{complaint_number}.pdf"
                if output_path.exists() and not overwrite:
                    skipped_existing += 1
                    continue

                write_pdf(
                    output_path=output_path,
                    complaint_number=complaint_number,
                    workbook_name=workbook_path.name,
                    sheet_name=worksheet.title,
                    row_index=row_index,
                    row_data=row_data,
                )
                created += 1
    finally:
        workbook.close()

    if skipped_existing:
        logger.info(
            "%s: skipped %d existing PDF(s). Use --overwrite to replace them.",
            workbook_path.name,
            skipped_existing,
        )
    return created


def convert_csv(csv_path: Path, output_dir: Path, overwrite: bool) -> int:
    logger = logging.getLogger(LOGGER_NAME)
    rows = read_csv_rows(csv_path)
    header_row_index, headers = find_header_row(rows)
    if header_row_index is None:
        logger.warning("Skipping %s: no complaint number header found.", csv_path.name)
        return 0

    complaint_index = find_complaint_column(headers)
    created = 0
    skipped_existing = 0

    for row_index, row in enumerate(
        rows[header_row_index:],
        start=header_row_index + 1,
    ):
        if row_is_empty(row):
            continue

        row_data = build_row_data(headers, row)
        complaint_number = get_complaint_number(row_data, row, complaint_index)
        if not complaint_number:
            logger.warning(
                "Skipping %s row %s: no complaint number found.",
                csv_path.name,
                row_index,
            )
            continue

        output_path = output_dir / f"{complaint_number}.pdf"
        if output_path.exists() and not overwrite:
            skipped_existing += 1
            continue

        write_pdf(
            output_path=output_path,
            complaint_number=complaint_number,
            workbook_name=csv_path.name,
            sheet_name="CSV",
            row_index=row_index,
            row_data=row_data,
        )
        created += 1

    if skipped_existing:
        logger.info(
            "%s: skipped %d existing PDF(s). Use --overwrite to replace them.",
            csv_path.name,
            skipped_existing,
        )
    return created


def read_csv_rows(csv_path: Path) -> list[tuple[str, ...]]:
    with csv_path.open("r", encoding="utf-8-sig", newline="") as handle:
        return [tuple(row) for row in csv.reader(handle)]


def find_header_row(rows: Iterable[tuple[object, ...]]) -> tuple[int | None, list[str]]:
    for index, row in enumerate(rows, start=1):
        values = [clean_text(value) for value in row]
        non_empty = [value for value in values if value]
        if not non_empty:
            continue
        if any(COMPLAINT_NO_RE.fullmatch(value) for value in non_empty):
            continue
        normalized_values = {normalize_header(value) for value in non_empty}
        if normalized_values.intersection({"complaintno", "complaintnumber"}):
            return index, make_unique_headers(values)
    return None, []


def make_unique_headers(headers: list[str]) -> list[str]:
    result: list[str] = []
    seen: dict[str, int] = {}

    for column_number, header in enumerate(headers, start=1):
        base = header or f"Column {column_number}"
        count = seen.get(base, 0) + 1
        seen[base] = count
        result.append(base if count == 1 else f"{base} ({count})")

    return result


def find_complaint_column(headers: list[str]) -> int | None:
    for index, header in enumerate(headers):
        if normalize_header(header) in {"complaintno", "complaintnumber"}:
            return index
    return None


def normalize_header(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", value.lower())


def build_row_data(headers: list[str], row: tuple[object, ...]) -> list[tuple[str, str]]:
    data: list[tuple[str, str]] = []
    width = max(len(headers), len(row))

    for index in range(width):
        header = headers[index] if index < len(headers) else f"Column {index + 1}"
        value = row[index] if index < len(row) else None
        data.append((header, clean_text(value)))

    return data


def get_complaint_number(
    row_data: list[tuple[str, str]],
    row: tuple[object, ...],
    complaint_index: int | None,
) -> str | None:
    if complaint_index is not None and complaint_index < len(row):
        match = COMPLAINT_NO_RE.search(clean_text(row[complaint_index]))
        if match:
            return match.group(0)

    for _, value in row_data:
        match = COMPLAINT_NO_RE.search(value)
        if match:
            return match.group(0)

    return None


def row_is_empty(row: tuple[object, ...]) -> bool:
    return not any(clean_text(value) for value in row)


def clean_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, Decimal):
        return format(value, "f")
    return str(value).strip()


def write_pdf(
    *,
    output_path: Path,
    complaint_number: str,
    workbook_name: str,
    sheet_name: str,
    row_index: int,
    row_data: list[tuple[str, str]],
) -> None:
    styles = getSampleStyleSheet()
    normal = ParagraphStyle(
        "RowValue",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=9,
        leading=12,
        alignment=TA_LEFT,
        wordWrap="CJK",
    )
    label = ParagraphStyle(
        "RowLabel",
        parent=normal,
        fontName="Helvetica-Bold",
        textColor=colors.HexColor("#1f2937"),
    )
    title = ParagraphStyle(
        "ComplaintTitle",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=16,
        leading=20,
        spaceAfter=8,
        textColor=colors.HexColor("#111827"),
    )
    meta = ParagraphStyle(
        "Meta",
        parent=normal,
        fontSize=8,
        leading=10,
        textColor=colors.HexColor("#4b5563"),
        spaceAfter=12,
    )
    long_item_label = ParagraphStyle(
        "LongItemLabel",
        parent=label,
        fontSize=11,
        spaceBefore=12,
        spaceAfter=4,
    )

    document = SimpleDocTemplate(
        str(output_path),
        pagesize=A4,
        rightMargin=16 * mm,
        leftMargin=16 * mm,
        topMargin=14 * mm,
        bottomMargin=14 * mm,
        title=f"Complaint {complaint_number}",
        author="complaints-management",
    )

    story = [
        Paragraph(f"Complaint {escape_text(complaint_number)}", title),
        Paragraph(
            escape_text(
                f"Source: {workbook_name} | Sheet: {sheet_name} | Row: {row_index}"
            ),
            meta,
        ),
    ]

    table_rows = []
    long_items = []

    for field, value in row_data:
        if value and (len(value) > 1500 or value.count("\n") > 25):
            long_items.append((field, value))
        else:
            table_rows.append(
                [
                    Paragraph(escape_text(field), label),
                    Paragraph(escape_text(value) if value else "-", normal),
                ]
            )

    if table_rows:
        table = Table(table_rows, colWidths=[48 * mm, 130 * mm], repeatRows=0)
        table.setStyle(
            TableStyle(
                [
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#f3f4f6")),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#d1d5db")),
                    ("LEFTPADDING", (0, 0), (-1, -1), 5),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ]
            )
        )
        story.append(table)

    if long_items:
        story.append(Spacer(1, 6 * mm))
        for field, value in long_items:
            story.append(Paragraph(escape_text(field), long_item_label))
            story.append(Paragraph(escape_text(value), normal))
            story.append(Spacer(1, 4 * mm))

    document.build(story)


def escape_text(value: str) -> str:
    return (
        value.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace("\n", "<br/>")
    )


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Convert CRM sheet complaint rows into one PDF per complaint."
    )
    parser.add_argument(
        "--input-dir",
        default=DEFAULT_INPUT_DIR,
        help="Folder containing Excel/CSV files, or one Excel/CSV file.",
    )
    parser.add_argument(
        "--output-dir",
        default=DEFAULT_OUTPUT_DIR,
        help="Folder where generated complaint PDFs are written.",
    )
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="Replace existing complaint PDFs in the output folder.",
    )
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    logger = configure_logging()
    project_root = Path(__file__).resolve().parent
    args = parse_args(argv or sys.argv[1:])
    input_dir = resolve_project_path(project_root, args.input_dir)
    output_dir = resolve_project_path(project_root, args.output_dir)

    if not input_dir.exists():
        logger.error("Sheet input path does not exist: %s", input_dir)
        return 1

    output_dir.mkdir(parents=True, exist_ok=True)
    input_files = list(find_input_files(input_dir))
    if not input_files:
        logger.info("No Excel or CSV files found in %s.", input_dir)
        return 0

    logger.info("Sheet input: %s", input_dir)
    logger.info("PDF output: %s", output_dir)
    total = 0
    for input_path in input_files:
        count = convert_input_file(input_path, output_dir, args.overwrite)
        total += count
        logger.info("%s: created %d PDF(s).", input_path.name, count)

    logger.info("Done. Created %d PDF(s) in %s.", total, output_dir)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
