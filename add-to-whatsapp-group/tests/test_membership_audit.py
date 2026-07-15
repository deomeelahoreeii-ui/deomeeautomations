from pathlib import Path

from openpyxl import Workbook, load_workbook

from whatsapp_group_importer.membership_audit import GroupSnapshot, generate_not_member_report


def test_generates_complete_not_member_report(tmp_path: Path):
    source = tmp_path / "applicants.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["name", "contact_no", "post"])
    sheet.append(["Member", "03001234567", "A"])
    sheet.append(["Missing", "03111234567", "B"])
    sheet.append(["Invalid", "abc", "C"])
    workbook.save(source)
    workbook.close()
    snapshot = GroupSnapshot(
        checked_at="2026-06-22T00:00:00Z",
        group_jid="group@g.us",
        group_name="Test",
        member_phones=frozenset({"923001234567"}),
        reported_size=1,
        unresolved_members=0,
        worker_id="default",
    )
    output = tmp_path / "report.xlsx"
    summary = generate_not_member_report(source, output, snapshot)
    report = load_workbook(output, read_only=True)
    rows = list(report["Not Members"].iter_rows(values_only=True))
    report.close()

    assert summary.matched_members == 1
    assert summary.not_members == 1
    assert summary.invalid_contacts == 1
    assert rows[1][0] == "Missing"
    assert rows[2][0] == "Invalid"
