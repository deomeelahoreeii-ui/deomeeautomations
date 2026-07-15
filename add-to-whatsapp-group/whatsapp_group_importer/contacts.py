from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


class ContactInputError(ValueError):
    pass


@dataclass(frozen=True)
class Contact:
    source_row: int
    raw_value: str
    phone: str

    @property
    def jid(self) -> str:
        return f"{self.phone}@s.whatsapp.net"


def normalize_phone(value: object, country_code: str = "92", trunk_prefix: str = "0") -> str:
    raw = str(value or "").strip()
    if raw.endswith(".0"):
        raw = raw[:-2]
    digits = re.sub(r"\D", "", raw)
    country_code = re.sub(r"\D", "", country_code)

    if trunk_prefix and digits.startswith(trunk_prefix):
        digits = digits[len(trunk_prefix) :]
        digits = country_code + digits
    elif country_code == "92" and len(digits) == 10 and digits.startswith("3"):
        digits = country_code + digits

    if not 8 <= len(digits) <= 15:
        raise ContactInputError(f"not a plausible international phone number: {raw!r}")
    return digits


def load_contacts(
    file_path: Path,
    *,
    column_name: str = "contact_no",
    sheet_name: str | None = None,
    country_code: str = "92",
    trunk_prefix: str = "0",
) -> tuple[list[Contact], list[str]]:
    if not file_path.is_file():
        raise ContactInputError(f"Excel file not found: {file_path}")

    workbook = load_workbook(file_path, read_only=True, data_only=True)
    try:
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise ContactInputError(
                    f"Sheet {sheet_name!r} not found; available: {', '.join(workbook.sheetnames)}"
                )
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.active

        rows = sheet.iter_rows(values_only=True)
        try:
            headers = next(rows)
        except StopIteration as exc:
            raise ContactInputError("Excel sheet is empty") from exc

        normalized_headers = [str(value or "").strip().casefold() for value in headers]
        wanted = column_name.strip().casefold()
        if wanted not in normalized_headers:
            available = ", ".join(str(value or "") for value in headers)
            raise ContactInputError(
                f"Required column {column_name!r} not found. Columns: {available}"
            )
        column_index = normalized_headers.index(wanted)

        contacts: list[Contact] = []
        warnings: list[str] = []
        seen: set[str] = set()
        for source_row, row in enumerate(rows, start=2):
            value = row[column_index] if column_index < len(row) else None
            if value is None or str(value).strip() == "":
                continue
            try:
                phone = normalize_phone(value, country_code, trunk_prefix)
            except ContactInputError as exc:
                warnings.append(f"Row {source_row}: {exc}")
                continue
            if phone in seen:
                warnings.append(f"Row {source_row}: duplicate contact {phone} skipped")
                continue
            seen.add(phone)
            contacts.append(Contact(source_row, str(value), phone))
        return contacts, warnings
    finally:
        workbook.close()


def select_contacts(
    contacts: list[Contact],
    selection: str,
    count: int,
    start: int | None,
    end: int | None,
) -> list[Contact]:
    if selection == "all":
        return contacts[:]
    if selection in {"first", "last"}:
        if count < 1:
            raise ContactInputError("count must be at least 1")
        return contacts[:count] if selection == "first" else contacts[-count:]
    if selection == "range":
        if start is None or end is None:
            raise ContactInputError("range selection requires --start and --end")
        if start < 1 or end < start:
            raise ContactInputError("range must use 1-based positions with end >= start")
        return contacts[start - 1 : end]
    raise ContactInputError(f"Unsupported selection: {selection}")
