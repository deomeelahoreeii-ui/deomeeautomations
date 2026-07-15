from __future__ import annotations

import asyncio
import json
import re
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Callable

import nats
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from .contacts import ContactInputError, normalize_phone


@dataclass(frozen=True)
class GroupSnapshot:
    checked_at: str
    group_jid: str
    group_name: str
    member_phones: frozenset[str]
    reported_size: int
    unresolved_members: int
    worker_id: str


@dataclass(frozen=True)
class AuditSummary:
    output_path: Path
    source_rows: int
    matched_members: int
    not_members: int
    invalid_contacts: int
    snapshot: GroupSnapshot


def digits_from_jid(value: object) -> str | None:
    text = str(value or "").strip()
    if not text.endswith("@s.whatsapp.net"):
        return None
    digits = re.sub(r"\D", "", text.split("@", 1)[0])
    return digits or None


async def fetch_group_snapshot(
    *,
    nats_url: str,
    health_subject: str,
    members_subject: str,
    group_jid: str,
    ready_timeout: int = 120,
    progress: Callable[[str], None] | None = None,
) -> GroupSnapshot:
    report = progress or (lambda _message: None)

    async def suppress_error(_error: Exception) -> None:
        return

    try:
        connection = await nats.connect(
            nats_url,
            connect_timeout=5,
            allow_reconnect=False,
            max_reconnect_attempts=1,
            reconnect_time_wait=0.1,
            error_cb=suppress_error,
        )
    except Exception as exc:
        raise RuntimeError(f"NATS is unavailable at {nats_url}") from exc

    try:
        deadline = asyncio.get_running_loop().time() + ready_timeout
        report(f"Waiting for worker health on {health_subject}...")
        while True:
            try:
                response = await connection.request(health_subject, b"health", timeout=3)
                health = json.loads(response.data.decode("utf-8"))
                if health.get("ready"):
                    break
            except Exception:
                pass
            remaining = int(deadline - asyncio.get_running_loop().time())
            if remaining <= 0:
                raise RuntimeError("Selected WhatsApp worker did not become ready")
            report(f"Worker not ready; {remaining} seconds remaining...")
            await asyncio.sleep(min(5, remaining))

        report("Fetching current group membership (read-only)...")
        response = await connection.request(
            members_subject,
            json.dumps({"group_jid": group_jid}).encode("utf-8"),
            timeout=60,
        )
        payload = json.loads(response.data.decode("utf-8"))
        if payload.get("error"):
            raise RuntimeError(str(payload["error"]))

        phones: set[str] = set()
        unresolved = 0
        for member in payload.get("members", []):
            phone = digits_from_jid(member.get("phoneNumber"))
            if phone is None:
                phone = digits_from_jid(member.get("id"))
            if phone:
                phones.add(phone)
            else:
                unresolved += 1
        return GroupSnapshot(
            checked_at=str(payload.get("checkedAt") or datetime.now(timezone.utc).isoformat()),
            group_jid=str(payload.get("groupJid") or group_jid),
            group_name=str(payload.get("groupName") or ""),
            member_phones=frozenset(phones),
            reported_size=int(payload.get("size") or len(payload.get("members", []))),
            unresolved_members=unresolved,
            worker_id=str(payload.get("workerId") or "default"),
        )
    finally:
        await connection.close()


def generate_not_member_report(
    source_file: Path,
    output_file: Path,
    snapshot: GroupSnapshot,
    *,
    column_name: str = "contact_no",
    sheet_name: str | None = None,
    country_code: str = "92",
    trunk_prefix: str = "0",
) -> AuditSummary:
    source = load_workbook(source_file, read_only=True, data_only=True)
    try:
        sheet = source[sheet_name] if sheet_name else source.active
        rows = sheet.iter_rows(values_only=True)
        try:
            headers = list(next(rows))
        except StopIteration as exc:
            raise ContactInputError("Excel sheet is empty") from exc
        normalized_headers = [str(value or "").strip().casefold() for value in headers]
        wanted = column_name.strip().casefold()
        if wanted not in normalized_headers:
            raise ContactInputError(f"Required column {column_name!r} was not found")
        contact_index = normalized_headers.index(wanted)

        output = Workbook()
        report_sheet = output.active
        report_sheet.title = "Not Members"
        report_headers = [*headers, "normalized_contact", "membership_status", "audit_note"]
        report_sheet.append(report_headers)
        header_fill = PatternFill("solid", fgColor="1F4E78")
        for cell in report_sheet[1]:
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = header_fill
        report_sheet.freeze_panes = "A2"
        report_sheet.auto_filter.ref = f"A1:{report_sheet.cell(1, len(report_headers)).coordinate}"

        source_rows = matched = not_members = invalid = 0
        for row in rows:
            source_rows += 1
            raw_contact = row[contact_index] if contact_index < len(row) else None
            try:
                phone = normalize_phone(raw_contact, country_code, trunk_prefix)
            except ContactInputError as exc:
                invalid += 1
                report_sheet.append([*row, "", "invalid_contact", str(exc)])
                continue
            if phone in snapshot.member_phones:
                matched += 1
                continue
            not_members += 1
            report_sheet.append(
                [*row, phone, "not_a_member", "Phone number not present in current group metadata"]
            )

        summary_sheet = output.create_sheet("Summary")
        summary_rows = [
            ("Source file", str(source_file.resolve())),
            ("Source sheet", sheet.title),
            ("Group", snapshot.group_name),
            ("Group JID", snapshot.group_jid),
            ("Worker account", snapshot.worker_id),
            ("Checked at", snapshot.checked_at),
            ("Group reported size", snapshot.reported_size),
            ("Group members with phone numbers", len(snapshot.member_phones)),
            ("Group members unresolved as LID only", snapshot.unresolved_members),
            ("Excel applicant rows", source_rows),
            ("Applicants matched as members", matched),
            ("Applicants not members", not_members),
            ("Applicants with invalid contacts", invalid),
        ]
        for item in summary_rows:
            summary_sheet.append(item)
        for cell in summary_sheet["A"]:
            cell.font = Font(bold=True)
        summary_sheet.column_dimensions["A"].width = 38
        summary_sheet.column_dimensions["B"].width = 80

        output_file.parent.mkdir(parents=True, exist_ok=True)
        output.save(output_file)
        output.close()
        return AuditSummary(
            output_path=output_file.resolve(),
            source_rows=source_rows,
            matched_members=matched,
            not_members=not_members,
            invalid_contacts=invalid,
            snapshot=snapshot,
        )
    finally:
        source.close()
