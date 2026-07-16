from __future__ import annotations

import hashlib
import json
import re
import uuid
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from PIL import Image, ImageDraw, ImageFont
from sqlalchemy import select
from sqlmodel import Session

from automation_core.models import Artifact, Job
from automation_core.config import get_settings
from master_data.models import Markaz, School, Tehsil, Wing


RENDERER_KEY = "antidengue.tehsil_dormant.v1"
WING_RENDERER_KEY = "antidengue.wing_dormant.v1"
PAKISTAN_TIME = ZoneInfo("Asia/Karachi")
REQUIRED_COLUMNS = {"school emis", "school name"}


@dataclass(frozen=True)
class DormantSourceRow:
    emis: str
    school_name: str


@dataclass(frozen=True)
class ScopedDormantSchool:
    emis: str
    school_name: str
    tehsil: str
    markaz: str


@dataclass
class RenderedTehsilDormantReport:
    message: str
    context: dict[str, str]
    schools: list[ScopedDormantSchool]
    issues: list[dict[str, Any]] = field(default_factory=list)
    attachment_paths: list[Path] = field(default_factory=list)
    source_artifact_id: int | None = None
    source_artifact_sha256: str = ""


@dataclass
class RenderedWingDormantReport:
    message: str
    context: dict[str, str]
    schools: list[ScopedDormantSchool]
    issues: list[dict[str, Any]] = field(default_factory=list)
    attachment_path: Path | None = None
    attachment_paths: list[Path] = field(default_factory=list)
    source_artifact_id: int | None = None
    source_artifact_sha256: str = ""


def _issue(code: str, severity: str, message: str, **details: Any) -> dict[str, Any]:
    return {"code": code, "severity": severity, "message": message, **details}


def _normalize_emis(value: object) -> str:
    text = str(value or "").strip()
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    return text


def _report_rows(path: Path) -> list[DormantSourceRow] | None:
    if path.suffix.lower() != ".xlsx" or not path.is_file():
        return None
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet in workbook.worksheets:
            rows = sheet.iter_rows(values_only=True)
            for raw_header in rows:
                header = [str(value or "").strip().lower() for value in raw_header]
                if not REQUIRED_COLUMNS.issubset(header):
                    continue
                emis_index = header.index("school emis")
                name_index = header.index("school name")
                result: list[DormantSourceRow] = []
                for raw_row in rows:
                    emis = _normalize_emis(
                        raw_row[emis_index] if emis_index < len(raw_row) else ""
                    )
                    if not emis:
                        continue
                    school_name = str(
                        raw_row[name_index] if name_index < len(raw_row) else ""
                    ).strip()
                    result.append(DormantSourceRow(emis=emis, school_name=school_name))
                return result
    finally:
        workbook.close()
    return None


def _canonical_report_artifact(
    session: Session, source_job: Job
) -> tuple[Artifact, Path, list[DormantSourceRow]]:
    artifacts = session.scalars(
        select(Artifact)
        .where(Artifact.job_id == source_job.id)
        .order_by(Artifact.created_at, Artifact.id)
    ).all()
    # The platform adapter registers the canonical dormant workbook as
    # ``report``. Header inspection is retained as a safe migration fallback.
    candidates = sorted(artifacts, key=lambda item: item.kind != "report")
    for artifact in candidates:
        path = Path(artifact.path).expanduser().resolve(strict=False)
        try:
            rows = _report_rows(path)
        except Exception:
            continue
        if rows is not None:
            return artifact, path, rows
    raise ValueError(
        "The source run has no canonical dormant-school report with School EMIS and School Name columns."
    )


def _school_count(count: int) -> str:
    return f"{count} {'school' if count == 1 else 'schools'}"


def normalize_presentation_policy(
    value: dict[str, Any] | None,
    *,
    report_key: str,
) -> dict[str, str]:
    value = value or {}
    legacy_mode = "none" if report_key == "tehsil_dormant_summary" else "excel"
    return {
        "message_style": value.get("message_style")
        if value.get("message_style") in {"summary", "detailed"}
        else ("detailed" if report_key == "tehsil_dormant_summary" else "summary"),
        "attachment_mode": value.get("attachment_mode")
        if value.get("attachment_mode") in {"none", "image", "excel", "image_excel"}
        else legacy_mode,
        "image_content": value.get("image_content")
        if value.get("image_content") in {"summary", "details"}
        else ("details" if report_key == "tehsil_dormant_summary" else "summary"),
    }


def _has_attachment(policy: dict[str, str], kind: str) -> bool:
    mode = policy["attachment_mode"]
    return mode == kind or mode == "image_excel"


def _wing_label(wing: Wing) -> str:
    value = (wing.code or wing.name).strip()
    return re.sub(r"^DEO[\s_-]+", "", value, flags=re.IGNORECASE) or value


def _wing_display_name(wing: Wing) -> str:
    label = _wing_label(wing).upper().replace("-", "").replace(" ", "")
    if label == "SE":
        return "Secondary Wing"
    return f"{_wing_label(wing)} Wing"


def _report_time(source_job: Job) -> datetime:
    value = source_job.finished_at or source_job.started_at or source_job.created_at
    if value.tzinfo is None:
        # PostgreSQL returns the configured server-local timestamp without its
        # offset in this deployment. Treat it as Pakistan time rather than
        # shifting an already-local clock by another five hours.
        value = value.replace(tzinfo=PAKISTAN_TIME)
    return value.astimezone(PAKISTAN_TIME)


def _build_message(
    *,
    source_job: Job,
    recipient_name: str,
    wing: Wing,
    tehsil: Tehsil,
    schools: list[ScopedDormantSchool],
    deadline: str,
    policy: dict[str, str] | None = None,
) -> tuple[str, dict[str, str]]:
    policy = policy or normalize_presentation_policy(None, report_key="tehsil_dormant_summary")
    generated_at = _report_time(source_job)
    title = f"Anti-Dengue Dormant Users - {generated_at:%d-%b-%Y - %I:%M %p}"
    wing_label = _wing_label(wing)
    lines = [
        f"*{title}*",
        f"*Group:* {recipient_name}",
        f"*Tehsil:* {tehsil.name}",
        f"*Wing:* {wing_label}",
        f"*Total dormant:* {_school_count(len(schools))}",
        "",
    ]
    by_markaz: dict[str, list[ScopedDormantSchool]] = defaultdict(list)
    for school in schools:
        by_markaz[school.markaz].append(school)
    if policy["message_style"] == "detailed":
        lines.extend([
            "*TEHSIL SUMMARY*", f"1. {tehsil.name}: {_school_count(len(schools))}",
            "", "*DETAILS BY MARKAZ*", "", f"*{tehsil.name} - {_school_count(len(schools))}*",
        ])
    else:
        lines.extend(["*MARKAZ SUMMARY*"])
        for index, markaz_name in enumerate(sorted(by_markaz, key=str.casefold), start=1):
            lines.append(f"{index}. {markaz_name}: {_school_count(len(by_markaz[markaz_name]))}")
    for markaz_name in sorted(by_markaz, key=str.casefold):
        if policy["message_style"] != "detailed":
            break
        markaz_schools = sorted(
            by_markaz[markaz_name],
            key=lambda item: (item.school_name.casefold(), item.emis),
        )
        lines.extend(["", f"*{markaz_name} - {_school_count(len(markaz_schools))}*"])
        lines.extend(
            f"{index}. {school.emis} - {school.school_name}"
            for index, school in enumerate(markaz_schools, start=1)
        )
    try:
        parsed_deadline = datetime.strptime(deadline.strip(), "%I:%M %p").time()
        deadline_passed = generated_at.time() > parsed_deadline
    except ValueError:
        deadline_passed = False
    deadline_message = (
        f"The {deadline} submission deadline has passed. Please update activities on the portal immediately."
        if deadline_passed
        else f"Please ensure activities are submitted before {deadline} today."
    )
    lines.extend(["", deadline_message])
    if policy["attachment_mode"] != "none":
        labels = []
        if _has_attachment(policy, "image"):
            labels.append("image report")
        if _has_attachment(policy, "excel"):
            labels.append("Excel workbook")
        lines.extend(["", f"Attached: {' and '.join(labels)}."])
    message = "\n".join(lines)
    context = {
        "title": title,
        "recipient_name": recipient_name,
        "tehsil": tehsil.name,
        "wing": wing_label,
        "total_schools": str(len(schools)),
        "report_body": message,
        "deadline_message": deadline_message,
        "renderer_key": RENDERER_KEY,
    }
    return message, context


def _build_wing_message(
    *,
    source_job: Job,
    recipient_name: str,
    wing: Wing,
    schools: list[ScopedDormantSchool],
    policy: dict[str, str] | None = None,
) -> tuple[str, dict[str, str]]:
    legacy_policy = policy is None
    policy = policy or normalize_presentation_policy(None, report_key="wing_summary")
    generated_at = _report_time(source_job)
    title = f"Anti-Dengue Dormant Users - {generated_at:%d-%b-%Y - %I:%M %p}"
    wing_label = _wing_label(wing)
    wing_display_name = _wing_display_name(wing)
    tehsil_count = len({school.tehsil for school in schools})
    lines = [
        "*Anti-Dengue Dormancy Summary*",
        f"*{wing_display_name} — {generated_at:%d-%b-%Y, %I:%M %p}*",
        "",
        f"*Total dormant:* {_school_count(len(schools))} across {tehsil_count} {'tehsil' if tehsil_count == 1 else 'tehsils'}",
        "",
        "*TEHSIL BREAKDOWN*",
    ]
    by_tehsil: dict[str, list[ScopedDormantSchool]] = defaultdict(list)
    for school in schools:
        by_tehsil[school.tehsil].append(school)
    lines.extend(
        f"{index}. {tehsil_name}: {_school_count(len(tehsil_schools))}"
        for index, (tehsil_name, tehsil_schools) in enumerate(
            sorted(by_tehsil.items(), key=lambda item: item[0].casefold()), start=1
        )
    )
    authoritative_markaz = any(
        school.markaz != "UNMAPPED MARKAZ" for school in schools
    )
    if authoritative_markaz:
        lines.extend(["", "*MARKAZ BREAKDOWN*"])
        for tehsil_name, tehsil_schools in sorted(
            by_tehsil.items(), key=lambda item: item[0].casefold()
        ):
            lines.extend(["", f"*{tehsil_name} - {_school_count(len(tehsil_schools))}*"])
            markaz_counts = Counter(school.markaz for school in tehsil_schools)
            lines.extend(
                f"{index}. {markaz_name}: {_school_count(count)}"
                for index, (markaz_name, count) in enumerate(
                    sorted(markaz_counts.items(), key=lambda item: item[0].casefold()), start=1
                )
            )
    lines.extend(
        [
            "",
            "Please ensure all pending activities are updated on the portal today.",
        ]
    )
    if legacy_policy:
        lines.extend(["", "The attached workbook contains the detailed school list."])
    elif policy["attachment_mode"] != "none":
        labels = []
        if _has_attachment(policy, "image"):
            labels.append("image report")
        if _has_attachment(policy, "excel"):
            labels.append("Excel workbook")
        lines.extend(["", f"Attached: {' and '.join(labels)}."])
    message = "\n".join(lines)
    context = {
        "title": title,
        "recipient_name": recipient_name,
        "wing": wing_label,
        "wing_display_name": wing_display_name,
        "scope": "Wing",
        "total_schools": str(len(schools)),
        "report_body": message,
        "renderer_key": WING_RENDERER_KEY,
    }
    return message, context


def _ensure_excel_attachment(
    session: Session,
    *,
    source_job: Job,
    wing: Wing,
    schools: list[ScopedDormantSchool],
    source_digest: str,
    scope_key: str,
    scope_label: str,
) -> tuple[Artifact, Path, str]:
    content_key = hashlib.sha256(
        json.dumps(
            {
                "source": source_digest,
                "wing": str(wing.id),
                "scope": [scope_key, scope_label],
                "schools": [
                    [school.emis, school.school_name, school.tehsil, school.markaz]
                    for school in schools
                ],
            },
            ensure_ascii=False,
            sort_keys=True,
        ).encode("utf-8")
    ).hexdigest()
    safe_wing = re.sub(r"[^A-Za-z0-9_-]+", "-", wing.code or wing.name).strip("-") or "wing"
    safe_scope = re.sub(r"[^A-Za-z0-9_-]+", "-", scope_label).strip("-") or scope_key
    filename = f"Anti-Dengue Dormant Users - {safe_wing} - {safe_scope}.xlsx"
    path = (
        get_settings().artifact_root.expanduser().resolve()
        / "antidengue"
        / "native-reports"
        / str(source_job.id)
        / f"{safe_wing}-{safe_scope}-{content_key[:16]}.xlsx"
    )
    artifact = session.scalar(
        select(Artifact).where(Artifact.job_id == source_job.id, Artifact.path == str(path))
    )
    if artifact is not None and path.is_file():
        return artifact, path, hashlib.sha256(path.read_bytes()).hexdigest()

    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(f".{path.name}.{uuid.uuid4().hex}.tmp")
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Dormant Schools"
    title = f"Anti-Dengue Dormant Schools — {scope_label} — {_wing_display_name(wing)}"
    sheet.merge_cells("A1:E1")
    sheet["A1"] = title
    sheet["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="087487")
    sheet["A1"].alignment = Alignment(horizontal="center")
    sheet.append(["School EMIS", "School Name", "Tehsil", "Markaz", "Wing"])
    for cell in sheet[2]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="26445D")
    for school in schools:
        sheet.append([school.emis, school.school_name, school.tehsil, school.markaz, wing.name])
    sheet.freeze_panes = "A3"
    sheet.auto_filter.ref = f"A2:E{sheet.max_row}"
    for column, width in {"A": 16, "B": 52, "C": 20, "D": 28, "E": 18}.items():
        sheet.column_dimensions[column].width = width
    for row in sheet.iter_rows(min_row=3):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    workbook.save(temporary)
    workbook.close()
    temporary.replace(path)

    if artifact is None:
        artifact = Artifact(
            job_id=source_job.id,
            kind="delivery",
            name=filename,
            path=str(path),
            size_bytes=path.stat().st_size,
        )
    else:
        artifact.name = filename
        artifact.size_bytes = path.stat().st_size
    session.add(artifact)
    session.flush()
    return artifact, path, hashlib.sha256(path.read_bytes()).hexdigest()


def _font(size: int, *, bold: bool = False) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
    candidates = [
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf" if bold else "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/TTF/DejaVuSans-Bold.ttf" if bold else "/usr/share/fonts/TTF/DejaVuSans.ttf",
    ]
    for candidate in candidates:
        try:
            return ImageFont.truetype(candidate, size=size)
        except OSError:
            continue
    return ImageFont.load_default()


def _wrap_text(draw: ImageDraw.ImageDraw, text: str, font: Any, max_width: int) -> list[str]:
    words = str(text).split()
    lines: list[str] = []
    current = ""
    for word in words:
        candidate = f"{current} {word}".strip()
        if not current or draw.textbbox((0, 0), candidate, font=font)[2] <= max_width:
            current = candidate
        else:
            lines.append(current)
            current = word
    if current:
        lines.append(current)
    return lines or [""]


def _ensure_image_attachment(
    session: Session,
    *,
    source_job: Job,
    wing: Wing,
    schools: list[ScopedDormantSchool],
    source_digest: str,
    scope_key: str,
    scope_label: str,
    image_content: str,
) -> tuple[Artifact, Path, str]:
    content_key = hashlib.sha256(
        json.dumps(
            {
                "source": source_digest,
                "wing": str(wing.id),
                "scope": [scope_key, scope_label],
                "content": image_content,
                "layout": 2,
                "schools": [[s.emis, s.school_name, s.tehsil, s.markaz] for s in schools],
            },
            ensure_ascii=False,
            sort_keys=True,
        ).encode("utf-8")
    ).hexdigest()
    safe_wing = re.sub(r"[^A-Za-z0-9_-]+", "-", wing.code or wing.name).strip("-") or "wing"
    safe_scope = re.sub(r"[^A-Za-z0-9_-]+", "-", scope_label).strip("-") or scope_key
    filename = f"Anti-Dengue Dormant Users - {safe_wing} - {safe_scope}.png"
    path = get_settings().artifact_root.expanduser().resolve() / "antidengue" / "native-reports" / str(source_job.id) / f"{safe_wing}-{safe_scope}-{content_key[:16]}.png"
    artifact = session.scalar(select(Artifact).where(Artifact.job_id == source_job.id, Artifact.path == str(path)))
    if artifact is not None and path.is_file():
        return artifact, path, hashlib.sha256(path.read_bytes()).hexdigest()

    width = 1080
    margin = 64
    title_font, heading_font, body_font, small_font = _font(34, bold=True), _font(24, bold=True), _font(20), _font(17)
    scratch = Image.new("RGB", (width, 100), "white")
    draw = ImageDraw.Draw(scratch)
    rows: list[tuple[str, str, str]] = []
    by_tehsil = Counter(s.tehsil for s in schools)
    for tehsil, count in sorted(by_tehsil.items(), key=lambda item: item[0].casefold()):
        rows.append((tehsil, _school_count(count), "summary"))
    if image_content == "details":
        rows.append(("SCHOOL DETAILS", "", "heading"))
        for school in schools:
            rows.append((f"{school.emis}  {school.school_name}", f"{school.tehsil} · {school.markaz}", "detail"))
    calculated = 300
    for left, right, kind in rows:
        if kind == "heading":
            calculated += 64
        else:
            lines = _wrap_text(draw, left, body_font if kind == "summary" else small_font, 700 if kind == "summary" else 900)
            calculated += max(58, len(lines) * (27 if kind == "summary" else 23) + (28 if kind == "summary" else 50))
    height = min(max(calculated + 80, 720), 14000)
    image = Image.new("RGB", (width, height), "#F4F7FA")
    draw = ImageDraw.Draw(image)
    draw.rounded_rectangle((32, 32, width - 32, height - 32), radius=22, fill="white", outline="#D9E2EC", width=2)
    draw.rounded_rectangle((32, 32, width - 32, 210), radius=22, fill="#087487")
    draw.text((margin, 62), "ANTI-DENGUE DORMANCY REPORT", font=title_font, fill="white")
    draw.text((margin, 115), f"{scope_label}  ·  {_wing_display_name(wing)}", font=heading_font, fill="#D8F5F5")
    draw.text((margin, 158), f"Total dormant: {_school_count(len(schools))}", font=body_font, fill="white")
    y = 246
    draw.text((margin, y), "TEHSIL SUMMARY", font=heading_font, fill="#26445D")
    y += 48
    for left, right, kind in rows:
        if y > height - 80:
            break
        if kind == "heading":
            y += 18
            draw.text((margin, y), left, font=heading_font, fill="#087487")
            y += 48
            continue
        font = body_font if kind == "summary" else small_font
        lines = _wrap_text(draw, left, font, 700 if kind == "summary" else 900)
        row_height = max(58, len(lines) * (27 if kind == "summary" else 23) + (28 if kind == "summary" else 50))
        draw.rounded_rectangle((margin, y, width - margin, y + row_height - 8), radius=10, fill="#F7FAFC")
        line_y = y + 12
        for line in lines:
            draw.text((margin + 18, line_y), line, font=font, fill="#1F2937")
            line_y += 27 if kind == "summary" else 23
        if right and kind == "summary":
            draw.text((width - margin - 18, y + 14), right, font=small_font, fill="#60758A", anchor="ra")
        elif right:
            draw.text((margin + 18, line_y + 4), right, font=small_font, fill="#60758A")
        y += row_height
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(f".{path.name}.{uuid.uuid4().hex}.tmp")
    image.save(temporary, format="PNG", optimize=True)
    temporary.replace(path)
    if artifact is None:
        artifact = Artifact(job_id=source_job.id, kind="delivery", name=filename, path=str(path), size_bytes=path.stat().st_size)
    else:
        artifact.name = filename
        artifact.size_bytes = path.stat().st_size
    session.add(artifact)
    session.flush()
    return artifact, path, hashlib.sha256(path.read_bytes()).hexdigest()


def render_wing_dormant_report(
    session: Session,
    *,
    source_job: Job,
    wing: Wing,
    recipient_name: str,
    presentation_policy: dict[str, Any] | None = None,
) -> RenderedWingDormantReport:
    policy = normalize_presentation_policy(presentation_policy, report_key="wing_summary")
    source_artifact, source_path, source_rows = _canonical_report_artifact(session, source_job)
    counts = Counter(row.emis for row in source_rows)
    duplicate_emis = sorted(emis for emis, count in counts.items() if count > 1)
    source_by_emis = {row.emis: row for row in source_rows}
    school_models = session.scalars(
        select(School).where(School.emis.in_(list(source_by_emis)))
    ).all()
    schools_by_emis = {school.emis: school for school in school_models}
    unknown_emis = sorted(set(source_by_emis) - set(schools_by_emis))
    scoped_models = [
        school for school in school_models if school.active and school.wing_id == wing.id
    ]
    tehsil_ids = {school.tehsil_id for school in scoped_models if school.tehsil_id}
    markaz_ids = {school.markaz_id for school in scoped_models if school.markaz_id}
    tehsils = {
        item.id: item
        for item in session.scalars(select(Tehsil).where(Tehsil.id.in_(tehsil_ids))).all()
    } if tehsil_ids else {}
    markazes = {
        item.id: item
        for item in session.scalars(select(Markaz).where(Markaz.id.in_(markaz_ids))).all()
    } if markaz_ids else {}

    scoped: list[ScopedDormantSchool] = []
    missing_markaz: list[str] = []
    missing_tehsil: list[str] = []
    for school in scoped_models:
        tehsil = tehsils.get(school.tehsil_id) if school.tehsil_id else None
        markaz = markazes.get(school.markaz_id) if school.markaz_id else None
        if tehsil is None:
            missing_tehsil.append(school.emis)
        if markaz is None:
            missing_markaz.append(school.emis)
        scoped.append(
            ScopedDormantSchool(
                emis=school.emis,
                school_name=school.name or source_by_emis[school.emis].school_name,
                tehsil=tehsil.name if tehsil else "UNMAPPED TEHSIL",
                markaz=markaz.name if markaz else "UNMAPPED MARKAZ",
            )
        )
    scoped.sort(
        key=lambda item: (
            item.tehsil.casefold(), item.markaz.casefold(), item.school_name.casefold(), item.emis
        )
    )

    issues: list[dict[str, Any]] = []
    if duplicate_emis:
        issues.append(
            _issue(
                "duplicate_source_emis",
                "blocked",
                f"The canonical report contains {len(duplicate_emis)} duplicate school EMIS value(s).",
                emis=duplicate_emis[:20],
            )
        )
    if unknown_emis:
        issues.append(
            _issue(
                "unknown_source_schools",
                "warning",
                f"{len(unknown_emis)} dormant school EMIS value(s) are absent from PostgreSQL and were excluded.",
                emis=unknown_emis[:20],
            )
        )
    if missing_tehsil:
        issues.append(
            _issue(
                "missing_tehsil",
                "warning",
                f"{len(missing_tehsil)} { _wing_label(wing) } school(s) have no authoritative Tehsil.",
                emis=missing_tehsil[:20],
            )
        )
    if missing_markaz:
        issues.append(
            _issue(
                "missing_markaz",
                "warning",
                f"{len(missing_markaz)} { _wing_label(wing) } school(s) have no authoritative Markaz and are summarized as UNMAPPED MARKAZ.",
                emis=missing_markaz[:20],
            )
        )
    if not scoped:
        issues.append(
            _issue(
                "nothing_to_dispatch",
                "warning",
                f"No dormant { _wing_label(wing) } schools were found.",
            )
        )

    message, context = _build_wing_message(
        source_job=source_job,
        recipient_name=recipient_name,
        wing=wing,
        schools=scoped,
        policy=policy,
    )
    source_digest = hashlib.sha256(source_path.read_bytes()).hexdigest()
    attachment_path: Path | None = None
    attachment_paths: list[Path] = []
    artifact_id: int | None = source_artifact.id
    artifact_digest = source_digest
    if scoped and _has_attachment(policy, "image"):
        derived_artifact, image_path, artifact_digest = _ensure_image_attachment(
            session,
            source_job=source_job,
            wing=wing,
            schools=scoped,
            source_digest=source_digest,
            scope_key="wing",
            scope_label=wing.name,
            image_content=policy["image_content"],
        )
        artifact_id = derived_artifact.id
        attachment_paths.append(image_path)
    if scoped and _has_attachment(policy, "excel"):
        derived_artifact, attachment_path, artifact_digest = _ensure_excel_attachment(
            session,
            source_job=source_job,
            wing=wing,
            schools=scoped,
            source_digest=source_digest,
            scope_key="wing",
            scope_label=wing.name,
        )
        artifact_id = derived_artifact.id
        attachment_paths.append(attachment_path)
    return RenderedWingDormantReport(
        message=message,
        context=context,
        schools=scoped,
        issues=issues,
        attachment_path=attachment_path,
        attachment_paths=attachment_paths,
        source_artifact_id=artifact_id,
        source_artifact_sha256=artifact_digest,
    )


def render_tehsil_dormant_report(
    session: Session,
    *,
    source_job: Job,
    wing: Wing,
    tehsil_id: uuid.UUID,
    recipient_name: str,
    deadline: str = "12:30 PM",
    presentation_policy: dict[str, Any] | None = None,
) -> RenderedTehsilDormantReport:
    policy = normalize_presentation_policy(
        presentation_policy,
        report_key="tehsil_dormant_summary",
    )
    tehsil = session.get(Tehsil, tehsil_id)
    if tehsil is None or not tehsil.active or tehsil.district_id != wing.district_id:
        raise ValueError("The audience target is not bound to an active tehsil in this wing's district.")

    artifact, artifact_path, source_rows = _canonical_report_artifact(session, source_job)
    counts = Counter(row.emis for row in source_rows)
    duplicate_emis = sorted(emis for emis, count in counts.items() if count > 1)
    source_by_emis = {row.emis: row for row in source_rows}
    schools = session.scalars(
        select(School).where(School.emis.in_(list(source_by_emis)))
    ).all()
    schools_by_emis = {school.emis: school for school in schools}
    unknown_emis = sorted(set(source_by_emis) - set(schools_by_emis))
    scoped_models = [
        school
        for school in schools
        if school.active and school.wing_id == wing.id and school.tehsil_id == tehsil.id
    ]
    markaz_ids = {school.markaz_id for school in scoped_models if school.markaz_id}
    markazes = {
        item.id: item
        for item in session.scalars(select(Markaz).where(Markaz.id.in_(markaz_ids))).all()
    } if markaz_ids else {}

    scoped: list[ScopedDormantSchool] = []
    missing_markaz: list[str] = []
    for school in scoped_models:
        markaz = markazes.get(school.markaz_id) if school.markaz_id else None
        if markaz is None:
            missing_markaz.append(school.emis)
        scoped.append(
            ScopedDormantSchool(
                emis=school.emis,
                school_name=school.name or source_by_emis[school.emis].school_name,
                tehsil=tehsil.name,
                markaz=markaz.name if markaz else "UNMAPPED MARKAZ",
            )
        )
    scoped.sort(key=lambda item: (item.markaz.casefold(), item.school_name.casefold(), item.emis))

    issues: list[dict[str, Any]] = []
    if duplicate_emis:
        issues.append(
            _issue(
                "duplicate_source_emis",
                "blocked",
                f"The canonical report contains {len(duplicate_emis)} duplicate school EMIS value(s).",
                emis=duplicate_emis[:20],
            )
        )
    if unknown_emis:
        issues.append(
            _issue(
                "unknown_source_schools",
                "warning",
                f"{len(unknown_emis)} dormant school EMIS value(s) are absent from PostgreSQL and were excluded.",
                emis=unknown_emis[:20],
            )
        )
    if missing_markaz:
        issues.append(
            _issue(
                "missing_markaz",
                "warning",
                f"{len(missing_markaz)} scoped school(s) have no authoritative Markaz and are listed under UNMAPPED MARKAZ.",
                emis=missing_markaz[:20],
            )
        )
    if not scoped:
        issues.append(
            _issue(
                "nothing_to_dispatch",
                "warning",
                f"No dormant { _wing_label(wing) } schools were found in {tehsil.name}.",
            )
        )

    message, context = _build_message(
        source_job=source_job,
        recipient_name=recipient_name,
        wing=wing,
        tehsil=tehsil,
        schools=scoped,
        deadline=deadline,
        policy=policy,
    )
    # A digest is calculated here without importing the preview service and
    # creating a circular dependency.
    import hashlib

    digest = hashlib.sha256(artifact_path.read_bytes()).hexdigest()
    attachment_paths: list[Path] = []
    if scoped and _has_attachment(policy, "image"):
        _, image_path, digest = _ensure_image_attachment(
            session,
            source_job=source_job,
            wing=wing,
            schools=scoped,
            source_digest=digest,
            scope_key="tehsil",
            scope_label=tehsil.name,
            image_content=policy["image_content"],
        )
        attachment_paths.append(image_path)
    if scoped and _has_attachment(policy, "excel"):
        _, excel_path, digest = _ensure_excel_attachment(
            session,
            source_job=source_job,
            wing=wing,
            schools=scoped,
            source_digest=hashlib.sha256(artifact_path.read_bytes()).hexdigest(),
            scope_key="tehsil",
            scope_label=tehsil.name,
        )
        attachment_paths.append(excel_path)
    return RenderedTehsilDormantReport(
        message=message,
        context=context,
        schools=scoped,
        issues=issues,
        attachment_paths=attachment_paths,
        source_artifact_id=artifact.id,
        source_artifact_sha256=digest,
    )
