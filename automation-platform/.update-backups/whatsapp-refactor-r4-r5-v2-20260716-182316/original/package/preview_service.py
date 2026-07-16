from __future__ import annotations

import copy
import hashlib
import json
import mimetypes
import re
import shutil
import uuid
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import or_, select
from sqlmodel import Session

from automation_core.config import get_settings
from automation_core.models import Artifact, Job, JobStatus, JobType
from automation_core.time import utcnow
from master_data.models import School, SchoolHead, Wing
from whatsapp_gateway.antidengue_renderer import (
    RENDERER_KEY as TEHSIL_DORMANT_RENDERER_KEY,
    WING_RENDERER_KEY,
    normalize_presentation_policy,
    render_tehsil_dormant_report,
    render_wing_dormant_report,
)
from whatsapp_gateway.models import (
    WhatsAppAccount,
    WhatsAppApplication,
    WhatsAppAudience,
    WhatsAppAudienceMember,
    WhatsAppContactLink,
    WhatsAppDirectoryContact,
    WhatsAppDirectoryGroup,
    WhatsAppDispatchPreview,
    WhatsAppDispatchPreviewArtifact,
    WhatsAppDispatchPreviewDelivery,
    WhatsAppDispatchProfile,
    WhatsAppGroup,
    WhatsAppRecipientScope,
    WhatsAppReportType,
    WhatsAppTemplate,
)


PLACEHOLDER_RE = re.compile(r"{{\s*([a-zA-Z0-9_]+)\s*}}")
PHONE_JID_SUFFIX = "@s.whatsapp.net"


class PreviewCompileError(ValueError):
    pass


def issue(code: str, severity: str, message: str, **details: Any) -> dict[str, Any]:
    return {
        "code": code,
        "severity": severity,
        "message": message,
        **details,
    }


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _managed_artifact_path(checksum: str, source: Path) -> Path:
    suffix = source.suffix.lower() if source.suffix else ".bin"
    return (
        get_settings().artifact_root.expanduser().resolve()
        / "dispatch-previews"
        / checksum[:2]
        / f"{checksum}{suffix}"
    )


def freeze_artifact(path: Path, checksum: str) -> Path:
    """Copy report bytes into content-addressed, platform-managed storage."""
    destination = _managed_artifact_path(checksum, path)
    destination.parent.mkdir(parents=True, exist_ok=True)
    if destination.is_file() and sha256_file(destination) == checksum:
        return destination
    temporary = destination.with_name(f".{destination.name}.{uuid.uuid4().hex}.tmp")
    shutil.copyfile(path, temporary)
    if sha256_file(temporary) != checksum:
        temporary.unlink(missing_ok=True)
        raise PreviewCompileError(f"Attachment changed while it was being frozen: {path.name}")
    temporary.replace(destination)
    return destination


def is_managed_preview_artifact(path: Path) -> bool:
    root = (
        get_settings().artifact_root.expanduser().resolve()
        / "dispatch-previews"
    )
    try:
        path.resolve(strict=False).relative_to(root)
    except ValueError:
        return False
    return True


def _preview_key() -> str:
    now = utcnow()
    return f"AD-{now:%Y%m%d-%H%M%S}-{uuid.uuid4().hex[:6].upper()}"


def _artifact_role(path: Path, referenced_paths: set[str]) -> str:
    normalized = str(path.resolve(strict=False))
    name = path.name.lower()
    if normalized in referenced_paths or "group-route-excels" in normalized:
        return "delivery"
    if name == "run_summary.json":
        return "manifest"
    if "audit" in name or "evidence" in name:
        return "audit"
    return "supporting"


def _attachment_paths(plan: dict[str, Any]) -> list[Path]:
    payload = plan.get("payload") if isinstance(plan.get("payload"), dict) else {}
    candidates: list[Any] = [
        payload.get("excel_path"),
        payload.get("image_path"),
        plan.get("excel_path"),
    ]
    attachment_paths = payload.get("attachment_paths") or []
    if isinstance(attachment_paths, list):
        candidates.extend(attachment_paths)
    documents = payload.get("documents") or []
    if isinstance(documents, list):
        for document in documents:
            candidates.append(document.get("path") if isinstance(document, dict) else document)
    result: list[Path] = []
    seen: set[str] = set()
    for value in candidates:
        if not value:
            continue
        path = Path(str(value)).expanduser().resolve(strict=False)
        key = str(path)
        if key not in seen:
            seen.add(key)
            result.append(path)
    return result


def _render_message(
    template: WhatsAppTemplate | None,
    legacy_message: str,
    context: dict[str, str],
) -> tuple[str, list[dict[str, Any]]]:
    if template is None:
        return legacy_message, []
    values = {**context, "message": legacy_message}
    missing = sorted({key for key in PLACEHOLDER_RE.findall(template.body) if key not in values})
    problems = [
        issue(
            "unknown_template_variable",
            "blocked",
            f"Template variable {{{{{key}}}}} has no preview value.",
        )
        for key in missing
    ]
    rendered = PLACEHOLDER_RE.sub(lambda match: values.get(match.group(1), match.group(0)), template.body)
    return rendered, problems


def _delivery_status(problems: list[dict[str, Any]]) -> str:
    if any(item["severity"] == "blocked" for item in problems):
        return "blocked"
    if any(item["severity"] == "warning" for item in problems):
        return "warning"
    return "ready"


def _canonical_contact_target(target: str) -> str:
    value = target.strip()
    if "@" in value:
        return value
    digits = re.sub(r"\D", "", value)
    return f"{digits}{PHONE_JID_SUFFIX}" if digits else value


def _plan_report_type(plan: dict[str, Any]) -> str:
    """Classify a legacy boundary record into a platform report type."""
    payload = plan.get("payload") if isinstance(plan.get("payload"), dict) else {}
    target = str(payload.get("target") or plan.get("target") or "").strip()
    target_type = str(payload.get("type") or plan.get("channel") or "contact").lower()
    if target_type not in {"group", "contact"}:
        target_type = "group" if target.endswith("@g.us") else "contact"
    if target_type == "contact":
        return "officer_summary"
    dispatch_route = (
        payload.get("dispatch_route")
        if isinstance(payload.get("dispatch_route"), dict)
        else {}
    )
    route_kind = str(
        dispatch_route.get("route_kind") or plan.get("route_kind") or ""
    ).lower()
    return "wing_summary" if route_kind == "district" else "school_activity"


def _plan_recipient_channel(plan: dict[str, Any]) -> str:
    payload = plan.get("payload") if isinstance(plan.get("payload"), dict) else {}
    target = str(payload.get("target") or plan.get("target") or "").strip()
    target_type = str(payload.get("type") or plan.get("channel") or "contact").lower()
    if target_type in {"individual", "contact"}:
        return "individual"
    if target_type == "group" or target.endswith("@g.us"):
        return "group"
    return "individual"


def _plan_recipient_scope(plan: dict[str, Any], report_type_key: str) -> str:
    payload = plan.get("payload") if isinstance(plan.get("payload"), dict) else {}
    dispatch_route = (
        payload.get("dispatch_route")
        if isinstance(payload.get("dispatch_route"), dict)
        else {}
    )
    explicit = str(
        dispatch_route.get("recipient_scope")
        or payload.get("recipient_scope")
        or payload.get("recipient_role")
        or plan.get("recipient_scope")
        or plan.get("role")
        or ""
    ).strip().lower()
    if explicit:
        return re.sub(r"[^a-z0-9]+", "_", explicit).strip("_")
    if _plan_recipient_channel(plan) == "group":
        if report_type_key == "wing_summary":
            return "wing"
        route_kind = str(
            dispatch_route.get("route_kind") or plan.get("route_kind") or ""
        ).lower()
        return route_kind if route_kind in {"district", "tehsil", "markaz"} else "other"
    message = str(payload.get("text") or "")
    role_match = re.search(r"\*?Role:\*?\s*([A-Za-z -]+)", message, re.IGNORECASE)
    role = role_match.group(1).strip().lower() if role_match else ""
    if "ddeo" in role:
        return "ddeo"
    if "aeo" in role:
        return "aeo"
    if role == "deo" or "district education officer" in role:
        return "deo"
    if "head" in role:
        return "school_head"
    return "other"


def _plan_target(plan: dict[str, Any]) -> str:
    payload = plan.get("payload") if isinstance(plan.get("payload"), dict) else {}
    value = str(payload.get("target") or plan.get("target") or "").strip()
    return value if _plan_recipient_channel(plan) == "group" else _canonical_contact_target(value)


def _plan_route_label(plan: dict[str, Any], scope_key: str) -> str:
    payload = plan.get("payload") if isinstance(plan.get("payload"), dict) else {}
    route = payload.get("dispatch_route") if isinstance(payload.get("dispatch_route"), dict) else {}
    if scope_key == "district":
        return "District"
    if scope_key == "wing":
        return "Wing"
    if scope_key == "tehsil":
        return str(route.get("tehsil") or "").strip()
    if scope_key == "markaz":
        return str(route.get("markaz") or "").strip()
    return str(route.get("markaz") or route.get("tehsil") or "").strip()


def _same_route_value(left: str, right: str) -> bool:
    normalize = lambda value: re.sub(r"[^a-z0-9]+", "", value.lower())
    return bool(normalize(left)) and normalize(left) == normalize(right)


def _plan_matches_audience_route(
    plan: dict[str, Any],
    *,
    scope_key: str,
    scope_label: str,
    report_type_key: str,
) -> bool:
    if _plan_recipient_channel(plan) != "group":
        return False
    if _plan_recipient_scope(plan, report_type_key) != scope_key:
        return False
    # District and wing source routes have no PostgreSQL area identifier. The
    # profile wing plus configured group authorization are their exact guards.
    if scope_key in {"district", "wing"}:
        return True
    return _same_route_value(_plan_route_label(plan, scope_key), scope_label)


def _retarget_group_plan(
    plan: dict[str, Any],
    *,
    target: str,
    recipient_name: str,
    scope_key: str,
    report_is_message_only: bool,
) -> dict[str, Any]:
    """Adapt source report data to a PostgreSQL-owned audience destination."""
    resolved = copy.deepcopy(plan)
    payload = resolved.get("payload") if isinstance(resolved.get("payload"), dict) else {}
    resolved["target"] = target
    resolved["recipient_name"] = recipient_name
    resolved["channel"] = "group"
    payload["target"] = target
    payload["recipient_name"] = recipient_name
    payload["type"] = "group"
    route = payload.get("dispatch_route") if isinstance(payload.get("dispatch_route"), dict) else {}
    route["recipient_scope"] = scope_key
    payload["dispatch_route"] = route
    if report_is_message_only:
        payload.pop("excel_path", None)
        payload.pop("image_path", None)
        payload.pop("documents", None)
        resolved.pop("excel_path", None)
    resolved["payload"] = payload
    return resolved


def _xlsx_emis_values(path: Path) -> set[str]:
    """Read school EMIS values from a report without trusting its filename."""
    if path.suffix.lower() != ".xlsx":
        return set()
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet in workbook.worksheets:
            rows = sheet.iter_rows(values_only=True)
            for row in rows:
                normalized = [str(value).strip().lower() if value is not None else "" for value in row]
                if "school emis" not in normalized:
                    continue
                index = normalized.index("school emis")
                values: set[str] = set()
                for data_row in rows:
                    if index >= len(data_row) or data_row[index] is None:
                        continue
                    value = str(data_row[index]).strip()
                    if value.endswith(".0") and value[:-2].isdigit():
                        value = value[:-2]
                    if value:
                        values.add(value)
                return values
    finally:
        workbook.close()
    return set()


def _classify_attachment_wings(
    session: Session,
    paths: list[Path],
) -> tuple[set[uuid.UUID], list[str], bool]:
    emis_values: set[str] = set()
    readable_report_found = False
    for path in paths:
        if path.suffix.lower() != ".xlsx" or not path.is_file():
            continue
        try:
            values = _xlsx_emis_values(path)
        except Exception:
            continue
        if values:
            readable_report_found = True
            emis_values.update(values)
    if not emis_values:
        return set(), [], readable_report_found
    schools = session.scalars(select(School).where(School.emis.in_(emis_values))).all()
    found = {school.emis for school in schools}
    return {school.wing_id for school in schools}, sorted(emis_values - found), True


def _classify_scoped_emis_wings(
    session: Session,
    values: list[Any],
) -> tuple[set[uuid.UUID], list[str], bool]:
    """Classify native outputs from the authoritative scope frozen by the renderer."""
    emis_values = {str(value).strip() for value in values if str(value).strip()}
    if not emis_values:
        return set(), [], False
    schools = session.scalars(select(School).where(School.emis.in_(emis_values))).all()
    found = {school.emis for school in schools}
    return {school.wing_id for school in schools}, sorted(emis_values - found), True


def compile_antidengue_preview(
    session: Session,
    *,
    source_job_id: uuid.UUID,
    dispatch_profile_id: uuid.UUID,
    created_by: str = "web",
) -> WhatsAppDispatchPreview:
    source_job = session.get(Job, source_job_id)
    if source_job is None:
        raise PreviewCompileError("Source dry run was not found")
    if source_job.type != JobType.antidengue_report.value:
        raise PreviewCompileError("Only AntiDengue report runs can create this preview")
    if source_job.status != JobStatus.succeeded.value:
        raise PreviewCompileError("The source run must finish successfully before previewing dispatch")
    if not bool(source_job.parameters.get("dry_run", False)):
        raise PreviewCompileError("Dispatch previews must be built from a dry run")

    profile = session.get(WhatsAppDispatchProfile, dispatch_profile_id)
    if profile is None or not profile.enabled:
        raise PreviewCompileError("Select an enabled dispatch profile")
    application = session.get(WhatsAppApplication, profile.application_id)
    report_type = session.get(WhatsAppReportType, profile.report_type_id)
    audience = session.get(WhatsAppAudience, profile.audience_id)
    account = session.get(WhatsAppAccount, profile.account_id)
    template = session.get(WhatsAppTemplate, profile.template_id) if profile.template_id else None
    wing = session.get(Wing, profile.wing_id) if profile.wing_id else None
    recipient_scope = (
        session.get(WhatsAppRecipientScope, profile.recipient_scope_id)
        if profile.recipient_scope_id
        else None
    )
    if not all((application, report_type, audience, account)):
        raise PreviewCompileError("The selected dispatch profile has incomplete configuration")
    if application.key != "antidengue":
        raise PreviewCompileError("Select an AntiDengue dispatch profile")
    if wing is None:
        raise PreviewCompileError("AntiDengue dispatch profiles require a wing")

    members = session.scalars(
        select(WhatsAppAudienceMember).where(
            WhatsAppAudienceMember.audience_id == audience.id,
            WhatsAppAudienceMember.enabled.is_(True),
        )
    ).all()
    audience_group_ids = {
        member.directory_group_id for member in members if member.directory_group_id
    }
    audience_contact_ids = {
        member.directory_contact_id for member in members if member.directory_contact_id
    }

    summary = (source_job.result or {}).get("summary") or {}
    whatsapp_summary = summary.get("whatsapp") if isinstance(summary, dict) else {}
    source_dispatch_plan = whatsapp_summary.get("dispatch_plan") if isinstance(whatsapp_summary, dict) else []
    if not isinstance(source_dispatch_plan, list):
        source_dispatch_plan = []
    batch_issues: list[dict[str, Any]] = []
    dispatch_plan: list[dict[str, Any]] = []
    source_plans = [plan for plan in source_dispatch_plan if isinstance(plan, dict)]
    native_renderer_used = False
    for member in members:
        if profile.recipient_channel == "group" and member.target_type != "group":
            continue
        if profile.recipient_channel == "individual" and member.target_type != "contact":
            continue
        if member.target_type == "group":
            directory_group = session.get(WhatsAppDirectoryGroup, member.directory_group_id)
            if directory_group is None:
                batch_issues.append(issue("missing_audience_target", "blocked", "An audience group no longer exists in the synchronized directory."))
                continue
            expected_scope = recipient_scope.key if recipient_scope else member.route_scope_key
            if report_type.key == "wing_summary":
                native_renderer_used = True
                if member.route_scope_key != expected_scope or expected_scope not in {"wing", "district"}:
                    batch_issues.append(
                        issue(
                            "audience_scope_mismatch",
                            "blocked",
                            f"{directory_group.name} must be bound to this profile's Wing or District scope.",
                        )
                    )
                    continue
                try:
                    area_id = uuid.UUID(member.route_scope_value)
                except ValueError:
                    batch_issues.append(
                        issue(
                            "invalid_audience_area",
                            "blocked",
                            f"{directory_group.name} has an invalid hierarchy-area binding.",
                        )
                    )
                    continue
                expected_area_id = wing.id if expected_scope == "wing" else wing.district_id
                if area_id != expected_area_id:
                    batch_issues.append(
                        issue(
                            "cross_wing_audience_area",
                            "blocked",
                            f"{directory_group.name} is not bound to {wing.name}.",
                        )
                    )
                    continue
                try:
                    rendered = render_wing_dormant_report(
                        session,
                        source_job=source_job,
                        wing=wing,
                        recipient_name=directory_group.name,
                        presentation_policy=profile.presentation_policy,
                    )
                except (ValueError, OSError) as exc:
                    batch_issues.append(issue("native_report_error", "blocked", str(exc)))
                    continue
                if not rendered.schools:
                    batch_issues.extend(rendered.issues)
                    continue
                dispatch_plan.append(
                    {
                        "job_id": f"{WING_RENDERER_KEY}:{member.id}",
                        "status": "planned",
                        "channel": "group",
                        "target": directory_group.jid,
                        "recipient_name": directory_group.name,
                        "route_kind": expected_scope,
                        "row_count": len(rendered.schools),
                        "native_renderer": WING_RENDERER_KEY,
                        "native_context": rendered.context,
                        "native_issues": rendered.issues,
                        "source_artifact_id": rendered.source_artifact_id,
                        "source_artifact_sha256": rendered.source_artifact_sha256,
                        "scoped_emis": [school.emis for school in rendered.schools],
                        "payload": {
                            "type": "group",
                            "target": directory_group.jid,
                            "recipient_name": directory_group.name,
                            "text": rendered.message,
                            "attachment_paths": [str(path) for path in rendered.attachment_paths],
                            "dispatch_route": {
                                "route_kind": expected_scope,
                                "recipient_scope": expected_scope,
                                "wing": wing.name,
                                "wing_id": str(wing.id),
                                "row_count": len(rendered.schools),
                            },
                        },
                    }
                )
                continue
            if report_type.key == "tehsil_dormant_summary":
                native_renderer_used = True
                if expected_scope != "tehsil" or member.route_scope_key != "tehsil":
                    batch_issues.append(
                        issue(
                            "audience_scope_mismatch",
                            "blocked",
                            f"{directory_group.name} must be bound to a Tehsil for {report_type.name}.",
                        )
                    )
                    continue
                try:
                    tehsil_id = uuid.UUID(member.route_scope_value)
                    rendered = render_tehsil_dormant_report(
                        session,
                        source_job=source_job,
                        wing=wing,
                        tehsil_id=tehsil_id,
                        recipient_name=directory_group.name,
                        deadline=get_settings().antidengue_submission_deadline,
                        presentation_policy=profile.presentation_policy,
                    )
                except (ValueError, OSError) as exc:
                    batch_issues.append(
                        issue("native_report_error", "blocked", str(exc))
                    )
                    continue
                if not rendered.schools:
                    batch_issues.extend(rendered.issues)
                    continue
                dispatch_plan.append(
                    {
                        "job_id": f"{TEHSIL_DORMANT_RENDERER_KEY}:{member.id}",
                        "status": "planned",
                        "channel": "group",
                        "target": directory_group.jid,
                        "recipient_name": directory_group.name,
                        "route_kind": "tehsil",
                        "row_count": len(rendered.schools),
                        "native_renderer": TEHSIL_DORMANT_RENDERER_KEY,
                        "native_context": rendered.context,
                        "native_issues": rendered.issues,
                        "source_artifact_id": rendered.source_artifact_id,
                        "source_artifact_sha256": rendered.source_artifact_sha256,
                        "scoped_emis": [school.emis for school in rendered.schools],
                        "payload": {
                            "type": "group",
                            "target": directory_group.jid,
                            "recipient_name": directory_group.name,
                            "text": rendered.message,
                            "attachment_paths": [str(path) for path in rendered.attachment_paths],
                            "dispatch_route": {
                                "route_kind": "tehsil",
                                "recipient_scope": "tehsil",
                                "tehsil": member.route_scope_label,
                                "tehsil_id": member.route_scope_value,
                                "row_count": len(rendered.schools),
                            },
                        },
                    }
                )
                continue
            exact_target_plan = next(
                (
                    plan
                    for plan in source_plans
                    if _plan_recipient_channel(plan) == "group"
                    and _plan_target(plan) == directory_group.jid
                    and (not expected_scope or _plan_recipient_scope(plan, report_type.key) == expected_scope)
                ),
                None,
            )
            if recipient_scope and member.route_scope_key != recipient_scope.key:
                batch_issues.append(
                    issue(
                        "audience_scope_mismatch",
                        "blocked",
                        f"{directory_group.name} is not bound to {recipient_scope.name}.",
                    )
                )
                continue
            if not member.route_scope_key or not member.route_scope_label:
                if exact_target_plan is None:
                    batch_issues.append(
                        issue(
                            "unbound_audience_route",
                            "blocked",
                            f"Bind {directory_group.name} to the hierarchy area it represents.",
                        )
                    )
                    continue
                selected_plan = exact_target_plan
                expected_scope = _plan_recipient_scope(selected_plan, report_type.key)
            else:
                selected_plan = next(
                    (
                        plan
                        for plan in source_plans
                        if _plan_matches_audience_route(
                            plan,
                            scope_key=member.route_scope_key,
                            scope_label=member.route_scope_label,
                            report_type_key=report_type.key,
                        )
                    ),
                    None,
                )
                if selected_plan is None:
                    batch_issues.append(
                        issue(
                            "missing_scope_report",
                            "blocked",
                            f"The dry run has no {member.route_scope_key} report for {member.route_scope_label}.",
                        )
                    )
                    continue
            dispatch_plan.append(
                _retarget_group_plan(
                    selected_plan,
                    target=directory_group.jid,
                    recipient_name=directory_group.name,
                    scope_key=expected_scope or "other",
                    report_is_message_only=report_type.artifact_kind == "message",
                )
            )
        else:
            directory_contact = session.get(WhatsAppDirectoryContact, member.directory_contact_id)
            if directory_contact is None:
                batch_issues.append(issue("missing_audience_target", "blocked", "An audience contact no longer exists in the synchronized directory."))
                continue
            targets = {
                _canonical_contact_target(value)
                for value in (
                    directory_contact.phone_jid,
                    directory_contact.primary_lid_jid,
                    directory_contact.canonical_key,
                )
                if value
            }
            selected_plan = next(
                (
                    plan
                    for plan in source_plans
                    if _plan_recipient_channel(plan) == "individual"
                    and _plan_target(plan) in targets
                    and (recipient_scope is None or _plan_recipient_scope(plan, report_type.key) == recipient_scope.key)
                ),
                None,
            )
            if selected_plan is None:
                batch_issues.append(
                    issue(
                        "missing_individual_report",
                        "blocked",
                        f"The dry run has no matching individual report for {directory_contact.display_name or directory_contact.canonical_key}.",
                    )
                )
                continue
            dispatch_plan.append(copy.deepcopy(selected_plan))

    quality = summary.get("quality_gate") if isinstance(summary, dict) else {}
    for warning in (quality.get("warnings") or []) if isinstance(quality, dict) else []:
        batch_issues.append(issue("report_quality_warning", "warning", str(warning)))
    zero_result = (
        (isinstance(quality, dict) and quality.get("final_school_count") == 0)
        or (
            isinstance(whatsapp_summary, dict)
            and str(whatsapp_summary.get("skipped_reason") or "").lower()
            == "zero inactive schools"
        )
    )
    if not dispatch_plan:
        if zero_result:
            batch_issues.append(
                issue(
                    "nothing_to_dispatch",
                    "warning",
                    "The dry run found no dormant schools, so there are no WhatsApp deliveries to preview.",
                )
            )
        elif native_renderer_used and batch_issues:
            pass
        elif source_dispatch_plan and not batch_issues:
            batch_issues.append(
                issue(
                    "no_audience_routes",
                    "blocked",
                    f"The selected audience has no routes for {report_type.name}.",
                )
            )
        else:
            batch_issues.append(
                issue(
                    "empty_dispatch_plan",
                    "blocked",
                    "The dry run did not produce an exact WhatsApp dispatch plan.",
                )
            )

    configuration_snapshot = {
        "application": {"id": str(application.id), "key": application.key, "name": application.name},
        "profile": {
            "id": str(profile.id),
            "key": profile.key,
            "name": profile.name,
            "version": profile.version,
            "delivery_mode": profile.delivery_mode,
            "require_approval": profile.require_approval,
            "recipient_channel": profile.recipient_channel,
            "recipient_scope_id": str(recipient_scope.id) if recipient_scope else None,
            "recipient_scope_key": recipient_scope.key if recipient_scope else None,
            "recipient_scope_name": recipient_scope.name if recipient_scope else None,
            "presentation_policy": profile.presentation_policy or {},
        },
        "report_type": {"id": str(report_type.id), "key": report_type.key, "name": report_type.name},
        "audience": {"id": str(audience.id), "key": audience.key, "name": audience.name},
        "account": {"id": str(account.id), "worker_key": account.worker_key, "name": account.name},
        "template": {
            "id": str(template.id) if template else None,
            "key": template.key if template else None,
            "name": template.name if template else None,
            "body": template.body if template else None,
        },
        "wing": {"id": str(wing.id), "code": wing.code, "name": wing.name},
        "renderer": {
            "key": (
                TEHSIL_DORMANT_RENDERER_KEY
                if report_type.key == "tehsil_dormant_summary"
                else WING_RENDERER_KEY
                if report_type.key == "wing_summary"
                else "legacy.dispatch_plan_adapter"
            ),
        },
    }
    configuration_snapshot["audience"]["target_keys"] = sorted(member.target_key for member in members)
    configuration_snapshot["audience"]["target_routes"] = sorted(
        f"{member.target_key}:{member.route_scope_key}:{member.route_scope_value}"
        for member in members
    )
    preview = WhatsAppDispatchPreview(
        preview_key=_preview_key(),
        application_id=application.id,
        source_job_id=source_job.id,
        dispatch_profile_id=profile.id,
        profile_version=profile.version,
        application_name=application.name,
        report_type_name=report_type.name,
        audience_name=audience.name,
        profile_name=profile.name,
        account_name=account.name,
        template_name=template.name if template else "Exact dry-run message",
        wing_name=wing.name,
        issues=batch_issues,
        configuration_snapshot=configuration_snapshot,
        created_by=created_by,
    )
    session.add(preview)
    session.flush()

    plans_with_paths = [(_plan, _attachment_paths(_plan)) for _plan in dispatch_plan]
    referenced_paths = {str(path) for _, paths in plans_with_paths for path in paths}
    job_artifacts = session.scalars(
        select(Artifact).where(Artifact.job_id == source_job.id).order_by(Artifact.created_at)
    ).all()
    artifacts_by_path = {
        str(Path(item.path).expanduser().resolve(strict=False)): item for item in job_artifacts
    }
    snapshot_artifacts: dict[str, WhatsAppDispatchPreviewArtifact] = {}

    def snapshot_artifact(path: Path, artifact: Artifact | None = None) -> WhatsAppDispatchPreviewArtifact:
        source_path = path.resolve(strict=False)
        key = str(source_path)
        existing = snapshot_artifacts.get(key)
        if existing:
            if key in referenced_paths:
                existing.role = "delivery"
            return existing
        problems: list[dict[str, Any]] = []
        checksum = ""
        size = 0
        artifact_status = "ready"
        frozen_path = source_path
        if not path.exists() or not path.is_file():
            artifact_status = "blocked"
            problems.append(issue("missing_attachment", "blocked", f"Attachment is missing: {path.name}"))
        else:
            size = path.stat().st_size
            checksum = sha256_file(path)
            if size == 0:
                artifact_status = "blocked"
                problems.append(issue("empty_attachment", "blocked", f"Attachment is empty: {path.name}"))
        if artifact is None and key in referenced_paths:
            artifact_status = "blocked"
            problems.append(
                issue(
                    "untracked_attachment",
                    "blocked",
                    f"Attachment is not registered to the source dry run: {path.name}",
                )
            )
        if artifact is not None and artifact_status == "ready":
            frozen_path = freeze_artifact(path, checksum)
        snapshot = WhatsAppDispatchPreviewArtifact(
            preview_id=preview.id,
            artifact_id=artifact.id if artifact else None,
            report_type_id=report_type.id,
            wing_id=wing.id,
            role=_artifact_role(path, referenced_paths),
            name=artifact.name if artifact else path.name,
            path_snapshot=str(frozen_path),
            mime_type=mimetypes.guess_type(path.name)[0] or "application/octet-stream",
            size_bytes=size,
            checksum_sha256=checksum,
            status=artifact_status,
            issues=problems,
        )
        session.add(snapshot)
        session.flush()
        snapshot_artifacts[key] = snapshot
        return snapshot

    for _, paths in plans_with_paths:
        for path in paths:
            snapshot_artifact(path, artifacts_by_path.get(str(path)))

    seen_signatures: set[str] = set()
    for sequence, (plan, attachment_paths) in enumerate(plans_with_paths, start=1):
        payload = plan.get("payload") if isinstance(plan.get("payload"), dict) else {}
        requested_target = str(payload.get("target") or plan.get("target") or "").strip()
        recipient_name = str(
            payload.get("recipient_name")
            or plan.get("recipient_name")
            or requested_target
        )
        target_type = str(payload.get("type") or plan.get("channel") or "contact").lower()
        target_type = "group" if target_type == "group" or requested_target.endswith("@g.us") else "contact"
        target = requested_target if target_type == "group" else _canonical_contact_target(requested_target)
        source_status = str(plan.get("status") or "planned").strip().lower()
        source_skipped = source_status == "skipped"
        legacy_message = str(payload.get("text") or "")
        dispatch_route = payload.get("dispatch_route") if isinstance(payload.get("dispatch_route"), dict) else {}
        route_kind = str(dispatch_route.get("route_kind") or plan.get("route_kind") or "")
        route_scope = str(
            dispatch_route.get("markaz")
            or dispatch_route.get("tehsil")
            or dispatch_route.get("wing")
            or ("District" if route_kind == "district" else "")
        )
        problems: list[dict[str, Any]] = []
        native_context = (
            plan.get("native_context")
            if isinstance(plan.get("native_context"), dict)
            else {}
        )
        native_issues = (
            plan.get("native_issues")
            if isinstance(plan.get("native_issues"), list)
            else []
        )
        directory_group: WhatsAppDirectoryGroup | None = None
        directory_contact: WhatsAppDirectoryContact | None = None
        contact_link: WhatsAppContactLink | None = None
        target_wing: Wing | None = None

        if source_skipped:
            problems.append(
                issue(
                    "source_route_skipped",
                    "warning",
                    str(plan.get("cause") or "The source dry run intentionally skipped this route."),
                )
            )
        elif source_status not in {"planned", "queued", "ready"}:
            problems.append(
                issue(
                    "source_route_not_planned",
                    "blocked",
                    str(plan.get("cause") or f"The source route has status: {source_status}"),
                )
            )
        else:
            problems.extend(
                item for item in native_issues if isinstance(item, dict)
            )

        if not target and not source_skipped:
            problems.append(issue("missing_target", "blocked", "The planned delivery has no WhatsApp target."))
        if not source_skipped and target_type == "group":
            if profile.delivery_mode == "individuals":
                problems.append(issue("delivery_mode_mismatch", "blocked", "A group route cannot use an individuals-only profile."))
            directory_group = session.scalar(
                select(WhatsAppDirectoryGroup).where(
                    WhatsAppDirectoryGroup.account_id == account.id,
                    WhatsAppDirectoryGroup.jid == target,
                )
            )
            if directory_group is None:
                problems.append(issue("unknown_group", "blocked", "The target group is not present in the synchronized directory."))
            else:
                if not directory_group.available:
                    problems.append(issue("unavailable_group", "blocked", "The target group is not currently available."))
                if directory_group.id not in audience_group_ids:
                    problems.append(issue("outside_audience", "blocked", "The target group is not in the selected audience."))
                configured_group = session.scalar(
                    select(WhatsAppGroup).where(
                        WhatsAppGroup.account_id == account.id,
                        or_(
                            WhatsAppGroup.directory_group_id == directory_group.id,
                            WhatsAppGroup.jid == target,
                        ),
                    )
                )
                if configured_group is None:
                    problems.append(issue("missing_wing_authorization", "blocked", "Assign this detected group to a wing before dispatch."))
                else:
                    target_wing = session.get(Wing, configured_group.wing_id)
                    if not configured_group.enabled:
                        problems.append(issue("disabled_route", "blocked", "The configured group route is disabled."))
                    if configured_group.wing_id != wing.id:
                        problems.append(
                            issue(
                                "cross_wing_route",
                                "blocked",
                                f"The group belongs to {target_wing.name if target_wing else 'another wing'}, not {wing.name}.",
                            )
                        )
        elif not source_skipped:
            if profile.delivery_mode == "groups":
                problems.append(issue("delivery_mode_mismatch", "blocked", "An individual route cannot use a groups-only profile."))
            directory_contact = session.scalar(
                select(WhatsAppDirectoryContact).where(
                    WhatsAppDirectoryContact.account_id == account.id,
                    or_(
                        WhatsAppDirectoryContact.phone_jid == target,
                        WhatsAppDirectoryContact.primary_lid_jid == target,
                        WhatsAppDirectoryContact.canonical_key == target,
                    ),
                )
            )
            if directory_contact is None:
                problems.append(issue("unknown_contact", "blocked", "The individual is not present in the synchronized directory."))
            else:
                if not directory_contact.active:
                    problems.append(issue("inactive_contact", "blocked", "The WhatsApp contact is inactive."))
                if directory_contact.id not in audience_contact_ids:
                    problems.append(issue("outside_audience", "blocked", "The individual is not in the selected audience."))
                contact_links = session.scalars(
                    select(WhatsAppContactLink).where(
                        WhatsAppContactLink.directory_contact_id == directory_contact.id,
                        WhatsAppContactLink.active.is_(True),
                        WhatsAppContactLink.status == "verified",
                    )
                ).all()
                if not contact_links:
                    problems.append(issue("missing_master_data_link", "blocked", "Link this contact to an officer or school head."))
                elif len(contact_links) > 1:
                    problems.append(issue("ambiguous_master_data_link", "blocked", "This contact has multiple verified master-data links; keep only the intended recipient link."))
                else:
                    contact_link = contact_links[0]
                    target_wing = session.get(Wing, contact_link.wing_id)
                    if contact_link.wing_id != wing.id:
                        problems.append(
                            issue(
                                "cross_wing_route",
                                "blocked",
                                f"The linked recipient belongs to {target_wing.name if target_wing else 'another wing'}, not {wing.name}.",
                            )
                        )

        if not source_skipped and not account.enabled:
            problems.append(issue("disabled_connection", "blocked", "The selected WhatsApp connection is disabled."))
        elif not source_skipped and not account.connected:
            problems.append(issue("connection_offline", "warning", "The connection is offline; the frozen preview remains reviewable."))

        attachment_snapshots = [snapshot_artifact(path, artifacts_by_path.get(str(path))) for path in attachment_paths]
        if not source_skipped:
            for attachment in attachment_snapshots:
                problems.extend(attachment.issues)
        if not source_skipped and attachment_paths:
            if plan.get("native_renderer"):
                attachment_wing_ids, unknown_emis, classified = _classify_scoped_emis_wings(
                    session, plan.get("scoped_emis") or []
                )
            else:
                attachment_wing_ids, unknown_emis, classified = _classify_attachment_wings(
                    session, attachment_paths
                )
            if unknown_emis:
                problems.append(
                    issue(
                        "unknown_report_schools",
                        "blocked",
                        f"{len(unknown_emis)} school EMIS value(s) are absent from PostgreSQL master data.",
                        emis=unknown_emis[:20],
                    )
                )
            if len(attachment_wing_ids) > 1:
                problems.append(
                    issue(
                        "mixed_wing_report",
                        "blocked",
                        "The attachment contains schools from more than one wing.",
                    )
                )
            elif attachment_wing_ids and wing.id not in attachment_wing_ids:
                source_wing = session.get(Wing, next(iter(attachment_wing_ids)))
                problems.append(
                    issue(
                        "cross_wing_report",
                        "blocked",
                        f"The report contains {source_wing.name if source_wing else 'another wing'} schools, not {wing.name} schools.",
                    )
                )
            elif not attachment_wing_ids:
                problems.append(
                    issue(
                        "unclassified_report_wing",
                        "blocked",
                        "The report wing could not be verified from school EMIS values.",
                        readable_report_found=classified,
                    )
                )
        attachment_required = report_type.artifact_kind != "message"
        if plan.get("native_renderer"):
            policy = normalize_presentation_policy(
                profile.presentation_policy,
                report_key=report_type.key,
            )
            attachment_required = policy["attachment_mode"] != "none"
        if not source_skipped and not attachment_paths and attachment_required:
            problems.append(
                issue(
                    "missing_report_attachment",
                    "blocked",
                    f"{report_type.name} requires a delivery attachment, but the source route has none.",
                )
            )
        context = {
            "recipient_name": recipient_name,
            "wing": wing.name,
            "report_name": report_type.name,
            "scope": route_scope or route_kind or "Configured scope",
            "row_count": str(dispatch_route.get("row_count") or plan.get("row_count") or ""),
            **{str(key): str(value) for key, value in native_context.items()},
        }
        template_for_render = template
        if plan.get("native_renderer"):
            template_variables = set(PLACEHOLDER_RE.findall(template.body)) if template else set()
            if template and not template_variables.intersection({"message", "report_body"}):
                # Native renderers own the complete, validated hierarchy body.
                # A template opts into wrapping it explicitly with {{report_body}}
                # or the backwards-compatible {{message}} variable.
                template_for_render = None
                problems.append(
                    issue(
                        "static_template_ignored",
                        "warning",
                        "The selected static template cannot include the native report body; add {{report_body}} to customize this message.",
                    )
                )
        message, template_issues = _render_message(template_for_render, legacy_message, context)
        if not source_skipped:
            problems.extend(template_issues)
        if not source_skipped and not message.strip():
            problems.append(issue("empty_message", "blocked", "The rendered WhatsApp message is empty."))
        elif not source_skipped and len(message) > 3500:
            problems.append(issue("long_message", "warning", "The rendered message exceeds 3,500 characters."))

        dedupe_payload = {
            "target": target,
            "message": message,
            "attachments": [item.checksum_sha256 or item.path_snapshot for item in attachment_snapshots],
        }
        dedupe_signature = hashlib.sha256(
            json.dumps(dedupe_payload, sort_keys=True, ensure_ascii=False).encode("utf-8")
        ).hexdigest()
        if not source_skipped and dedupe_signature in seen_signatures:
            problems.append(issue("duplicate_route", "blocked", "This exact recipient, message and attachment route appears more than once."))
        if not source_skipped:
            seen_signatures.add(dedupe_signature)
        idempotency_key = hashlib.sha256(
            f"{source_job.id}:{sequence}:{dedupe_signature}".encode("utf-8")
        ).hexdigest()
        status = "skipped" if source_skipped else _delivery_status(problems)
        delivery = WhatsAppDispatchPreviewDelivery(
            preview_id=preview.id,
            sequence=sequence,
            source_route_key=str(plan.get("job_id") or payload.get("job_id") or ""),
            target_type=target_type,
            target_name=recipient_name,
            target_jid=target,
            directory_group_id=directory_group.id if directory_group else None,
            directory_contact_id=directory_contact.id if directory_contact else None,
            contact_link_id=contact_link.id if contact_link else None,
            wing_id=target_wing.id if target_wing else None,
            wing_name=target_wing.name if target_wing else "Unresolved",
            route_kind=route_kind,
            route_scope=route_scope,
            message=message,
            attachment_ids=[str(item.id) for item in attachment_snapshots],
            routing_snapshot={
                "dispatch_route": dispatch_route,
                "profile_version": profile.version,
                "audience_id": str(audience.id),
                "account_id": str(account.id),
                "template_id": str(template.id) if template else None,
                "legacy_plan_job_id": str(plan.get("job_id") or payload.get("job_id") or ""),
                "source_status": source_status,
                "source_cause": str(plan.get("cause") or ""),
                "requested_target": requested_target,
                "recipient_channel": profile.recipient_channel,
                "recipient_scope_key": (
                    _plan_recipient_scope(plan, report_type.key)
                    if isinstance(plan, dict)
                    else None
                ),
                "native_renderer": str(plan.get("native_renderer") or ""),
                "source_artifact_id": plan.get("source_artifact_id"),
                "source_artifact_sha256": str(plan.get("source_artifact_sha256") or ""),
                "scoped_emis": plan.get("scoped_emis") or [],
                "contact_entity_type": contact_link.entity_type if contact_link else None,
                "contact_entity_key": contact_link.entity_key if contact_link else None,
            },
            issues=problems,
            status=status,
            idempotency_key=idempotency_key,
        )
        session.add(delivery)

    session.flush()
    deliveries = session.scalars(
        select(WhatsAppDispatchPreviewDelivery).where(
            WhatsAppDispatchPreviewDelivery.preview_id == preview.id
        )
    ).all()
    preview.ready_count = sum(item.status == "ready" for item in deliveries)
    preview.warning_count = sum(
        problem.get("severity") == "warning"
        for item in deliveries
        for problem in item.issues
    ) + sum(item["severity"] == "warning" for item in batch_issues)
    preview.blocked_count = sum(
        problem.get("severity") == "blocked"
        for item in deliveries
        for problem in item.issues
    ) + sum(item["severity"] == "blocked" for item in batch_issues)
    preview.skipped_count = sum(item.status == "skipped" for item in deliveries)
    preview.delivery_count = len(deliveries)
    preview.artifact_count = len(snapshot_artifacts)
    active_attachment_ids = {
        artifact_id
        for delivery in deliveries
        if delivery.status != "skipped"
        for artifact_id in delivery.attachment_ids
    }
    artifact_blocked = any(
        str(item.id) in active_attachment_ids and item.status == "blocked"
        for item in snapshot_artifacts.values()
    )
    batch_blocked = any(item["severity"] == "blocked" for item in batch_issues)
    preview.status = "blocked" if preview.blocked_count or artifact_blocked or batch_blocked else "ready"
    artifacts_by_snapshot_id = {str(item.id): item for item in snapshot_artifacts.values()}
    frozen_content = {
        "configuration": configuration_snapshot,
        "deliveries": [
            {
                "sequence": item.sequence,
                "target": item.target_jid,
                "message": item.message,
                "attachments": [
                    artifacts_by_snapshot_id[artifact_id].checksum_sha256
                    for artifact_id in item.attachment_ids
                    if artifact_id in artifacts_by_snapshot_id
                ],
                "status": item.status,
                "issues": item.issues,
                "routing": item.routing_snapshot,
            }
            for item in sorted(deliveries, key=lambda value: value.sequence)
        ],
    }
    preview.content_sha256 = hashlib.sha256(
        json.dumps(frozen_content, sort_keys=True, ensure_ascii=False).encode("utf-8")
    ).hexdigest()
    session.add(preview)
    session.commit()
    session.refresh(preview)
    return preview


def preview_is_stale(session: Session, preview: WhatsAppDispatchPreview, *, check_files: bool = False) -> bool:
    profile = session.get(WhatsAppDispatchProfile, preview.dispatch_profile_id)
    if profile is None or profile.version != preview.profile_version or not profile.enabled:
        return True
    snapshot = preview.configuration_snapshot or {}
    audience_snapshot = snapshot.get("audience") or {}
    audience = session.get(WhatsAppAudience, profile.audience_id)
    if audience is None or not audience.enabled:
        return True
    current_targets = sorted(
        session.scalars(
            select(WhatsAppAudienceMember.target_key).where(
                WhatsAppAudienceMember.audience_id == audience.id,
                WhatsAppAudienceMember.enabled.is_(True),
            )
        ).all()
    )
    if current_targets != sorted(audience_snapshot.get("target_keys") or []):
        return True
    current_routes = sorted(
        f"{member.target_key}:{member.route_scope_key}:{member.route_scope_value}"
        for member in session.scalars(
            select(WhatsAppAudienceMember).where(
                WhatsAppAudienceMember.audience_id == audience.id,
                WhatsAppAudienceMember.enabled.is_(True),
            )
        ).all()
    )
    if current_routes != sorted(audience_snapshot.get("target_routes") or []):
        return True
    account_snapshot = snapshot.get("account") or {}
    account = session.get(WhatsAppAccount, profile.account_id)
    if account is None or not account.enabled or account.worker_key != account_snapshot.get("worker_key"):
        return True
    template_snapshot = snapshot.get("template") or {}
    template_id = template_snapshot.get("id")
    if template_id:
        template = session.get(WhatsAppTemplate, uuid.UUID(template_id))
        if template is None or not template.enabled or template.body != template_snapshot.get("body"):
            return True
    elif profile.template_id is not None:
        return True
    if not check_files:
        return False
    artifacts = session.scalars(
        select(WhatsAppDispatchPreviewArtifact).where(
            WhatsAppDispatchPreviewArtifact.preview_id == preview.id,
            WhatsAppDispatchPreviewArtifact.role == "delivery",
        )
    ).all()
    for artifact in artifacts:
        path = Path(artifact.path_snapshot)
        if not path.is_file() or path.stat().st_size != artifact.size_bytes:
            return True
        if artifact.checksum_sha256 and sha256_file(path) != artifact.checksum_sha256:
            return True
    return False


def preview_dict(session: Session, preview: WhatsAppDispatchPreview, *, check_files: bool = False) -> dict[str, Any]:
    stale = preview_is_stale(session, preview, check_files=check_files)
    return {
        "id": str(preview.id),
        "preview_key": preview.preview_key,
        "application_id": str(preview.application_id),
        "source_job_id": str(preview.source_job_id),
        "dispatch_profile_id": str(preview.dispatch_profile_id),
        "status": "stale" if stale else preview.status,
        "stored_status": preview.status,
        "stale": stale,
        "profile_version": preview.profile_version,
        "application_name": preview.application_name,
        "report_type_name": preview.report_type_name,
        "audience_name": preview.audience_name,
        "profile_name": preview.profile_name,
        "account_name": preview.account_name,
        "template_name": preview.template_name,
        "wing_name": preview.wing_name,
        "ready_count": preview.ready_count,
        "warning_count": preview.warning_count,
        "blocked_count": preview.blocked_count,
        "skipped_count": preview.skipped_count,
        "delivery_count": preview.delivery_count,
        "artifact_count": preview.artifact_count,
        "issues": preview.issues,
        "configuration_snapshot": preview.configuration_snapshot,
        "content_sha256": preview.content_sha256,
        "created_by": preview.created_by,
        "created_at": preview.created_at,
        "frozen_at": preview.frozen_at,
    }


def delete_preview_records(
    session: Session,
    preview: WhatsAppDispatchPreview,
) -> set[Path]:
    """Delete one immutable preview and return managed files eligible for cleanup."""
    deliveries = session.scalars(
        select(WhatsAppDispatchPreviewDelivery).where(
            WhatsAppDispatchPreviewDelivery.preview_id == preview.id
        )
    ).all()
    snapshots = session.scalars(
        select(WhatsAppDispatchPreviewArtifact).where(
            WhatsAppDispatchPreviewArtifact.preview_id == preview.id
        )
    ).all()
    paths = {
        Path(item.path_snapshot)
        for item in snapshots
        if is_managed_preview_artifact(Path(item.path_snapshot))
    }
    for item in deliveries:
        session.delete(item)
    for item in snapshots:
        session.delete(item)
    session.flush()
    session.delete(preview)
    session.flush()
    return paths


def cleanup_unreferenced_preview_files(session: Session, paths: set[Path]) -> None:
    for path in paths:
        remaining = session.scalar(
            select(WhatsAppDispatchPreviewArtifact.id).where(
                WhatsAppDispatchPreviewArtifact.path_snapshot == str(path)
            )
        )
        if remaining is None:
            path.unlink(missing_ok=True)


def entity_link_details(session: Session, link: WhatsAppContactLink) -> dict[str, Any]:
    wing = session.get(Wing, link.wing_id)
    if link.entity_type == "officer":
        from master_data.models import Officer

        entity = session.get(Officer, link.officer_id)
        name = entity.name if entity else "Deleted officer"
        detail = entity.role.upper() if entity else "Officer"
    else:
        head = session.get(SchoolHead, link.school_head_id)
        school = session.get(School, head.school_id) if head else None
        name = head.name if head else "Deleted school head"
        detail = f"{school.emis} · {school.name}" if school else "School head"
    return {
        "id": str(link.id),
        "contact_id": str(link.directory_contact_id),
        "entity_type": link.entity_type,
        "entity_key": link.entity_key,
        "entity_name": name,
        "entity_detail": detail,
        "wing_id": str(link.wing_id),
        "wing_name": wing.name if wing else "Unknown wing",
        "status": link.status,
        "source": link.source,
        "confidence": link.confidence,
        "active": link.active,
        "updated_at": link.updated_at,
    }
