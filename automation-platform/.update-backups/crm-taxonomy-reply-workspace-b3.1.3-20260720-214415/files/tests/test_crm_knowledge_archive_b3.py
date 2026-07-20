from __future__ import annotations

import io
import json
import uuid
import zipfile
from datetime import datetime, timezone

from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy.pool import StaticPool
from sqlmodel import SQLModel, Session, create_engine

import automation_core.models  # noqa: F401
import master_data.models  # noqa: F401
import whatsapp_gateway.models  # noqa: F401
from automation_core.config import Settings
from automation_core.database import get_session
from crm_domain.knowledge import ComplaintKnowledgeArchive, KnowledgeFilters
from crm_domain.knowledge_api import router
from crm_domain.models import ComplaintCase, ComplaintReply


def engine():
    value = create_engine("sqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool)
    SQLModel.metadata.create_all(value)
    return value


def settings() -> Settings:
    return Settings(
        _env_file=None,
        database_url="sqlite://",
        frappe_helpdesk_approved_reply_statuses="Approved,Issued",
        frappe_helpdesk_crm_public_url="http://crm.test",
        paperless_url="http://paperless.test",
    )


def add_case(
    session: Session,
    number: str,
    *,
    category: str,
    subcategory: str = "General",
    complaint: str = "Complaint details",
    approval: str = "Not Prepared",
    reply: str | None = None,
    ai_eligible: bool = False,
    source: str = "crm_portal",
) -> ComplaintCase:
    case = ComplaintCase(
        complaint_number=number,
        state="published",
        category=category,
        sub_category=subcategory,
        remarks=complaint,
        source_system=source,
        canonical_paperless_document_id=100,
        frappe_ticket_id=f"T-{number}",
        frappe_ticket_url=f"http://helpdesk.test/{number}",
        frappe_sync_status="synchronized",
        frappe_reply_approval_status=approval,
        frappe_ai_eligible=ai_eligible,
        frappe_inquiry_findings="Field inquiry completed.",
        frappe_school_version="School denied the allegation.",
        frappe_applicable_policy="Applicable departmental policy.",
        frappe_disposal_outcome="Addressed",
        frappe_workflow_status="Closed",
        created_at=datetime(2026, 7, 20, 5, 0, tzinfo=timezone.utc),
    )
    session.add(case)
    session.flush()
    if reply is not None:
        session.add(
            ComplaintReply(
                complaint_case_id=case.id,
                reply_text=reply,
                source_filename=f"frappe-helpdesk/{number}.txt",
                source_row=0,
            )
        )
    session.commit()
    return case


def seed(session: Session) -> None:
    add_case(
        session,
        "104-7000001",
        category="Administrative",
        subcategory="School timing",
        complaint="School opened late.",
        approval="Approved",
        reply="The matter was examined and the school was sensitized.",
        ai_eligible=True,
    )
    add_case(
        session,
        "104-7000002",
        category="Teacher behaviour",
        subcategory="Conduct",
        complaint="Teacher used inappropriate language.",
        approval="Draft",
        reply="Unapproved draft reply must not leak.",
    )
    add_case(
        session,
        "104-7000003",
        category="Administrative",
        subcategory="Fees",
        complaint="Advance fee was demanded.",
        approval="Not Prepared",
        reply=None,
    )
    add_case(
        session,
        "104-7000004",
        category="Others",
        complaint="=HYPERLINK(\"https://example.invalid\",\"unsafe\")",
        approval="Issued",
        reply="+This approved reply begins with a spreadsheet operator.",
    )


def test_default_archive_includes_only_approved_or_issued_pairs() -> None:
    db = engine()
    with Session(db) as session:
        seed(session)
        archive = ComplaintKnowledgeArchive(session, settings())
        records = archive.records()
        assert [record.complaint_number for record in records] == ["104-7000001", "104-7000004"]
        assert all(record.ready_for_ai for record in records)
        assert records[0].reply_editor_url.startswith("http://crm.test/crm/replies/")
        assert records[0].case_url.startswith("http://crm.test/crm/cases/")
        assert records[0].paperless_url == "http://paperless.test/documents/100/details"
        assert "Unapproved draft reply" not in archive.render_text(records, KnowledgeFilters()).decode()


def test_filters_facets_and_coverage_statistics_are_category_aware() -> None:
    db = engine()
    with Session(db) as session:
        seed(session)
        archive = ComplaintKnowledgeArchive(session, settings())
        filtered = KnowledgeFilters(category="Administrative", reply_scope="approved")
        records = archive.records(filtered)
        assert [record.complaint_number for record in records] == ["104-7000001"]
        stats = archive.statistics(filtered)
        assert stats["total_records"] == 1
        assert stats["approved_pairs"] == 1
        assert stats["awaiting_reply"] == 1
        facets = archive.facets()
        assert facets["category_subcategories"]["Administrative"] == ["Fees", "School timing"]


def test_ai_eligible_filter_and_search() -> None:
    db = engine()
    with Session(db) as session:
        seed(session)
        archive = ComplaintKnowledgeArchive(session, settings())
        eligible = archive.records(KnowledgeFilters(ai_eligible_only=True))
        assert [record.complaint_number for record in eligible] == ["104-7000001"]
        searched = archive.records(KnowledgeFilters(search="sensitized"))
        assert [record.complaint_number for record in searched] == ["104-7000001"]


def test_csv_and_xlsx_prevent_formula_execution_and_have_category_sheets() -> None:
    db = engine()
    with Session(db) as session:
        seed(session)
        archive = ComplaintKnowledgeArchive(session, settings())
        records = archive.records()
        csv_text = archive.render_csv(records).decode("utf-8-sig")
        assert "'=HYPERLINK" in csv_text
        assert "'+This approved reply" in csv_text
        workbook = load_workbook(io.BytesIO(archive.render_xlsx(records, KnowledgeFilters())), data_only=False)
        assert "Summary" in workbook.sheetnames
        assert "All Approved Records" in workbook.sheetnames
        assert "Administrative" in workbook.sheetnames
        assert "Others" in workbook.sheetnames
        all_sheet = workbook["All Approved Records"]
        values = [cell.value for cell in all_sheet[3]]
        assert any(isinstance(value, str) and value.startswith("'=HYPERLINK") for value in values)
        assert all(not (isinstance(value, str) and value.startswith("=HYPERLINK")) for value in values)


def test_markdown_json_and_zip_preserve_full_approved_text_by_category() -> None:
    db = engine()
    with Session(db) as session:
        seed(session)
        archive = ComplaintKnowledgeArchive(session, settings())
        filters = KnowledgeFilters()
        records = archive.records(filters)
        markdown = archive.render_markdown(records, filters).decode()
        assert "School opened late." in markdown
        assert "The matter was examined" in markdown
        assert "Unapproved draft reply" not in markdown
        payload = json.loads(archive.render_json(records, filters))
        assert payload["record_count"] == 2
        bundle = archive.render_bundle(records, filters)
        with zipfile.ZipFile(io.BytesIO(bundle)) as zipped:
            names = set(zipped.namelist())
            assert "all/complaint-reply-archive.xlsx" in names
            assert "all/complaint-reply-chatgpt.md" in names
            assert "categories/administrative/complaint-reply-chatgpt.txt" in names
            assert "categories/others/complaint-reply.csv" in names
            manifest = json.loads(zipped.read("manifest.json"))
            assert manifest["record_count"] == 2
            assert manifest["contains_unredacted_text"] is True


def test_api_lists_and_exports_the_same_approved_dataset() -> None:
    db = engine()
    with Session(db) as session:
        seed(session)

    app = FastAPI()
    app.include_router(router)

    def local_session():
        with Session(db) as session:
            yield session

    app.dependency_overrides[get_session] = local_session
    client = TestClient(app)
    records = client.get("/api/v1/crm/knowledge/records").json()
    assert records["total"] == 2
    assert {row["complaint_number"] for row in records["items"]} == {"104-7000001", "104-7000004"}
    response = client.get("/api/v1/crm/knowledge/exports/zip")
    assert response.status_code == 200
    assert response.headers["content-type"].startswith("application/zip")
    with zipfile.ZipFile(io.BytesIO(response.content)) as zipped:
        assert json.loads(zipped.read("manifest.json"))["record_count"] == 2


def test_all_scope_may_preview_unapproved_records_but_default_export_does_not() -> None:
    db = engine()
    with Session(db) as session:
        seed(session)
        archive = ComplaintKnowledgeArchive(session, settings())
        all_records = archive.records(KnowledgeFilters(reply_scope="all"))
        assert len(all_records) == 4
        draft = next(record for record in all_records if record.complaint_number == "104-7000002")
        assert draft.ready_for_ai is False
        assert "reply_not_approved_or_issued" in draft.quality_issues
        default_content, _media, _name = archive.render("json", KnowledgeFilters())
        assert "Unapproved draft reply" not in default_content.decode()
