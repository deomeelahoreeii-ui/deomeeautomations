from __future__ import annotations

import uuid
from pathlib import Path
from types import SimpleNamespace

from openpyxl import Workbook, load_workbook
from sqlalchemy.pool import StaticPool
from sqlmodel import SQLModel, Session, create_engine, select

import antidengue_automation.models  # noqa: F401
import automation_core.models  # noqa: F401
import master_data.models  # noqa: F401
import whatsapp_gateway.models  # noqa: F401
from automation_core.models import Artifact, Job, JobStatus, JobType
from master_data.models import (
    Department, District, Markaz, Officer, OfficerJurisdiction, School, Tehsil, Wing,
)
from antidengue_automation.scheduling import validate_dispatch_profiles
from whatsapp_gateway.configuration.defaults import ensure_consolidated_digest_profiles
from whatsapp_gateway.models import (
    WhatsAppAccount, WhatsAppApplication, WhatsAppAudience, WhatsAppDispatchProfile,
    WhatsAppDispatchPreviewArtifact, WhatsAppDispatchPreviewDelivery,
    WhatsAppRecipientScope, WhatsAppReportType, WhatsAppTemplate,
)
from whatsapp_gateway.persistence.audience_sources import WhatsAppAudienceSource
from whatsapp_gateway.previews.compiler.orchestrator import compile_antidengue_preview
from whatsapp_gateway.rendering.antidengue.digest_report import render_consolidated_action_digest


def _engine():
    engine = create_engine(
        "sqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool
    )
    SQLModel.metadata.create_all(engine)
    return engine


def _workbook(path: Path, headers: list[str], rows: list[list[object]], title: str = "Report") -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = title
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    workbook.save(path)
    workbook.close()


def _seed(session: Session, tmp_path: Path):
    district = District(name="Lahore")
    department = Department(name="School Education")
    session.add_all([district, department]); session.flush()
    wing = Wing(
        district_id=district.id, department_id=department.id, name="DEO MEE", code="DEO MEE"
    )
    tehsil = Tehsil(district_id=district.id, name="CITY")
    session.add_all([wing, tehsil]); session.flush()
    markaz_a = Markaz(wing_id=wing.id, tehsil_id=tehsil.id, name="LAHORE KHAS - MALE")
    markaz_b = Markaz(wing_id=wing.id, tehsil_id=tehsil.id, name="NIAZBAIG - MALE")
    session.add_all([markaz_a, markaz_b]); session.flush()
    schools = [
        School(emis="1001", name="GES CHAH MIRAN", district_id=district.id, department_id=department.id, wing_id=wing.id, tehsil_id=tehsil.id, markaz_id=markaz_a.id),
        School(emis="1002", name="GPS KHAN PUR", district_id=district.id, department_id=department.id, wing_id=wing.id, tehsil_id=tehsil.id, markaz_id=markaz_b.id),
        School(emis="1003", name="GES MODEL", district_id=district.id, department_id=department.id, wing_id=wing.id, tehsil_id=tehsil.id, markaz_id=markaz_b.id),
    ]
    session.add_all(schools)
    job = Job(
        type=JobType.antidengue_report.value, title="Dry run", status=JobStatus.succeeded.value,
        parameters={"dry_run": True},
        result={"summary": {"portal_acquisition": {"window": {"dateto": "2026-07-18T11:00:00"}}}},
    )
    session.add(job); session.flush()
    dormant = tmp_path / "Anti-Dengue App Dormant Users.xlsx"
    _workbook(dormant, ["School EMIS", "School Name"], [["1001", "GES CHAH MIRAN"], ["1002", "GPS KHAN PUR"]])
    hotspot = tmp_path / "Hotspot Distance Review - source.xlsx"
    activity_headers = [
        "School EMIS", "School Name", "Wing ID", "Tehsil ID", "Markaz ID",
        "Tehsil", "Markaz", "Distance Difference (In meters)", "Activity(Lat,Long)", "ID",
    ]
    _workbook(hotspot, activity_headers, [
        ["1002", "GPS KHAN PUR", str(wing.id), str(tehsil.id), str(markaz_b.id), "CITY", markaz_b.name, 91.5, "31.5,74.3", "h1"],
        ["1003", "GES MODEL", str(wing.id), str(tehsil.id), str(markaz_b.id), "CITY", markaz_b.name, 74.0, "31.6,74.4", "h2"],
    ], "Review Required")
    timing = tmp_path / "Simple Activity Timing Review - source.xlsx"
    timing_headers = [
        "School EMIS", "School Name", "Wing ID", "Tehsil ID", "Markaz ID",
        "Tehsil", "Markaz", "Time Difference(Sec)", "ID",
    ]
    _workbook(timing, timing_headers, [
        ["1002", "GPS KHAN PUR", str(wing.id), str(tehsil.id), str(markaz_b.id), "CITY", markaz_b.name, 245, "t1"],
    ], "Review Required")
    for name, path in ((dormant.name, dormant), (hotspot.name, hotspot), (timing.name, timing)):
        session.add(Artifact(
            job_id=job.id, module_key="antidengue", kind="report", name=name,
            path=str(path), size_bytes=path.stat().st_size,
        ))
    session.commit()
    return job, wing, tehsil, markaz_b


def test_digest_merges_by_emis_and_creates_one_four_sheet_workbook(tmp_path: Path, monkeypatch) -> None:
    import whatsapp_gateway.rendering.antidengue.digest_report as report_module
    import whatsapp_gateway.rendering.antidengue.digest_workbook as workbook_module

    settings = SimpleNamespace(
        artifact_root=tmp_path / "artifacts", antidengue_submission_deadline="12:30 PM"
    )
    monkeypatch.setattr(report_module, "get_settings", lambda: settings)
    monkeypatch.setattr(workbook_module, "get_settings", lambda: settings)
    with Session(_engine()) as session:
        job, wing, tehsil, _ = _seed(session, tmp_path)

        rendered = render_consolidated_action_digest(
            session, source_job=job, wing=wing, recipient_name="Head Teachers — CITY",
            scope_key="tehsil", scope_value=str(tehsil.id), scope_label="CITY",
        )

        assert len(rendered.schools) == 3
        by_emis = {item.emis: item for item in rendered.schools}
        assert by_emis["1001"].issue_codes == ["DORMANT"]
        assert by_emis["1002"].issue_codes == ["DORMANT", "DISTANCE", "TIMING"]
        assert by_emis["1003"].issue_codes == ["DISTANCE"]
        assert "`1002`" not in rendered.message
        assert "GPS KHAN PUR" not in rendered.message
        assert "*MARKAZ SUMMARY*" in rendered.message
        assert "NIAZBAIG - MALE" in rendered.message
        assert "*3 UNIQUE SCHOOLS REQUIRE ATTENTION*" in rendered.message
        assert "Dormant activity: 2 schools" in rendered.message
        assert "Hotspot distance: 2 schools" in rendered.message
        assert "Activity timing: 1 school" in rendered.message
        assert len(rendered.attachment_paths) == 1
        workbook = load_workbook(rendered.attachment_paths[0], read_only=False, data_only=True)
        try:
            assert workbook.sheetnames == ["Overview", "Dormant Schools", "Hotspot Distance", "Activity Timing"]
            overview = workbook["Overview"]
            assert overview.freeze_panes == "A5"
            assert overview["A3"].value == "Action deadline: 12:30 PM (Asia/Karachi)"
            rows = list(overview.iter_rows(min_row=5, values_only=True))
            issue_by_emis = {str(row[0]): row[5] for row in rows}
            assert issue_by_emis["1002"] == "DORMANT | DISTANCE | TIMING"
            assert workbook["Dormant Schools"].max_row == 6
            assert workbook["Hotspot Distance"].max_row == 6
            assert workbook["Activity Timing"].max_row == 5
        finally:
            workbook.close()


def test_digest_blocks_when_one_required_source_workbook_is_missing(tmp_path: Path, monkeypatch) -> None:
    import whatsapp_gateway.rendering.antidengue.digest_report as report_module
    import whatsapp_gateway.rendering.antidengue.digest_workbook as workbook_module

    settings = SimpleNamespace(
        artifact_root=tmp_path / "artifacts", antidengue_submission_deadline="12:30 PM"
    )
    monkeypatch.setattr(report_module, "get_settings", lambda: settings)
    monkeypatch.setattr(workbook_module, "get_settings", lambda: settings)
    with Session(_engine()) as session:
        job, wing, tehsil, _ = _seed(session, tmp_path)
        timing = next(tmp_path.glob("Simple Activity Timing Review*.xlsx"))
        timing.unlink()

        try:
            render_consolidated_action_digest(
                session, source_job=job, wing=wing, recipient_name="CITY",
                scope_key="tehsil", scope_value=str(tehsil.id), scope_label="CITY",
            )
        except ValueError as exc:
            assert "no classified Simple Activity timing workbook" in str(exc)
        else:
            raise AssertionError("A digest must not interpret a missing source workbook as zero findings")


def test_digest_keeps_empty_markaz_route_with_all_clear_workbook(tmp_path: Path, monkeypatch) -> None:
    import whatsapp_gateway.rendering.antidengue.digest_report as report_module
    import whatsapp_gateway.rendering.antidengue.digest_workbook as workbook_module

    settings = SimpleNamespace(
        artifact_root=tmp_path / "artifacts", antidengue_submission_deadline="12:30 PM"
    )
    monkeypatch.setattr(report_module, "get_settings", lambda: settings)
    monkeypatch.setattr(workbook_module, "get_settings", lambda: settings)
    with Session(_engine()) as session:
        job, wing, tehsil, _ = _seed(session, tmp_path)
        clear_markaz = Markaz(
            wing_id=wing.id, tehsil_id=tehsil.id, name="CLEAR MARKAZ - MALE"
        )
        session.add(clear_markaz); session.commit()

        rendered = render_consolidated_action_digest(
            session, source_job=job, wing=wing, recipient_name="AEO Clear",
            scope_key="markaz", scope_value=str(clear_markaz.id),
            scope_values=[str(clear_markaz.id)], scope_label=clear_markaz.name,
        )

        assert rendered.schools == []
        assert "ACTION DIGEST — ALL CLEAR" in rendered.message
        assert len(rendered.attachment_paths) == 1
        workbook = load_workbook(rendered.attachment_paths[0], read_only=True)
        try:
            assert workbook.sheetnames == [
                "Overview", "Dormant Schools", "Hotspot Distance", "Activity Timing",
            ]
        finally:
            workbook.close()


def test_markaz_digest_fans_out_one_delivery_and_workbook_per_live_scope(
    tmp_path: Path, monkeypatch,
) -> None:
    import whatsapp_gateway.previews.artifact_storage as storage_module
    import whatsapp_gateway.previews.compiler.artifact_snapshots as snapshots_module
    import whatsapp_gateway.rendering.antidengue.digest_report as report_module
    import whatsapp_gateway.rendering.antidengue.digest_workbook as workbook_module

    settings = SimpleNamespace(
        artifact_root=tmp_path / "artifacts",
        antidengue_submission_deadline="12:30 PM",
        object_storage_enabled=False,
    )
    for module in (storage_module, snapshots_module, report_module, workbook_module):
        monkeypatch.setattr(module, "get_settings", lambda: settings)

    with Session(_engine()) as session:
        job, wing, tehsil, _ = _seed(session, tmp_path)
        markazes = list(session.scalars(select(Markaz).order_by(Markaz.name)).all())
        district = session.get(District, wing.district_id)
        department = session.get(Department, wing.department_id)
        officer = Officer(
            role="aeo", name="Additional Charge AEO", mobile="03001234567",
            normalized_mobile="923001234567", district_id=district.id,
            department_id=department.id, wing_id=wing.id, tehsil_id=tehsil.id,
            markaz_id=markazes[0].id,
        )
        application = WhatsAppApplication(key="antidengue", name="AntiDengue")
        account = WhatsAppAccount(name="Primary", worker_key="digest-scope-test")
        session.add_all([officer, application, account]); session.flush()
        for markaz in markazes:
            session.add(OfficerJurisdiction(
                role="aeo", officer_id=officer.id, district_id=district.id,
                department_id=department.id, wing_id=wing.id,
                tehsil_id=tehsil.id, markaz_id=markaz.id,
            ))
        report = WhatsAppReportType(
            application_id=application.id, key="consolidated_action_digest",
            name="Consolidated Action Digest",
        )
        scope = WhatsAppRecipientScope(
            application_id=application.id, channel="individual", key="aeo", name="AEO",
        )
        audience = WhatsAppAudience(
            application_id=application.id, key="dynamic-aeo", name="Dynamic AEO",
        )
        session.add_all([report, scope, audience]); session.flush()
        session.add(WhatsAppAudienceSource(
            audience_id=audience.id, recipient_role="aeo", wing_id=wing.id,
            route_scope_key="markaz", aggregate_by_recipient=True,
        ))
        template = WhatsAppTemplate(
            application_id=application.id, report_type_id=report.id,
            recipient_scope_id=scope.id, recipient_channel="individual",
            key="digest-scope-template", name="Digest", category="report",
            body="{{report_body}}",
        )
        session.add(template); session.flush()
        profile = WhatsAppDispatchProfile(
            application_id=application.id, key="digest-scope", name="Digest scope",
            report_type_id=report.id, audience_id=audience.id, account_id=account.id,
            template_id=template.id, recipient_scope_id=scope.id,
            recipient_channel="individual", wing_id=wing.id,
            delivery_mode="individuals", delivery_granularity="scope",
            presentation_policy={"attachment_mode": "excel"},
        )
        session.add(profile); session.commit()

        preview = compile_antidengue_preview(
            session, source_job_id=job.id, dispatch_profile_id=profile.id,
            deadline_snapshot={
                "time": "13:45", "label": "1:45 PM", "timezone": "Asia/Karachi",
                "policy_version": 7, "source": "schedule_override",
            },
        )
        deliveries = list(session.scalars(select(WhatsAppDispatchPreviewDelivery).where(
            WhatsAppDispatchPreviewDelivery.preview_id == preview.id
        ).order_by(WhatsAppDispatchPreviewDelivery.sequence)).all())
        artifacts = list(session.scalars(select(WhatsAppDispatchPreviewArtifact).where(
            WhatsAppDispatchPreviewArtifact.preview_id == preview.id,
            WhatsAppDispatchPreviewArtifact.role == "delivery",
        )).all())

        assert len(deliveries) == len(markazes) == 2
        assert {item.target_jid for item in deliveries} == {"923001234567@s.whatsapp.net"}
        assert {
            tuple(item.routing_snapshot["dynamic_audience"]["markaz_ids"])
            for item in deliveries
        } == {(str(markazes[0].id),), (str(markazes[1].id),)}
        assert len({item.source_route_key for item in deliveries}) == 2
        assert len(artifacts) == 2
        assert len({item.checksum_sha256 for item in artifacts}) == 2
        assert preview.configuration_snapshot["deadline"] == {
            "time": "13:45", "label": "1:45 PM", "timezone": "Asia/Karachi",
            "policy_version": 7, "source": "schedule_override",
        }
        assert all("1:45 PM" in item.message for item in deliveries)


def test_digest_profiles_are_additive_idempotent_and_reuse_live_audiences() -> None:
    with Session(_engine()) as session:
        district = District(name="Lahore")
        department = Department(name="School Education")
        session.add_all([district, department]); session.flush()
        wing = Wing(district_id=district.id, department_id=department.id, name="DEO MEE", code="DEO MEE")
        application = WhatsAppApplication(key="antidengue", name="AntiDengue")
        account = WhatsAppAccount(name="Primary", worker_key="digest-default-test")
        session.add_all([wing, application, account]); session.flush()
        dormant = WhatsAppReportType(application_id=application.id, key="wing_summary", name="Wing summary")
        digest = WhatsAppReportType(application_id=application.id, key="consolidated_action_digest", name="Consolidated Action Digest")
        scope = WhatsAppRecipientScope(application_id=application.id, channel="group", key="wing", name="Wing group")
        audience = WhatsAppAudience(application_id=application.id, key="wing-groups", name="Wing groups")
        session.add_all([dormant, digest, scope, audience]); session.flush()
        source_template = WhatsAppTemplate(
            application_id=application.id, report_type_id=dormant.id,
            recipient_scope_id=scope.id, recipient_channel="group", key="wing-source",
            name="Wing source", category="report", body="{{report_body}}",
        )
        session.add(source_template); session.flush()
        source = WhatsAppDispatchProfile(
            application_id=application.id, key="wing-source", name="Wing detailed",
            report_type_id=dormant.id, audience_id=audience.id, account_id=account.id,
            template_id=source_template.id, recipient_scope_id=scope.id,
            recipient_channel="group", wing_id=wing.id, delivery_mode="groups",
            presentation_policy={"message_style": "detailed", "attachment_mode": "image_excel"},
        )
        session.add(source); session.commit()

        ensure_consolidated_digest_profiles(session, application=application)
        session.flush()
        ensure_consolidated_digest_profiles(session, application=application)
        session.flush()

        profiles = list(session.scalars(select(WhatsAppDispatchProfile).where(
            WhatsAppDispatchProfile.report_type_id == digest.id
        )).all())
        assert len(profiles) == 1
        created = profiles[0]
        assert created.audience_id == source.audience_id
        assert created.account_id == source.account_id
        assert created.recipient_scope_id == source.recipient_scope_id
        assert created.wing_id == source.wing_id
        assert created.presentation_policy["attachment_mode"] == "excel"
        assert created.presentation_policy["digest_source_profile_id"] == str(source.id)
        assert created.template_id != source.template_id
        created_template = session.get(WhatsAppTemplate, created.template_id)
        assert created_template is not None
        assert "{{detail_heading}}" in created_template.body

        created_template.body = created_template.body.replace(
            "⚠️ Distance and timing findings are rule-based review candidates, not confirmed violations.",
            "Operator-specific closing line.",
        ).replace("*{{detail_heading}}*", "*DETAILS BY MARKAZ*")
        session.add(created_template); session.flush()
        ensure_consolidated_digest_profiles(session, application=application)
        session.flush(); session.refresh(created_template)
        assert "*{{detail_heading}}*" in created_template.body
        assert "Operator-specific closing line." in created_template.body

        session.refresh(source)
        assert source.presentation_policy["attachment_mode"] == "image_excel"

        try:
            validate_dispatch_profiles(session, [source.id, created.id])
        except ValueError as exc:
            assert "either the Consolidated Action Digest or detailed profiles" in str(exc)
        else:
            raise AssertionError("Selecting digest and detailed delivery for one audience must be rejected")


def test_digest_uses_scope_specific_message_detail(tmp_path: Path, monkeypatch) -> None:
    import whatsapp_gateway.rendering.antidengue.digest_report as report_module
    import whatsapp_gateway.rendering.antidengue.digest_workbook as workbook_module

    settings = SimpleNamespace(
        artifact_root=tmp_path / "artifacts", antidengue_submission_deadline="12:30 PM"
    )
    monkeypatch.setattr(report_module, "get_settings", lambda: settings)
    monkeypatch.setattr(workbook_module, "get_settings", lambda: settings)
    with Session(_engine()) as session:
        job, wing, tehsil, markaz = _seed(session, tmp_path)

        wing_rendered = render_consolidated_action_digest(
            session, source_job=job, wing=wing, recipient_name="MEE Heads",
            scope_key="wing", scope_value=str(wing.id), scope_label=wing.name,
        )
        assert "*TEHSIL AND MARKAZ SUMMARY*" in wing_rendered.message
        assert "*TEHSIL SUMMARY*" in wing_rendered.message
        assert "*MARKAZ SUMMARY*" in wing_rendered.message
        assert "GES CHAH MIRAN" not in wing_rendered.message
        assert "`1001`" not in wing_rendered.message

        tehsil_rendered = render_consolidated_action_digest(
            session, source_job=job, wing=wing, recipient_name="CITY Heads",
            scope_key="tehsil", scope_value=str(tehsil.id), scope_label=tehsil.name,
        )
        assert "*MARKAZ SUMMARY*" in tehsil_rendered.message
        assert "NIAZBAIG - MALE" in tehsil_rendered.message
        assert "GPS KHAN PUR" not in tehsil_rendered.message
        assert "`1002`" not in tehsil_rendered.message

        markaz_rendered = render_consolidated_action_digest(
            session, source_job=job, wing=wing, recipient_name="AEO Niazbaig",
            scope_key="markaz", scope_value=str(markaz.id),
            scope_values=[str(markaz.id)], scope_label=markaz.name,
        )
        assert "*SCHOOL ACTION DETAILS*" in markaz_rendered.message
        assert "GPS KHAN PUR" in markaz_rendered.message
        assert "`1002`" in markaz_rendered.message


def test_parent_digest_attachment_covers_timing_sample_without_false_blocker(
    tmp_path: Path, monkeypatch,
) -> None:
    import whatsapp_gateway.rendering.antidengue.digest_report as report_module
    import whatsapp_gateway.rendering.antidengue.digest_workbook as workbook_module

    settings = SimpleNamespace(
        artifact_root=tmp_path / "artifacts", antidengue_submission_deadline="12:30 PM"
    )
    monkeypatch.setattr(report_module, "get_settings", lambda: settings)
    monkeypatch.setattr(workbook_module, "get_settings", lambda: settings)
    with Session(_engine()) as session:
        job, wing, _, markaz = _seed(session, tmp_path)
        timing_path = next(tmp_path.glob("Simple Activity Timing Review*.xlsx"))
        workbook = load_workbook(timing_path)
        try:
            sheet = workbook["Review Required"]
            for index in range(2, 8):
                sheet.append([
                    "1002", "GPS KHAN PUR", str(wing.id), str(markaz.tehsil_id),
                    str(markaz.id), "CITY", markaz.name, 245 + index, f"t{index}",
                ])
            workbook.save(timing_path)
        finally:
            workbook.close()
        artifact = session.scalar(select(Artifact).where(Artifact.job_id == job.id, Artifact.name.like("Simple Activity Timing Review%")))
        assert artifact is not None
        artifact.size_bytes = timing_path.stat().st_size
        session.add(artifact)
        session.commit()

        rendered = render_consolidated_action_digest(
            session, source_job=job, wing=wing, recipient_name="AEO Niazbaig",
            scope_key="markaz", scope_value=str(markaz.id),
            scope_values=[str(markaz.id)], scope_label=markaz.name,
        )

        assert len(rendered.attachment_paths) == 1
        assert not any(issue["severity"] == "blocked" for issue in rendered.issues)
        summarized = [
            issue for issue in rendered.issues
            if issue["code"] == "message_evidence_summarized"
        ]
        assert summarized
        assert summarized[0]["evidence_attachment_source"] == "parent"
        assert summarized[0]["omitted_activity_count"] > 0


def test_parent_digest_downgrades_component_only_presentation_blockers() -> None:
    from whatsapp_gateway.rendering.antidengue.digest_report import _parent_issue

    normalized = _parent_issue({
        "code": "message_only_evidence_truncated",
        "severity": "blocked",
        "category": "presentation",
        "message": "Child-only message sample omitted evidence.",
    })
    assert normalized["severity"] == "info"
    assert normalized["code"] == "message_evidence_summarized"
    assert normalized["component_code"] == "message_only_evidence_truncated"
    assert normalized["evidence_attachment_source"] == "parent"

    routing_issue = {"code": "missing_recipient", "severity": "blocked", "category": "routing"}
    assert _parent_issue(routing_issue) is routing_issue


def test_consolidated_digest_modules_preserve_renderer_boundaries() -> None:
    renderer_root = (
        Path(__file__).resolve().parents[1]
        / "packages"
        / "whatsapp_gateway"
        / "whatsapp_gateway"
        / "rendering"
        / "antidengue"
    )
    relevant = [
        renderer_root / "digest_report.py",
        renderer_root / "digest_scope.py",
        renderer_root / "digest_presentation.py",
    ]
    oversized = {
        path.name: len(path.read_text(encoding="utf-8").splitlines())
        for path in relevant
        if len(path.read_text(encoding="utf-8").splitlines()) > 250
    }
    assert oversized == {}
