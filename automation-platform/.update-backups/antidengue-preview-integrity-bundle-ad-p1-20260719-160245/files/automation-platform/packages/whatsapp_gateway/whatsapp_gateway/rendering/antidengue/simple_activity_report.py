from __future__ import annotations

import hashlib
import re
import uuid
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from sqlmodel import Session, select

from automation_core.config import get_settings
from automation_core.models import Artifact, Job
from master_data.models import Wing
from whatsapp_gateway.rendering.antidengue.hotspot_report import _plain_whatsapp_text, _window_label
from whatsapp_gateway.rendering.antidengue.issues import _issue

SIMPLE_ACTIVITY_RENDERER_KEY = "antidengue.simple_activity_timing.v1"
SIMPLE_ACTIVITY_REPORT_KEY = "simple_activity_timing"
SIMPLE_ACTIVITY_DEFAULT_TEMPLATE = """⏱️ *Anti-Dengue Activity Timing Review*
👥 *Group:* {{group_name}}
📌 *Scope:* {{scope_name}}
🕒 *Report window:* {{report_window}}

🏫 *School-wise timing evidence*

{{school_activity_details}}

{{attachment_note}}
{{disclaimer}}"""


@dataclass(frozen=True)
class SimpleActivityRow:
    values: dict[str, Any]
    emis: str
    school_name: str
    wing_id: str
    tehsil_id: str
    markaz_id: str
    seconds: float | None


@dataclass
class RenderedSimpleActivityReport:
    message: str
    context: dict[str, str]
    rows: list[SimpleActivityRow]
    issues: list[dict[str, Any]] = field(default_factory=list)
    attachment_paths: list[Path] = field(default_factory=list)
    source_artifact_id: int | None = None
    source_artifact_sha256: str = ""
    source_issues: list[dict[str, Any]] = field(default_factory=list)


def _details(rows: list[SimpleActivityRow], *, school_limit: int = 8, activity_limit: int = 4) -> tuple[str, int, int]:
    grouped: dict[str, list[SimpleActivityRow]] = {}
    for row in rows: grouped.setdefault(row.emis, []).append(row)
    lines: list[str] = []
    ordered = sorted(grouped.items(), key=lambda item: (str(item[1][0].values.get("Tehsil") or ""), item[1][0].school_name))
    omitted_schools = max(0, len(ordered) - school_limit)
    omitted_activities = 0
    for number, (emis, school_rows) in enumerate(ordered[:school_limit], 1):
        first = school_rows[0]; values = first.values
        lines.extend([
            f"🏫 *{number}. {_plain_whatsapp_text(first.school_name, 'Unknown school')}*",
            f"🆔 *EMIS:* `{_plain_whatsapp_text(emis)}`",
            f"📍 *Tehsil:* {_plain_whatsapp_text(values.get('Tehsil'))}  |  *Markaz:* {_plain_whatsapp_text(values.get('Markaz'))}",
        ])
        ordered_activities = sorted(school_rows, key=lambda item: item.seconds or 0)
        omitted_activities += max(0, len(ordered_activities) - activity_limit)
        for activity_number, row in enumerate(ordered_activities[:activity_limit], 1):
            if len(school_rows) > 1: lines.append(f"🔸 *Activity {activity_number}*")
            if row.seconds is None:
                interval = "Not available"
            else:
                minutes, seconds = divmod(int(round(row.seconds)), 60)
                friendly = f"{minutes} min {seconds} sec" if minutes else f"{seconds} sec"
                interval = f"{row.seconds:,.0f} seconds ({friendly})"
            lines.append(f"⏱️ *Time Difference:* {interval}")
        lines.append("")
    if omitted_schools or omitted_activities:
        lines.append(
            "📎 *Complete evidence:* "
            f"{omitted_schools} additional school(s) and {omitted_activities} additional activity row(s) "
            "are included in the attached scope-filtered Excel file."
        )
    return "\n".join(lines).strip(), omitted_schools, omitted_activities


def _source(session: Session, job: Job) -> tuple[Artifact, Path, list[str], list[SimpleActivityRow]]:
    artifacts = session.scalars(select(Artifact).where(Artifact.job_id == job.id).order_by(Artifact.created_at, Artifact.id)).all()
    for artifact in artifacts:
        if not artifact.name.lower().split("/")[-1].startswith("simple activity timing review"): continue
        path = Path(artifact.path).expanduser().resolve(strict=False)
        if not path.is_file(): continue
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook["Review Required"] if "Review Required" in workbook.sheetnames else workbook.active
            values = sheet.iter_rows(values_only=True); header = [str(value or "").strip() for value in next(values, ())]
            required = {"School EMIS", "School Name", "Wing ID", "Tehsil ID", "Markaz ID", "Time Difference(Sec)"}
            if not required.issubset(header): continue
            rows = []
            for raw in values:
                item = {name: raw[index] if index < len(raw) else None for index, name in enumerate(header)}
                try: seconds = float(item.get("Time Difference(Sec)"))
                except (TypeError, ValueError): seconds = None
                rows.append(SimpleActivityRow(item, str(item.get("School EMIS") or "").strip(), str(item.get("School Name") or "").strip(), str(item.get("Wing ID") or "").strip(), str(item.get("Tehsil ID") or "").strip(), str(item.get("Markaz ID") or "").strip(), seconds))
            return artifact, path, header, rows
        finally: workbook.close()
    raise ValueError("The source run has no classified Simple Activity timing workbook. Publish and enable a timing rule, then run Test Run again.")


def _attachment(session: Session, job: Job, source: Artifact, header: list[str], rows: list[SimpleActivityRow], scope: str) -> tuple[Artifact, Path, str]:
    digest = hashlib.sha256((str(source.id) + ":" + ",".join(str(row.values.get("ID") or index) for index, row in enumerate(rows))).encode()).hexdigest()
    safe = re.sub(r"[^A-Za-z0-9_-]+", "-", scope).strip("-") or "scope"
    path = get_settings().artifact_root.expanduser().resolve() / "antidengue" / "native-reports" / str(job.id) / f"simple-timing-{safe}-{digest[:16]}.xlsx"
    artifact = session.scalar(select(Artifact).where(Artifact.job_id == job.id, Artifact.path == str(path)))
    if not path.is_file():
        path.parent.mkdir(parents=True, exist_ok=True); temporary = path.with_name(f".{path.name}.{uuid.uuid4().hex}.tmp")
        workbook = Workbook(); sheet = workbook.active; sheet.title = "Review Required"; sheet.append(header)
        for cell in sheet[1]: cell.font = Font(bold=True, color="FFFFFF"); cell.fill = PatternFill("solid", fgColor="0E7490")
        for row in rows: sheet.append([row.values.get(name) for name in header])
        sheet.freeze_panes = "A2"; sheet.auto_filter.ref = sheet.dimensions; workbook.save(temporary); workbook.close(); temporary.replace(path)
    if artifact is None: artifact = Artifact(job_id=job.id, module_key="antidengue", kind="delivery", name=f"Simple Activity Timing Review - {scope}.xlsx", path=str(path), size_bytes=path.stat().st_size)
    else: artifact.size_bytes = path.stat().st_size
    session.add(artifact); session.flush()
    return artifact, path, hashlib.sha256(path.read_bytes()).hexdigest()


def render_simple_activity_report(session: Session, *, source_job: Job, wing: Wing, recipient_name: str, scope_key: str, scope_value: str, scope_label: str, presentation_policy: dict[str, Any] | None = None, scope_values: list[str] | None = None) -> RenderedSimpleActivityReport:
    source, source_path, header, all_rows = _source(session, source_job)
    rows = [row for row in all_rows if row.wing_id == str(wing.id)]
    accepted_scope_values = set(scope_values or [scope_value])
    if scope_key == "tehsil": rows = [row for row in rows if row.tehsil_id in accepted_scope_values]
    elif scope_key == "markaz": rows = [row for row in rows if row.markaz_id in accepted_scope_values]
    elif scope_key not in {"district", "wing"}: raise ValueError(f"Simple Activity timing delivery does not support {scope_key or 'an unbound'} scope")
    issues = []
    source_issues = []
    unmapped = [row for row in all_rows if not row.wing_id or not row.emis]
    if unmapped:
        counts: dict[str, int] = {}
        for row in unmapped:
            key = row.emis or "missing EMIS"
            counts[key] = counts.get(key, 0) + 1
        identities = ", ".join(
            f"{emis} ({count} row{'s' if count != 1 else ''})"
            for emis, count in sorted(counts.items())
        )
        source_issues.append(_issue(
            "unmapped_simple_activity_submitters",
            "warning",
            (
                f"{len(counts)} portal submitter{'s are' if len(counts) != 1 else ' is'} absent from Master Data; "
                f"{len(unmapped)} timing row{'s were' if len(unmapped) != 1 else ' was'} excluded from routed deliveries: "
                f"{identities}. Add or classify the submitter in Master Data, then create a new Test Run."
            ),
        ))
    digest = hashlib.sha256(source_path.read_bytes()).hexdigest()
    if not rows:
        issues.append(_issue("nothing_to_dispatch", "info", f"No Simple Activity timing exceptions were found for {scope_label}."))
        return RenderedSimpleActivityReport(
            "", {}, [], issues,
            source_artifact_id=source.id,
            source_artifact_sha256=digest,
            source_issues=source_issues,
        )
    window = dict(((source_job.result or {}).get("summary") or {}).get("portal_acquisition") or {}).get("window") or {}
    has_attachment = str((presentation_policy or {}).get("attachment_mode") or "excel") != "none"
    details, omitted_schools, omitted_activities = _details(rows)
    if omitted_schools or omitted_activities:
        issues.append(_issue(
            "message_evidence_summarized" if has_attachment else "message_only_evidence_truncated",
            "info" if has_attachment else "blocked",
            (
                "WhatsApp text shows a concise school sample; the attached Excel contains all scoped evidence."
                if has_attachment else
                "Message-only delivery cannot safely omit the remaining timing evidence. Enable an Excel attachment or narrow the rule."
            ),
        ))
    context = {
        "group_name": _plain_whatsapp_text(recipient_name), "scope_name": _plain_whatsapp_text(scope_label),
        "report_window": _window_label(window.get("dateto")) or "Not available",
        "school_activity_details": details,
        "attachment_note": "📎 The attached Excel file contains every selected activity and its timing evidence." if has_attachment else "",
        "disclaimer": "⚠️ These are rule-based review candidates, not confirmed fake activities. Please verify each finding.",
        "activity_count": str(len(rows)), "school_count": str(len({row.emis for row in rows if row.emis})),
    }
    message = re.sub(r"{{\s*([a-zA-Z0-9_]+)\s*}}", lambda match: context.get(match.group(1), match.group(0)), SIMPLE_ACTIVITY_DEFAULT_TEMPLATE).strip(); context["report_body"] = message
    attachments: list[Path] = []; artifact_id = source.id
    if has_attachment:
        derived, path, digest = _attachment(session, source_job, source, header, rows, scope_label); artifact_id = derived.id; attachments.append(path)
    return RenderedSimpleActivityReport(
        message, context, rows, issues, attachments, artifact_id, digest, source_issues
    )


__all__ = ["SIMPLE_ACTIVITY_DEFAULT_TEMPLATE", "SIMPLE_ACTIVITY_RENDERER_KEY", "SIMPLE_ACTIVITY_REPORT_KEY", "render_simple_activity_report"]
