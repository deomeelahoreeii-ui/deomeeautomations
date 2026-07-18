from __future__ import annotations

import uuid
from pathlib import Path
from openpyxl import Workbook
from sqlalchemy.pool import StaticPool
from sqlmodel import SQLModel, Session, create_engine

import antidengue_automation.models  # noqa: F401
import automation_core.models  # noqa: F401
import master_data.models  # noqa: F401
import whatsapp_gateway.models  # noqa: F401
from antidengue_automation.simple_activity_routing import configure_simple_activity_routes, expand_simple_activity_profile_ids
from automation_core.models import Artifact, Job, JobStatus, JobType
from master_data.models import Wing
from whatsapp_gateway.models import WhatsAppAccount, WhatsAppApplication, WhatsAppAudience, WhatsAppDispatchProfile, WhatsAppReportType
from whatsapp_gateway.rendering.antidengue.simple_activity_report import render_simple_activity_report
from whatsapp_gateway.previews.compiler.errors import extend_unique_issues


def _engine():
    engine = create_engine("sqlite://", connect_args={"check_same_thread":False}, poolclass=StaticPool)
    SQLModel.metadata.create_all(engine); return engine


def test_simple_activity_renderer_scopes_rows_and_hides_attachment_claim(tmp_path: Path) -> None:
    engine = _engine(); wing = Wing(id=uuid.uuid4(), district_id=uuid.uuid4(), department_id=uuid.uuid4(), name="MEE", code="MEE"); tehsil=uuid.uuid4()
    path = tmp_path / "Simple Activity Timing Review - test.xlsx"
    workbook=Workbook(); sheet=workbook.active; sheet.title="Review Required"
    sheet.append(["School EMIS","School Name","Wing ID","Tehsil ID","Markaz ID","Tehsil","Markaz","Time Difference(Sec)","Activity(Lat,Long)","ID"])
    sheet.append(["35210001","School A",str(wing.id),str(tehsil),"","CITY","M1",245,"31.52,74.35","a1"])
    sheet.append(["35210002","School B",str(uuid.uuid4()),str(tehsil),"","CITY","M2",100,"31.50,74.30","a2"])
    workbook.save(path); workbook.close()
    with Session(engine) as session:
        job=Job(type=JobType.antidengue_report.value,title="Run",status=JobStatus.succeeded.value,result={"summary":{"portal_acquisition":{"window":{"dateto":"2026-07-18T17:11"}}}});session.add(job);session.flush()
        session.add(Artifact(job_id=job.id,module_key="antidengue",kind="report",name=path.name,path=str(path),size_bytes=path.stat().st_size));session.commit()
        rendered=render_simple_activity_report(session,source_job=job,wing=wing,recipient_name="MEE Heads",scope_key="wing",scope_value=str(wing.id),scope_label="MEE",presentation_policy={"attachment_mode":"none"})
        assert [row.emis for row in rendered.rows] == ["35210001"]
        assert "245 seconds (4 min 5 sec)" in rendered.message
        assert "Submitted:" not in rendered.message
        assert "Google Maps:" not in rendered.message
        assert "attached Excel" not in rendered.message
        assert rendered.attachment_paths == []


def test_unmapped_simple_activity_rows_are_a_detailed_source_issue_not_a_delivery_issue(tmp_path: Path) -> None:
    engine = _engine()
    wing = Wing(id=uuid.uuid4(), district_id=uuid.uuid4(), department_id=uuid.uuid4(), name="MEE", code="MEE")
    path = tmp_path / "Simple Activity Timing Review - unmapped.xlsx"
    workbook = Workbook(); sheet = workbook.active; sheet.title = "Review Required"
    sheet.append(["School EMIS", "School Name", "Wing ID", "Tehsil ID", "Markaz ID", "Tehsil", "Markaz", "Time Difference(Sec)", "ID"])
    sheet.append(["35210001", "School A", str(wing.id), "", "", "CITY", "M1", 100, "a1"])
    sheet.append(["35210443", "", "", "", "", "", "", 89, "a2"])
    sheet.append(["35210443", "", "", "", "", "", "", 76, "a3"])
    workbook.save(path); workbook.close()
    with Session(engine) as session:
        job = Job(type=JobType.antidengue_report.value, title="Run", status=JobStatus.succeeded.value, result={})
        session.add(job); session.flush()
        session.add(Artifact(job_id=job.id, module_key="antidengue", kind="report", name=path.name, path=str(path), size_bytes=path.stat().st_size)); session.commit()
        rendered = render_simple_activity_report(session, source_job=job, wing=wing, recipient_name="MEE", scope_key="wing", scope_value=str(wing.id), scope_label="MEE")
        assert not any(item["code"] == "unmapped_simple_activity_submitters" for item in rendered.issues)
        assert len(rendered.source_issues) == 1
        assert "35210443 (2 rows)" in rendered.source_issues[0]["message"]
        batch: list[dict] = []
        extend_unique_issues(batch, rendered.source_issues)
        extend_unique_issues(batch, rendered.source_issues)
        assert len(batch) == 1


def test_simple_activity_message_caps_text_but_preserves_full_attachment(tmp_path: Path) -> None:
    engine = _engine(); wing = Wing(id=uuid.uuid4(), district_id=uuid.uuid4(), department_id=uuid.uuid4(), name="MEE", code="MEE")
    path = tmp_path / "Simple Activity Timing Review - many.xlsx"
    workbook=Workbook(); sheet=workbook.active; sheet.title="Review Required"
    sheet.append(["School EMIS","School Name","Wing ID","Tehsil ID","Markaz ID","Tehsil","Markaz","Time Difference(Sec)","Latitude","Longitude","ID"])
    for index in range(25):
        sheet.append([f"3521{index:04d}",f"School {index}",str(wing.id),"","","CITY","M1",100+index,31.5,74.3,f"a{index}"])
    workbook.save(path); workbook.close()
    with Session(engine) as session:
        job=Job(type=JobType.antidengue_report.value,title="Run",status=JobStatus.succeeded.value,result={});session.add(job);session.flush()
        session.add(Artifact(job_id=job.id,module_key="antidengue",kind="report",name=path.name,path=str(path),size_bytes=path.stat().st_size));session.commit()
        rendered=render_simple_activity_report(session,source_job=job,wing=wing,recipient_name="MEE",scope_key="wing",scope_value=str(wing.id),scope_label="MEE",presentation_policy={"attachment_mode":"excel"})
        assert len(rendered.rows) == 25
        assert len(rendered.message) < 3500
        assert "17 additional school(s)" in rendered.message
        assert len(rendered.attachment_paths) == 1
        assert any(issue["severity"] == "info" for issue in rendered.issues)


def test_simple_activity_routes_expand_from_any_dormant_source() -> None:
    engine=_engine()
    with Session(engine) as session:
        account=WhatsAppAccount(name="Primary",worker_key="simple-routing"); app=WhatsAppApplication(key="antidengue",name="AntiDengue");session.add_all([account,app]);session.flush()
        dormant=WhatsAppReportType(application_id=app.id,key="tehsil_dormant_summary",name="Dormant"); simple=WhatsAppReportType(application_id=app.id,key="simple_activity_timing",name="Simple timing"); audience=WhatsAppAudience(application_id=app.id,key="a",name="A");session.add_all([dormant,simple,audience]);session.flush()
        sources=[WhatsAppDispatchProfile(application_id=app.id,key=f"s{i}",name=f"Source {i}",report_type_id=dormant.id,audience_id=audience.id,account_id=account.id,recipient_channel="group",delivery_mode="groups") for i in (1,2)];session.add_all(sources);session.commit()
        configured=configure_simple_activity_routes(session,[item.id for item in sources]); assert configured["enabled_count"]==2
        derived={uuid.UUID(item["simple_activity_profile_id"]) for item in configured["items"]}
        assert set(expand_simple_activity_profile_ids(session,[sources[0].id])) == {sources[0].id,*derived}
