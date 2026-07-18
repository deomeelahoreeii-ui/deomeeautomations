from __future__ import annotations

import hashlib
import json
import re
import uuid
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from sqlmodel import Session, select

from automation_core.config import get_settings
from automation_core.models import Artifact, Job
from whatsapp_gateway.rendering.antidengue.digest_models import DigestSchool

HEADER_FILL = "0F6F83"
TITLE_FILL = "123B4A"
ISSUE_FILLS = {"DORMANT": "FFF2CC", "DISTANCE": "FCE4D6", "TIMING": "DDEBF7"}


def _safe(value: Any) -> Any:
    if isinstance(value, str) and value[:1] in {"=", "+", "-", "@"}:
        return f"'{value}"
    return value


def _style_sheet(
    sheet, *, title: str, generated_label: str, deadline_label: str,
    deadline_timezone: str, table_name: str,
) -> None:
    last_column = max(1, sheet.max_column)
    sheet.insert_rows(1, 3)
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_column)
    title_cell = sheet.cell(1, 1, title)
    title_cell.font = Font(bold=True, color="FFFFFF", size=14)
    title_cell.fill = PatternFill("solid", fgColor=TITLE_FILL)
    title_cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 25
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_column)
    sheet.cell(2, 1, f"Generated: {generated_label}").font = Font(italic=True, color="5B6573")
    sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=last_column)
    sheet.cell(3, 1, f"Action deadline: {deadline_label} ({deadline_timezone})").font = Font(
        italic=True, color="5B6573"
    )
    for cell in sheet[4]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    sheet.freeze_panes = "A5"
    sheet.auto_filter.ref = f"A4:{sheet.cell(sheet.max_row, last_column).coordinate}"
    if sheet.max_row >= 5:
        table = Table(displayName=table_name, ref=f"A4:{sheet.cell(sheet.max_row, last_column).coordinate}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
            showRowStripes=True, showColumnStripes=False,
        )
        sheet.add_table(table)
    for column in sheet.columns:
        cells = [cell for cell in column if cell.coordinate not in sheet.merged_cells]
        if not cells:
            continue
        letter = cells[0].column_letter
        width = min(52, max(12, max(len(str(cell.value or "")) for cell in cells) + 2))
        sheet.column_dimensions[letter].width = width


def _append_overview(sheet, schools: Iterable[DigestSchool]) -> None:
    sheet.append([
        "School EMIS", "School Name", "Wing", "Tehsil", "Markaz", "Issue Codes",
        "Dormant", "Hotspot Distance", "Activity Timing",
    ])
    for school in schools:
        sheet.append([
            _safe(school.emis), _safe(school.school_name), _safe(school.wing),
            _safe(school.tehsil), _safe(school.markaz), " | ".join(school.issue_codes),
            "YES" if school.dormant else "", "YES" if school.hotspot_distance else "",
            "YES" if school.activity_timing else "",
        ])
        row = sheet.max_row
        for code, column in (("DORMANT", 7), ("DISTANCE", 8), ("TIMING", 9)):
            if sheet.cell(row, column).value:
                sheet.cell(row, column).fill = PatternFill("solid", fgColor=ISSUE_FILLS[code])


def _append_dormant(sheet, schools: Iterable[Any]) -> None:
    sheet.append(["School EMIS", "School Name", "Tehsil", "Markaz", "Issue"])
    for school in schools:
        sheet.append([
            _safe(school.emis), _safe(school.school_name), _safe(school.tehsil),
            _safe(school.markaz), "No qualifying activity submitted",
        ])


def _append_activity(sheet, rows: Iterable[Any], *, issue: str) -> None:
    materialized = list(rows)
    source_headers: list[str] = []
    for row in materialized:
        for key in row.values:
            if key not in source_headers:
                source_headers.append(key)
    sheet.append(["Issue"] + source_headers)
    for row in materialized:
        sheet.append([issue] + [_safe(row.values.get(name)) for name in source_headers])


def ensure_digest_workbook(
    session: Session,
    *,
    source_job: Job,
    scope_key: str,
    scope_label: str,
    scope_ids: list[str],
    generated_label: str,
    schools: list[DigestSchool],
    dormant_rows: list[Any],
    hotspot_rows: list[Any],
    timing_rows: list[Any],
    source_hashes: dict[str, str],
    deadline_label: str,
    deadline_timezone: str,
) -> tuple[Artifact, Path, str]:
    content_key = hashlib.sha256(json.dumps({
        "sources": source_hashes,
        "deadline": [deadline_label, deadline_timezone],
        "scope": [scope_key, scope_label, sorted(scope_ids)],
        "schools": [[item.emis, item.issue_codes] for item in schools],
        "hotspot": [str(item.values.get("ID") or index) for index, item in enumerate(hotspot_rows)],
        "timing": [str(item.values.get("ID") or index) for index, item in enumerate(timing_rows)],
    }, sort_keys=True).encode()).hexdigest()
    safe_scope = re.sub(r"[^A-Za-z0-9_-]+", "-", scope_label).strip("-") or scope_key
    path = (
        get_settings().artifact_root.expanduser().resolve()
        / "antidengue" / "native-reports" / str(source_job.id)
        / f"action-digest-{safe_scope}-{content_key[:16]}.xlsx"
    )
    artifact = session.scalar(select(Artifact).where(
        Artifact.job_id == source_job.id, Artifact.path == str(path)
    ))
    if not path.is_file():
        path.parent.mkdir(parents=True, exist_ok=True)
        temporary = path.with_name(f".{path.name}.{uuid.uuid4().hex}.tmp")
        workbook = Workbook()
        overview = workbook.active
        overview.title = "Overview"
        _append_overview(overview, schools)
        dormant = workbook.create_sheet("Dormant Schools")
        _append_dormant(dormant, dormant_rows)
        hotspot = workbook.create_sheet("Hotspot Distance")
        _append_activity(hotspot, hotspot_rows, issue="Distance outside permitted threshold")
        timing = workbook.create_sheet("Activity Timing")
        _append_activity(timing, timing_rows, issue="Before/after interval below minimum")
        for sheet, title, table_name in (
            (overview, "Anti-Dengue Action Digest — Overview", "DigestOverview"),
            (dormant, "Dormant Schools", "DigestDormant"),
            (hotspot, "Hotspot Distance Evidence", "DigestDistance"),
            (timing, "Activity Timing Evidence", "DigestTiming"),
        ):
            _style_sheet(
                sheet, title=title, generated_label=generated_label,
                deadline_label=deadline_label, deadline_timezone=deadline_timezone,
                table_name=table_name,
            )
        workbook.save(temporary)
        workbook.close()
        temporary.replace(path)
    if artifact is None:
        artifact = Artifact(
            job_id=source_job.id, module_key="antidengue", kind="delivery",
            name=f"Anti-Dengue Action Digest - {scope_label}.xlsx", path=str(path),
            size_bytes=path.stat().st_size,
        )
    else:
        artifact.size_bytes = path.stat().st_size
    session.add(artifact)
    session.flush()
    digest = hashlib.sha256(path.read_bytes()).hexdigest()
    return artifact, path, digest


__all__ = ["ensure_digest_workbook"]
