from __future__ import annotations

import hashlib
import json
import re
import uuid
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import select
from sqlmodel import Session

from automation_core.config import get_settings
from automation_core.models import Artifact, Job
from master_data.models import Wing
from whatsapp_gateway.rendering.antidengue.formatting import _wing_display_name
from whatsapp_gateway.rendering.antidengue.models import ScopedDormantSchool


def _ensure_excel_attachment(
    session: Session,
    *,
    source_job: Job,
    wing: Wing,
    schools: list[ScopedDormantSchool],
    source_digest: str,
    scope_key: str,
    scope_label: str,
) -> tuple[Artifact, Path, str]:
    content_key = hashlib.sha256(
        json.dumps(
            {
                "source": source_digest,
                "wing": str(wing.id),
                "scope": [scope_key, scope_label],
                "schools": [
                    [school.emis, school.school_name, school.tehsil, school.markaz]
                    for school in schools
                ],
            },
            ensure_ascii=False,
            sort_keys=True,
        ).encode("utf-8")
    ).hexdigest()
    safe_wing = re.sub(r"[^A-Za-z0-9_-]+", "-", wing.code or wing.name).strip("-") or "wing"
    safe_scope = re.sub(r"[^A-Za-z0-9_-]+", "-", scope_label).strip("-") or scope_key
    filename = f"Anti-Dengue Dormant Users - {safe_wing} - {safe_scope}.xlsx"
    path = (
        get_settings().artifact_root.expanduser().resolve()
        / "antidengue"
        / "native-reports"
        / str(source_job.id)
        / f"{safe_wing}-{safe_scope}-{content_key[:16]}.xlsx"
    )
    artifact = session.scalar(
        select(Artifact).where(Artifact.job_id == source_job.id, Artifact.path == str(path))
    )
    if artifact is not None and path.is_file():
        return artifact, path, hashlib.sha256(path.read_bytes()).hexdigest()

    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(f".{path.name}.{uuid.uuid4().hex}.tmp")
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Dormant Schools"
    title = f"Anti-Dengue Dormant Schools — {scope_label} — {_wing_display_name(wing)}"
    sheet.merge_cells("A1:E1")
    sheet["A1"] = title
    sheet["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="087487")
    sheet["A1"].alignment = Alignment(horizontal="center")
    sheet.append(["School EMIS", "School Name", "Tehsil", "Markaz", "Wing"])
    for cell in sheet[2]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="26445D")
    for school in schools:
        sheet.append([school.emis, school.school_name, school.tehsil, school.markaz, wing.name])
    sheet.freeze_panes = "A3"
    sheet.auto_filter.ref = f"A2:E{sheet.max_row}"
    for column, width in {"A": 16, "B": 52, "C": 20, "D": 28, "E": 18}.items():
        sheet.column_dimensions[column].width = width
    for row in sheet.iter_rows(min_row=3):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    workbook.save(temporary)
    workbook.close()
    temporary.replace(path)

    if artifact is None:
        artifact = Artifact(
            job_id=source_job.id,
            module_key="antidengue",
            kind="delivery",
            name=filename,
            path=str(path),
            size_bytes=path.stat().st_size,
        )
    else:
        artifact.name = filename
        artifact.size_bytes = path.stat().st_size
    session.add(artifact)
    session.flush()
    return artifact, path, hashlib.sha256(path.read_bytes()).hexdigest()
