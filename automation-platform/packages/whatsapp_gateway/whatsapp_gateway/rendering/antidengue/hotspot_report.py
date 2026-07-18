from __future__ import annotations

import hashlib
import json
import re
import uuid
from datetime import datetime
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import select
from sqlmodel import Session

from automation_core.config import get_settings
from automation_core.models import Artifact, Job
from master_data.models import Wing
from whatsapp_gateway.rendering.antidengue.issues import _issue

HOTSPOT_RENDERER_KEY = "antidengue.hotspot_distance.v1"
HOTSPOT_REPORT_KEY = "hotspot_distance_activity"
HOTSPOT_DEFAULT_TEMPLATE = """🚨 *Anti-Dengue Activities Requiring Review*
👥 *Group:* {{group_name}}
📌 *Scope:* {{scope_name}}
🕒 *Report window:* {{report_window}}

🏫 *School-wise activity evidence*

{{school_activity_details}}

{{attachment_note}}
{{disclaimer}}"""


@dataclass(frozen=True)
class HotspotReviewRow:
    values: dict[str, Any]
    emis: str
    school_name: str
    wing_id: str
    tehsil_id: str
    markaz_id: str
    distance_meters: float | None


@dataclass
class RenderedHotspotReport:
    message: str
    context: dict[str, str]
    rows: list[HotspotReviewRow]
    issues: list[dict[str, Any]] = field(default_factory=list)
    attachment_paths: list[Path] = field(default_factory=list)
    source_artifact_id: int | None = None
    source_artifact_sha256: str = ""


def _plain_whatsapp_text(value: Any, fallback: str = "Not available") -> str:
    text = re.sub(r"[\r\n\t]+", " ", str(value or "")).strip()
    # Do not allow source data to open or close WhatsApp formatting markers.
    text = re.sub(r"[*_~`]", "", text)
    return text or fallback


def _coordinates(value: Any) -> tuple[float, float] | None:
    numbers = re.findall(r"[-+]?\d+(?:\.\d+)?", str(value or ""))
    if len(numbers) < 2:
        return None
    latitude, longitude = float(numbers[0]), float(numbers[1])
    if not (-90 <= latitude <= 90 and -180 <= longitude <= 180):
        return None
    return latitude, longitude


def _maps_link(value: Any) -> str:
    point = _coordinates(value)
    if point is None:
        return ""
    latitude, longitude = point
    return f"https://www.google.com/maps?q={latitude:.7f},{longitude:.7f}"


def _window_label(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    try:
        cutoff = datetime.fromisoformat(text)
    except ValueError:
        return text.replace("T", " ")
    return f"12:00 AM – {cutoff.strftime('%I:%M %p')} ({cutoff.strftime('%d %b %Y')})"


def _school_summary_lines(rows: list[HotspotReviewRow]) -> list[str]:
    grouped: dict[str, list[HotspotReviewRow]] = {}
    for row in rows:
        grouped.setdefault(row.emis, []).append(row)

    def group_key(item: tuple[str, list[HotspotReviewRow]]) -> tuple[str, str, str]:
        _, school_rows = item
        values = school_rows[0].values
        return (
            str(values.get("Tehsil") or ""),
            str(values.get("Markaz") or ""),
            school_rows[0].school_name,
        )

    lines: list[str] = []
    for index, (emis, school_rows) in enumerate(sorted(grouped.items(), key=group_key), start=1):
        representative = school_rows[0]
        values = representative.values
        lines.extend([
            f"🏫 *{index}. {_plain_whatsapp_text(representative.school_name, 'Unknown school')}*",
            f"🆔 *EMIS:* `{_plain_whatsapp_text(emis)}`",
            (
                f"📍 *Tehsil:* {_plain_whatsapp_text(values.get('Tehsil'))}"
                f"  |  *Markaz:* {_plain_whatsapp_text(values.get('Markaz'))}"
            ),
        ])
        ordered_rows = sorted(
            school_rows,
            key=lambda row: row.distance_meters if row.distance_meters is not None else -1,
            reverse=True,
        )
        for activity_index, row in enumerate(ordered_rows, start=1):
            if len(ordered_rows) > 1:
                lines.append(f"🔸 *Activity {activity_index}*")
            distance = (
                f"{row.distance_meters:,.2f}"
                if row.distance_meters is not None
                else "Not available"
            )
            lines.append(f"📏 *Distance Difference (In meters):* {distance}")
            link = _maps_link(row.values.get("Activity(Lat,Long)"))
            if link:
                lines.append(f"🗺️ *Google Maps:* {link}")
            else:
                lines.append("🗺️ *Google Maps:* Coordinates unavailable in the portal report")
        lines.append("")
    return lines


def _review_artifact(session: Session, source_job: Job) -> tuple[Artifact, Path, list[str], list[HotspotReviewRow]]:
    artifacts = session.scalars(
        select(Artifact).where(Artifact.job_id == source_job.id).order_by(Artifact.created_at, Artifact.id)
    ).all()
    for artifact in artifacts:
        if not artifact.name.lower().split("/")[-1].startswith("hotspot distance review"):
            continue
        path = Path(artifact.path).expanduser().resolve(strict=False)
        if not path.is_file():
            continue
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook["Review Required"] if "Review Required" in workbook.sheetnames else workbook.active
            values = sheet.iter_rows(values_only=True)
            header = [str(value or "").strip() for value in next(values, ())]
            required = {"School EMIS", "School Name", "Wing ID", "Tehsil ID", "Markaz ID", "Distance Difference (In meters)"}
            if not required.issubset(header):
                continue
            rows: list[HotspotReviewRow] = []
            for raw in values:
                item = {name: raw[index] if index < len(raw) else None for index, name in enumerate(header)}
                try:
                    distance = float(item.get("Distance Difference (In meters)"))
                except (TypeError, ValueError):
                    distance = None
                rows.append(HotspotReviewRow(
                    values=item,
                    emis=str(item.get("School EMIS") or "").strip(),
                    school_name=str(item.get("School Name") or "").strip(),
                    wing_id=str(item.get("Wing ID") or "").strip(),
                    tehsil_id=str(item.get("Tehsil ID") or "").strip(),
                    markaz_id=str(item.get("Markaz ID") or "").strip(),
                    distance_meters=distance,
                ))
            return artifact, path, header, rows
        finally:
            workbook.close()
    raise ValueError(
        "The source run has no classified hotspot-distance review workbook. Publish and enable an Activity Rule, then run Test Run again."
    )


def _ensure_attachment(
    session: Session,
    *,
    source_job: Job,
    wing: Wing,
    source_digest: str,
    header: list[str],
    rows: list[HotspotReviewRow],
    scope_key: str,
    scope_label: str,
) -> tuple[Artifact, Path, str]:
    content_key = hashlib.sha256(json.dumps({
        "source": source_digest,
        "wing": str(wing.id),
        "scope": [scope_key, scope_label],
        "activities": [str(row.values.get("ID") or "") for row in rows],
    }, sort_keys=True).encode()).hexdigest()
    safe_scope = re.sub(r"[^A-Za-z0-9_-]+", "-", scope_label).strip("-") or scope_key
    path = (
        get_settings().artifact_root.expanduser().resolve()
        / "antidengue" / "native-reports" / str(source_job.id)
        / f"hotspot-{safe_scope}-{content_key[:16]}.xlsx"
    )
    artifact = session.scalar(select(Artifact).where(Artifact.job_id == source_job.id, Artifact.path == str(path)))
    if artifact is not None and path.is_file():
        return artifact, path, hashlib.sha256(path.read_bytes()).hexdigest()
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(f".{path.name}.{uuid.uuid4().hex}.tmp")
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Review Required"
    sheet.append(header)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="B45309")
        cell.alignment = Alignment(wrap_text=True)
    for row in rows:
        sheet.append([row.values.get(name) for name in header])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column in sheet.columns:
        letter = column[0].column_letter
        width = min(55, max(12, max(len(str(cell.value or "")) for cell in column) + 2))
        sheet.column_dimensions[letter].width = width
    workbook.save(temporary)
    workbook.close()
    temporary.replace(path)
    if artifact is None:
        artifact = Artifact(
            job_id=source_job.id, module_key="antidengue", kind="delivery",
            name=f"Hotspot Distance Review - {scope_label}.xlsx", path=str(path),
            size_bytes=path.stat().st_size,
        )
    else:
        artifact.size_bytes = path.stat().st_size
    session.add(artifact)
    session.flush()
    return artifact, path, hashlib.sha256(path.read_bytes()).hexdigest()


def render_hotspot_distance_report(
    session: Session,
    *,
    source_job: Job,
    wing: Wing,
    recipient_name: str,
    scope_key: str,
    scope_value: str,
    scope_label: str,
    presentation_policy: dict[str, Any] | None = None,
) -> RenderedHotspotReport:
    source_artifact, source_path, header, source_rows = _review_artifact(session, source_job)
    scoped = [row for row in source_rows if row.wing_id == str(wing.id)]
    if scope_key == "tehsil":
        scoped = [row for row in scoped if row.tehsil_id == scope_value]
    elif scope_key == "markaz":
        scoped = [row for row in scoped if row.markaz_id == scope_value]
    elif scope_key not in {"district", "wing"}:
        raise ValueError(f"Hotspot-distance delivery does not support {scope_key or 'an unbound'} scope")

    issues: list[dict[str, Any]] = []
    unmapped = [row for row in source_rows if not row.wing_id or not row.emis]
    if unmapped:
        issues.append(_issue(
            "unmapped_hotspot_submitters", "warning",
            f"{len(unmapped)} classified activity row(s) could not be mapped to authoritative school routing and were excluded.",
        ))
    if not scoped:
        issues.append(_issue(
            "nothing_to_dispatch", "warning",
            f"No activities requiring review were found for {scope_label}.",
        ))
        return RenderedHotspotReport(
            message="", context={}, rows=[], issues=issues,
            source_artifact_id=source_artifact.id,
            source_artifact_sha256=hashlib.sha256(source_path.read_bytes()).hexdigest(),
        )

    unique_schools = len({row.emis for row in scoped if row.emis})
    window = dict(((source_job.result or {}).get("summary") or {}).get("portal_acquisition") or {}).get("window") or {}
    window_label = _window_label(window.get("dateto"))
    school_activity_details = "\n".join(_school_summary_lines(scoped)).strip()
    attachment_mode = str((presentation_policy or {}).get("attachment_mode") or "excel")
    has_attachment = attachment_mode != "none"
    context = {
        "group_name": _plain_whatsapp_text(recipient_name),
        "scope_name": _plain_whatsapp_text(scope_label),
        "report_window": window_label or "Not available",
        "school_activity_details": school_activity_details,
        "attachment_note": (
            "📎 The attached Excel file contains every selected activity and its evidence."
            if has_attachment
            else ""
        ),
        "disclaimer": "⚠️ These are rule-based review candidates, not confirmed fake activities. Please verify each finding.",
        "activity_count": str(len(scoped)),
        "school_count": str(unique_schools),
    }
    message = re.sub(
        r"{{\s*([a-zA-Z0-9_]+)\s*}}",
        lambda match: context.get(match.group(1), match.group(0)),
        HOTSPOT_DEFAULT_TEMPLATE,
    ).strip()
    context["report_body"] = message
    attachment_paths: list[Path] = []
    artifact_id = source_artifact.id
    digest = hashlib.sha256(source_path.read_bytes()).hexdigest()
    if has_attachment:
        derived, attachment, digest = _ensure_attachment(
            session, source_job=source_job, wing=wing, source_digest=digest,
            header=header, rows=scoped, scope_key=scope_key, scope_label=scope_label,
        )
        artifact_id = derived.id
        attachment_paths.append(attachment)
    return RenderedHotspotReport(
        message=message,
        context=context,
        rows=scoped,
        issues=issues,
        attachment_paths=attachment_paths,
        source_artifact_id=artifact_id,
        source_artifact_sha256=digest,
    )


__all__ = [
    "HOTSPOT_DEFAULT_TEMPLATE", "HOTSPOT_RENDERER_KEY", "HOTSPOT_REPORT_KEY",
    "HotspotReviewRow", "RenderedHotspotReport", "render_hotspot_distance_report",
]
