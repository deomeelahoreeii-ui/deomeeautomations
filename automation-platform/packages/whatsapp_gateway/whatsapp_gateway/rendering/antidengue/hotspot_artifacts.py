from __future__ import annotations

import hashlib
import json
import re
import uuid
from pathlib import Path
from typing import TYPE_CHECKING

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import select
from sqlmodel import Session

from automation_core.config import get_settings
from automation_core.models import Artifact, Job
from automation_core.storage_catalog import ensure_artifact_local
from master_data.models import Wing

if TYPE_CHECKING:
    from whatsapp_gateway.rendering.antidengue.hotspot_report import HotspotReviewRow


def review_artifact(
    session: Session, source_job: Job,
) -> tuple[Artifact, Path, list[str], list[HotspotReviewRow]]:
    from whatsapp_gateway.rendering.antidengue.hotspot_report import HotspotReviewRow

    artifacts = session.scalars(
        select(Artifact).where(Artifact.job_id == source_job.id).order_by(
            Artifact.created_at, Artifact.id
        )
    ).all()
    for artifact in artifacts:
        if not artifact.name.lower().split("/")[-1].startswith("hotspot distance review"):
            continue
        try:
            path = ensure_artifact_local(session, artifact)
        except Exception:
            continue
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook["Review Required"] if "Review Required" in workbook.sheetnames else workbook.active
            values = sheet.iter_rows(values_only=True)
            header = [str(value or "").strip() for value in next(values, ())]
            required = {
                "School EMIS", "School Name", "Wing ID", "Tehsil ID", "Markaz ID",
                "Distance Difference (In meters)",
            }
            if not required.issubset(header):
                continue
            rows: list[HotspotReviewRow] = []
            for raw in values:
                item = {
                    name: raw[index] if index < len(raw) else None
                    for index, name in enumerate(header)
                }
                try:
                    distance = float(item.get("Distance Difference (In meters)"))
                except (TypeError, ValueError):
                    distance = None
                rows.append(HotspotReviewRow(
                    values=item, emis=str(item.get("School EMIS") or "").strip(),
                    school_name=str(item.get("School Name") or "").strip(),
                    wing_id=str(item.get("Wing ID") or "").strip(),
                    tehsil_id=str(item.get("Tehsil ID") or "").strip(),
                    markaz_id=str(item.get("Markaz ID") or "").strip(),
                    distance_meters=distance,
                ))
            return artifact, path, header, rows
        finally:
            workbook.close()
    raise ValueError(
        "The source run has no classified hotspot-distance review workbook. "
        "Publish and enable an Activity Rule, then run Test Run again."
    )


def ensure_attachment(
    session: Session, *, source_job: Job, wing: Wing, source_digest: str,
    header: list[str], rows: list[HotspotReviewRow], scope_key: str,
    scope_label: str,
) -> tuple[Artifact, Path, str]:
    content_key = hashlib.sha256(json.dumps({
        "source": source_digest, "wing": str(wing.id),
        "scope": [scope_key, scope_label],
        "activities": [str(row.values.get("ID") or "") for row in rows],
    }, sort_keys=True).encode()).hexdigest()
    safe_scope = re.sub(r"[^A-Za-z0-9_-]+", "-", scope_label).strip("-") or scope_key
    path = (
        get_settings().artifact_root.expanduser().resolve()
        / "antidengue" / "native-reports" / str(source_job.id)
        / f"hotspot-{safe_scope}-{content_key[:16]}.xlsx"
    )
    artifact = session.scalar(
        select(Artifact).where(Artifact.job_id == source_job.id, Artifact.path == str(path))
    )
    if artifact is not None and path.is_file():
        return artifact, path, hashlib.sha256(path.read_bytes()).hexdigest()
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(f".{path.name}.{uuid.uuid4().hex}.tmp")
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Review Required"
    sheet.append(header)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="B45309")
        cell.alignment = Alignment(wrap_text=True)
    for row in rows:
        sheet.append([row.values.get(name) for name in header])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column in sheet.columns:
        letter = column[0].column_letter
        width = min(55, max(12, max(len(str(cell.value or "")) for cell in column) + 2))
        sheet.column_dimensions[letter].width = width
    workbook.save(temporary)
    workbook.close()
    temporary.replace(path)
    if artifact is None:
        artifact = Artifact(
            job_id=source_job.id, module_key="antidengue", kind="delivery",
            name=f"Hotspot Distance Review - {scope_label}.xlsx", path=str(path),
            size_bytes=path.stat().st_size,
        )
    else:
        artifact.size_bytes = path.stat().st_size
    session.add(artifact)
    session.flush()
    return artifact, path, hashlib.sha256(path.read_bytes()).hexdigest()


__all__ = ["ensure_attachment", "review_artifact"]
