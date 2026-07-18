from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select
from sqlmodel import Session

from automation_core.models import Artifact, Job
from whatsapp_gateway.rendering.antidengue.models import DormantSourceRow, REQUIRED_COLUMNS


def _normalize_emis(value: object) -> str:
    text = str(value or "").strip()
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    return text


def _report_rows(path: Path) -> list[DormantSourceRow] | None:
    if path.suffix.lower() != ".xlsx" or not path.is_file():
        return None
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet in workbook.worksheets:
            rows = sheet.iter_rows(values_only=True)
            for raw_header in rows:
                header = [str(value or "").strip().lower() for value in raw_header]
                if not REQUIRED_COLUMNS.issubset(header):
                    continue
                emis_index = header.index("school emis")
                name_index = header.index("school name")
                result: list[DormantSourceRow] = []
                for raw_row in rows:
                    emis = _normalize_emis(
                        raw_row[emis_index] if emis_index < len(raw_row) else ""
                    )
                    if not emis:
                        continue
                    school_name = str(
                        raw_row[name_index] if name_index < len(raw_row) else ""
                    ).strip()
                    result.append(DormantSourceRow(emis=emis, school_name=school_name))
                return result
    finally:
        workbook.close()
    return None


def _canonical_report_artifact(
    session: Session, source_job: Job
) -> tuple[Artifact, Path, list[DormantSourceRow]]:
    artifacts = session.scalars(
        select(Artifact)
        .where(Artifact.job_id == source_job.id)
        .order_by(Artifact.created_at, Artifact.id)
    ).all()
    # The platform adapter registers the canonical dormant workbook as
    # ``report``. Header inspection is retained as a safe migration fallback.
    candidates = sorted(
        (
            item
            for item in artifacts
            if "hotspot distance review" not in item.name.lower()
            and "simple activity timing review" not in item.name.lower()
        ),
        key=lambda item: (
            "anti-dengue app dormant users" not in item.name.lower(),
            item.kind != "report",
            item.created_at,
        ),
    )
    for artifact in candidates:
        path = Path(artifact.path).expanduser().resolve(strict=False)
        try:
            rows = _report_rows(path)
        except Exception:
            continue
        if rows is not None:
            return artifact, path, rows
    raise ValueError(
        "The source run has no canonical dormant-school report with School EMIS and School Name columns."
    )
