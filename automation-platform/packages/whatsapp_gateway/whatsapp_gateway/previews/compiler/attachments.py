from __future__ import annotations

import uuid
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select
from sqlmodel import Session

from master_data.models import School

def _artifact_role(path: Path, referenced_paths: set[str]) -> str:
    normalized = str(path.resolve(strict=False))
    name = path.name.lower()
    if normalized in referenced_paths or "group-route-excels" in normalized:
        return "delivery"
    if name == "run_summary.json":
        return "manifest"
    if "audit" in name or "evidence" in name:
        return "audit"
    return "supporting"

def _attachment_paths(plan: dict[str, Any]) -> list[Path]:
    payload = plan.get("payload") if isinstance(plan.get("payload"), dict) else {}
    candidates: list[Any] = [
        payload.get("excel_path"),
        payload.get("image_path"),
        plan.get("excel_path"),
    ]
    attachment_paths = payload.get("attachment_paths") or []
    if isinstance(attachment_paths, list):
        candidates.extend(attachment_paths)
    documents = payload.get("documents") or []
    if isinstance(documents, list):
        for document in documents:
            candidates.append(document.get("path") if isinstance(document, dict) else document)
    result: list[Path] = []
    seen: set[str] = set()
    for value in candidates:
        if not value:
            continue
        path = Path(str(value)).expanduser().resolve(strict=False)
        key = str(path)
        if key not in seen:
            seen.add(key)
            result.append(path)
    return result

def _xlsx_emis_values(path: Path) -> set[str]:
    """Read school EMIS values from a report without trusting its filename."""
    if path.suffix.lower() != ".xlsx":
        return set()
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet in workbook.worksheets:
            rows = sheet.iter_rows(values_only=True)
            for row in rows:
                normalized = [str(value).strip().lower() if value is not None else "" for value in row]
                if "school emis" not in normalized:
                    continue
                index = normalized.index("school emis")
                values: set[str] = set()
                for data_row in rows:
                    if index >= len(data_row) or data_row[index] is None:
                        continue
                    value = str(data_row[index]).strip()
                    if value.endswith(".0") and value[:-2].isdigit():
                        value = value[:-2]
                    if value:
                        values.add(value)
                return values
    finally:
        workbook.close()
    return set()

def _classify_attachment_wings(
    session: Session,
    paths: list[Path],
) -> tuple[set[uuid.UUID], list[str], bool]:
    emis_values: set[str] = set()
    readable_report_found = False
    for path in paths:
        if path.suffix.lower() != ".xlsx" or not path.is_file():
            continue
        try:
            values = _xlsx_emis_values(path)
        except Exception:
            continue
        if values:
            readable_report_found = True
            emis_values.update(values)
    if not emis_values:
        return set(), [], readable_report_found
    schools = session.scalars(select(School).where(School.emis.in_(emis_values))).all()
    found = {school.emis for school in schools}
    return {school.wing_id for school in schools}, sorted(emis_values - found), True

def _classify_scoped_emis_wings(
    session: Session,
    values: list[Any],
) -> tuple[set[uuid.UUID], list[str], bool]:
    """Classify native outputs from the authoritative scope frozen by the renderer."""
    emis_values = {str(value).strip() for value in values if str(value).strip()}
    if not emis_values:
        return set(), [], False
    schools = session.scalars(select(School).where(School.emis.in_(emis_values))).all()
    found = {school.emis for school in schools}
    return {school.wing_id for school in schools}, sorted(emis_values - found), True

