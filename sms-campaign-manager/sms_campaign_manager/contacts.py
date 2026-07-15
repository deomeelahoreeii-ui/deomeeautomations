from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


NUMBER_HINTS = ("contact_no", "contact", "phone", "phone_no", "mobile", "number", "cell")
MAX_CONSECUTIVE_EMPTY_ROWS = 1000


@dataclass(frozen=True)
class Contact:
    row: int
    phone: str


def workbook_columns(path: Path, sheet: str = "") -> tuple[list[str], str]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if not workbook.sheetnames:
            raise ValueError("Workbook has no sheets")
        selected = sheet or workbook.sheetnames[0]
        if selected not in workbook.sheetnames:
            raise ValueError(f"Sheet not found: {selected}")
        row = next(workbook[selected].iter_rows(min_row=1, max_row=1, values_only=True), ())
        columns = [str(value).strip() if value is not None else f"Column {index}" for index, value in enumerate(row, 1)]
        return columns, selected
    finally:
        workbook.close()


def suggested_number_column(columns: list[str]) -> str:
    normalized = {re.sub(r"[^a-z0-9]+", "_", name.lower()).strip("_"): name for name in columns}
    for hint in NUMBER_HINTS:
        if hint in normalized:
            return normalized[hint]
    for key, name in normalized.items():
        if any(hint in key for hint in NUMBER_HINTS):
            return name
    return columns[0] if columns else ""


def normalize_phone(value: object, country_code: str = "92") -> str | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    raw = str(value).strip()
    digits = re.sub(r"\D", "", raw)
    if raw.startswith("+"):
        normalized = digits
    elif digits.startswith("00"):
        normalized = digits[2:]
    elif digits.startswith(country_code):
        normalized = digits
    elif digits.startswith("0"):
        normalized = country_code + digits[1:]
    else:
        normalized = country_code + digits
    if not 10 <= len(normalized) <= 15:
        return None
    return "+" + normalized


def load_contacts(path: Path, column: str, sheet: str = "", country_code: str = "92") -> tuple[list[Contact], int]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        selected = sheet or workbook.sheetnames[0]
        worksheet = workbook[selected]
        rows = worksheet.iter_rows(values_only=True)
        headers = next(rows, ())
        names = [str(value).strip() if value is not None else "" for value in headers]
        if column not in names:
            raise ValueError(f"Column '{column}' not found. Available: {', '.join(names)}")
        index = names.index(column)
        contacts: list[Contact] = []
        seen: set[str] = set()
        invalid = 0
        empty_streak = 0
        for row_number, values in enumerate(rows, 2):
            if not any(value not in (None, "") for value in values):
                empty_streak += 1
                if empty_streak >= MAX_CONSECUTIVE_EMPTY_ROWS:
                    break
                continue
            empty_streak = 0
            phone = normalize_phone(values[index] if index < len(values) else None, country_code)
            if not phone:
                invalid += 1
                continue
            if phone in seen:
                continue
            seen.add(phone)
            contacts.append(Contact(row_number, phone))
        return contacts, invalid
    finally:
        workbook.close()


def sms_segments(text: str) -> tuple[int, str]:
    gsm_basic = set("@£$¥èéùìòÇ\nØø\rÅåΔ_ΦΓΛΩΠΨΣΘΞ !\"#¤%&'()*+,-./0123456789:;<=>?¡ABCDEFGHIJKLMNOPQRSTUVWXYZÄÖÑÜ`¿abcdefghijklmnopqrstuvwxyzäöñüà")
    gsm_extended = set("^{}\\[~]|€")
    gsm = all(character in gsm_basic or character in gsm_extended for character in text)
    units = sum(2 if character in gsm_extended else 1 for character in text) if gsm else len(text)
    single, multipart = (160, 153) if gsm else (70, 67)
    segments = 1 if units <= single else (units + multipart - 1) // multipart
    return segments, "GSM-7" if gsm else "Unicode"
