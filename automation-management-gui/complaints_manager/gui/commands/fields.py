from __future__ import annotations

import csv
import hashlib
import json
import re
import shlex
import shutil
import sqlite3
import subprocess
import tempfile
import threading
import time
import urllib.request
import uuid
from collections import Counter
from contextlib import contextmanager
from datetime import date, datetime, timedelta
from pathlib import Path

from imgui_bundle import imgui
from openpyxl import load_workbook

from complaints_manager.core.commands import (
    ArgField,
    COMMANDS,
    SYSTEMS,
    CommandSpec,
    SystemSpec,
    command_by_id,
    commands_by_system,
)
from complaints_manager.core.services import normalize_whatsapp_account_id
from complaints_manager.gui.paths import (
    apply_selected_folder,
    apply_selected_file,
    consume_file_dialog_result,
    consume_folder_dialog_result,
    existing_initial_dir,
    resolve_gui_path,
    start_file_dialog,
    start_folder_dialog,
)
from complaints_manager.gui.state import GuiState

from complaints_manager.gui.common.context import disabled_when
from complaints_manager.gui.common.desktop import open_folder_from_gui
from complaints_manager.gui.common.formatting import combine_datetime_value, parse_datetime_parts
from complaints_manager.gui.sms.campaigns import draw_sms_campaign_selector
from complaints_manager.gui.whatsapp.groups import draw_whatsapp_account_selector, draw_whatsapp_group_selector

def draw_field(
    state: GuiState,
    command: CommandSpec,
    values: dict[str, str],
    field: ArgField,
) -> None:
    if field.kind.startswith("hidden_env:"):
        return

    current_value = values.get(field.key, field.default)

    if field.kind == "sms_campaign":
        draw_sms_campaign_selector(state, command, values, field)
        return

    if command.id in {"whatsapp_group_import", "whatsapp_group_membership_audit"} and field.key == "group_jid":
        draw_whatsapp_group_selector(state, command, values, field)
        return
    if command.id == "whatsapp_group_membership_audit" and field.key == "worker_id":
        draw_whatsapp_account_selector(state, command, values, field)
        return
    if field.kind == "bool_flag":
        checked = current_value.strip().lower() in {"1", "true", "yes", "y", "on"}
        changed, checked = imgui.checkbox(field.label, checked)
        if changed:
            values[field.key] = "true" if checked else "false"
            state.persist_command_values(command)
        return

    if field.kind == "select":
        labels = [option.label for option in field.options]
        option_values = [option.value for option in field.options]
        try:
            index = option_values.index(current_value)
        except ValueError:
            index = 0
        changed, index = imgui.combo(field.label, index, labels)
        if changed:
            values[field.key] = option_values[index]
            state.persist_command_values(command)
        return

    if field.kind == "excel_column":
        file_value = values.get("file", "").strip()
        sheet_value = values.get("sheet", "").strip()
        columns: list[str] = []
        if file_value:
            path = resolve_gui_path(state.command_root(command), file_value)
            try:
                workbook = load_workbook(path, read_only=True, data_only=True)
                selected_sheet = sheet_value if sheet_value in workbook.sheetnames else workbook.sheetnames[0]
                header = next(workbook[selected_sheet].iter_rows(min_row=1, max_row=1, values_only=True), ())
                columns = [str(item).strip() for item in header if item is not None and str(item).strip()]
                workbook.close()
            except (OSError, ValueError, KeyError):
                columns = []
        if columns:
            if current_value not in columns:
                normalized = {name.lower().replace(" ", "_"): name for name in columns}
                values[field.key] = normalized.get("contact_no", columns[0])
                current_value = values[field.key]
            index = columns.index(current_value)
            imgui.set_next_item_width(420)
            changed, index = imgui.combo(field.label, index, columns)
            if changed:
                values[field.key] = columns[index]
                state.persist_command_values(command)
        else:
            imgui.text_disabled("Select a readable Excel file to load number columns.")
        return

    if field.kind == "excel_sheet":
        file_value = values.get("file", "").strip()
        sheets: list[str] = []
        if file_value:
            try:
                workbook = load_workbook(resolve_gui_path(state.command_root(command), file_value), read_only=True, data_only=True)
                sheets = list(workbook.sheetnames)
                workbook.close()
            except (OSError, ValueError, KeyError):
                sheets = []
        if sheets:
            selected = current_value if current_value in sheets else sheets[0]
            if current_value != selected:
                values[field.key] = selected
            index = sheets.index(selected)
            imgui.set_next_item_width(420)
            changed, index = imgui.combo(field.label, index, sheets)
            if changed:
                values[field.key] = sheets[index]
                values["column"] = ""
                state.persist_command_values(command)
        else:
            imgui.text_disabled("Select a readable Excel file to load sheets.")
        return

    if field.kind == "multiline":
        imgui.set_next_item_width(720)
        changed, value = imgui.input_text_multiline(field.label, current_value, (720, 110))
        if changed:
            values[field.key] = value
        return

    if field.kind in {"path", "file"}:
        imgui.set_next_item_width(520)
        changed, value = imgui.input_text(field.label, current_value)
        if changed:
            values[field.key] = value
            state.persist_command_values(command)

        if field.kind == "path":
            consume_folder_dialog_result(state, command, values, field.key, field.label)
            imgui.same_line()
            dialog_is_open = field.key in state.folder_dialogs
            with disabled_when(dialog_is_open):
                if imgui.button(f"Select Folder##{command.id}_{field.key}", (130, 0)):
                    selected = start_folder_dialog(
                        state=state,
                        key=field.key,
                        title=f"Select {field.label}",
                        initial_dir=existing_initial_dir(
                            state.command_root(command), values[field.key]
                        ),
                    )
                    if selected:
                        apply_selected_folder(
                            state, command, values, field.key, field.label, selected
                        )
            imgui.same_line()
            if imgui.button(f"Open##{command.id}_{field.key}", (70, 0)):
                open_folder_from_gui(state, command, values[field.key])
            return

        consume_file_dialog_result(state, command, values, field.key, field.label)
        imgui.same_line()
        dialog_is_open = field.key in state.file_dialogs
        with disabled_when(dialog_is_open):
            if imgui.button(f"Select File##{command.id}_{field.key}", (115, 0)):
                selected = start_file_dialog(
                    state=state,
                    key=field.key,
                    title=f"Select {field.label}",
                    initial_dir=existing_initial_dir(
                        state.command_root(command), values[field.key]
                    ),
                )
                if selected:
                    apply_selected_file(
                        state, command, values, field.key, field.label, selected
                    )
        return

    if field.kind == "date":
        if not current_value.strip() and field.default == "today":
            current_value = date.today().isoformat()
            values[field.key] = current_value
            state.persist_command_values(command)
        imgui.set_next_item_width(140)
        changed, value = imgui.input_text(field.label, current_value)
        if changed:
            values[field.key] = value
            state.persist_command_values(command)
        imgui.same_line()
        if imgui.button(f"Today##{command.id}_{field.key}", (70, 0)):
            values[field.key] = date.today().isoformat()
            state.persist_command_values(command)
        imgui.same_line()
        if imgui.button(f"Clear##{command.id}_{field.key}", (65, 0)):
            values[field.key] = ""
            state.persist_command_values(command)
        return

    if field.kind == "datetime":
        date_value, hour, minute = parse_datetime_parts(current_value)
        imgui.set_next_item_width(140)
        changed, new_date = imgui.input_text(field.label, date_value)
        if changed:
            values[field.key] = combine_datetime_value(new_date, hour, minute)
            state.persist_command_values(command)

        hours = [f"{item:02d}" for item in range(24)]
        minutes = [f"{item:02d}" for item in range(60)]

        imgui.same_line()
        imgui.set_next_item_width(70)
        changed, hour = imgui.combo(f"Hour##{command.id}_{field.key}", hour, hours)
        if changed:
            values[field.key] = combine_datetime_value(date_value, hour, minute)
            state.persist_command_values(command)

        imgui.same_line()
        imgui.set_next_item_width(70)
        changed, minute = imgui.combo(
            f"Min##{command.id}_{field.key}", minute, minutes
        )
        if changed:
            values[field.key] = combine_datetime_value(date_value, hour, minute)
            state.persist_command_values(command)

        imgui.same_line()
        if imgui.button(f"Today##{command.id}_{field.key}", (70, 0)):
            values[field.key] = combine_datetime_value(date.today().isoformat(), hour, minute)
            state.persist_command_values(command)
        imgui.same_line()
        if imgui.button(f"+15m##{command.id}_{field.key}", (65, 0)):
            target = datetime.now() + timedelta(minutes=15)
            values[field.key] = target.strftime("%Y-%m-%d %H:%M")
            state.persist_command_values(command)
        imgui.same_line()
        if imgui.button(f"Clear##{command.id}_{field.key}", (65, 0)):
            values[field.key] = ""
            state.persist_command_values(command)
        return

    imgui.set_next_item_width(260)
    changed, value = imgui.input_text(field.label, current_value)
    if changed:
        values[field.key] = value
        state.persist_command_values(command)

def draw_env_field(
    state: GuiState,
    command: CommandSpec,
    values: dict[str, str],
    key: str,
    label: str,
    kind: str,
) -> None:
    current_value = values.get(key, "")
    if kind == "bool":
        checked = current_value.strip().lower() in {"1", "true", "yes", "y", "on"}
        changed, checked = imgui.checkbox(label, checked)
        if changed:
            values[key] = "true" if checked else "false"
            state.propagate_env_value(key, values[key])
            state.persist_command_values(command)
        return

    imgui.set_next_item_width(520 if kind == "path" else 220)
    changed, value = imgui.input_text(label, current_value)
    if changed:
        values[key] = value
        state.propagate_env_value(key, value)
        state.persist_command_values(command)
    if kind == "path":
        consume_folder_dialog_result(state, command, values, key, label)

        imgui.same_line()
        dialog_is_open = key in state.folder_dialogs
        with disabled_when(dialog_is_open):
            if imgui.button(f"Select Folder##{key}", (130, 0)):
                selected = start_folder_dialog(
                    state=state,
                    key=key,
                    title=f"Select {label}",
                    initial_dir=existing_initial_dir(state.command_root(command), values[key]),
                )
                if selected:
                    apply_selected_folder(state, command, values, key, label, selected)
        imgui.same_line()
        if imgui.button(f"Create Folder##{key}", (120, 0)):
            path = resolve_gui_path(state.command_root(command), values[key])
            path.mkdir(parents=True, exist_ok=True)
            state.runner.append_log(f"Folder ready: {path}")
        imgui.same_line()
        if imgui.button(f"Open##{key}", (70, 0)):
            open_folder_from_gui(state, command, values[key])
