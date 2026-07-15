from __future__ import annotations

import csv
import hashlib
import json
import re
import shlex
import shutil
import sqlite3
import subprocess
import tempfile
import threading
import time
import urllib.request
import uuid
from collections import Counter
from contextlib import contextmanager
from datetime import date, datetime, timedelta
from pathlib import Path

from imgui_bundle import imgui
from openpyxl import load_workbook

from complaints_manager.core.commands import (
    ArgField,
    COMMANDS,
    SYSTEMS,
    CommandSpec,
    SystemSpec,
    command_by_id,
    commands_by_system,
)
from complaints_manager.core.services import normalize_whatsapp_account_id
from complaints_manager.gui.paths import (
    apply_selected_folder,
    apply_selected_file,
    consume_file_dialog_result,
    consume_folder_dialog_result,
    existing_initial_dir,
    resolve_gui_path,
    start_file_dialog,
    start_folder_dialog,
)
from complaints_manager.gui.state import GuiState


@contextmanager
def disabled_when(disabled: bool):
    if disabled:
        imgui.begin_disabled()
    try:
        yield
    finally:
        if disabled:
            imgui.end_disabled()


def system_by_id(system_id: str) -> SystemSpec:
    for system in SYSTEMS:
        if system.id == system_id:
            return system
    raise KeyError(system_id)


def command_validation_error(
    state: GuiState,
    command: CommandSpec,
    values: dict[str, str],
) -> str:
    if command.id in {"sms_campaign_create", "sms_campaign_update"}:
        if not values.get("campaign_id", "").strip():
            return "Enter a campaign name/reference." if command.id == "sms_campaign_create" else "Select an existing campaign."
        if not values.get("file", "").strip():
            return "Select an Excel file."
        if not values.get("column", "").strip():
            return "Select the column containing phone numbers."
        if not values.get("message", "").strip():
            return "Enter the SMS message."
        return ""
    if command.id in {"sms_campaign_show", "sms_campaign_failed", "sms_campaign_delete"}:
        if not values.get("campaign_id", "").strip():
            return "Select a campaign."
        return ""
    if command.id == "sms_campaign_send":
        has_campaign = bool(values.get("campaign_id", "").strip())
        if not has_campaign:
            if not values.get("file", "").strip():
                return "Select a campaign or Excel file."
            if not values.get("column", "").strip():
                return "Select the column containing phone numbers."
            if not values.get("message", "").strip():
                return "Enter the SMS message."
        try:
            minute = int(values.get("minute_limit", "0"))
            hour = int(values.get("hour_limit", "0"))
            day = int(values.get("day_limit", "0"))
            delay = float(values.get("delay_seconds", "0"))
            watch_timeout = int(values.get("watch_timeout_seconds", "1"))
            watch_poll = int(values.get("watch_poll_seconds", "30"))
        except ValueError:
            return "Rate limits and delay must be numbers."
        if min(minute, hour, day) <= 0 or delay < 0:
            return "Rate limits must be positive and delay cannot be negative."
        if watch_timeout <= 0 or watch_poll < 5:
            return "Watch timeout must be positive and watch polling interval must be at least 5 seconds."
        if values.get("retry_failed", "false").lower() in {"1", "true", "yes", "on"} and delay < 15:
            return "Failed-message retries require at least 15 seconds between recipients."
        if not minute <= hour <= day:
            return "Limits must satisfy minute <= hour <= day."
        if values.get("execute", "false").lower() in {"1", "true", "yes", "on"}:
            if values.get("consent_confirmed", "false").lower() not in {"1", "true", "yes", "on"}:
                return "Execution requires consent and suppression/DNCR confirmation."
            if not (state.command_root(command) / ".env").is_file():
                return "Create sms-campaign-manager/.env from .env.example and add gateway credentials."
        return ""
    if command.id in {"whatsapp_group_import", "whatsapp_group_membership_audit"}:
        file_value = values.get("file", "").strip()
        if not file_value:
            return "Select an Excel file."
        if command.id == "whatsapp_group_membership_audit":
            group_jid = values.get("group_jid", "").strip()
            if not group_jid.endswith("@g.us"):
                return "Select a WhatsApp group JID ending in @g.us."
            return ""
        if values.get("execute", "false").lower() in {"1", "true", "yes", "on"}:
            return (
                "Direct additions are disabled: WhatsApp invalidated linked sessions "
                "after participant-add operations. Use Audit Group Membership and an invite link."
            )
            group_jid = values.get("group_jid", "").strip()
            if not group_jid.endswith("@g.us"):
                return "Execute mode requires a WhatsApp group JID ending in @g.us."
            if values.get("consent_confirmed", "false").lower() not in {
                "1", "true", "yes", "on"
            }:
                return (
                    "Direct addition requires confirmation that recipients were informed "
                    "and consented. Use an office-distributed invite link otherwise."
                )
            try:
                delay_ms = int(values.get("delay_ms", "0"))
                wait_timeout = int(values.get("wait_timeout", "0"))
            except ValueError:
                return "Delay and completion timeout must be whole numbers."
            if delay_ms < 10_000:
                return "Execute mode requires at least 10 seconds between additions."
            if wait_timeout <= 0:
                return "Execute mode requires a positive completion timeout."
        if values.get("selection") == "range" and (
            not values.get("start", "").strip() or not values.get("end", "").strip()
        ):
            return "Range selection requires both start and end positions."
        return ""
    if command.id != "crm_filter_pdf_duplicates":
        return ""
    command_root = state.command_root(command)
    input_dir = resolve_gui_path(command_root, values.get("input_dir", ""))
    output_dir = resolve_gui_path(command_root, values.get("output_dir", ""))
    try:
        input_dir.relative_to(output_dir)
    except ValueError:
        return ""
    return (
        "Input folder is inside the filtered output folder. "
        "Select the raw PDF folder, usually phase1-crm/unprocessed-crm/pdf."
    )


def open_folder_from_gui(state: GuiState, command: CommandSpec, value: str) -> None:
    path = resolve_gui_path(state.command_root(command), value)
    path.mkdir(parents=True, exist_ok=True)
    opener = shutil.which("xdg-open")
    if not opener:
        state.runner.append_log(f"Cannot open folder; xdg-open not found: {path}")
        return
    subprocess.Popen(
        [opener, str(path)],
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        start_new_session=True,
    )
    state.runner.append_log(f"Opened folder: {path}")


def open_external_file(
    state: GuiState, path: Path, label: str, service_id: str = "whatsapp"
) -> None:
    whatsapp_service = state.services.services.get(service_id)
    if not path.is_file():
        if whatsapp_service:
            whatsapp_service.append_log(f"{label} is not available yet: {path}")
        return
    opener = shutil.which("xdg-open")
    if not opener:
        if whatsapp_service:
            whatsapp_service.append_log(
                f"Cannot open {label}; xdg-open was not found: {path}"
            )
        return
    subprocess.Popen(
        [opener, str(path)],
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        start_new_session=True,
    )


def reset_whatsapp_authentication(state: GuiState, service_id: str) -> None:
    service = state.services.services[service_id]
    if service.is_running:
        service.append_log("Stop the WhatsApp Worker before resetting its login.")
        return
    service_env = dict(service.spec.env)
    auth_dir = service.spec.cwd / service_env.get("WA_AUTH_DIR", "auth_info_baileys")
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    backup_dir = auth_dir.with_name(f"{auth_dir.name}.manual-backup-{timestamp}")
    try:
        if auth_dir.exists():
            auth_dir.rename(backup_dir)
        auth_dir.mkdir(parents=True, exist_ok=True)
        qr_path = service.spec.cwd / service_env.get(
            "WA_QR_IMAGE_PATH", "data/whatsapp-login-qr.png"
        )
        qr_path.unlink(missing_ok=True)
    except OSError as exc:
        service.append_log(f"Could not reset WhatsApp login: {exc}")
        service.status_message = "Login reset failed"
        return
    service.append_log(f"Previous WhatsApp login archived safely: {backup_dir}")
    service.append_log("Starting a clean login; wait for the Open QR Code button.")
    state.services.start(service_id)


def count_path_items(path: Path) -> int:
    if path.is_dir() and not path.is_symlink():
        return 1 + sum(1 for _item in path.rglob("*"))
    return 1


def clear_duplicate_filter_results(
    state: GuiState,
    command: CommandSpec,
    values: dict[str, str],
) -> None:
    output_value = values.get("output_dir", "").strip()
    if not output_value:
        state.runner.append_log(f"Cannot clear {command.title} results; output folder is empty.")
        return

    command_root = state.command_root(command)
    output_dir = resolve_gui_path(command_root, output_value)
    protected_paths = {
        Path("/").resolve(strict=False),
        Path.home().resolve(strict=False),
        command_root.resolve(strict=False),
        state.project_root.resolve(strict=False),
    }
    if output_dir in protected_paths:
        state.runner.append_log(f"Refusing to clear protected folder: {output_dir}")
        return

    removed_count = 0
    try:
        output_dir.mkdir(parents=True, exist_ok=True)
        for child in output_dir.iterdir():
            removed_count += count_path_items(child)
            if child.is_dir() and not child.is_symlink():
                shutil.rmtree(child)
            else:
                child.unlink()
    except OSError as exc:
        state.runner.append_log(f"Failed to clear results folder {output_dir}: {exc}")
        state.runner.status_message = f"Failed: Clear {command.title} results"
        return

    state.runner.append_log(
        f"Cleared results folder for {command.title}: {output_dir} ({removed_count} item(s) removed)"
    )
    state.runner.status_message = f"Finished: Clear {command.title} results"


def copy_excel_filter_values_to_notice_command(
    state: GuiState,
    filter_values: dict[str, str],
    notice_command: CommandSpec,
    status: str,
) -> dict[str, str]:
    notice_values = state.command_values(notice_command)
    notice_values["input_file"] = filter_values.get("input_file", "")
    notice_values["output_dir"] = filter_values.get(
        "output_dir", "phase1-crm/unprocessed-crm/filtered"
    )
    notice_values["status"] = status
    notice_values.setdefault("dispatch_dir", "")
    if notice_command.id == "crm_send_duplicate_notices":
        notice_values.setdefault("dispatch_csv", "")
        notice_values.setdefault("dry_run", "false")
    state.persist_command_values(notice_command)
    return notice_values


def status_color(state: GuiState) -> tuple[float, float, float, float]:
    if state.runner.is_running:
        return (0.28, 0.64, 0.96, 1.0)
    if state.runner.last_exit_code == 0:
        return (0.26, 0.72, 0.40, 1.0)
    if state.runner.last_exit_code is not None:
        return (0.95, 0.35, 0.28, 1.0)
    return (0.82, 0.82, 0.82, 1.0)


def service_status_color(is_running: bool, exit_code: int | None) -> tuple[float, float, float, float]:
    if is_running:
        return (0.26, 0.72, 0.40, 1.0)
    if exit_code not in (None, 0):
        return (0.95, 0.35, 0.28, 1.0)
    return (0.78, 0.78, 0.78, 1.0)


def sms_campaign_status_color(status: str) -> tuple[float, float, float, float]:
    status = status.strip().lower()
    if status == "verified":
        return (0.26, 0.72, 0.40, 1.0)
    if status in {"running", "submitted"}:
        return (0.28, 0.64, 0.96, 1.0)
    if status == "needs_attention":
        return (0.95, 0.58, 0.22, 1.0)
    if status == "deleted":
        return (0.65, 0.65, 0.65, 1.0)
    return (0.82, 0.82, 0.82, 1.0)


def apply_sms_campaign_to_command(
    state: GuiState,
    command_id: str,
    campaign: dict[str, str],
) -> dict[str, str]:
    command = command_by_id(command_id)
    values = state.command_values(command)
    values["campaign_id"] = campaign.get("campaign_name", "")
    if command_id in {"sms_campaign_update", "sms_campaign_send"}:
        values["file"] = campaign.get("source_file", values.get("file", ""))
        values["sheet"] = campaign.get("sheet", values.get("sheet", ""))
        values["column"] = campaign.get("number_column", values.get("column", ""))
        values["message"] = campaign.get("message", values.get("message", ""))
        values["description"] = campaign.get("description", values.get("description", ""))
    state.persist_command_values(command)
    return values


def format_elapsed(seconds: int) -> str:
    return f"{seconds // 3600:02d}:{(seconds % 3600) // 60:02d}:{seconds % 60:02d}"


def parse_datetime_parts(value: str) -> tuple[str, int, int]:
    now = datetime.now()
    cleaned = value.strip().replace("T", " ")
    if cleaned:
        try:
            parsed = datetime.strptime(cleaned, "%Y-%m-%d %H:%M")
            return parsed.date().isoformat(), parsed.hour, parsed.minute
        except ValueError:
            pass
    return "", now.hour, now.minute


def combine_datetime_value(date_value: str, hour: int, minute: int) -> str:
    date_value = date_value.strip()
    if not date_value:
        return ""
    return f"{date_value} {hour:02d}:{minute:02d}"


def draw_top_bar(state: GuiState) -> None:
    imgui.text_colored((0.92, 0.92, 0.92, 1.0), "Automation Management Console")
    imgui.same_line()
    imgui.text_disabled(f"GUI: {state.project_root}")
    imgui.separator()


def draw_metric_row(state: GuiState, system_id: str) -> None:
    metrics = state.metrics(system_id) or []
    if not metrics:
        imgui.text_disabled("No metrics available.")
        return
    if imgui.begin_table(f"{system_id}_metrics", max(1, len(metrics))):
        for metric in metrics:
            imgui.table_setup_column(metric.label)
        imgui.table_headers_row()
        imgui.table_next_row()
        for metric in metrics:
            imgui.table_next_column()
            imgui.text(str(metric.value))
        imgui.end_table()


def draw_services_panel(state: GuiState) -> None:
    imgui.separator_text("Infrastructure Services")
    imgui.text_wrapped(
        "Manage NATS, WhatsApp workers, and PocketBase here. WhatsApp account creation and QR login are handled from WhatsApp Accounts."
    )
    if imgui.begin_table("service_dashboard", 6):
        imgui.table_setup_column("Service")
        imgui.table_setup_column("Type")
        imgui.table_setup_column("Status")
        imgui.table_setup_column("Uptime")
        imgui.table_setup_column("Controls")
        imgui.table_setup_column("Logs")
        imgui.table_headers_row()

        for service in state.services.services.values():
            imgui.push_id(service.spec.id)
            imgui.table_next_row()

            imgui.table_next_column()
            imgui.text(service.spec.title)
            imgui.text_disabled(service.spec.summary)

            imgui.table_next_column()
            if service.spec.kind == "whatsapp":
                imgui.text_colored((0.18, 0.78, 0.48, 1.0), "WhatsApp")
                imgui.text_disabled(service.spec.profile_id or "default")
            elif service.spec.id == "nats":
                imgui.text_colored((0.72, 0.62, 0.95, 1.0), "NATS")
            elif service.spec.kind == "database":
                imgui.text_colored((0.32, 0.70, 0.92, 1.0), "Database")
            else:
                imgui.text_disabled(service.spec.kind)

            imgui.table_next_column()
            imgui.text_colored(
                service_status_color(service.is_running, service.last_exit_code),
                service.status_message,
            )

            imgui.table_next_column()
            imgui.text(format_elapsed(service.elapsed_seconds()) if service.is_running else "-")

            imgui.table_next_column()
            with disabled_when(service.is_running):
                if imgui.button("Start", (72, 0)):
                    state.services.start(service.spec.id)
            imgui.same_line()
            with disabled_when(not service.is_running):
                if imgui.button("Stop", (72, 0)):
                    state.services.stop(service.spec.id)
            imgui.same_line()
            if imgui.button("Restart", (82, 0)):
                state.services.restart(service.spec.id)

            imgui.table_next_column()
            if imgui.button("Details", (88, 0)):
                state.current_service_id = service.spec.id

            imgui.pop_id()

        imgui.end_table()

    imgui.text_disabled("Tip: open Details for full logs, WhatsApp QR state, auth folder, and NATS subjects.")


def draw_command_list(state: GuiState, system_id: str) -> None:
    grouped: dict[str, list[CommandSpec]] = {}
    for command in commands_by_system()[system_id]:
        grouped.setdefault(command.group, []).append(command)

    for group, commands in grouped.items():
        imgui.separator_text(group)
        for command in commands:
            selected = state.selected_command_id == command.id
            if imgui.selectable(command.title, selected)[0]:
                state.selected_command_id = command.id


def draw_field(
    state: GuiState,
    command: CommandSpec,
    values: dict[str, str],
    field: ArgField,
) -> None:
    if field.kind.startswith("hidden_env:"):
        return

    current_value = values.get(field.key, field.default)

    if field.kind == "sms_campaign":
        draw_sms_campaign_selector(state, command, values, field)
        return

    if command.id in {"whatsapp_group_import", "whatsapp_group_membership_audit"} and field.key == "group_jid":
        draw_whatsapp_group_selector(state, command, values, field)
        return
    if command.id == "whatsapp_group_membership_audit" and field.key == "worker_id":
        draw_whatsapp_account_selector(state, command, values, field)
        return
    if field.kind == "bool_flag":
        checked = current_value.strip().lower() in {"1", "true", "yes", "y", "on"}
        changed, checked = imgui.checkbox(field.label, checked)
        if changed:
            values[field.key] = "true" if checked else "false"
            state.persist_command_values(command)
        return

    if field.kind == "select":
        labels = [option.label for option in field.options]
        option_values = [option.value for option in field.options]
        try:
            index = option_values.index(current_value)
        except ValueError:
            index = 0
        changed, index = imgui.combo(field.label, index, labels)
        if changed:
            values[field.key] = option_values[index]
            state.persist_command_values(command)
        return

    if field.kind == "excel_column":
        file_value = values.get("file", "").strip()
        sheet_value = values.get("sheet", "").strip()
        columns: list[str] = []
        if file_value:
            path = resolve_gui_path(state.command_root(command), file_value)
            try:
                workbook = load_workbook(path, read_only=True, data_only=True)
                selected_sheet = sheet_value if sheet_value in workbook.sheetnames else workbook.sheetnames[0]
                header = next(workbook[selected_sheet].iter_rows(min_row=1, max_row=1, values_only=True), ())
                columns = [str(item).strip() for item in header if item is not None and str(item).strip()]
                workbook.close()
            except (OSError, ValueError, KeyError):
                columns = []
        if columns:
            if current_value not in columns:
                normalized = {name.lower().replace(" ", "_"): name for name in columns}
                values[field.key] = normalized.get("contact_no", columns[0])
                current_value = values[field.key]
            index = columns.index(current_value)
            imgui.set_next_item_width(420)
            changed, index = imgui.combo(field.label, index, columns)
            if changed:
                values[field.key] = columns[index]
                state.persist_command_values(command)
        else:
            imgui.text_disabled("Select a readable Excel file to load number columns.")
        return

    if field.kind == "excel_sheet":
        file_value = values.get("file", "").strip()
        sheets: list[str] = []
        if file_value:
            try:
                workbook = load_workbook(resolve_gui_path(state.command_root(command), file_value), read_only=True, data_only=True)
                sheets = list(workbook.sheetnames)
                workbook.close()
            except (OSError, ValueError, KeyError):
                sheets = []
        if sheets:
            selected = current_value if current_value in sheets else sheets[0]
            if current_value != selected:
                values[field.key] = selected
            index = sheets.index(selected)
            imgui.set_next_item_width(420)
            changed, index = imgui.combo(field.label, index, sheets)
            if changed:
                values[field.key] = sheets[index]
                values["column"] = ""
                state.persist_command_values(command)
        else:
            imgui.text_disabled("Select a readable Excel file to load sheets.")
        return

    if field.kind == "multiline":
        imgui.set_next_item_width(720)
        changed, value = imgui.input_text_multiline(field.label, current_value, (720, 110))
        if changed:
            values[field.key] = value
        return

    if field.kind in {"path", "file"}:
        imgui.set_next_item_width(520)
        changed, value = imgui.input_text(field.label, current_value)
        if changed:
            values[field.key] = value
            state.persist_command_values(command)

        if field.kind == "path":
            consume_folder_dialog_result(state, command, values, field.key, field.label)
            imgui.same_line()
            dialog_is_open = field.key in state.folder_dialogs
            with disabled_when(dialog_is_open):
                if imgui.button(f"Select Folder##{command.id}_{field.key}", (130, 0)):
                    selected = start_folder_dialog(
                        state=state,
                        key=field.key,
                        title=f"Select {field.label}",
                        initial_dir=existing_initial_dir(
                            state.command_root(command), values[field.key]
                        ),
                    )
                    if selected:
                        apply_selected_folder(
                            state, command, values, field.key, field.label, selected
                        )
            imgui.same_line()
            if imgui.button(f"Open##{command.id}_{field.key}", (70, 0)):
                open_folder_from_gui(state, command, values[field.key])
            return

        consume_file_dialog_result(state, command, values, field.key, field.label)
        imgui.same_line()
        dialog_is_open = field.key in state.file_dialogs
        with disabled_when(dialog_is_open):
            if imgui.button(f"Select File##{command.id}_{field.key}", (115, 0)):
                selected = start_file_dialog(
                    state=state,
                    key=field.key,
                    title=f"Select {field.label}",
                    initial_dir=existing_initial_dir(
                        state.command_root(command), values[field.key]
                    ),
                )
                if selected:
                    apply_selected_file(
                        state, command, values, field.key, field.label, selected
                    )
        return

    if field.kind == "date":
        if not current_value.strip() and field.default == "today":
            current_value = date.today().isoformat()
            values[field.key] = current_value
            state.persist_command_values(command)
        imgui.set_next_item_width(140)
        changed, value = imgui.input_text(field.label, current_value)
        if changed:
            values[field.key] = value
            state.persist_command_values(command)
        imgui.same_line()
        if imgui.button(f"Today##{command.id}_{field.key}", (70, 0)):
            values[field.key] = date.today().isoformat()
            state.persist_command_values(command)
        imgui.same_line()
        if imgui.button(f"Clear##{command.id}_{field.key}", (65, 0)):
            values[field.key] = ""
            state.persist_command_values(command)
        return

    if field.kind == "datetime":
        date_value, hour, minute = parse_datetime_parts(current_value)
        imgui.set_next_item_width(140)
        changed, new_date = imgui.input_text(field.label, date_value)
        if changed:
            values[field.key] = combine_datetime_value(new_date, hour, minute)
            state.persist_command_values(command)

        hours = [f"{item:02d}" for item in range(24)]
        minutes = [f"{item:02d}" for item in range(60)]

        imgui.same_line()
        imgui.set_next_item_width(70)
        changed, hour = imgui.combo(f"Hour##{command.id}_{field.key}", hour, hours)
        if changed:
            values[field.key] = combine_datetime_value(date_value, hour, minute)
            state.persist_command_values(command)

        imgui.same_line()
        imgui.set_next_item_width(70)
        changed, minute = imgui.combo(
            f"Min##{command.id}_{field.key}", minute, minutes
        )
        if changed:
            values[field.key] = combine_datetime_value(date_value, hour, minute)
            state.persist_command_values(command)

        imgui.same_line()
        if imgui.button(f"Today##{command.id}_{field.key}", (70, 0)):
            values[field.key] = combine_datetime_value(date.today().isoformat(), hour, minute)
            state.persist_command_values(command)
        imgui.same_line()
        if imgui.button(f"+15m##{command.id}_{field.key}", (65, 0)):
            target = datetime.now() + timedelta(minutes=15)
            values[field.key] = target.strftime("%Y-%m-%d %H:%M")
            state.persist_command_values(command)
        imgui.same_line()
        if imgui.button(f"Clear##{command.id}_{field.key}", (65, 0)):
            values[field.key] = ""
            state.persist_command_values(command)
        return

    imgui.set_next_item_width(260)
    changed, value = imgui.input_text(field.label, current_value)
    if changed:
        values[field.key] = value
        state.persist_command_values(command)


def discovered_whatsapp_groups(state: GuiState) -> list[tuple[str, str, str]]:
    data_dir = state.project_root.parent / "whatsappbot" / "data"
    registries = sorted(data_dir.glob("discovered-groups*.csv"))
    rows: list[dict] = []
    for registry in registries:
        try:
            with registry.open("r", encoding="utf-8-sig", newline="") as handle:
                rows.extend(csv.DictReader(handle))
        except (OSError, csv.Error) as exc:
            state.runner.append_log(f"Could not read WhatsApp group registry {registry}: {exc}")

    groups = [
        (
            str(row.get("group_id", "")).strip(),
            str(row.get("group_name", "")).strip(),
            str(row.get("discovered_at", "")).strip(),
        )
        for row in rows
        if str(row.get("group_id", "")).strip().endswith("@g.us")
    ]
    groups.sort(key=lambda item: item[2], reverse=True)
    deduplicated: list[tuple[str, str, str]] = []
    seen: set[str] = set()
    for group in groups:
        if group[0] in seen:
            continue
        seen.add(group[0])
        deduplicated.append(group)
    return deduplicated


def draw_whatsapp_group_selector(
    state: GuiState,
    command: CommandSpec,
    values: dict[str, str],
    field: ArgField,
) -> None:
    groups = discovered_whatsapp_groups(state)
    current = values.get(field.key, "").strip()
    ids = [""] + [group_id for group_id, _name, _when in groups]
    labels = ["Select a discovered group..."] + [
        f"{name or '(unnamed group)'}  |  {group_id}"
        for group_id, name, _when in groups
    ]
    try:
        index = ids.index(current)
    except ValueError:
        index = 0

    imgui.set_next_item_width(620)
    changed, index = imgui.combo(field.label, index, labels)
    if changed and index > 0:
        values[field.key] = ids[index]
        state.persist_command_values(command)

    imgui.same_line()
    with disabled_when(not groups):
        if imgui.button(f"Use Latest##{command.id}_{field.key}", (105, 0)):
            values[field.key] = groups[0][0]
            state.persist_command_values(command)
            state.runner.append_log(
                f"Selected latest WhatsApp group: {groups[0][1] or groups[0][0]} ({groups[0][0]})"
            )

    imgui.same_line()
    if imgui.button(f"Refresh##{command.id}_{field.key}", (85, 0)):
        state.runner.append_log(f"Refreshed WhatsApp groups: {len(groups)} discovered")

    imgui.set_next_item_width(620)
    changed, manual_value = imgui.input_text("Manual group JID", current)
    if changed:
        values[field.key] = manual_value
        state.persist_command_values(command)

    if groups:
        latest_id, latest_name, latest_when = groups[0]
        imgui.text_disabled(
            f"Latest: {latest_name or '(unnamed group)'} | {latest_id} | {latest_when}"
        )
    else:
        imgui.text_disabled(
            "No groups discovered yet. Start or restart the WhatsApp Worker, then refresh."
        )


def draw_whatsapp_account_selector(
    state: GuiState,
    command: CommandSpec,
    values: dict[str, str],
    field: ArgField,
) -> None:
    profiles = [
        (service.spec.profile_id, service.spec.title)
        for service in state.services.services.values()
        if service.spec.kind == "whatsapp"
    ]
    if not profiles:
        profiles = [("default", "Primary WhatsApp")]
    ids = [item[0] for item in profiles]
    labels = [item[1] for item in profiles]
    current = values.get(field.key, field.default)
    try:
        index = ids.index(current)
    except ValueError:
        index = 0
    imgui.set_next_item_width(360)
    changed, index = imgui.combo(field.label, index, labels)
    if changed:
        profile_id = ids[index]
        values[field.key] = profile_id
        values["health_subject"] = (
            "whatsapp.worker.health"
            if profile_id == "default"
            else f"whatsapp.worker.{profile_id}.health"
        )
        values["members_subject"] = (
            "whatsapp.worker.group-members"
            if profile_id == "default"
            else f"whatsapp.worker.{profile_id}.group-members"
        )
        state.persist_command_values(command)


def sms_campaign_database_path(command_root: Path) -> Path:
    env_path = command_root / ".env"
    database = "data/sms-campaigns.sqlite3"
    try:
        for line in env_path.read_text(encoding="utf-8").splitlines():
            stripped = line.strip()
            if stripped.startswith("SMS_GATE_DATABASE="):
                database = stripped.split("=", 1)[1].strip().strip("\"'")
                break
    except OSError:
        pass
    path = Path(database).expanduser()
    if not path.is_absolute():
        path = command_root / path
    return path


def discovered_sms_campaigns(state: GuiState, command: CommandSpec, *, force: bool = False) -> list[dict[str, str]]:
    database = sms_campaign_database_path(state.command_root(command))
    cache_key = str(database)
    now = time.monotonic()
    if (
        not force
        and state.sms_campaign_cache_path == cache_key
        and now - state.sms_campaign_cache_at < 5
    ):
        return state.sms_campaign_cache
    if not database.is_file():
        state.sms_campaign_cache = []
        state.sms_campaign_cache_path = cache_key
        state.sms_campaign_cache_at = now
        return []
    try:
        uri = f"file:{database}?mode=ro"
        with sqlite3.connect(uri, uri=True, timeout=2) as connection:
            connection.row_factory = sqlite3.Row
            rows = connection.execute(
                """SELECT c.campaign_name,c.source_file,c.sheet,c.number_column,c.message,c.description,c.status,
                          COALESCE(c.updated_at,c.created_at) AS updated_at,
                          COUNT(r.id) AS recipients,
                          SUM(CASE WHEN r.status='pending' THEN 1 ELSE 0 END) AS pending,
                          SUM(CASE WHEN r.status='sent' THEN 1 ELSE 0 END) AS sent,
                          SUM(CASE WHEN r.status='delivered' THEN 1 ELSE 0 END) AS delivered,
                          SUM(CASE WHEN r.status='failed' THEN 1 ELSE 0 END) AS failed,
                          SUM(CASE WHEN r.status='unknown' THEN 1 ELSE 0 END) AS unknown
                   FROM campaigns c LEFT JOIN recipients r ON r.campaign_id=c.id
                   WHERE c.campaign_name IS NOT NULL AND c.campaign_name<>'' AND c.status<>'deleted'
                   GROUP BY c.id
                   ORDER BY COALESCE(c.updated_at,c.created_at) DESC,c.id DESC"""
            ).fetchall()
    except sqlite3.Error as exc:
        try:
            output = subprocess.check_output(
                [shutil.which("uv") or "uv", "run", "sms-campaign", "campaigns", "list", "--json"],
                cwd=state.command_root(command),
                text=True,
                stderr=subprocess.DEVNULL,
                timeout=5,
            )
            payload = json.loads(output)
            if isinstance(payload, list):
                campaigns = [
                    {str(key): "" if value is None else str(value) for key, value in item.items()}
                    for item in payload
                    if isinstance(item, dict)
                ]
                state.sms_campaign_cache = campaigns
                state.sms_campaign_cache_path = cache_key
                state.sms_campaign_cache_at = now
                state.sms_campaign_last_error = ""
                return campaigns
        except (OSError, subprocess.SubprocessError, json.JSONDecodeError):
            pass
        message = f"Could not read SMS campaign registry {database}: {exc}"
        if state.sms_campaign_last_error != message:
            state.runner.append_log(message)
            state.sms_campaign_last_error = message
        state.sms_campaign_cache_path = cache_key
        state.sms_campaign_cache_at = now
        return state.sms_campaign_cache if state.sms_campaign_cache_path == cache_key else []
    campaigns = [{key: "" if row[key] is None else str(row[key]) for key in row.keys()} for row in rows]
    state.sms_campaign_cache = campaigns
    state.sms_campaign_cache_path = cache_key
    state.sms_campaign_cache_at = now
    state.sms_campaign_last_error = ""
    return campaigns


def draw_sms_campaign_selector(
    state: GuiState,
    command: CommandSpec,
    values: dict[str, str],
    field: ArgField,
) -> None:
    campaigns = discovered_sms_campaigns(state, command)
    current = values.get(field.key, "").strip()
    ids = [""] + [item["campaign_name"] for item in campaigns]
    labels = ["Select saved campaign..."] + [
        (
            f"{item['campaign_name']}  |  {item.get('status','-')}  |  recipients={item.get('recipients','0')} "
            f"pending={item.get('pending','0')} sent={item.get('sent','0')} "
            f"delivered={item.get('delivered','0')} failed={item.get('failed','0')}"
        )
        for item in campaigns
    ]
    try:
        index = ids.index(current)
    except ValueError:
        index = 0

    imgui.set_next_item_width(760)
    changed, index = imgui.combo(field.label, index, labels)
    if changed:
        values[field.key] = ids[index]
        if index > 0 and command.id in {"sms_campaign_send", "sms_campaign_update"}:
            selected = campaigns[index - 1]
            values["file"] = selected.get("source_file", values.get("file", ""))
            values["sheet"] = selected.get("sheet", values.get("sheet", ""))
            values["column"] = selected.get("number_column", values.get("column", ""))
            values["message"] = selected.get("message", values.get("message", ""))
            values["description"] = selected.get("description", values.get("description", ""))
        state.persist_command_values(command)

    imgui.same_line()
    if imgui.button(f"Refresh##{command.id}_{field.key}", (85, 0)):
        campaigns = discovered_sms_campaigns(state, command, force=True)
        state.runner.append_log(f"Refreshed SMS campaigns: {len(campaigns)} saved")

    imgui.set_next_item_width(420)
    changed, manual = imgui.input_text(f"Manual campaign name##{command.id}_{field.key}", current)
    if changed:
        values[field.key] = manual
        state.persist_command_values(command)

    if campaigns:
        latest = campaigns[0]
        imgui.text_disabled(
            f"Latest: {latest['campaign_name']} | {latest.get('status','-')} | recipients={latest.get('recipients','0')} | updated={latest.get('updated_at','')}"
        )
    else:
        imgui.text_disabled("No saved SMS campaigns yet. Use Create SMS Campaign first.")


def draw_env_field(
    state: GuiState,
    command: CommandSpec,
    values: dict[str, str],
    key: str,
    label: str,
    kind: str,
) -> None:
    current_value = values.get(key, "")
    if kind == "bool":
        checked = current_value.strip().lower() in {"1", "true", "yes", "y", "on"}
        changed, checked = imgui.checkbox(label, checked)
        if changed:
            values[key] = "true" if checked else "false"
            state.propagate_env_value(key, values[key])
            state.persist_command_values(command)
        return

    imgui.set_next_item_width(520 if kind == "path" else 220)
    changed, value = imgui.input_text(label, current_value)
    if changed:
        values[key] = value
        state.propagate_env_value(key, value)
        state.persist_command_values(command)
    if kind == "path":
        consume_folder_dialog_result(state, command, values, key, label)

        imgui.same_line()
        dialog_is_open = key in state.folder_dialogs
        with disabled_when(dialog_is_open):
            if imgui.button(f"Select Folder##{key}", (130, 0)):
                selected = start_folder_dialog(
                    state=state,
                    key=key,
                    title=f"Select {label}",
                    initial_dir=existing_initial_dir(state.command_root(command), values[key]),
                )
                if selected:
                    apply_selected_folder(state, command, values, key, label, selected)
        imgui.same_line()
        if imgui.button(f"Create Folder##{key}", (120, 0)):
            path = resolve_gui_path(state.command_root(command), values[key])
            path.mkdir(parents=True, exist_ok=True)
            state.runner.append_log(f"Folder ready: {path}")
        imgui.same_line()
        if imgui.button(f"Open##{key}", (70, 0)):
            open_folder_from_gui(state, command, values[key])


def start_command_with_services(
    state: GuiState,
    command: CommandSpec,
    values: dict[str, str],
    *,
    source: str = "",
) -> bool:
    if state.runner.is_running or state.command_start_in_progress:
        state.runner.append_log(
            f"Cannot start {command.title}; another command is already running."
        )
        return False
    validation_error = command_validation_error(state, command, values)
    if validation_error:
        state.runner.append_log(f"Cannot start {command.title}: {validation_error}")
        state.runner.status_message = f"Failed: {command.title} (validation)"
        return False

    state.persist_command_values(command)
    state.command_start_in_progress = True
    state.runner.status_message = f"Starting: {command.title}"
    state.runner.append_log(f"Preparing {command.title}...")
    values_snapshot = dict(values)
    thread = threading.Thread(
        target=_start_command_with_services_worker,
        args=(state, command, values_snapshot, source),
        daemon=True,
    )
    thread.start()
    return True


def _start_command_with_services_worker(
    state: GuiState,
    command: CommandSpec,
    values: dict[str, str],
    source: str,
) -> None:
    started_services = False
    try:
        for service_id in command.required_services:
            service = state.services.services.get(service_id)
            if service is None:
                continue
            ready, detail = state.services.readiness(service_id)
            if not ready and not service.is_running:
                state.runner.append_log(
                    f"Starting required service for {command.title}: {service.spec.title}"
                )
                state.services.start(service_id)
                started_services = True
            timeout_seconds = 120.0 if service.spec.kind == "whatsapp" else 15.0
            state.runner.append_log(
                f"Waiting for required service: {service.spec.title}"
            )
            ready, detail = state.services.wait_until_ready(
                service_id,
                timeout_seconds=timeout_seconds,
            )
            if not ready:
                state.runner.append_log(
                    f"Cannot start {command.title}; required service is not ready: "
                    f"{service.spec.title} ({detail})"
                )
                state.runner.status_message = f"Failed: {command.title} (service not ready)"
                return

        if started_services:
            state.services.drain_output()

        if source:
            state.runner.append_log(source)
        if command.id in {"sms_campaign_create", "sms_campaign_update", "sms_campaign_delete"}:
            state.sms_campaign_cache_at = 0.0
            state.sms_campaign_cache = []
        state.runner.start(command, values)
    except Exception as exc:
        state.runner.status_message = f"Failed: {command.title} (startup error)"
        state.runner.append_log(f"Cannot start {command.title}: {exc}")
    finally:
        state.command_start_in_progress = False


WEEKDAYS = ("Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun")
PORTAL_LOGIN_MODES = (
    ("manual", "Manual review"),
    ("auto", "Auto click Sign In"),
    ("remote_approve", "Remote approval"),
)


def parse_schedule_times(value: str) -> list[str]:
    times: list[str] = []
    for raw in value.replace(";", ",").split(","):
        item = raw.strip()
        if not item:
            continue
        try:
            parsed = datetime.strptime(item, "%H:%M")
        except ValueError:
            continue
        formatted = parsed.strftime("%H:%M")
        if formatted not in times:
            times.append(formatted)
    return sorted(times)


def parse_weekdays(value: str) -> set[int]:
    result = set()
    for raw in value.split(","):
        try:
            day = int(raw.strip())
        except ValueError:
            continue
        if 0 <= day <= 6:
            result.add(day)
    return result


def format_weekdays(value: str) -> str:
    days = parse_weekdays(value)
    if not days:
        return "-"
    return ", ".join(WEEKDAYS[day] for day in sorted(days))


def schedule_matches_now(schedule: dict, now: datetime) -> tuple[bool, str]:
    if not schedule.get("enabled", True):
        return False, ""

    current_time = now.strftime("%H:%M")
    times = parse_schedule_times(str(schedule.get("times", "")))
    if current_time not in times:
        return False, ""

    mode = str(schedule.get("mode", "daily"))
    if mode == "once" and str(schedule.get("date", "")) != now.date().isoformat():
        return False, ""
    if mode == "weekly" and now.weekday() not in parse_weekdays(
        str(schedule.get("weekdays", ""))
    ):
        return False, ""

    return True, f"{now.date().isoformat()} {current_time}"


def next_schedule_run(schedule: dict, now: datetime) -> str:
    times = parse_schedule_times(str(schedule.get("times", "")))
    if not times:
        return "No valid times"

    mode = str(schedule.get("mode", "daily"))
    for day_offset in range(0, 15):
        candidate_date = now.date() + timedelta(days=day_offset)
        if mode == "once" and str(schedule.get("date", "")) != candidate_date.isoformat():
            continue
        if mode == "weekly" and candidate_date.weekday() not in parse_weekdays(
            str(schedule.get("weekdays", ""))
        ):
            continue
        for item in times:
            hour, minute = [int(part) for part in item.split(":", 1)]
            candidate = datetime.combine(
                candidate_date, datetime.min.time()
            ).replace(hour=hour, minute=minute)
            if candidate >= now:
                return candidate.strftime("%Y-%m-%d %H:%M")
    return "-"


def process_antidengue_schedules(state: GuiState) -> None:
    if state.runner.is_running:
        return

    now = datetime.now().replace(second=0, microsecond=0)
    command = command_by_id("antidengue_run")
    for schedule in state.antidengue_schedules():
        due, run_key = schedule_matches_now(schedule, now)
        if not due or str(schedule.get("last_run_key", "")) == run_key:
            continue
        schedule["last_run_key"] = run_key
        if schedule.get("mode") == "once":
            schedule["enabled"] = False
        state.update_antidengue_schedule(schedule)
        start_command_with_services(
            state,
            command,
            state.command_values(command),
            source=f"Scheduled AntiDengue run due: {run_key}",
        )
        break


def antidengue_root(state: GuiState) -> Path:
    root = state.project_root.resolve(strict=False)
    for candidate in [root, *root.parents][:6]:
        if (candidate / "antidengue").exists():
            return (candidate / "antidengue").resolve(strict=False)
    return (root.parent / "antidengue").resolve(strict=False)


def antidengue_pocketbase_root(state: GuiState) -> Path:
    root = state.project_root.resolve(strict=False)
    for candidate in [root, *root.parents][:6]:
        if (candidate / "antidengue-pocketbase").exists():
            return (candidate / "antidengue-pocketbase").resolve(strict=False)
    return (root.parent / "antidengue-pocketbase").resolve(strict=False)


def whatsappbot_root(state: GuiState) -> Path:
    root = state.project_root.resolve(strict=False)
    for candidate in [root, *root.parents][:6]:
        if (candidate / "whatsappbot").exists():
            return (candidate / "whatsappbot").resolve(strict=False)
    return (root.parent / "whatsappbot").resolve(strict=False)


def whatsapp_auth_dir(state: GuiState) -> Path:
    return (whatsappbot_root(state) / "auth_info_baileys").resolve(strict=False)


def latest_antidengue_summary(state: GuiState) -> tuple[Path | None, dict, str]:
    output_root = antidengue_root(state) / "output-files"
    summaries = sorted(
        output_root.glob("*/run_summary.json"),
        key=lambda path: path.stat().st_mtime if path.exists() else 0,
        reverse=True,
    )
    if not summaries:
        return None, {}, "No AntiDengue run summary found yet."

    summary_path = summaries[0]
    try:
        payload = json.loads(summary_path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        return summary_path, {}, f"Could not read latest summary: {exc}"
    return summary_path, payload if isinstance(payload, dict) else {}, ""


def read_json_file(path: Path):
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return None


def bailey_token_data_length(value) -> int:
    if isinstance(value, list):
        return len(value)
    if isinstance(value, str):
        return len(value)
    return 0


def is_baileys_tctoken_expired(timestamp: object) -> bool:
    try:
        parsed = int(str(timestamp))
    except (TypeError, ValueError):
        return True
    bucket_duration = 604800
    bucket_count = 4
    current_bucket = int(time.time()) // bucket_duration
    cutoff_timestamp = (current_bucket - (bucket_count - 1)) * bucket_duration
    return parsed < cutoff_timestamp


def lid_for_phone(auth_dir: Path, phone: str) -> str:
    cleaned_phone = "".join(char for char in str(phone) if char.isdigit())
    if not cleaned_phone:
        return ""

    direct = read_json_file(auth_dir / f"lid-mapping-{cleaned_phone}.json")
    if isinstance(direct, str) and direct.endswith("@lid"):
        return direct

    try:
        reverse_files = list(auth_dir.glob("lid-mapping-*_reverse.json"))
    except OSError:
        return ""

    for path in reverse_files:
        mapped_phone = "".join(char for char in str(read_json_file(path) or "") if char.isdigit())
        if mapped_phone != cleaned_phone:
            continue
        lid_user = path.name.removeprefix("lid-mapping-").removesuffix("_reverse.json")
        return f"{lid_user}@lid" if lid_user else ""
    return ""


def antidengue_token_status_for_mobile(state: GuiState, mobile: object) -> dict[str, str]:
    phone = "".join(char for char in str(mobile or "") if char.isdigit())
    if not phone:
        return {"status": "unknown", "label": "no mobile", "lid": "", "detail": ""}

    cached = state.antidengue_token_cache.get(phone)
    if cached:
        return cached

    auth_dir = whatsapp_auth_dir(state)
    lid = lid_for_phone(auth_dir, phone)
    if not lid:
        result = {"status": "missing", "label": "no LID", "lid": "", "detail": "No Baileys LID mapping found."}
        state.antidengue_token_cache[phone] = result
        return result

    token_path = auth_dir / f"tctoken-{lid}.json"
    entry = read_json_file(token_path) or {}
    token_length = bailey_token_data_length((entry.get("token") or {}).get("data"))
    timestamp = entry.get("timestamp") or ""
    sender_timestamp = entry.get("senderTimestamp") or ""

    if token_length <= 0:
        result = {
            "status": "missing",
            "label": "missing",
            "lid": lid,
            "detail": f"No usable token stored. senderTimestamp={sender_timestamp or '-'}",
        }
    elif is_baileys_tctoken_expired(timestamp):
        result = {
            "status": "expired",
            "label": "expired",
            "lid": lid,
            "detail": f"Token exists but expired. timestamp={timestamp}",
        }
    else:
        result = {
            "status": "usable",
            "label": "usable",
            "lid": lid,
            "detail": f"Token is usable. timestamp={timestamp}",
        }

    state.antidengue_token_cache[phone] = result
    return result


def refresh_antidengue_token_status(state: GuiState) -> None:
    state.antidengue_token_cache = {}
    state.antidengue_token_cache_at = time.time()
    state.runner.status_message = "Refreshed AntiDengue WhatsApp token status"
    state.runner.append_log(state.runner.status_message)


def start_or_check_antidengue_worker(state: GuiState) -> None:
    nats = state.services.services.get("nats")
    whatsapp = state.services.services.get("whatsapp")

    if nats and not nats.is_running:
        state.services.start("nats")
    if whatsapp and not whatsapp.is_running:
        state.services.start("whatsapp")

    state.services.drain_output()
    refresh_antidengue_token_status(state)

    if whatsapp:
        ready, detail = state.services.readiness("whatsapp")
    else:
        ready, detail = False, ""

    if whatsapp and ready:
        state.runner.status_message = "Primary WhatsApp worker is ready for AntiDengue"
    elif whatsapp:
        state.runner.status_message = f"Primary WhatsApp worker is not ready: {detail or whatsapp.status_message}"
    else:
        state.runner.status_message = "Primary WhatsApp worker service is not configured"
    state.runner.append_log(state.runner.status_message)


def start_or_check_control_api(state: GuiState) -> None:
    service = state.services.services.get("automation_control_api")
    if service and not service.is_running:
        state.services.start("automation_control_api")
    state.services.drain_output()

    if is_control_api_reachable():
        state.runner.status_message = "Automation Control API is running on http://127.0.0.1:8787"
    elif service:
        state.runner.status_message = f"Automation Control API is {service.status_message}"
    else:
        state.runner.status_message = "Automation Control API service is not configured"
    state.runner.append_log(state.runner.status_message)


def is_control_api_reachable() -> bool:
    try:
        with urllib.request.urlopen("http://127.0.0.1:8787/health", timeout=0.35) as response:
            return 200 <= response.status < 500
    except Exception:
        return False


def draw_antidengue_worker_controls(state: GuiState) -> None:
    nats = state.services.services.get("nats")
    whatsapp = state.services.services.get("whatsapp")
    control_api = state.services.services.get("automation_control_api")
    nats_ready, _nats_detail = (
        state.services.readiness("nats") if nats else (False, "")
    )
    whatsapp_ready, whatsapp_detail = (
        state.services.readiness("whatsapp")
        if whatsapp and nats_ready
        else (False, "")
    )
    whatsapp_running = bool(whatsapp and whatsapp.is_running)
    control_api_running = bool(control_api and control_api.is_running) or is_control_api_reachable()

    imgui.text("Worker:")
    imgui.same_line()
    whatsapp_label = "Primary WhatsApp ready"
    whatsapp_status = "ok"
    if not whatsapp_ready:
        whatsapp_label = (
            "Primary WhatsApp starting"
            if whatsapp_running
            else "Primary WhatsApp stopped"
        )
        whatsapp_status = "warning" if whatsapp_running else "failed"
    imgui.text_colored(
        status_badge_color(whatsapp_status),
        whatsapp_label,
    )
    if whatsapp_running and not whatsapp_ready and whatsapp_detail:
        imgui.same_line()
        imgui.text_disabled(whatsapp_detail)
    imgui.same_line()
    imgui.text_disabled("NATS ready" if nats_ready else "NATS stopped")

    if imgui.button("Start / Check WhatsApp Worker", (230, 0)):
        start_or_check_antidengue_worker(state)
    imgui.same_line()
    if imgui.button("Refresh Tokens", (125, 0)):
        refresh_antidengue_token_status(state)
    if state.antidengue_token_cache_at:
        imgui.same_line()
        checked_at = datetime.fromtimestamp(state.antidengue_token_cache_at).strftime("%H:%M:%S")
        imgui.text_disabled(f"checked {checked_at}")

    imgui.text("Control API:")
    imgui.same_line()
    imgui.text_colored(
        status_badge_color("ok" if control_api_running else "failed"),
        "running" if control_api_running else "stopped",
    )
    imgui.same_line()
    imgui.text_disabled("http://127.0.0.1:8787")
    imgui.same_line()
    if imgui.button("Start / Check Control API", (205, 0)):
        start_or_check_control_api(state)


def open_desktop_path(state: GuiState, path: Path, label: str) -> None:
    if not path.exists():
        state.runner.append_log(f"{label} is not available: {path}")
        return
    opener = shutil.which("xdg-open")
    if not opener:
        state.runner.append_log(f"Cannot open {label}; xdg-open was not found: {path}")
        return
    subprocess.Popen(
        [opener, str(path)],
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        start_new_session=True,
    )


def open_url_from_gui(state: GuiState, url: str) -> None:
    opener = shutil.which("xdg-open")
    if not opener:
        state.runner.append_log(f"Cannot open URL; xdg-open was not found: {url}")
        return
    subprocess.Popen(
        [opener, url],
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        start_new_session=True,
    )
    state.runner.append_log(f"Opened URL: {url}")


def antidengue_pocketbase_counts(state: GuiState) -> tuple[dict[str, int], str]:
    db_path = antidengue_pocketbase_root(state) / "pb_data" / "data.db"
    if not db_path.is_file():
        return {}, f"PocketBase database has not been initialized yet: {db_path}"
    tables = (
        "districts",
        "departments",
        "wings",
        "tehsils",
        "markazes",
        "schools",
        "ddeo_officers",
        "aeo_officers",
        "ddeo_jurisdictions",
        "aeo_jurisdictions",
        "school_ddeo_overrides",
        "school_aeo_overrides",
        "whatsapp_recipients",
        "officer_import_batches",
        "report_runs",
        "dormant_records",
        "delivery_events",
        "data_quality_issues",
        "dispatch_settings",
        "whatsapp_groups",
        "message_templates",
        "dispatch_events",
    )
    counts: dict[str, int] = {}
    try:
        with open_pocketbase_read_connection(db_path, state) as conn:
            existing = {
                row[0]
                for row in conn.execute(
                    "select name from sqlite_master where type = 'table'"
                ).fetchall()
            }
            for table in tables:
                if table in existing:
                    counts[table] = int(
                        conn.execute(f"select count(*) from {table}").fetchone()[0]
                    )
                else:
                    counts[table] = 0
    except sqlite3.Error as exc:
        return {}, (
            f"Could not read PocketBase database: {exc}. "
            f"path={db_path}, exists={db_path.exists()}, parent_exists={db_path.parent.exists()}, "
            f"project_root={state.project_root}"
        )
    return counts, ""


def antidengue_pocketbase_health_checks(state: GuiState) -> tuple[list[dict], str]:
    db_path = antidengue_pocketbase_root(state) / "pb_data" / "data.db"
    if not db_path.is_file():
        return [], f"PocketBase database has not been initialized yet: {db_path}"

    checks: list[dict] = []

    def add_check(severity: str, name: str, count: int, message: str) -> None:
        checks.append(
            {
                "severity": severity,
                "name": name,
                "count": count,
                "message": message,
            }
        )

    def severity_for_count(count: int, *, warning: bool = False) -> str:
        if count <= 0:
            return "ok"
        return "warning" if warning else "failed"

    try:
        with open_pocketbase_read_connection(db_path, state) as conn:
            existing = {
                row[0]
                for row in conn.execute(
                    "select name from sqlite_master where type = 'table'"
                ).fetchall()
            }
            required_tables = {
                "schools",
                "districts",
                "departments",
                "wings",
                "tehsils",
                "markazes",
                "ddeo_officers",
                "aeo_officers",
                "ddeo_jurisdictions",
                "aeo_jurisdictions",
                "school_ddeo_overrides",
                "school_aeo_overrides",
                "whatsapp_recipients",
                "dispatch_settings",
                "whatsapp_groups",
                "message_templates",
                "dispatch_events",
            }
            missing_tables = sorted(required_tables.difference(existing))
            if missing_tables:
                add_check(
                    "failed",
                    "Required collections",
                    len(missing_tables),
                    "Missing: " + ", ".join(missing_tables),
                )
                return checks, ""

            required_migration = "202607020007_required_refs.js"
            migration_count = int(
                conn.execute(
                    "select count(*) from _migrations where file = ?",
                    (required_migration,),
                ).fetchone()[0]
            )
            add_check(
                "ok" if migration_count else "warning",
                "Required-ref migration",
                0 if migration_count else 1,
                "Applied" if migration_count else "Open Database > Apply Migrations.",
            )

            query_checks = [
                (
                    "School hierarchy refs",
                    """
                    select count(*) from schools
                    where active = 1
                      and (
                        district_ref = '' or department_ref = '' or wing_ref = ''
                        or tehsil_ref = '' or markaz_ref = ''
                      )
                    """,
                    "Active schools must have district/department/wing/tehsil/markaz refs.",
                    False,
                ),
                (
                    "School report fields",
                    """
                    select count(*) from schools
                    where active = 1
                      and (
                        school_type = '' or deos_wise = '' or school_level = ''
                      )
                    """,
                    "Active schools should have report-facing type/wing/level fields.",
                    True,
                ),
                (
                    "School head contacts",
                    """
                    select count(*) from schools
                    where active = 1
                      and (head_name = '' or head_contact = '')
                    """,
                    "Active schools should have head name and contact for follow-up.",
                    True,
                ),
                (
                    "DDEO jurisdiction refs",
                    """
                    select count(*) from ddeo_jurisdictions
                    where active = 1
                      and (
                        ddeo_ref = '' or district_ref = '' or department_ref = ''
                        or wing_ref = '' or tehsil_ref = ''
                      )
                    """,
                    "Active DDEO jurisdictions must be fully linked.",
                    False,
                ),
                (
                    "AEO jurisdiction refs",
                    """
                    select count(*) from aeo_jurisdictions
                    where active = 1
                      and (
                        aeo_ref = '' or district_ref = '' or department_ref = ''
                        or wing_ref = '' or tehsil_ref = '' or markaz_ref = ''
                      )
                    """,
                    "Active AEO jurisdictions must be fully linked.",
                    False,
                ),
                (
                    "Schools without DDEO",
                    """
                    select count(*) from schools s
                    where s.active = 1
                      and not exists (
                        select 1
                        from school_ddeo_overrides o
                        join ddeo_officers d on d.id = o.ddeo_ref and d.active = 1
                        where o.school_ref = s.id and o.active = 1
                      )
                      and not exists (
                        select 1
                        from ddeo_jurisdictions j
                        join ddeo_officers d on d.id = j.ddeo_ref and d.active = 1
                        where j.wing_ref = s.wing_ref
                          and j.tehsil_ref = s.tehsil_ref
                          and j.active = 1
                      )
                    """,
                    "Every active school should resolve to a DDEO.",
                    False,
                ),
                (
                    "Schools without AEO",
                    """
                    select count(*) from schools s
                    where s.active = 1
                      and not exists (
                        select 1
                        from school_aeo_overrides o
                        join aeo_officers a on a.id = o.aeo_ref and a.active = 1
                        where o.school_ref = s.id and o.active = 1
                      )
                      and not exists (
                        select 1
                        from aeo_jurisdictions j
                        join aeo_officers a on a.id = j.aeo_ref and a.active = 1
                        where j.wing_ref = s.wing_ref
                          and j.markaz_ref = s.markaz_ref
                          and (j.tehsil_ref = '' or j.tehsil_ref = s.tehsil_ref)
                          and j.active = 1
                      )
                    """,
                    "Every active school should resolve to an AEO.",
                    False,
                ),
                (
                    "Invalid officer mobiles",
                    """
                    select
                      (select count(*) from ddeo_officers
                       where active = 1
                         and (length(normalized_mobile) != 12 or substr(normalized_mobile, 1, 3) != '923'))
                      +
                      (select count(*) from aeo_officers
                       where active = 1
                         and (length(normalized_mobile) != 12 or substr(normalized_mobile, 1, 3) != '923'))
                    """,
                    "Officer mobiles must be normalized Pakistani WhatsApp numbers.",
                    False,
                ),
                (
                    "Invalid recipients",
                    """
                    select count(*) from whatsapp_recipients
                    where enabled = 1
                      and (
                        target = ''
                        or (type = 'group' and target not like '%@g.us')
                      )
                    """,
                    "Enabled fixed/group recipients must have valid targets.",
                    False,
                ),
                (
                    "Dispatch settings",
                    """
                    select count(*) from dispatch_settings
                    where active = 1
                    """,
                    "At least one active Dispatch Center settings row should exist.",
                    False,
                ),
                (
                    "Enabled dispatch groups",
                    """
                    select count(*) from whatsapp_groups
                    where enabled = 1
                    """,
                    "At least one group route should be enabled for group/fallback dispatch.",
                    True,
                ),
            ]

            for name, sql, message, warning in query_checks:
                count = int(conn.execute(sql).fetchone()[0])
                if name in {"Dispatch settings", "Enabled dispatch groups"}:
                    severity = "ok" if count else ("warning" if warning else "failed")
                else:
                    severity = severity_for_count(count, warning=warning)
                add_check(severity, name, count, message)

            duplicate_ddeo_rows = conn.execute(
                """
                select
                  w.name as wing,
                  t.name as tehsil,
                  count(distinct j.ddeo_ref) as owners,
                  group_concat(distinct d.name || ' ' || d.normalized_mobile) as officers
                from ddeo_jurisdictions j
                join wings w on w.id = j.wing_ref
                join tehsils t on t.id = j.tehsil_ref
                join ddeo_officers d on d.id = j.ddeo_ref
                where j.active = 1
                group by j.wing_ref, j.tehsil_ref
                having owners > 1
                order by t.name
                """
            ).fetchall()
            duplicate_ddeo_message = "A wing/tehsil should not have multiple active DDEO owners."
            if duplicate_ddeo_rows:
                samples = [
                    f"{row['tehsil']} ({row['owners']} owners: {row['officers']})"
                    for row in duplicate_ddeo_rows[:2]
                ]
                duplicate_ddeo_message += " " + " | ".join(samples)
            add_check(
                severity_for_count(len(duplicate_ddeo_rows)),
                "Duplicate DDEO jurisdictions",
                len(duplicate_ddeo_rows),
                duplicate_ddeo_message,
            )

            duplicate_aeo_rows = conn.execute(
                """
                select
                  w.name as wing,
                  t.name as tehsil,
                  m.name as markaz,
                  count(distinct j.aeo_ref) as owners,
                  group_concat(distinct a.name || ' ' || a.normalized_mobile) as officers
                from aeo_jurisdictions j
                join wings w on w.id = j.wing_ref
                join tehsils t on t.id = j.tehsil_ref
                join markazes m on m.id = j.markaz_ref
                join aeo_officers a on a.id = j.aeo_ref
                where j.active = 1
                group by j.wing_ref, j.tehsil_ref, j.markaz_ref
                having owners > 1
                order by t.name, m.name
                """
            ).fetchall()
            duplicate_aeo_message = "A wing/tehsil/markaz should not have multiple active AEO owners."
            if duplicate_aeo_rows:
                samples = [
                    f"{row['tehsil']} / {row['markaz']} ({row['owners']} owners: {row['officers']})"
                    for row in duplicate_aeo_rows[:2]
                ]
                duplicate_aeo_message += " " + " | ".join(samples)
            add_check(
                severity_for_count(len(duplicate_aeo_rows)),
                "Duplicate AEO jurisdictions",
                len(duplicate_aeo_rows),
                duplicate_aeo_message,
            )

            enabled_recipients = int(
                conn.execute(
                    "select count(*) from whatsapp_recipients where enabled = 1"
                ).fetchone()[0]
            )
            add_check(
                "ok" if enabled_recipients else "failed",
                "Enabled fixed recipients",
                enabled_recipients,
                "At least one enabled fixed/group recipient should exist.",
            )
    except sqlite3.Error as exc:
        return [], (
            f"Could not run PocketBase health checks: {exc}. "
            f"path={db_path}, exists={db_path.exists()}, parent_exists={db_path.parent.exists()}"
        )

    return checks, ""


@contextmanager
def open_pocketbase_read_connection(db_path: Path, state: GuiState):
    live_error = ""
    conn = None
    try:
        conn = sqlite3.connect(str(db_path), timeout=2)
        conn.row_factory = sqlite3.Row
        conn.execute("pragma query_only = 1")
    except sqlite3.Error as exc:
        live_error = str(exc)
    else:
        try:
            yield conn
        finally:
            conn.close()
        return

    snapshot_path = Path(tempfile.gettempdir()) / f"antidengue-pocketbase-{uuid.uuid4().hex}.db"
    snapshot_conn = None
    try:
        shutil.copy2(db_path, snapshot_path)
        snapshot_conn = sqlite3.connect(str(snapshot_path), timeout=2)
        snapshot_conn.row_factory = sqlite3.Row
        snapshot_conn.execute("pragma query_only = 1")
    except (OSError, sqlite3.Error) as exc:
        raise sqlite3.OperationalError(
            "unable to open live database or snapshot. "
            f"live_error={live_error}; snapshot_error={exc}; "
            f"path={db_path}; snapshot={snapshot_path}; "
            f"exists={db_path.exists()}; parent_exists={db_path.parent.exists()}; "
            f"project_root={state.project_root}"
        ) from exc
    try:
        yield snapshot_conn
    finally:
        snapshot_conn.close()
        try:
            snapshot_path.unlink(missing_ok=True)
        except OSError:
            pass


def _pocketbase_report_runs_has_lifecycle(conn: sqlite3.Connection) -> bool:
    columns = {
        str(row[1])
        for row in conn.execute("pragma table_info(report_runs)").fetchall()
    }
    return {"archived", "archived_at"}.issubset(columns)


def _now_iso() -> str:
    return datetime.now().isoformat(timespec="seconds")


DISPATCH_SETTING_ID = "b50691f465c2e1a"
DISPATCH_MESSAGE_MODES = (
    ("full_report", "Full report"),
    ("tehsil_summary", "Tehsil summary"),
    ("markaz_summary", "Markaz summary"),
    ("tehsil_markaz_summary", "Tehsil + Markaz summary"),
    ("ddeo_wise", "DDEO-wise"),
    ("aeo_wise", "AEO-wise"),
    ("failed_fallback", "Failed fallback"),
)
DISPATCH_ROUTE_KINDS = (
    ("district", "District"),
    ("tehsil", "Tehsil"),
    ("markaz", "Markaz"),
    ("fallback", "Fallback"),
    ("custom", "Custom"),
)


def antidengue_record_id(prefix: str, *parts: object) -> str:
    key = ":".join(str(part) for part in parts if part is not None)
    return hashlib.sha1(f"{prefix}:{key}".encode("utf-8")).hexdigest()[:15]


@contextmanager
def open_pocketbase_write_connection(db_path: Path):
    conn = sqlite3.connect(str(db_path), timeout=5)
    conn.row_factory = sqlite3.Row
    try:
        yield conn
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def _pb_bool(value: object, default: bool = False) -> bool:
    if value is None:
        return default
    text = str(value).strip().lower()
    if not text:
        return default
    return text not in {"0", "false", "no", "off"}


def _one_or_zero(value: object) -> int:
    return 1 if _pb_bool(value) else 0


def _dispatch_settings_defaults() -> dict[str, object]:
    return {
        "id": DISPATCH_SETTING_ID,
        "active": True,
        "name": "default",
        "allow_individual": True,
        "allow_groups": True,
        "manual_only": False,
        "attach_excel": True,
        "send_failed_only": False,
        "require_preview": False,
        "notes": "Default AntiDengue dispatch controls.",
    }


def load_antidengue_dispatch_settings(state: GuiState) -> tuple[dict[str, object], str]:
    db_path = antidengue_pocketbase_root(state) / "pb_data" / "data.db"
    defaults = _dispatch_settings_defaults()
    if not db_path.is_file():
        return defaults, f"PocketBase database has not been initialized yet: {db_path}"
    try:
        with open_pocketbase_read_connection(db_path, state) as conn:
            tables = {
                row[0]
                for row in conn.execute(
                    "select name from sqlite_master where type = 'table'"
                ).fetchall()
            }
            if "dispatch_settings" not in tables:
                return defaults, "Dispatch Center schema is missing. Apply PocketBase migrations."
            row = conn.execute(
                """
                select id, active, name, allow_individual, allow_groups, manual_only,
                       attach_excel, send_failed_only, require_preview, notes
                from dispatch_settings
                where active = 1
                order by name
                limit 1
                """
            ).fetchone()
            if row is None:
                return defaults, ""
            settings = dict(defaults)
            settings.update(dict(row))
            for key in (
                "active",
                "allow_individual",
                "allow_groups",
                "manual_only",
                "attach_excel",
                "send_failed_only",
                "require_preview",
            ):
                settings[key] = _pb_bool(settings.get(key), bool(defaults[key]))
            return settings, ""
    except sqlite3.Error as exc:
        return defaults, f"Could not read dispatch settings: {exc}"


def save_antidengue_dispatch_settings(
    state: GuiState,
    settings: dict[str, object],
) -> tuple[bool, str]:
    db_path = antidengue_pocketbase_root(state) / "pb_data" / "data.db"
    try:
        with open_pocketbase_write_connection(db_path) as conn:
            tables = {
                row[0]
                for row in conn.execute(
                    "select name from sqlite_master where type = 'table'"
                ).fetchall()
            }
            if "dispatch_settings" not in tables:
                return False, "Dispatch Center schema is missing. Apply PocketBase migrations."
            values = _dispatch_settings_defaults()
            values.update(settings)
            conn.execute(
                """
                insert into dispatch_settings
                (id, active, name, allow_individual, allow_groups, manual_only,
                 attach_excel, send_failed_only, require_preview, notes)
                values
                (:id, :active, :name, :allow_individual, :allow_groups, :manual_only,
                 :attach_excel, :send_failed_only, :require_preview, :notes)
                on conflict(name) do update set
                  active = excluded.active,
                  allow_individual = excluded.allow_individual,
                  allow_groups = excluded.allow_groups,
                  manual_only = excluded.manual_only,
                  attach_excel = excluded.attach_excel,
                  send_failed_only = excluded.send_failed_only,
                  require_preview = excluded.require_preview,
                  notes = excluded.notes
                """,
                {
                    "id": str(values["id"]),
                    "active": _one_or_zero(values["active"]),
                    "name": str(values["name"] or "default"),
                    "allow_individual": _one_or_zero(values["allow_individual"]),
                    "allow_groups": _one_or_zero(values["allow_groups"]),
                    "manual_only": _one_or_zero(values["manual_only"]),
                    "attach_excel": _one_or_zero(values["attach_excel"]),
                    "send_failed_only": _one_or_zero(values["send_failed_only"]),
                    "require_preview": _one_or_zero(values["require_preview"]),
                    "notes": str(values.get("notes") or ""),
                },
            )
        return True, "Saved AntiDengue dispatch settings."
    except sqlite3.Error as exc:
        return False, f"Could not save dispatch settings: {exc}"


def load_antidengue_dispatch_groups(state: GuiState) -> tuple[list[dict], str]:
    db_path = antidengue_pocketbase_root(state) / "pb_data" / "data.db"
    if not db_path.is_file():
        return [], f"PocketBase database has not been initialized yet: {db_path}"
    try:
        with open_pocketbase_read_connection(db_path, state) as conn:
            tables = {
                row[0]
                for row in conn.execute(
                    "select name from sqlite_master where type = 'table'"
                ).fetchall()
            }
            if "whatsapp_groups" not in tables:
                return [], "Dispatch group schema is missing. Apply PocketBase migrations."
            rows = conn.execute(
                """
                select
                  wg.id,
                  wg.enabled,
                  wg.name,
                  wg.target,
                  coalesce(wg.route_kind, 'district') as route_kind,
                  coalesce(wg.message_mode, 'full_report') as message_mode,
                  coalesce(wg.attach_excel, 1) as attach_excel,
                  coalesce(wg.manual_only, 0) as manual_only,
                  coalesce(wg.attachment_text_mode, '') as attachment_text_mode,
                  coalesce(wg.delay_ms, 0) as delay_ms,
                  coalesce(wg.tehsil_ref, '') as tehsil_ref,
                  coalesce(wg.markaz_ref, '') as markaz_ref,
                  coalesce(t.name, '') as tehsil,
                  coalesce(m.name, '') as markaz,
                  coalesce(wg.notes, '') as notes
                from whatsapp_groups wg
                left join tehsils t on t.id = wg.tehsil_ref
                left join markazes m on m.id = wg.markaz_ref
                order by wg.enabled desc, wg.route_kind, wg.name
                """
            ).fetchall()
            return [dict(row) for row in rows], ""
    except sqlite3.Error as exc:
        return [], f"Could not read dispatch groups: {exc}"


def load_antidengue_scope_options(state: GuiState) -> tuple[list[dict], list[dict], str]:
    db_path = antidengue_pocketbase_root(state) / "pb_data" / "data.db"
    try:
        with open_pocketbase_read_connection(db_path, state) as conn:
            tehsils = [
                dict(row)
                for row in conn.execute(
                    "select id, name from tehsils where active = 1 order by name"
                ).fetchall()
            ]
            markazes = [
                dict(row)
                for row in conn.execute(
                    """
                    select m.id, m.name, coalesce(m.tehsil_ref, '') as tehsil_ref, coalesce(t.name, '') as tehsil
                    from markazes m
                    left join tehsils t on t.id = m.tehsil_ref
                    where m.active = 1
                    order by t.name, m.name
                    """
                ).fetchall()
            ]
            return tehsils, markazes, ""
    except sqlite3.Error as exc:
        return [], [], f"Could not read scope options: {exc}"


MASTER_SCHOOL_SHIFTS = ("Single", "Morning", "Evening", "Double")
MASTER_SCHOOL_TYPES = ("Male", "Female", "Co-Education")
MASTER_SCHOOL_LEVELS = ("Primary", "Middle", "High", "Higher Secondary", "sMosque")


def _normalized_emis(value: object) -> str:
    return "".join(ch for ch in str(value or "") if ch.isdigit())


def normalize_pk_mobile_for_gui(value: object) -> str:
    digits = "".join(ch for ch in str(value or "") if ch.isdigit())
    if not digits:
        return ""
    if digits.startswith("0092"):
        digits = digits[2:]
    if digits.startswith("92"):
        return digits
    if digits.startswith("0"):
        return "92" + digits[1:]
    if len(digits) == 10 and digits.startswith("3"):
        return "92" + digits
    return digits


def _school_level_from_name_gui(name: object) -> str:
    prefix = str(name or "").strip().upper().split(" ", 1)[0]
    if prefix == "GPS":
        return "Primary"
    if prefix == "GES":
        return "Middle"
    if prefix == "GMMS":
        return "sMosque"
    return ""


def _pocketbase_table_columns(conn: sqlite3.Connection, table: str) -> set[str]:
    return {str(row[1]) for row in conn.execute(f"pragma table_info({table})").fetchall()}


def _master_school_form_defaults() -> dict[str, object]:
    return {
        "emis": "",
        "name": "",
        "district_ref": "",
        "department_ref": "",
        "wing_ref": "",
        "tehsil_ref": "",
        "markaz_ref": "",
        "shift": "Single",
        "school_type": "Male",
        "deos_wise": "M-EE",
        "school_level": "Primary",
        "head_name": "",
        "head_contact": "",
        "active": True,
        "notes": "",
        "last_result": "",
        "last_success": False,
        "last_saved_emis": "",
    }


def antidengue_new_school_form(state: GuiState) -> dict[str, object]:
    form = state.app_settings.setdefault("antidengue_new_school_form", {})
    if not isinstance(form, dict):
        form = {}
        state.app_settings["antidengue_new_school_form"] = form
    for key, value in _master_school_form_defaults().items():
        form.setdefault(key, value)
    return form


def load_antidengue_master_data_options(state: GuiState) -> tuple[dict[str, list[dict]], str]:
    db_path = antidengue_pocketbase_root(state) / "pb_data" / "data.db"
    if not db_path.is_file():
        return {}, f"PocketBase database has not been initialized yet: {db_path}"
    try:
        with open_pocketbase_read_connection(db_path, state) as conn:
            tables = {
                row[0]
                for row in conn.execute(
                    "select name from sqlite_master where type = 'table'"
                ).fetchall()
            }
            required = {"districts", "departments", "wings", "tehsils", "markazes", "schools"}
            missing = sorted(required.difference(tables))
            if missing:
                return {}, "PocketBase is missing master collections: " + ", ".join(missing)

            options = {
                "districts": [
                    dict(row)
                    for row in conn.execute(
                        "select id, name from districts where active = 1 order by name"
                    ).fetchall()
                ],
                "departments": [
                    dict(row)
                    for row in conn.execute(
                        "select id, name from departments where active = 1 order by name"
                    ).fetchall()
                ],
                "wings": [
                    dict(row)
                    for row in conn.execute(
                        """
                        select id, name, district_ref, department_ref
                        from wings
                        where active = 1
                        order by name
                        """
                    ).fetchall()
                ],
                "tehsils": [
                    dict(row)
                    for row in conn.execute(
                        """
                        select id, name, district_ref
                        from tehsils
                        where active = 1
                        order by name
                        """
                    ).fetchall()
                ],
                "markazes": [
                    dict(row)
                    for row in conn.execute(
                        """
                        select m.id, m.name, m.tehsil_ref, m.wing_ref, coalesce(t.name, '') as tehsil
                        from markazes m
                        left join tehsils t on t.id = m.tehsil_ref
                        where m.active = 1
                        order by t.name, m.name
                        """
                    ).fetchall()
                ],
            }
            return options, ""
    except sqlite3.Error as exc:
        return {}, f"Could not read master data options: {exc}"


def _relation_options(rows: list[dict], placeholder: str, *, include_parent: bool = False) -> list[tuple[str, str]]:
    options = [("", placeholder)]
    for row in rows:
        label = str(row.get("name") or "-")
        if include_parent and row.get("tehsil"):
            label = f"{row.get('tehsil')} / {label}"
        options.append((str(row.get("id") or ""), label))
    return options


def _set_single_default(form: dict[str, object], key: str, rows: list[dict]) -> None:
    if not form.get(key) and len(rows) == 1:
        form[key] = str(rows[0].get("id") or "")


def _draw_relation_combo(
    label: str,
    key: str,
    form: dict[str, object],
    options: list[tuple[str, str]],
    *,
    width: int,
) -> bool:
    valid_values = {value for value, _ in options}
    current = str(form.get(key) or "")
    if current not in valid_values:
        current = ""
        form[key] = ""
    index = _option_index(options, current)
    imgui.set_next_item_width(width)
    changed, index = imgui.combo(label, index, [option_label for _, option_label in options])
    if changed:
        form[key] = options[index][0]
    return changed


def resolve_antidengue_school_owners(
    state: GuiState,
    *,
    district_ref: str,
    department_ref: str,
    wing_ref: str,
    tehsil_ref: str,
    markaz_ref: str,
) -> tuple[dict[str, list[dict]], str]:
    db_path = antidengue_pocketbase_root(state) / "pb_data" / "data.db"
    try:
        with open_pocketbase_read_connection(db_path, state) as conn:
            ddeo_rows = [
                dict(row)
                for row in conn.execute(
                    """
                    select distinct
                      d.id,
                      d.name,
                      d.normalized_mobile,
                      j.id as jurisdiction_id
                    from ddeo_jurisdictions j
                    join ddeo_officers d on d.id = j.ddeo_ref and d.active = 1
                    where j.active = 1
                      and j.district_ref = ?
                      and j.department_ref = ?
                      and j.wing_ref = ?
                      and j.tehsil_ref = ?
                    order by d.name, d.normalized_mobile
                    """,
                    (district_ref, department_ref, wing_ref, tehsil_ref),
                ).fetchall()
            ]
            aeo_rows = [
                dict(row)
                for row in conn.execute(
                    """
                    select distinct
                      a.id,
                      a.name,
                      a.normalized_mobile,
                      j.id as jurisdiction_id
                    from aeo_jurisdictions j
                    join aeo_officers a on a.id = j.aeo_ref and a.active = 1
                    where j.active = 1
                      and j.district_ref = ?
                      and j.department_ref = ?
                      and j.wing_ref = ?
                      and j.markaz_ref = ?
                      and (j.tehsil_ref = '' or j.tehsil_ref = ?)
                    order by a.name, a.normalized_mobile
                    """,
                    (district_ref, department_ref, wing_ref, markaz_ref, tehsil_ref),
                ).fetchall()
            ]
            return {"ddeo": ddeo_rows, "aeo": aeo_rows}, ""
    except sqlite3.Error as exc:
        return {"ddeo": [], "aeo": []}, f"Could not resolve school officers: {exc}"


def validate_antidengue_new_school_form(
    state: GuiState,
    form: dict[str, object],
) -> tuple[list[str], dict[str, list[dict]], str]:
    errors: list[str] = []
    emis = _normalized_emis(form.get("emis"))
    if not emis:
        errors.append("EMIS is required.")
    elif len(emis) != 8:
        errors.append("EMIS must be 8 digits.")
    if not str(form.get("name") or "").strip():
        errors.append("School name is required.")

    required_refs = {
        "district_ref": "District",
        "department_ref": "Department",
        "wing_ref": "Wing",
        "tehsil_ref": "Tehsil",
        "markaz_ref": "Markaz",
    }
    for key, label in required_refs.items():
        if not str(form.get(key) or "").strip():
            errors.append(f"{label} is required.")

    if not str(form.get("shift") or "").strip():
        errors.append("Shift is required.")
    if not str(form.get("school_type") or "").strip():
        errors.append("School type is required.")
    if not str(form.get("school_level") or "").strip():
        errors.append("School level is required.")
    if not str(form.get("head_name") or "").strip():
        errors.append("Head name is required.")

    head_contact = normalize_pk_mobile_for_gui(form.get("head_contact"))
    if not head_contact:
        errors.append("Head contact is required.")
    elif len(head_contact) != 12 or not head_contact.startswith("923"):
        errors.append("Head contact must normalize to a Pakistani WhatsApp number, e.g. 923001234567.")

    owners = {"ddeo": [], "aeo": []}
    owner_error = ""
    refs_complete = all(str(form.get(key) or "").strip() for key in required_refs)
    if refs_complete:
        owners, owner_error = resolve_antidengue_school_owners(
            state,
            district_ref=str(form.get("district_ref") or ""),
            department_ref=str(form.get("department_ref") or ""),
            wing_ref=str(form.get("wing_ref") or ""),
            tehsil_ref=str(form.get("tehsil_ref") or ""),
            markaz_ref=str(form.get("markaz_ref") or ""),
        )
        if owner_error:
            errors.append(owner_error)
        if len(owners.get("ddeo") or []) != 1:
            errors.append("Selected hierarchy must resolve to exactly one active DDEO.")
        if len(owners.get("aeo") or []) != 1:
            errors.append("Selected hierarchy must resolve to exactly one active AEO.")

    return errors, owners, owner_error


def create_antidengue_school_from_form(
    state: GuiState,
    form: dict[str, object],
) -> tuple[bool, str]:
    errors, _owners, _owner_error = validate_antidengue_new_school_form(state, form)
    if errors:
        return False, "Cannot save school: " + " ".join(errors[:3])

    db_path = antidengue_pocketbase_root(state) / "pb_data" / "data.db"
    emis = _normalized_emis(form.get("emis"))
    school_name = " ".join(str(form.get("name") or "").strip().upper().split())
    head_contact = normalize_pk_mobile_for_gui(form.get("head_contact"))
    source_hash_payload = {
        "emis": emis,
        "name": school_name,
        "district_ref": str(form.get("district_ref") or ""),
        "department_ref": str(form.get("department_ref") or ""),
        "wing_ref": str(form.get("wing_ref") or ""),
        "tehsil_ref": str(form.get("tehsil_ref") or ""),
        "markaz_ref": str(form.get("markaz_ref") or ""),
        "source": "manual_gui",
    }
    values: dict[str, object] = {
        "id": antidengue_record_id("school", emis),
        "active": _one_or_zero(form.get("active")),
        "emis": emis,
        "name": school_name,
        "shift": str(form.get("shift") or "").strip(),
        "head_name": " ".join(str(form.get("head_name") or "").strip().split()),
        "head_contact": head_contact,
        "school_type": str(form.get("school_type") or "").strip(),
        "deos_wise": str(form.get("deos_wise") or "").strip() or "M-EE",
        "school_level": str(form.get("school_level") or "").strip(),
        "source_row_hash": hashlib.sha256(
            json.dumps(source_hash_payload, sort_keys=True).encode("utf-8")
        ).hexdigest(),
        "district_ref": str(form.get("district_ref") or ""),
        "department_ref": str(form.get("department_ref") or ""),
        "wing_ref": str(form.get("wing_ref") or ""),
        "tehsil_ref": str(form.get("tehsil_ref") or ""),
        "markaz_ref": str(form.get("markaz_ref") or ""),
        "source": "manual_gui",
        "notes": str(form.get("notes") or "").strip(),
    }

    try:
        with open_pocketbase_write_connection(db_path) as conn:
            existing = conn.execute(
                "select id, name, active from schools where emis = ?",
                (emis,),
            ).fetchone()
            if existing:
                status = "active" if _pb_bool(existing["active"]) else "inactive"
                return False, f"EMIS {emis} already exists as {existing['name']} ({status})."

            columns_in_table = _pocketbase_table_columns(conn, "schools")
            filtered = {
                key: value
                for key, value in values.items()
                if key in columns_in_table
            }
            columns = ", ".join(filtered)
            placeholders = ", ".join(f":{key}" for key in filtered)
            conn.execute(
                f"insert into schools ({columns}) values ({placeholders})",
                filtered,
            )
        state.cached_metrics.pop("antidengue", None)
        return True, f"Saved school {emis} - {school_name}."
    except sqlite3.Error as exc:
        return False, f"Could not save school: {exc}"


def update_antidengue_dispatch_group(
    state: GuiState,
    group_id: str,
    updates: dict[str, object],
) -> tuple[bool, str]:
    allowed = {
        "enabled",
        "name",
        "target",
        "route_kind",
        "message_mode",
        "attach_excel",
        "manual_only",
        "attachment_text_mode",
        "delay_ms",
        "tehsil_ref",
        "markaz_ref",
        "notes",
    }
    clean_updates = {key: value for key, value in updates.items() if key in allowed}
    if not clean_updates:
        return True, "No group changes to save."

    for key in ("enabled", "attach_excel", "manual_only"):
        if key in clean_updates:
            clean_updates[key] = _one_or_zero(clean_updates[key])
    db_path = antidengue_pocketbase_root(state) / "pb_data" / "data.db"
    try:
        with open_pocketbase_write_connection(db_path) as conn:
            assignments = ", ".join(f"{key} = :{key}" for key in clean_updates)
            conn.execute(
                f"update whatsapp_groups set {assignments} where id = :id",
                {**clean_updates, "id": group_id},
            )
        return True, "Saved dispatch group."
    except sqlite3.Error as exc:
        return False, f"Could not save dispatch group: {exc}"


def add_antidengue_dispatch_group(
    state: GuiState,
    values: dict[str, object],
) -> tuple[bool, str]:
    target = str(values.get("target") or "").strip()
    name = str(values.get("name") or "").strip()
    if not name:
        return False, "Enter a group name."
    if not target.endswith("@g.us"):
        return False, "Group target must be a WhatsApp group JID ending with @g.us."

    route_kind = str(values.get("route_kind") or "district")
    message_mode = str(values.get("message_mode") or "full_report")
    db_path = antidengue_pocketbase_root(state) / "pb_data" / "data.db"
    try:
        with open_pocketbase_write_connection(db_path) as conn:
            conn.execute(
                """
                insert into whatsapp_groups
                (id, enabled, name, target, route_kind, message_mode, attach_excel,
                 manual_only, attachment_text_mode, delay_ms, tehsil_ref, markaz_ref, notes)
                values
                (:id, :enabled, :name, :target, :route_kind, :message_mode, :attach_excel,
                 :manual_only, :attachment_text_mode, :delay_ms, :tehsil_ref, :markaz_ref, :notes)
                """,
                {
                    "id": antidengue_record_id("whatsapp_group", target),
                    "enabled": _one_or_zero(values.get("enabled", True)),
                    "name": name,
                    "target": target,
                    "route_kind": route_kind,
                    "message_mode": message_mode,
                    "attach_excel": _one_or_zero(values.get("attach_excel", True)),
                    "manual_only": _one_or_zero(values.get("manual_only", False)),
                    "attachment_text_mode": str(
                        values.get("attachment_text_mode") or "separate"
                    ),
                    "delay_ms": int(values.get("delay_ms") or 1500),
                    "tehsil_ref": str(values.get("tehsil_ref") or ""),
                    "markaz_ref": str(values.get("markaz_ref") or ""),
                    "notes": str(values.get("notes") or "Created from Dispatch Center."),
                },
            )
        return True, f"Added dispatch group: {name}."
    except sqlite3.IntegrityError:
        return False, "A dispatch group with this target already exists."
    except (ValueError, sqlite3.Error) as exc:
        return False, f"Could not add dispatch group: {exc}"


def record_antidengue_dispatch_event(
    state: GuiState,
    *,
    run_key: str,
    channel: str,
    action: str,
    target: str,
    recipient_name: str,
    status: str,
    message: str,
    payload: dict,
) -> None:
    db_path = antidengue_pocketbase_root(state) / "pb_data" / "data.db"
    try:
        with open_pocketbase_write_connection(db_path) as conn:
            tables = {
                row[0]
                for row in conn.execute(
                    "select name from sqlite_master where type = 'table'"
                ).fetchall()
            }
            if "dispatch_events" not in tables:
                return
            event_id = antidengue_record_id(
                "dispatch_event",
                run_key,
                channel,
                action,
                target,
                time.time_ns(),
            )
            conn.execute(
                """
                insert into dispatch_events
                (id, run_key, channel, action, target, recipient_name, status,
                 message_preview, payload, created_at, notes)
                values
                (:id, :run_key, :channel, :action, :target, :recipient_name, :status,
                 :message_preview, :payload, :created_at, :notes)
                """,
                {
                    "id": event_id,
                    "run_key": run_key,
                    "channel": channel,
                    "action": action,
                    "target": target,
                    "recipient_name": recipient_name,
                    "status": status,
                    "message_preview": message[:1000],
                    "payload": json.dumps(payload, ensure_ascii=False, sort_keys=True),
                    "created_at": _now_iso(),
                    "notes": "Recorded from Automation Management GUI.",
                },
            )
    except sqlite3.Error:
        return


def load_antidengue_latest_dispatch_context(state: GuiState) -> tuple[dict, str]:
    db_path = antidengue_pocketbase_root(state) / "pb_data" / "data.db"
    if not db_path.is_file():
        return {}, f"PocketBase database has not been initialized yet: {db_path}"
    try:
        with open_pocketbase_read_connection(db_path, state) as conn:
            tables = {
                row[0]
                for row in conn.execute(
                    "select name from sqlite_master where type = 'table'"
                ).fetchall()
            }
            required = {"report_runs", "dormant_records", "schools", "tehsils", "markazes"}
            missing = required.difference(tables)
            if missing:
                return {}, "PocketBase is missing collections: " + ", ".join(sorted(missing))
            has_lifecycle = _pocketbase_report_runs_has_lifecycle(conn)
            row = conn.execute(
                f"""
                select started_at, raw_file_name, summary
                from report_runs
                {"where coalesce(archived, 0) = 0" if has_lifecycle else ""}
                order by started_at desc
                limit 1
                """
            ).fetchone()
            if row is None:
                return {}, "No AntiDengue run history is available yet."
            run_key = str(row["started_at"] or "")
            dormant_rows = [
                dict(item)
                for item in conn.execute(
                    """
                    select
                      s.id as school_ref,
                      s.emis,
                      s.name as school_name,
                      s.tehsil_ref,
                      s.markaz_ref,
                      t.name as tehsil,
                      m.name as markaz,
                      coalesce(override_ddeo.name, jurisdiction_ddeo.name, '') as ddeo_name,
                      coalesce(override_ddeo.normalized_mobile, jurisdiction_ddeo.normalized_mobile, '') as ddeo_mobile,
                      coalesce(override_aeo.name, jurisdiction_aeo.name, '') as aeo_name,
                      coalesce(override_aeo.normalized_mobile, jurisdiction_aeo.normalized_mobile, '') as aeo_mobile
                    from dormant_records dr
                    left join schools s on s.id = dr.school_ref
                    left join tehsils t on t.id = s.tehsil_ref
                    left join markazes m on m.id = s.markaz_ref
                    left join school_ddeo_overrides sdo
                      on sdo.school_ref = s.id and sdo.active = 1
                    left join ddeo_officers override_ddeo
                      on override_ddeo.id = sdo.ddeo_ref and override_ddeo.active = 1
                    left join ddeo_jurisdictions dj
                      on dj.wing_ref = s.wing_ref
                     and dj.tehsil_ref = s.tehsil_ref
                     and dj.active = 1
                     and sdo.id is null
                    left join ddeo_officers jurisdiction_ddeo
                      on jurisdiction_ddeo.id = dj.ddeo_ref and jurisdiction_ddeo.active = 1
                    left join school_aeo_overrides sao
                      on sao.school_ref = s.id and sao.active = 1
                    left join aeo_officers override_aeo
                      on override_aeo.id = sao.aeo_ref and override_aeo.active = 1
                    left join aeo_jurisdictions aj
                      on aj.wing_ref = s.wing_ref
                     and aj.markaz_ref = s.markaz_ref
                     and (aj.tehsil_ref = '' or aj.tehsil_ref = s.tehsil_ref)
                     and aj.active = 1
                     and sao.id is null
                    left join aeo_officers jurisdiction_aeo
                      on jurisdiction_aeo.id = aj.aeo_ref and jurisdiction_aeo.active = 1
                    where dr.run_key = ?
                    order by t.name, m.name, s.name
                    """,
                    (run_key,),
                ).fetchall()
            ]
    except sqlite3.Error as exc:
        return {}, f"Could not load dispatch context: {exc}"

    summary = _safe_json_loads(row["summary"])
    return {
        "run_key": run_key,
        "raw_file_name": str(row["raw_file_name"] or ""),
        "summary": summary,
        "rows": dormant_rows,
    }, ""


def _group_scope_label(group: dict) -> str:
    if group.get("markaz"):
        return f"{group.get('tehsil') or '-'} / {group.get('markaz')}"
    if group.get("tehsil"):
        return str(group.get("tehsil"))
    return str(group.get("route_kind") or "district").replace("_", " ").title()


def _rows_for_dispatch_group(group: dict, rows: list[dict]) -> list[dict]:
    route_kind = str(group.get("route_kind") or "district").strip().lower()
    tehsil_ref = str(group.get("tehsil_ref") or "")
    markaz_ref = str(group.get("markaz_ref") or "")
    if route_kind == "tehsil" and not tehsil_ref:
        return []
    if route_kind == "markaz" and not markaz_ref:
        return []

    filtered = rows
    if tehsil_ref:
        filtered = [row for row in filtered if str(row.get("tehsil_ref") or "") == tehsil_ref]
    if markaz_ref:
        filtered = [row for row in filtered if str(row.get("markaz_ref") or "") == markaz_ref]
    return filtered


def _school_count_text(count: int) -> str:
    return f"{count} {'school' if count == 1 else 'schools'}"


def _school_line(index: int, row: dict) -> str:
    emis = str(row.get("emis") or "").strip()
    emis_part = f"{emis} - " if emis else ""
    return f"{index}. {emis_part}{row.get('school_name') or 'Unknown School'}"


def _tehsil_summary_lines(rows: list[dict]) -> list[str]:
    tehsil_counts = Counter(str(row.get("tehsil") or "Unmapped") for row in rows)
    if not tehsil_counts:
        return []
    return [
        "",
        "*TEHSIL SUMMARY*",
        *[
            f"{index}. {name}: {_school_count_text(count)}"
            for index, (name, count) in enumerate(sorted(tehsil_counts.items()), start=1)
        ],
    ]


def _markaz_summary_lines(rows: list[dict]) -> list[str]:
    if not rows:
        return []

    grouped_by_tehsil: dict[str, Counter[str]] = {}
    for row in rows:
        tehsil = str(row.get("tehsil") or "Unmapped")
        markaz = str(row.get("markaz") or "Unmapped")
        grouped_by_tehsil.setdefault(tehsil, Counter())[markaz] += 1

    lines = ["", "*MARKAZ SUMMARY*"]
    for tehsil, markaz_counts in sorted(grouped_by_tehsil.items()):
        tehsil_total = sum(markaz_counts.values())
        lines.extend(["", f"*{tehsil} - {_school_count_text(tehsil_total)}*"])
        lines.extend(
            f"{index}. {markaz}: {_school_count_text(count)}"
            for index, (markaz, count) in enumerate(sorted(markaz_counts.items()), start=1)
        )
    return lines


def _summary_only_sections(
    rows: list[dict],
    *,
    include_tehsil: bool,
    include_markaz: bool,
) -> list[str]:
    if not rows:
        return ["", "No dormant schools in this group's current scope."]

    lines: list[str] = []
    if include_tehsil:
        lines.extend(_tehsil_summary_lines(rows))
    if include_markaz:
        lines.extend(_markaz_summary_lines(rows))
    return lines


def _hierarchy_school_sections(
    rows: list[dict],
    *,
    include_heading: bool = True,
) -> list[str]:
    if not rows:
        return ["", "No dormant schools in this group's current scope."]

    grouped_by_tehsil: dict[str, list[dict]] = {}
    for row in rows:
        grouped_by_tehsil.setdefault(str(row.get("tehsil") or "Unmapped"), []).append(row)

    lines: list[str] = []
    if include_heading:
        lines.extend(["", "*DETAILS BY TEHSIL / MARKAZ*"])

    for tehsil, tehsil_rows in sorted(grouped_by_tehsil.items()):
        sorted_tehsil_rows = sorted(
            tehsil_rows,
            key=lambda item: (str(item.get("markaz") or ""), str(item.get("school_name") or "")),
        )
        lines.extend(["", f"*{tehsil} - {_school_count_text(len(sorted_tehsil_rows))}*"])

        grouped_by_markaz: dict[str, list[dict]] = {}
        for row in sorted_tehsil_rows:
            grouped_by_markaz.setdefault(str(row.get("markaz") or "Unmapped"), []).append(row)

        for markaz, markaz_rows in sorted(grouped_by_markaz.items()):
            sorted_markaz_rows = sorted(
                markaz_rows,
                key=lambda item: str(item.get("school_name") or ""),
            )
            lines.extend(
                [
                    "",
                    f"*{markaz} - {_school_count_text(len(sorted_markaz_rows))}*",
                ]
            )
            lines.extend(
                _school_line(index, row)
                for index, row in enumerate(sorted_markaz_rows, start=1)
            )

    return lines


def _markaz_school_sections(rows: list[dict]) -> list[str]:
    if not rows:
        return ["", "No dormant schools in this group's current scope."]

    grouped_by_markaz: dict[str, list[dict]] = {}
    for row in rows:
        grouped_by_markaz.setdefault(str(row.get("markaz") or "Unmapped"), []).append(row)

    lines: list[str] = []
    for markaz, markaz_rows in sorted(grouped_by_markaz.items()):
        sorted_markaz_rows = sorted(
            markaz_rows,
            key=lambda item: str(item.get("school_name") or ""),
        )
        lines.extend(["", f"*{markaz} - {_school_count_text(len(sorted_markaz_rows))}*"])
        lines.extend(
            _school_line(index, row)
            for index, row in enumerate(sorted_markaz_rows, start=1)
        )
    return lines


def _grouped_school_sections(rows: list[dict], key_fields: tuple[str, str]) -> list[str]:
    grouped: dict[tuple[str, str], list[dict]] = {}
    for row in rows:
        key = (
            str(row.get(key_fields[0]) or "Unmapped"),
            str(row.get(key_fields[1]) or ""),
        )
        grouped.setdefault(key, []).append(row)

    lines: list[str] = []
    for (name, mobile), group_rows in sorted(grouped.items()):
        title = name if not mobile else f"{name} ({mobile})"
        lines.extend(["", f"*{title} - {_school_count_text(len(group_rows))}*"])
        lines.extend(_hierarchy_school_sections(group_rows, include_heading=False))
    return lines


def build_antidengue_group_dispatch_message(group: dict, context: dict) -> str:
    rows = _rows_for_dispatch_group(group, context.get("rows") or [])
    summary = context.get("summary") or {}
    run_time = str(context.get("run_key") or summary.get("run_started_at") or "")[:19].replace("T", " ")
    mode = str(group.get("message_mode") or "full_report")
    lines = [
        f"*Anti-Dengue Dormant Schools Report - {run_time}*",
        f"*Group:* {group.get('name') or group.get('target') or '-'}",
        f"*Scope:* {_group_scope_label(group)}",
        f"*Total dormant:* {_school_count_text(len(rows))}",
    ]

    if mode == "failed_fallback":
        failed = failed_officer_rows(
            ((summary.get("whatsapp") or {}).get("officer_delivery_audit") or [])
        )
        lines.extend(["", "Failed direct-delivery officers:"])
        if not failed:
            lines.append("No failed individual officer deliveries in the latest run.")
        for index, row in enumerate(failed, start=1):
            lines.append(
                f"{index}. {row.get('Officer Name') or '-'} ({row.get('Roles') or '-'}) "
                f"{row.get('Mobile Number') or '-'} - schools: {row.get('School Count') or 0}"
            )
    elif mode == "ddeo_wise":
        lines.extend(_grouped_school_sections(rows, ("ddeo_name", "ddeo_mobile")))
    elif mode == "aeo_wise":
        lines.extend(_grouped_school_sections(rows, ("aeo_name", "aeo_mobile")))
    elif mode == "tehsil_summary":
        lines.extend(
            _summary_only_sections(rows, include_tehsil=True, include_markaz=False)
        )
    elif mode == "markaz_summary":
        lines.extend(
            _summary_only_sections(rows, include_tehsil=False, include_markaz=True)
        )
    elif mode == "tehsil_markaz_summary":
        lines.extend(
            _summary_only_sections(rows, include_tehsil=True, include_markaz=True)
        )
    else:
        lines.extend(_tehsil_summary_lines(rows))
        lines.extend(_hierarchy_school_sections(rows))

    if _pb_bool(group.get("attach_excel"), True):
        lines.extend(["", "Excel report should be attached with this dispatch."])
    lines.extend(["", "Please ensure activities are submitted/updated on the portal today."])
    return "\n".join(lines)


def set_antidengue_run_archived(
    state: GuiState,
    run_key: str,
    archived: bool,
) -> tuple[bool, str]:
    db_path = antidengue_pocketbase_root(state) / "pb_data" / "data.db"
    if not db_path.is_file():
        return False, f"PocketBase database not found: {db_path}"
    try:
        with sqlite3.connect(str(db_path), timeout=5) as conn:
            if not _pocketbase_report_runs_has_lifecycle(conn):
                return False, "Run lifecycle migration is not applied. Open Database > Apply Migrations."
            conn.execute(
                """
                update report_runs
                   set archived = ?,
                       archived_at = ?
                 where started_at = ?
                """,
                (1 if archived else 0, _now_iso() if archived else "", run_key),
            )
    except sqlite3.Error as exc:
        return False, f"Could not update run archive state: {exc}"
    return True, "Archived run." if archived else "Restored run."


def delete_antidengue_run_history(state: GuiState, run_key: str) -> tuple[bool, str]:
    db_path = antidengue_pocketbase_root(state) / "pb_data" / "data.db"
    if not db_path.is_file():
        return False, f"PocketBase database not found: {db_path}"
    try:
        with sqlite3.connect(str(db_path), timeout=5) as conn:
            tables = {
                row[0]
                for row in conn.execute(
                    "select name from sqlite_master where type = 'table'"
                ).fetchall()
            }
            for table in (
                "dormant_records",
                "delivery_events",
                "data_quality_issues",
                "run_stage_events",
                "dispatch_plan_items",
            ):
                if table not in tables:
                    continue
                conn.execute(f"delete from {table} where run_key = ?", (run_key,))
            conn.execute("delete from report_runs where started_at = ?", (run_key,))
    except sqlite3.Error as exc:
        return False, f"Could not delete run history: {exc}"
    if state.antidengue_history_selected_run_key == run_key:
        state.antidengue_history_selected_run_key = ""
    if state.antidengue_history_pending_delete_run_key == run_key:
        state.antidengue_history_pending_delete_run_key = ""
    return True, "Deleted run history and related records."


def clear_archived_antidengue_runs(state: GuiState) -> tuple[bool, str]:
    db_path = antidengue_pocketbase_root(state) / "pb_data" / "data.db"
    if not db_path.is_file():
        return False, f"PocketBase database not found: {db_path}"
    try:
        with sqlite3.connect(str(db_path), timeout=5) as conn:
            if not _pocketbase_report_runs_has_lifecycle(conn):
                return False, "Run lifecycle migration is not applied. Open Database > Apply Migrations."
            run_keys = [
                str(row[0])
                for row in conn.execute(
                    "select started_at from report_runs where archived = 1"
                ).fetchall()
            ]
            tables = {
                row[0]
                for row in conn.execute(
                    "select name from sqlite_master where type = 'table'"
                ).fetchall()
            }
            for run_key in run_keys:
                for table in (
                    "dormant_records",
                    "delivery_events",
                    "data_quality_issues",
                    "run_stage_events",
                    "dispatch_plan_items",
                ):
                    if table not in tables:
                        continue
                    conn.execute(f"delete from {table} where run_key = ?", (run_key,))
                conn.execute("delete from report_runs where started_at = ?", (run_key,))
    except sqlite3.Error as exc:
        return False, f"Could not clear archived runs: {exc}"
    state.antidengue_history_selected_run_key = ""
    state.antidengue_history_pending_delete_run_key = ""
    return True, f"Deleted {len(run_keys)} archived run(s)."


def _safe_json_loads(value) -> dict:
    if isinstance(value, dict):
        return value
    if value is None:
        return {}
    try:
        parsed = json.loads(str(value))
    except (TypeError, json.JSONDecodeError):
        return {}
    return parsed if isinstance(parsed, dict) else {}


def _antidengue_run_key(started_at: object) -> str:
    return re.sub(r"[^0-9A-Za-z_.:-]+", "_", str(started_at or ""))[:150]


def load_antidengue_run_history(state: GuiState, limit: int = 50) -> tuple[list[dict], str]:
    db_path = antidengue_pocketbase_root(state) / "pb_data" / "data.db"
    if not db_path.is_file():
        return [], f"PocketBase database has not been initialized yet: {db_path}"

    try:
        with open_pocketbase_read_connection(db_path, state) as conn:
            tables = {
                row[0]
                for row in conn.execute(
                    "select name from sqlite_master where type = 'table'"
                ).fetchall()
            }
            required = {
                "report_runs",
                "dormant_records",
                "delivery_events",
                "data_quality_issues",
            }
            missing = required.difference(tables)
            if missing:
                return [], "PocketBase is missing collections: " + ", ".join(sorted(missing))

            has_lifecycle = _pocketbase_report_runs_has_lifecycle(conn)
            has_stage_events = "run_stage_events" in tables
            has_dispatch_plan = "dispatch_plan_items" in tables
            filter_clause = ""
            filter_params: tuple = ()
            if has_lifecycle:
                if state.antidengue_history_filter == "archived":
                    filter_clause = "where coalesce(rr.archived, 0) = 1"
                elif state.antidengue_history_filter != "all":
                    filter_clause = "where coalesce(rr.archived, 0) = 0"

            rows = conn.execute(
                f"""
                select
                  rr.id,
                  rr.started_at,
                  rr.finished_at,
                  rr.source,
                  rr.status,
                  rr.raw_file_name,
                  rr.raw_file_sha256,
                  rr.summary,
                  {"coalesce(rr.archived, 0)" if has_lifecycle else "0"} as archived,
                  {"rr.archived_at" if has_lifecycle else "''"} as archived_at,
                  (select count(*) from dormant_records dr where dr.run_key = rr.started_at) as dormant_count,
                  (select count(*) from delivery_events de where de.run_key = rr.started_at) as delivery_count,
                  (select count(*) from delivery_events de where de.run_key = rr.started_at and lower(de.status) = 'delivered') as delivered_count,
                  (select count(*) from delivery_events de where de.run_key = rr.started_at and lower(de.status) != 'delivered') as failed_count,
                  (select count(*) from data_quality_issues qi where qi.run_key = rr.started_at) as issue_count,
                  {"(select count(*) from run_stage_events se where se.run_key = rr.started_at)" if has_stage_events else "0"} as stage_count,
                  {"(select count(*) from dispatch_plan_items dpi where dpi.run_key = rr.started_at)" if has_dispatch_plan else "0"} as dispatch_plan_count
                from report_runs rr
                {filter_clause}
                order by rr.started_at desc
                limit ?
                """,
                (*filter_params, limit),
            ).fetchall()
    except sqlite3.Error as exc:
        return [], (
            f"Could not read PocketBase run history: {exc}. "
            f"path={db_path}, exists={db_path.exists()}, parent_exists={db_path.parent.exists()}, "
            f"project_root={state.project_root}"
        )

    history: list[dict] = []
    for row in rows:
        item = dict(row)
        item["run_key"] = _antidengue_run_key(item.get("started_at"))
        item["summary_dict"] = _safe_json_loads(item.get("summary"))
        history.append(item)
    return history, ""


def load_antidengue_run_details(
    state: GuiState,
    run_key: str,
) -> tuple[dict[str, list[dict]], str]:
    db_path = antidengue_pocketbase_root(state) / "pb_data" / "data.db"
    if not db_path.is_file():
        return {}, f"PocketBase database has not been initialized yet: {db_path}"

    try:
        with open_pocketbase_read_connection(db_path, state) as conn:
            tables = {
                row[0]
                for row in conn.execute(
                    "select name from sqlite_master where type = 'table'"
                ).fetchall()
            }
            summary_row = conn.execute(
                "select summary from report_runs where started_at = ?",
                (run_key,),
            ).fetchone()
            summary_dict = _safe_json_loads(summary_row["summary"]) if summary_row else {}
            dormant_columns = {
                str(row[1])
                for row in conn.execute("pragma table_info(dormant_records)").fetchall()
            }
            if "school_ref" in dormant_columns:
                dormant_sql = """
                    select
                      s.emis as school_emis,
                      s.name as school_name,
                      t.name as tehsil,
                      m.name as markaz,
                      dr.row_data
                    from dormant_records dr
                    left join schools s
                      on s.id = dr.school_ref
                    left join tehsils t
                      on t.id = s.tehsil_ref
                    left join markazes m
                      on m.id = s.markaz_ref
                    where dr.run_key = ?
                    order by t.name, m.name, s.name
                    limit 80
                """
            else:
                dormant_sql = """
                    select school_emis, school_name, tehsil, markaz, row_data
                    from dormant_records
                    where run_key = ?
                    order by tehsil, markaz, school_name
                    limit 80
                """
            dormant_rows = [
                dict(row)
                for row in conn.execute(dormant_sql, (run_key,)).fetchall()
            ]
            delivery_rows = [
                dict(row)
                for row in conn.execute(
                    """
                    select job_id, target, recipient_name, role, status, cause, payload
                    from delivery_events
                    where run_key = ?
                    order by
                      case when lower(status) = 'delivered' then 1 else 0 end,
                      role,
                      recipient_name
                    limit 80
                    """,
                    (run_key,),
                ).fetchall()
            ]
            quality_rows = [
                dict(row)
                for row in conn.execute(
                    """
                    select severity, category, message, context
                    from data_quality_issues
                    where run_key = ?
                    order by severity desc, category, message
                    limit 100
                    """,
                    (run_key,),
                ).fetchall()
            ]
            if "run_stage_events" in tables:
                lifecycle_rows = [
                    dict(row)
                    for row in conn.execute(
                        """
                        select stage, status, message, payload, created_at
                        from run_stage_events
                        where run_key = ?
                        order by created_at, id
                        limit 120
                        """,
                        (run_key,),
                    ).fetchall()
                ]
            else:
                lifecycle_rows = []
            if not lifecycle_rows:
                lifecycle_rows = [
                    row
                    for row in summary_dict.get("lifecycle") or []
                    if isinstance(row, dict)
                ][:120]

            if "dispatch_plan_items" in tables:
                dispatch_plan_rows = [
                    dict(row)
                    for row in conn.execute(
                        """
                        select
                          job_id,
                          channel,
                          recipient_name,
                          target,
                          route_kind,
                          message_mode,
                          row_count,
                          excel_path,
                          status,
                          cause,
                          payload,
                          created_at,
                          updated_at
                        from dispatch_plan_items
                        where run_key = ?
                        order by
                          case channel
                            when 'group' then 0
                            when 'fallback' then 1
                            else 2
                          end,
                          recipient_name,
                          target
                        limit 160
                        """,
                        (run_key,),
                    ).fetchall()
                ]
            else:
                dispatch_plan_rows = []
            if not dispatch_plan_rows:
                dispatch_plan_rows = [
                    row
                    for row in (
                        (summary_dict.get("whatsapp") or {}).get("dispatch_plan") or []
                    )
                    if isinstance(row, dict)
                ][:160]
    except sqlite3.Error as exc:
        return {}, (
            f"Could not read selected run details: {exc}. "
            f"path={db_path}, exists={db_path.exists()}, parent_exists={db_path.parent.exists()}, "
            f"project_root={state.project_root}"
        )

    for row in delivery_rows:
        row["payload_dict"] = _safe_json_loads(row.get("payload"))
    for row in dormant_rows:
        row["row_data_dict"] = _safe_json_loads(row.get("row_data"))
    for row in lifecycle_rows:
        row["payload_dict"] = _safe_json_loads(row.get("payload"))
    for row in dispatch_plan_rows:
        row["payload_dict"] = _safe_json_loads(row.get("payload"))
    return {
        "dormant": dormant_rows,
        "delivery": delivery_rows,
        "quality": quality_rows,
        "lifecycle": lifecycle_rows,
        "dispatch_plan": dispatch_plan_rows,
    }, ""


def status_badge_color(status: str) -> tuple[float, float, float, float]:
    normalized = status.strip().lower().replace("_", " ")
    if normalized in {"ok", "passed", "delivered", "queued", "success", "manual sent", "started"}:
        return (0.26, 0.72, 0.40, 1.0)
    if normalized in {
        "warning",
        "partial",
        "pending",
        "copied",
        "manual pending",
        "preview required",
        "sent pending confirmation",
        "missing status",
    }:
        return (0.95, 0.58, 0.22, 1.0)
    if normalized in {"failed", "error", "blocked", "not queued", "skipped"}:
        return (0.95, 0.35, 0.28, 1.0)
    return (0.82, 0.82, 0.82, 1.0)


def delivery_failure_cause(error: str, ack_statuses: str, target: str) -> str:
    text = f"{error} {ack_statuses}".lower()
    if "tctoken" in text or "privacy token" in text:
        return "Direct WhatsApp send is blocked by a missing or expired Baileys privacy token for this contact; fallback escalation should carry the report."
    if "463" in text:
        return "WhatsApp Web rejected this direct chat as a reachout/privacy-token case. Avoid repeated retries; refresh the contact token or use fallback delivery."
    if "status was pending" in text or " pending" in text:
        if target.endswith("@g.us"):
            return "Group message stayed pending; verify the sender is still a member and the group accepts messages."
        return "WhatsApp accepted the job locally, but no server acknowledgement arrived."
    if "whatsapp reported error" in text or " error" in text:
        return "WhatsApp rejected this recipient. Reauth alone usually does not fix a recipient/token/privacy rejection."
    if "missing status" in text:
        return "The worker did not send a terminal status for this job."
    return "Delivery did not complete cleanly."


def build_antidengue_manual_message(row: dict) -> str:
    officer_name = str(row.get("Officer Name") or "Officer").strip()
    roles = str(row.get("Roles") or "").strip()
    school_count = str(row.get("School Count") or "").strip()
    tehsils = ", ".join(
        line.strip()
        for line in str(row.get("Tehsils") or "").splitlines()
        if line.strip()
    )
    markazes = ", ".join(
        line.strip()
        for line in str(row.get("Markazes") or "").splitlines()
        if line.strip()
    )
    schools = [
        line.strip()
        for line in str(row.get("Schools") or "").splitlines()
        if line.strip()
    ]

    header_name = officer_name if officer_name and officer_name != "-" else "Officer"
    role_text = f" ({roles})" if roles else ""
    count_text = f"{school_count} " if school_count else ""
    context_parts = []
    if tehsils:
        context_parts.append(f"Tehsil: {tehsils}")
    if markazes:
        context_parts.append(f"Markaz: {markazes}")

    lines = [
        f"Assalam o Alaikum {header_name}{role_text},",
        "",
        f"Anti-Dengue portal is still showing {count_text}school(s) as dormant/inactive under your supervision.",
    ]

    if context_parts:
        lines.append(" | ".join(context_parts))

    if schools:
        lines.extend(["", "Schools:"])
        lines.extend(f"{index}. {school}" for index, school in enumerate(schools, start=1))

    lines.extend(
        [
            "",
            "Please ensure activities are submitted/updated on the portal today.",
            "After checking, please reply OK here so the automated WhatsApp system can refresh direct delivery for your number.",
        ]
    )
    return "\n".join(lines)


def draw_delivery_metric(label: str, value: object, status: str = "") -> None:
    imgui.text_disabled(label)
    if status:
        imgui.text_colored(status_badge_color(status), str(value))
    else:
        imgui.text(str(value))


def _record_dispatch_action_result(state: GuiState, success: bool, message: str) -> None:
    state.runner.status_message = message
    state.runner.append_log(message)
    if success:
        state.cached_metrics.pop("antidengue", None)


def draw_antidengue_dispatch_settings_controls(state: GuiState) -> dict[str, object] | None:
    settings, error = load_antidengue_dispatch_settings(state)
    if error:
        imgui.text_colored(status_badge_color("warning"), error)
        return None

    imgui.separator_text("Dispatch Controls")
    changed_any = False
    for index, (key, label) in enumerate(
        (
            ("allow_individual", "Individual messages"),
            ("allow_groups", "Group messages"),
            ("manual_only", "Manual only"),
            ("attach_excel", "Attach Excel"),
            ("require_preview", "Require preview"),
        )
    ):
        if index:
            imgui.same_line()
        current = bool(settings.get(key))
        changed, current = imgui.checkbox(label, current)
        if changed:
            settings[key] = current
            changed_any = True

    if changed_any:
        _record_dispatch_action_result(
            state,
            *save_antidengue_dispatch_settings(state, settings),
        )

    if settings.get("manual_only"):
        imgui.text_colored(
            status_badge_color("warning"),
            "Manual-only is active: runs build reports and messages, but do not queue WhatsApp jobs.",
        )
    elif settings.get("require_preview"):
        imgui.text_colored(
            status_badge_color("warning"),
            "Preview is required: runs build dispatch items, then wait for manual/operator action.",
        )
    return settings


def draw_antidengue_dispatch_overview(state: GuiState, context: dict, settings: dict[str, object]) -> None:
    rows = context.get("rows") or []
    summary = context.get("summary") or {}
    whatsapp = summary.get("whatsapp") or {}
    audit_rows = whatsapp.get("officer_delivery_audit") or []
    failed_rows = failed_officer_rows(audit_rows)
    groups, group_error = load_antidengue_dispatch_groups(state)
    enabled_groups = [group for group in groups if _pb_bool(group.get("enabled"))]

    if imgui.begin_table("antidengue_dispatch_overview", 6):
        for label in ("Run", "Dormant", "Officers", "Failed/Pending", "Groups", "Mode"):
            imgui.table_setup_column(label)
        imgui.table_headers_row()
        imgui.table_next_row()
        imgui.table_next_column()
        imgui.text(str(context.get("run_key") or "-").replace("T", " ")[:19])
        imgui.table_next_column()
        imgui.text(str(len(rows)))
        imgui.table_next_column()
        imgui.text(str(len(audit_rows)))
        imgui.table_next_column()
        imgui.text_colored(
            status_badge_color("failed" if failed_rows else "ok"),
            str(len(failed_rows)),
        )
        imgui.table_next_column()
        imgui.text_colored(
            status_badge_color("ok" if enabled_groups else "warning"),
            f"{len(enabled_groups)} / {len(groups)}",
        )
        imgui.table_next_column()
        mode = "Manual only" if settings.get("manual_only") else "Preview required" if settings.get("require_preview") else "Auto allowed"
        imgui.text_colored(
            status_badge_color("warning" if settings.get("manual_only") or settings.get("require_preview") else "ok"),
            mode,
        )
        imgui.end_table()

    if group_error:
        imgui.text_colored(status_badge_color("warning"), group_error)

    imgui.text_disabled(
        "Use Individuals for failed direct messages. Use Groups to enable routes, choose message shape, and copy group-ready text."
    )


def draw_antidengue_dispatch_individuals(state: GuiState, context: dict) -> None:
    summary = context.get("summary") or {}
    whatsapp = summary.get("whatsapp") or {}
    audit_rows = whatsapp.get("officer_delivery_audit") or []
    run_key = str(context.get("run_key") or summary.get("run_started_at") or "")

    show_all = _pb_bool(
        state.app_settings.get("antidengue_dispatch_show_all_individuals"),
        False,
    )
    changed, show_all = imgui.checkbox("Show delivered too", show_all)
    if changed:
        state.app_settings["antidengue_dispatch_show_all_individuals"] = show_all

    rows = audit_rows if show_all else failed_officer_rows(audit_rows)
    if not rows:
        imgui.text_colored(status_badge_color("ok"), "No individual rows need manual dispatch in the latest run.")
        return

    if imgui.begin_table("antidengue_dispatch_individuals", 9):
        for label in ("Officer", "Role", "Schools", "Mobile", "Status", "Token", "Copy", "Manual", "Cause"):
            imgui.table_setup_column(label)
        imgui.table_headers_row()
        for index, row in enumerate(rows[:80]):
            imgui.push_id(f"dispatch_individual_{index}")
            status = str(row.get("Delivery Status") or "pending")
            message = build_antidengue_manual_message(row)
            imgui.table_next_row()
            imgui.table_next_column()
            draw_clipped_text(row.get("Officer Name"), 28)
            imgui.table_next_column()
            imgui.text(str(row.get("Roles") or "-"))
            imgui.table_next_column()
            imgui.text(str(row.get("School Count") or "-"))
            imgui.table_next_column()
            imgui.text(str(row.get("Mobile Number") or "-"))
            imgui.table_next_column()
            imgui.text_colored(status_badge_color(status), status)
            imgui.table_next_column()
            token_status = antidengue_token_status_for_mobile(state, row.get("Mobile Number"))
            imgui.text_colored(
                status_badge_color({"usable": "ok", "expired": "warning", "missing": "failed"}.get(str(token_status.get("status") or ""), "")),
                str(token_status.get("label") or "-"),
            )
            imgui.table_next_column()
            if imgui.button("Copy", (70, 0)):
                imgui.set_clipboard_text(message)
                record_antidengue_dispatch_event(
                    state,
                    run_key=run_key,
                    channel="individual",
                    action="copied",
                    target=str(row.get("Target") or row.get("Mobile Number") or ""),
                    recipient_name=str(row.get("Officer Name") or ""),
                    status="copied",
                    message=message,
                    payload=row,
                )
                state.runner.status_message = f"Copied manual message for {row.get('Officer Name') or row.get('Mobile Number') or 'officer'}"
                state.runner.append_log(state.runner.status_message)
            imgui.table_next_column()
            if imgui.button("Mark Sent", (95, 0)):
                record_antidengue_dispatch_event(
                    state,
                    run_key=run_key,
                    channel="individual",
                    action="marked_sent",
                    target=str(row.get("Target") or row.get("Mobile Number") or ""),
                    recipient_name=str(row.get("Officer Name") or ""),
                    status="manual_sent",
                    message=message,
                    payload=row,
                )
                state.runner.status_message = f"Marked manual dispatch sent for {row.get('Officer Name') or row.get('Mobile Number') or 'officer'}"
                state.runner.append_log(state.runner.status_message)
            imgui.table_next_column()
            draw_clipped_text(
                delivery_failure_cause(
                    str(row.get("Delivery Error") or ""),
                    str(row.get("WhatsApp Ack Statuses") or ""),
                    str(row.get("Target") or ""),
                ),
                54,
            )
            imgui.pop_id()
        imgui.end_table()
    if len(rows) > 80:
        imgui.text_disabled(f"Showing first 80 of {len(rows)} individual dispatch rows.")


def _option_index(options: tuple[tuple[str, str], ...] | list[tuple[str, str]], value: str) -> int:
    values = [item[0] for item in options]
    try:
        return values.index(value)
    except ValueError:
        return 0


def _draw_group_scope_selector(
    state: GuiState,
    form: dict[str, object],
    tehsils: list[dict],
    markazes: list[dict],
) -> None:
    route_kind = str(form.get("route_kind") or "district")
    if route_kind in {"tehsil", "markaz"}:
        tehsil_options = [("", "All tehsils"), *[(str(row["id"]), str(row["name"])) for row in tehsils]]
        tehsil_index = _option_index(tehsil_options, str(form.get("tehsil_ref") or ""))
        imgui.set_next_item_width(220)
        changed, tehsil_index = imgui.combo("Tehsil", tehsil_index, [label for _, label in tehsil_options])
        if changed:
            form["tehsil_ref"] = tehsil_options[tehsil_index][0]
            form["markaz_ref"] = ""

    if route_kind == "markaz":
        selected_tehsil = str(form.get("tehsil_ref") or "")
        scoped_markazes = [
            row for row in markazes
            if not selected_tehsil or str(row.get("tehsil_ref") or "") == selected_tehsil
        ]
        markaz_options = [
            ("", "All markazes"),
            *[
                (
                    str(row["id"]),
                    f"{row.get('tehsil') or '-'} / {row.get('name') or '-'}",
                )
                for row in scoped_markazes
            ],
        ]
        markaz_index = _option_index(markaz_options, str(form.get("markaz_ref") or ""))
        imgui.set_next_item_width(320)
        changed, markaz_index = imgui.combo("Markaz", markaz_index, [label for _, label in markaz_options])
        if changed:
            form["markaz_ref"] = markaz_options[markaz_index][0]


def _tehsil_name_by_id(tehsils: list[dict], tehsil_id: str) -> str:
    return next(
        (str(row.get("name") or "") for row in tehsils if str(row.get("id") or "") == tehsil_id),
        "",
    )


def draw_existing_group_scope_selector(
    state: GuiState,
    group: dict,
    tehsils: list[dict],
    markazes: list[dict],
) -> None:
    route_kind = str(group.get("route_kind") or "district").strip().lower()
    group_id = str(group.get("id") or "")
    if route_kind == "district":
        imgui.text("District")
        return
    if route_kind == "fallback":
        imgui.text("Fallback")
        return
    if route_kind == "custom":
        imgui.text_disabled("Custom")
        return

    tehsil_options = [("", "Select tehsil"), *[(str(row["id"]), str(row["name"])) for row in tehsils]]
    tehsil_ref = str(group.get("tehsil_ref") or "")
    tehsil_index = _option_index(tehsil_options, tehsil_ref)
    imgui.set_next_item_width(165)
    changed, tehsil_index = imgui.combo(
        "##existing_tehsil",
        tehsil_index,
        [label for _, label in tehsil_options],
    )
    if changed:
        selected_tehsil = tehsil_options[tehsil_index][0]
        group["tehsil_ref"] = selected_tehsil
        group["tehsil"] = _tehsil_name_by_id(tehsils, selected_tehsil)
        group["markaz_ref"] = ""
        group["markaz"] = ""
        _record_dispatch_action_result(
            state,
            *update_antidengue_dispatch_group(
                state,
                group_id,
                {"tehsil_ref": selected_tehsil, "markaz_ref": ""},
            ),
        )

    if route_kind != "markaz":
        return

    selected_tehsil = str(group.get("tehsil_ref") or "")
    scoped_markazes = [
        row
        for row in markazes
        if not selected_tehsil or str(row.get("tehsil_ref") or "") == selected_tehsil
    ]
    markaz_options = [
        ("", "Select markaz"),
        *[
            (
                str(row["id"]),
                f"{row.get('tehsil') or '-'} / {row.get('name') or '-'}",
            )
            for row in scoped_markazes
        ],
    ]
    markaz_ref = str(group.get("markaz_ref") or "")
    markaz_index = _option_index(markaz_options, markaz_ref)
    imgui.set_next_item_width(230)
    changed, markaz_index = imgui.combo(
        "##existing_markaz",
        markaz_index,
        [label for _, label in markaz_options],
    )
    if changed:
        selected_markaz = markaz_options[markaz_index][0]
        group["markaz_ref"] = selected_markaz
        selected_markaz_row = next(
            (row for row in markazes if str(row.get("id") or "") == selected_markaz),
            {},
        )
        group["markaz"] = str(selected_markaz_row.get("name") or "")
        if selected_markaz_row.get("tehsil_ref"):
            group["tehsil_ref"] = str(selected_markaz_row.get("tehsil_ref") or "")
            group["tehsil"] = str(selected_markaz_row.get("tehsil") or "")
        _record_dispatch_action_result(
            state,
            *update_antidengue_dispatch_group(
                state,
                group_id,
                {
                    "tehsil_ref": group.get("tehsil_ref") or "",
                    "markaz_ref": selected_markaz,
                },
            ),
        )


def draw_antidengue_group_add_form(state: GuiState) -> None:
    tehsils, markazes, scope_error = load_antidengue_scope_options(state)
    if scope_error:
        imgui.text_colored(status_badge_color("warning"), scope_error)
        return

    form = state.app_settings.setdefault(
        "antidengue_new_dispatch_group",
        {
            "name": "",
            "target": "",
            "route_kind": "district",
            "message_mode": "full_report",
            "tehsil_ref": "",
            "markaz_ref": "",
            "enabled": True,
            "attach_excel": True,
            "manual_only": False,
            "delay_ms": "1500",
        },
    )

    imgui.separator_text("Add Group Route")
    imgui.set_next_item_width(260)
    changed, value = imgui.input_text("Group name", str(form.get("name") or ""))
    if changed:
        form["name"] = value
    imgui.same_line()
    imgui.set_next_item_width(300)
    changed, value = imgui.input_text("Group JID", str(form.get("target") or ""))
    if changed:
        form["target"] = value

    route_index = _option_index(DISPATCH_ROUTE_KINDS, str(form.get("route_kind") or "district"))
    imgui.set_next_item_width(160)
    changed, route_index = imgui.combo("Route", route_index, [label for _, label in DISPATCH_ROUTE_KINDS])
    if changed:
        form["route_kind"] = DISPATCH_ROUTE_KINDS[route_index][0]
        if form["route_kind"] == "district":
            form["tehsil_ref"] = ""
            form["markaz_ref"] = ""
    imgui.same_line()
    mode_index = _option_index(DISPATCH_MESSAGE_MODES, str(form.get("message_mode") or "full_report"))
    imgui.set_next_item_width(190)
    changed, mode_index = imgui.combo("Message", mode_index, [label for _, label in DISPATCH_MESSAGE_MODES])
    if changed:
        form["message_mode"] = DISPATCH_MESSAGE_MODES[mode_index][0]

    _draw_group_scope_selector(state, form, tehsils, markazes)

    for key, label in (("enabled", "Enabled"), ("attach_excel", "Attach Excel"), ("manual_only", "Manual only")):
        current = _pb_bool(form.get(key), key != "manual_only")
        changed, current = imgui.checkbox(label, current)
        if changed:
            form[key] = current
        imgui.same_line()
    imgui.new_line()

    with disabled_when(not str(form.get("target") or "").strip()):
        if imgui.button("Add Group Route", (150, 0)):
            success, message = add_antidengue_dispatch_group(state, form)
            _record_dispatch_action_result(state, success, message)
            if success:
                form.update({"name": "", "target": "", "tehsil_ref": "", "markaz_ref": ""})


def draw_antidengue_dispatch_groups(state: GuiState, context: dict) -> None:
    groups, error = load_antidengue_dispatch_groups(state)
    if error:
        imgui.text_colored(status_badge_color("warning"), error)
        return

    summary = context.get("summary") or {}
    outputs = summary.get("outputs") or {}
    excel_path = str(outputs.get("excel_report") or "")
    run_key = str(context.get("run_key") or summary.get("run_started_at") or "")
    tehsils, markazes, scope_error = load_antidengue_scope_options(state)
    if scope_error:
        imgui.text_colored(status_badge_color("warning"), scope_error)

    draw_antidengue_group_add_form(state)
    imgui.separator_text("Group Routes")
    if not groups:
        imgui.text_disabled("No group routes configured yet.")
        return

    if excel_path:
        if imgui.button("Copy Excel Path", (130, 0)):
            imgui.set_clipboard_text(excel_path)
            state.runner.status_message = "Copied latest AntiDengue Excel report path."
            state.runner.append_log(state.runner.status_message)
        imgui.same_line()
        output_dir = Path(excel_path).parent
        with disabled_when(not output_dir.exists()):
            if imgui.button("Open Output Folder", (155, 0)):
                open_desktop_path(state, output_dir, "AntiDengue dispatch output folder")

    if imgui.begin_table("antidengue_dispatch_groups", 10):
        for label in ("Enabled", "Group", "Route", "Scope", "Message", "Excel", "Manual", "Target", "Copy", "Rows"):
            imgui.table_setup_column(label)
        imgui.table_headers_row()
        for index, group in enumerate(groups):
            imgui.push_id(f"dispatch_group_{group.get('id') or index}")
            rows = _rows_for_dispatch_group(group, context.get("rows") or [])
            message = build_antidengue_group_dispatch_message(group, context)
            imgui.table_next_row()

            imgui.table_next_column()
            enabled = _pb_bool(group.get("enabled"))
            changed, enabled = imgui.checkbox("##enabled", enabled)
            if changed:
                _record_dispatch_action_result(
                    state,
                    *update_antidengue_dispatch_group(state, str(group["id"]), {"enabled": enabled}),
                )

            imgui.table_next_column()
            draw_clipped_text(group.get("name"), 28)

            imgui.table_next_column()
            route_index = _option_index(DISPATCH_ROUTE_KINDS, str(group.get("route_kind") or "district"))
            imgui.set_next_item_width(115)
            changed, route_index = imgui.combo("##route", route_index, [label for _, label in DISPATCH_ROUTE_KINDS])
            if changed:
                next_route = DISPATCH_ROUTE_KINDS[route_index][0]
                updates: dict[str, object] = {"route_kind": next_route}
                if next_route in {"district", "fallback", "custom"}:
                    updates.update({"tehsil_ref": "", "markaz_ref": ""})
                    group["tehsil_ref"] = ""
                    group["markaz_ref"] = ""
                    group["tehsil"] = ""
                    group["markaz"] = ""
                elif next_route == "tehsil":
                    updates["markaz_ref"] = ""
                    group["markaz_ref"] = ""
                    group["markaz"] = ""
                group["route_kind"] = next_route
                _record_dispatch_action_result(
                    state,
                    *update_antidengue_dispatch_group(
                        state,
                        str(group["id"]),
                        updates,
                    ),
                )

            imgui.table_next_column()
            draw_existing_group_scope_selector(state, group, tehsils, markazes)

            imgui.table_next_column()
            mode_index = _option_index(DISPATCH_MESSAGE_MODES, str(group.get("message_mode") or "full_report"))
            imgui.set_next_item_width(190)
            changed, mode_index = imgui.combo("##mode", mode_index, [label for _, label in DISPATCH_MESSAGE_MODES])
            if changed:
                _record_dispatch_action_result(
                    state,
                    *update_antidengue_dispatch_group(
                        state,
                        str(group["id"]),
                        {"message_mode": DISPATCH_MESSAGE_MODES[mode_index][0]},
                    ),
                )

            imgui.table_next_column()
            attach_excel = _pb_bool(group.get("attach_excel"), True)
            changed, attach_excel = imgui.checkbox("##excel", attach_excel)
            if changed:
                _record_dispatch_action_result(
                    state,
                    *update_antidengue_dispatch_group(state, str(group["id"]), {"attach_excel": attach_excel}),
                )

            imgui.table_next_column()
            manual_only = _pb_bool(group.get("manual_only"))
            changed, manual_only = imgui.checkbox("##manual", manual_only)
            if changed:
                _record_dispatch_action_result(
                    state,
                    *update_antidengue_dispatch_group(state, str(group["id"]), {"manual_only": manual_only}),
                )

            imgui.table_next_column()
            draw_clipped_text(group.get("target"), 30)

            imgui.table_next_column()
            if imgui.button("Copy", (70, 0)):
                imgui.set_clipboard_text(message)
                record_antidengue_dispatch_event(
                    state,
                    run_key=run_key,
                    channel="group",
                    action="copied",
                    target=str(group.get("target") or ""),
                    recipient_name=str(group.get("name") or ""),
                    status="copied",
                    message=message,
                    payload={"group": group, "row_count": len(rows)},
                )
                state.runner.status_message = f"Copied group message for {group.get('name') or group.get('target') or 'group'}"
                state.runner.append_log(state.runner.status_message)
            imgui.same_line()
            if imgui.button("Sent", (62, 0)):
                record_antidengue_dispatch_event(
                    state,
                    run_key=run_key,
                    channel="group",
                    action="marked_sent",
                    target=str(group.get("target") or ""),
                    recipient_name=str(group.get("name") or ""),
                    status="manual_sent",
                    message=message,
                    payload={"group": group, "row_count": len(rows)},
                )
                state.runner.status_message = f"Marked group dispatch sent for {group.get('name') or group.get('target') or 'group'}"
                state.runner.append_log(state.runner.status_message)

            imgui.table_next_column()
            imgui.text(str(len(rows)))
            imgui.pop_id()
        imgui.end_table()


def draw_antidengue_dispatch_center_section(state: GuiState) -> None:
    settings = draw_antidengue_dispatch_settings_controls(state)
    if settings is None:
        return

    context, context_error = load_antidengue_latest_dispatch_context(state)
    if context_error:
        imgui.text_colored(status_badge_color("warning"), context_error)
        return

    imgui.separator()
    subviews = (
        ("overview", "Overview", 110),
        ("individuals", "Individuals", 125),
        ("groups", "Groups", 100),
    )
    current = state.antidengue_dispatch_view
    if current not in {item[0] for item in subviews}:
        current = "overview"
    for index, (view_id, label, width) in enumerate(subviews):
        if index:
            imgui.same_line()
        selected = current == view_id
        with disabled_when(selected):
            if imgui.button(f"{label}{' *' if selected else ''}", (width, 0)):
                current = view_id
                state.antidengue_dispatch_view = view_id
    imgui.separator()

    if current == "individuals":
        draw_antidengue_dispatch_individuals(state, context)
    elif current == "groups":
        draw_antidengue_dispatch_groups(state, context)
    else:
        draw_antidengue_dispatch_overview(state, context, settings)


def draw_antidengue_health_section(
    summary: dict,
    whatsapp: dict,
    delivery: dict,
    quality_gate: dict,
    filter_info: dict,
    fallback: dict,
    fallback_delivery: dict,
) -> None:
    queued = bool(whatsapp.get("queued"))
    error_text = str(whatsapp.get("error") or "")
    skipped_text = str(whatsapp.get("skipped_reason") or "")
    status_label = (
        "delivered"
        if queued
        else "failed"
        if error_text or skipped_text
        else "queued"
    )

    if imgui.begin_table("antidengue_latest_metrics", 6):
        for label in [
            "Run",
            "Dormant",
            "Quality",
            "Delivered",
            "Failed",
            "Fallback",
        ]:
            imgui.table_setup_column(label)
        imgui.table_headers_row()
        imgui.table_next_row()

        imgui.table_next_column()
        draw_delivery_metric(
            "Started",
            str(summary.get("run_started_at", "-")).replace("T", " ")[:19],
            status_label,
        )

        imgui.table_next_column()
        draw_delivery_metric(
            "Rows",
            filter_info.get("dormant_raw_rows")
            or quality_gate.get("final_school_count")
            or "-",
        )

        imgui.table_next_column()
        draw_delivery_metric(
            "Gate",
            "passed" if quality_gate.get("passed") else "blocked",
            "passed" if quality_gate.get("passed") else "failed",
        )

        imgui.table_next_column()
        draw_delivery_metric(
            "WhatsApp",
            delivery.get("delivered", "-"),
            "delivered" if delivery.get("failed", 0) == 0 else "partial",
        )

        imgui.table_next_column()
        draw_delivery_metric(
            "WhatsApp",
            delivery.get("failed", "-"),
            "failed" if delivery.get("failed", 0) else "ok",
        )

        imgui.table_next_column()
        if fallback:
            fallback_failed = int(fallback_delivery.get("failed") or 0)
            fallback_delivered = int(fallback_delivery.get("delivered") or 0)
            fallback_value = f"{fallback_delivered} ok / {fallback_failed} failed"
            draw_delivery_metric(
                "Escalation",
                fallback_value,
                "failed" if fallback_failed else "delivered",
            )
        else:
            draw_delivery_metric("Escalation", "-", "")

        imgui.end_table()

    if error_text or skipped_text:
        issue_text = error_text or skipped_text
        issue_color = "failed" if error_text else "warning"
        imgui.text_colored(
            status_badge_color(issue_color),
            "What went wrong:" if error_text else "Not sent:",
        )
        imgui.same_line()
        imgui.text_wrapped(issue_text)
    else:
        imgui.text_colored(status_badge_color("ok"), "Latest run completed without delivery errors.")

    warnings = quality_gate.get("warnings") or []
    if warnings:
        imgui.text_colored(status_badge_color("warning"), "Data warnings:")
        for warning in warnings[:3]:
            imgui.bullet_text(str(warning))


def failed_officer_rows(audit_rows: list[dict]) -> list[dict]:
    return [
        row
        for row in audit_rows
        if str(row.get("Delivery Status", "")).strip().lower() != "delivered"
    ]


def failed_delivery_statuses(delivery: dict) -> list[dict]:
    return [
        item
        for item in (delivery.get("statuses") or {}).values()
        if str(item.get("status", "")).strip().lower() != "delivered"
    ]


def draw_antidengue_failed_officers_section(
    state: GuiState,
    audit_rows: list[dict],
    delivery: dict,
) -> None:
    failed_rows = [
        row
        for row in audit_rows
        if str(row.get("Delivery Status", "")).strip().lower() != "delivered"
    ]

    if not failed_rows:
        imgui.text_colored(status_badge_color("ok"), "No failed officer direct deliveries in the latest run.")
        return

    if imgui.begin_table("antidengue_failed_officers", 8):
        imgui.table_setup_column("Officer")
        imgui.table_setup_column("Role")
        imgui.table_setup_column("Schools")
        imgui.table_setup_column("Mobile")
        imgui.table_setup_column("WhatsApp")
        imgui.table_setup_column("Token")
        imgui.table_setup_column("Manual")
        imgui.table_setup_column("Cause")
        imgui.table_headers_row()
        for index, row in enumerate(failed_rows[:30]):
            imgui.push_id(f"failed_officer_{index}")
            target = str(row.get("Target") or "")
            error_value = str(row.get("Delivery Error") or "")
            ack_value = str(row.get("WhatsApp Ack Statuses") or "")
            token_status = antidengue_token_status_for_mobile(
                state,
                row.get("Mobile Number"),
            )
            imgui.table_next_row()

            imgui.table_next_column()
            imgui.text_wrapped(str(row.get("Officer Name") or "-"))
            imgui.table_next_column()
            imgui.text(str(row.get("Roles") or "-"))
            imgui.table_next_column()
            imgui.text(str(row.get("School Count") or "-"))
            imgui.table_next_column()
            imgui.text(str(row.get("Mobile Number") or "-"))
            imgui.table_next_column()
            imgui.text_colored(status_badge_color("failed"), ack_value or "failed")
            imgui.table_next_column()
            token_color_key = {
                "usable": "ok",
                "expired": "warning",
                "missing": "failed",
            }.get(str(token_status.get("status") or ""), "")
            imgui.text_colored(
                status_badge_color(token_color_key),
                str(token_status.get("label") or "-"),
            )
            if imgui.is_item_hovered():
                detail_parts = [
                    str(token_status.get("detail") or ""),
                    f"LID: {token_status.get('lid') or '-'}",
                ]
                imgui.set_tooltip("\n".join(part for part in detail_parts if part))
            imgui.table_next_column()
            if imgui.button("Copy##manual_msg", (70, 0)):
                imgui.set_clipboard_text(build_antidengue_manual_message(row))
                state.runner.status_message = (
                    "Copied manual WhatsApp message for "
                    f"{row.get('Officer Name') or row.get('Mobile Number') or 'officer'}"
                )
                state.runner.append_log(state.runner.status_message)
            imgui.table_next_column()
            imgui.text_wrapped(delivery_failure_cause(error_value, ack_value, target))
            imgui.pop_id()
        imgui.end_table()
    if len(failed_rows) > 30:
        imgui.text_disabled(f"Showing 30 of {len(failed_rows)} failed officer rows. Open the delivery audit for all rows.")


def draw_antidengue_group_delivery_section(delivery: dict, fallback: dict) -> None:
    failed_statuses = failed_delivery_statuses(delivery)
    fixed_failures = [
        item
        for item in failed_statuses
        if str(item.get("target") or "").endswith("@g.us")
    ]
    fallback_delivery = fallback.get("delivery") or {}
    fallback_statuses = [
        item
        for item in (fallback_delivery.get("statuses") or {}).values()
        if str(item.get("target") or "").endswith("@g.us")
    ]
    rows = [*fixed_failures, *fallback_statuses]
    if not rows:
        imgui.text_colored(status_badge_color("ok"), "No fixed/group delivery failures in the latest run.")
        return

    if imgui.begin_table("antidengue_group_failures", 4):
        imgui.table_setup_column("Target")
        imgui.table_setup_column("Status")
        imgui.table_setup_column("Messages")
        imgui.table_setup_column("Cause")
        imgui.table_headers_row()
        for index, item in enumerate(rows[:12]):
            imgui.push_id(f"group_delivery_{index}")
            target = str(item.get("target") or "")
            error_value = str(item.get("error") or "")
            operation_result = item.get("operationResult") or []
            imgui.table_next_row()
            imgui.table_next_column()
            imgui.text_colored(status_badge_color("failed"), target)
            imgui.table_next_column()
            imgui.text(str(item.get("status") or "failed"))
            imgui.table_next_column()
            imgui.text(str(len(operation_result)) if isinstance(operation_result, list) else "-")
            imgui.table_next_column()
            imgui.text_wrapped(delivery_failure_cause(error_value, "PENDING", target))
            imgui.pop_id()
        imgui.end_table()


def draw_antidengue_files_section(
    state: GuiState,
    summary_path: Path | None,
    run_dir: Path | None,
    outputs: dict,
) -> None:
    if run_dir:
        if imgui.button("Open Run Folder", (150, 0)):
            open_desktop_path(state, run_dir, "AntiDengue run folder")
        imgui.same_line()
    if summary_path and imgui.button("Open Summary JSON", (165, 0)):
        open_desktop_path(state, summary_path, "AntiDengue run summary")
    imgui.same_line()
    delivery_audit_value = str(outputs.get("officer_delivery_audit") or "").strip()
    delivery_audit = Path(delivery_audit_value) if delivery_audit_value else None
    with disabled_when(delivery_audit is None or not delivery_audit.exists()):
        if imgui.button("Open Delivery Audit", (170, 0)):
            assert delivery_audit is not None
            open_desktop_path(state, delivery_audit, "Officer delivery audit")


def draw_antidengue_manual_report_section(state: GuiState) -> None:
    command = command_by_id("antidengue_manual_file")
    values = state.command_values(command)
    for field in command.arg_fields:
        values.setdefault(field.key, field.default)

    imgui.text_wrapped(
        "Use this when portal scraping is broken or unavailable. Select the report you downloaded manually; the app will copy it into AntiDengue staging and run the normal processing, audit, and WhatsApp delivery workflow."
    )
    imgui.separator()

    for field in command.arg_fields:
        draw_field(state, command, values, field)

    selected_file = resolve_gui_path(command.working_dir(state.project_root), values.get("file", ""))
    file_ok = selected_file.is_file()
    if values.get("file", "").strip():
        if file_ok:
            size_kb = selected_file.stat().st_size / 1024
            imgui.text_colored(status_badge_color("ok"), f"Ready: {selected_file.name} ({size_kb:.1f} KB)")
        else:
            imgui.text_colored(status_badge_color("failed"), f"File not found: {selected_file}")
    else:
        imgui.text_disabled("Select the downloaded portal file first.")

    imgui.separator()
    draw_antidengue_worker_controls(state)

    with disabled_when(state.runner.is_running or state.command_start_in_progress or not file_ok):
        if imgui.button("Process Manual Report & Send", (230, 0)):
            start_command_with_services(state, command, values)
    imgui.same_line()
    with disabled_when(state.runner.is_running or state.command_start_in_progress or not file_ok):
        if imgui.button("Dry Run Manual Report", (175, 0)):
            run_values = dict(values)
            run_values["dry_run"] = "true"
            start_command_with_services(state, command, run_values)

    imgui.text_disabled(
        "The selected file is copied before processing, so the original download remains untouched."
    )


def _draw_master_owner_preview(owners: dict[str, list[dict]], owner_error: str) -> None:
    imgui.separator_text("Officer Preview")
    if owner_error:
        imgui.text_colored(status_badge_color("warning"), owner_error)
        return

    if imgui.begin_table("antidengue_master_owner_preview", 4):
        for label in ("Role", "Status", "Officer", "Mobile"):
            imgui.table_setup_column(label)
        imgui.table_headers_row()
        for role, rows in (("DDEO", owners.get("ddeo") or []), ("AEO", owners.get("aeo") or [])):
            status = "ok" if len(rows) == 1 else "failed"
            if not rows:
                display_rows = [{"name": "-", "normalized_mobile": "-"}]
            else:
                display_rows = rows
            for index, row in enumerate(display_rows):
                imgui.push_id(f"master_owner_{role}_{index}")
                imgui.table_next_row()
                imgui.table_next_column()
                imgui.text(role)
                imgui.table_next_column()
                if index == 0:
                    label = "ready" if len(rows) == 1 else "missing" if not rows else "duplicate"
                    imgui.text_colored(status_badge_color(status), label)
                else:
                    imgui.text_colored(status_badge_color("failed"), "duplicate")
                imgui.table_next_column()
                draw_clipped_text(row.get("name"), 38)
                imgui.table_next_column()
                imgui.text(str(row.get("normalized_mobile") or "-"))
                imgui.pop_id()
        imgui.end_table()


def _draw_master_school_choice_fields(
    state: GuiState,
    form: dict[str, object],
    options: dict[str, list[dict]],
) -> None:
    districts = options.get("districts") or []
    departments = options.get("departments") or []
    wings = options.get("wings") or []
    tehsils = options.get("tehsils") or []
    markazes = options.get("markazes") or []

    _set_single_default(form, "district_ref", districts)
    _set_single_default(form, "department_ref", departments)

    selected_district = str(form.get("district_ref") or "")
    selected_department = str(form.get("department_ref") or "")
    filtered_wings = [
        row
        for row in wings
        if (not selected_district or str(row.get("district_ref") or "") == selected_district)
        and (not selected_department or str(row.get("department_ref") or "") == selected_department)
    ]
    _set_single_default(form, "wing_ref", filtered_wings)

    selected_wing = str(form.get("wing_ref") or "")
    filtered_tehsils = [
        row
        for row in tehsils
        if not selected_district or str(row.get("district_ref") or "") == selected_district
    ]
    selected_tehsil = str(form.get("tehsil_ref") or "")
    filtered_markazes = [
        row
        for row in markazes
        if selected_tehsil
        and selected_wing
        and str(row.get("tehsil_ref") or "") == selected_tehsil
        and str(row.get("wing_ref") or "") == selected_wing
    ]

    if _draw_relation_combo(
        "District##master_school_district",
        "district_ref",
        form,
        _relation_options(districts, "Select district"),
        width=250,
    ):
        form["wing_ref"] = ""
        form["tehsil_ref"] = ""
        form["markaz_ref"] = ""
    imgui.same_line()
    if _draw_relation_combo(
        "Department##master_school_department",
        "department_ref",
        form,
        _relation_options(departments, "Select department"),
        width=320,
    ):
        form["wing_ref"] = ""
        form["markaz_ref"] = ""

    if _draw_relation_combo(
        "Wing##master_school_wing",
        "wing_ref",
        form,
        _relation_options(filtered_wings, "Select wing"),
        width=250,
    ):
        form["markaz_ref"] = ""
    imgui.same_line()
    if _draw_relation_combo(
        "Tehsil##master_school_tehsil",
        "tehsil_ref",
        form,
        _relation_options(filtered_tehsils, "Select tehsil"),
        width=250,
    ):
        form["markaz_ref"] = ""
    imgui.same_line()
    _draw_relation_combo(
        "Markaz##master_school_markaz",
        "markaz_ref",
        form,
        _relation_options(filtered_markazes, "Select markaz", include_parent=True),
        width=360,
    )


def draw_antidengue_master_data_section(state: GuiState) -> None:
    options, error = load_antidengue_master_data_options(state)
    if error:
        imgui.text_colored(status_badge_color("warning"), error)
        return

    form = antidengue_new_school_form(state)
    imgui.separator_text("Add School")

    imgui.set_next_item_width(160)
    changed, value = imgui.input_text("EMIS##master_school_emis", str(form.get("emis") or ""))
    if changed:
        form["emis"] = _normalized_emis(value)
    imgui.same_line()
    imgui.set_next_item_width(520)
    changed, value = imgui.input_text("School name##master_school_name", str(form.get("name") or ""))
    if changed:
        form["name"] = value
        if not form.get("school_level"):
            form["school_level"] = _school_level_from_name_gui(value) or "Primary"

    _draw_master_school_choice_fields(state, form, options)

    shift_index = _option_index([(item, item) for item in MASTER_SCHOOL_SHIFTS], str(form.get("shift") or "Single"))
    imgui.set_next_item_width(160)
    changed, shift_index = imgui.combo("Shift##master_school_shift", shift_index, list(MASTER_SCHOOL_SHIFTS))
    if changed:
        form["shift"] = MASTER_SCHOOL_SHIFTS[shift_index]
    imgui.same_line()

    type_index = _option_index([(item, item) for item in MASTER_SCHOOL_TYPES], str(form.get("school_type") or "Male"))
    imgui.set_next_item_width(180)
    changed, type_index = imgui.combo("Type##master_school_type", type_index, list(MASTER_SCHOOL_TYPES))
    if changed:
        form["school_type"] = MASTER_SCHOOL_TYPES[type_index]
    imgui.same_line()

    level_index = _option_index([(item, item) for item in MASTER_SCHOOL_LEVELS], str(form.get("school_level") or "Primary"))
    imgui.set_next_item_width(190)
    changed, level_index = imgui.combo("Level##master_school_level", level_index, list(MASTER_SCHOOL_LEVELS))
    if changed:
        form["school_level"] = MASTER_SCHOOL_LEVELS[level_index]
    imgui.same_line()

    imgui.set_next_item_width(130)
    changed, value = imgui.input_text("DEOs wise##master_school_deos_wise", str(form.get("deos_wise") or "M-EE"))
    if changed:
        form["deos_wise"] = value

    imgui.set_next_item_width(300)
    changed, value = imgui.input_text("Head name##master_school_head_name", str(form.get("head_name") or ""))
    if changed:
        form["head_name"] = value
    imgui.same_line()
    imgui.set_next_item_width(230)
    changed, value = imgui.input_text("Head contact##master_school_head_contact", str(form.get("head_contact") or ""))
    if changed:
        form["head_contact"] = value
    imgui.same_line()
    changed, active = imgui.checkbox("Active##master_school_active", _pb_bool(form.get("active"), True))
    if changed:
        form["active"] = active

    changed, value = imgui.input_text_multiline(
        "Notes##master_school_notes",
        str(form.get("notes") or ""),
        (760, 70),
    )
    if changed:
        form["notes"] = value

    errors, owners, owner_error = validate_antidengue_new_school_form(state, form)
    _draw_master_owner_preview(owners, owner_error)

    if errors:
        imgui.separator_text("Validation")
        for index, message in enumerate(errors[:6]):
            imgui.push_id(f"master_school_error_{index}")
            imgui.text_colored(status_badge_color("warning"), message)
            imgui.pop_id()
        if len(errors) > 6:
            imgui.text_disabled(f"{len(errors) - 6} more validation issue(s).")

    imgui.separator()
    with disabled_when(bool(errors)):
        if imgui.button("Save School", (130, 0)):
            success, message = create_antidengue_school_from_form(state, form)
            form["last_result"] = message
            form["last_success"] = success
            _record_dispatch_action_result(state, success, message)
            if success:
                form["last_saved_emis"] = _normalized_emis(form.get("emis"))
                for key in ("emis", "name", "head_name", "head_contact", "notes"):
                    form[key] = ""
    imgui.same_line()
    if imgui.button("Clear Form", (120, 0)):
        preserved = {
            key: form.get(key)
            for key in ("district_ref", "department_ref", "wing_ref", "tehsil_ref", "markaz_ref")
        }
        form.clear()
        form.update(_master_school_form_defaults())
        form.update(preserved)
    result = str(form.get("last_result") or "")
    if result:
        imgui.same_line()
        color_key = "ok" if _pb_bool(form.get("last_success")) else "failed"
        imgui.text_colored(status_badge_color(color_key), result)


def draw_antidengue_database_section(state: GuiState) -> None:
    root = antidengue_pocketbase_root(state)
    service = state.services.services.get("antidengue_pocketbase")
    migrate_command = command_by_id("antidengue_pocketbase_migrate")
    import_command = command_by_id("antidengue_pocketbase_import_officers")

    imgui.text_wrapped(
        "PocketBase stores AntiDengue schools, officers, jurisdictions, fixed WhatsApp recipients, report runs, delivery events, and data-quality issues in one local database with an admin dashboard."
    )
    imgui.separator()

    if service:
        imgui.text("Database service:")
        imgui.same_line()
        imgui.text_colored(
            service_status_color(service.is_running, service.last_exit_code),
            service.status_message,
        )
        imgui.same_line()
        imgui.text_disabled("http://127.0.0.1:8090/_/")

        with disabled_when(service.is_running):
            if imgui.button("Start Database", (140, 0)):
                state.services.start("antidengue_pocketbase")
        imgui.same_line()
        with disabled_when(not service.is_running):
            if imgui.button("Stop Database", (130, 0)):
                state.services.stop("antidengue_pocketbase")
        imgui.same_line()
        if imgui.button("Open Admin", (120, 0)):
            open_url_from_gui(state, "http://127.0.0.1:8090/_/")
    else:
        imgui.text_colored(status_badge_color("failed"), "PocketBase service is not registered.")

    imgui.same_line()
    if imgui.button("Open DB Folder", (140, 0)):
        open_desktop_path(state, root, "AntiDengue PocketBase folder")

    imgui.separator_text("Schema and Sync")
    with disabled_when(state.runner.is_running):
        if imgui.button("Apply Migrations", (155, 0)):
            start_command_with_services(state, migrate_command, {})
    imgui.same_line()
    with disabled_when(state.runner.is_running):
        if imgui.button("Sync DB Data", (155, 0)):
            start_command_with_services(state, import_command, {})
    imgui.same_line()
    imgui.text_disabled("officers_list.csv + whatsapp_recipients.csv")

    counts, error = antidengue_pocketbase_counts(state)
    if error:
        imgui.text_colored(status_badge_color("warning"), error)
        return

    if imgui.begin_table("antidengue_pocketbase_counts", 7):
        for label in ("Schools", "DDEOs", "AEOs", "Jurisdictions", "Overrides", "Recipients", "Import batches"):
            imgui.table_setup_column(label)
        imgui.table_headers_row()
        imgui.table_next_row()
        values = (
            counts.get("schools", 0),
            counts.get("ddeo_officers", 0),
            counts.get("aeo_officers", 0),
            counts.get("ddeo_jurisdictions", 0)
            + counts.get("aeo_jurisdictions", 0),
            counts.get("school_ddeo_overrides", 0)
            + counts.get("school_aeo_overrides", 0),
            counts.get("whatsapp_recipients", 0),
            counts.get("officer_import_batches", 0),
        )
        for value in values:
            imgui.table_next_column()
            imgui.text(str(value))
        imgui.end_table()

        if imgui.begin_table("antidengue_pocketbase_audit_counts", 4):
            for label in ("Report runs", "Dormant rows", "Delivery events", "Data-quality issues"):
                imgui.table_setup_column(label)
            imgui.table_headers_row()
            imgui.table_next_row()
            for key in ("report_runs", "dormant_records", "delivery_events", "data_quality_issues"):
                imgui.table_next_column()
                imgui.text(str(counts.get(key, 0)))
            imgui.end_table()

        if imgui.begin_table("antidengue_pocketbase_dispatch_counts", 4):
            for label in ("Settings", "Groups", "Templates", "Dispatch events"):
                imgui.table_setup_column(label)
            imgui.table_headers_row()
            imgui.table_next_row()
            for key in ("dispatch_settings", "whatsapp_groups", "message_templates", "dispatch_events"):
                imgui.table_next_column()
                imgui.text(str(counts.get(key, 0)))
            imgui.end_table()

    imgui.separator_text("Database Health Checks")
    health_checks, health_error = antidengue_pocketbase_health_checks(state)
    if health_error:
        imgui.text_colored(status_badge_color("warning"), health_error)
        return

    has_problem = any(
        str(row.get("severity") or "").lower() in {"failed", "warning"}
        for row in health_checks
    )
    if not has_problem:
        imgui.text_colored(status_badge_color("ok"), "All database health checks passed.")

    if imgui.begin_table("antidengue_pocketbase_health_checks", 4):
        for label in ("Status", "Check", "Count", "Message"):
            imgui.table_setup_column(label)
        imgui.table_headers_row()
        for index, row in enumerate(health_checks):
            imgui.push_id(f"db_health_{index}")
            severity = str(row.get("severity") or "")
            imgui.table_next_row()
            imgui.table_next_column()
            imgui.text_colored(status_badge_color(severity), severity)
            imgui.table_next_column()
            imgui.text(str(row.get("name") or "-"))
            imgui.table_next_column()
            imgui.text(str(row.get("count") or 0))
            imgui.table_next_column()
            imgui.text_wrapped(str(row.get("message") or "-"))
            imgui.pop_id()
        imgui.end_table()


def _format_history_time(value: object) -> str:
    return str(value or "-").replace("T", " ")[:19]


def draw_clipped_text(value: object, max_chars: int = 48) -> None:
    text = str(value or "-")
    normalized = " ".join(part.strip() for part in text.splitlines() if part.strip())
    if not normalized:
        normalized = "-"
    clipped = normalized
    if len(clipped) > max_chars:
        clipped = clipped[: max_chars - 3].rstrip() + "..."
    imgui.text(clipped)
    if clipped != normalized and imgui.is_item_hovered():
        imgui.set_tooltip(normalized[:2000])


def _history_summary_output_dir(row: dict) -> Path | None:
    summary = row.get("summary_dict") or {}
    outputs = summary.get("outputs") or {}
    output_dir = str(outputs.get("output_dir") or "").strip()
    return Path(output_dir) if output_dir else None


def _run_is_dry(row: dict) -> bool:
    summary = row.get("summary_dict") or {}
    whatsapp = summary.get("whatsapp") or {}
    return bool(summary.get("dry_run") or whatsapp.get("dry_run"))


def _run_mode_label(row: dict) -> str:
    return "Dry Run" if _run_is_dry(row) else "Live Send"


def draw_history_filter_button(state: GuiState, value: str, label: str, width: int) -> None:
    selected = state.antidengue_history_filter == value
    with disabled_when(selected):
        if imgui.button(f"{label}{' *' if selected else ''}", (width, 0)):
            state.antidengue_history_filter = value
            state.antidengue_history_selected_run_key = ""
            state.antidengue_history_pending_delete_run_key = ""


def _record_history_action_result(state: GuiState, success: bool, message: str) -> None:
    state.runner.status_message = message
    state.runner.append_log(message)
    if success:
        state.cached_metrics.pop("antidengue", None)


def draw_antidengue_run_history_section(state: GuiState) -> None:
    draw_history_filter_button(state, "active", "Active", 90)
    imgui.same_line()
    draw_history_filter_button(state, "archived", "Archived", 105)
    imgui.same_line()
    draw_history_filter_button(state, "all", "All", 70)
    imgui.same_line()
    if state.antidengue_history_pending_delete_run_key == "__clear_archived__":
        if imgui.button("Confirm Clear Archived", (185, 0)):
            _record_history_action_result(state, *clear_archived_antidengue_runs(state))
        imgui.same_line()
        if imgui.button("Cancel", (80, 0)):
            state.antidengue_history_pending_delete_run_key = ""
    else:
        if imgui.button("Clear Archived", (140, 0)):
            state.antidengue_history_pending_delete_run_key = "__clear_archived__"
    imgui.text_disabled(
        "Archive hides runs from Active. Delete/Clear remove PocketBase history records only; output files remain on disk."
    )

    history, error = load_antidengue_run_history(state)
    if error:
        imgui.text_colored(status_badge_color("warning"), error)
        return
    if not history:
        imgui.text_disabled("No PocketBase run history has been saved yet.")
        return

    if not state.antidengue_history_selected_run_key:
        state.antidengue_history_selected_run_key = str(history[0].get("run_key") or "")

    selected = next(
        (
            row
            for row in history
            if str(row.get("run_key") or "") == state.antidengue_history_selected_run_key
        ),
        history[0],
    )
    state.antidengue_history_selected_run_key = str(selected.get("run_key") or "")

    imgui.text_disabled(
        f"Showing latest {len(history)} PocketBase run(s) in {state.antidengue_history_filter} view."
    )
    imgui.same_line()
    if imgui.button("Select Latest", (120, 0)):
        state.antidengue_history_selected_run_key = str(history[0].get("run_key") or "")
        selected = history[0]

    if imgui.begin_table("antidengue_run_history", 13):
        for label in (
            "Run",
            "Status",
            "Mode",
            "Source",
            "Dormant",
            "Delivered",
            "Failed",
            "Issues",
            "Stages",
            "Plan",
            "Archived",
            "Raw File",
            "Action",
        ):
            imgui.table_setup_column(label)
        imgui.table_headers_row()

        for index, row in enumerate(history):
            run_key = str(row.get("run_key") or "")
            is_selected = run_key == state.antidengue_history_selected_run_key
            imgui.push_id(f"history_{index}")
            imgui.table_next_row()

            imgui.table_next_column()
            imgui.text_colored(
                status_badge_color("ok" if is_selected else ""),
                _format_history_time(row.get("started_at")),
            )
            imgui.table_next_column()
            status = str(row.get("status") or "-")
            imgui.text_colored(status_badge_color(status), status)
            imgui.table_next_column()
            mode_label = _run_mode_label(row)
            imgui.text_colored(
                status_badge_color("warning" if _run_is_dry(row) else "ok"),
                mode_label,
            )
            imgui.table_next_column()
            imgui.text(str(row.get("source") or "-"))
            imgui.table_next_column()
            imgui.text(str(row.get("dormant_count") or 0))
            imgui.table_next_column()
            imgui.text_colored(status_badge_color("delivered"), str(row.get("delivered_count") or 0))
            imgui.table_next_column()
            failed_count = int(row.get("failed_count") or 0)
            imgui.text_colored(
                status_badge_color("failed" if failed_count else "ok"),
                str(failed_count),
            )
            imgui.table_next_column()
            issue_count = int(row.get("issue_count") or 0)
            imgui.text_colored(
                status_badge_color("warning" if issue_count else "ok"),
                str(issue_count),
            )
            imgui.table_next_column()
            imgui.text(str(row.get("stage_count") or 0))
            imgui.table_next_column()
            imgui.text(str(row.get("dispatch_plan_count") or 0))
            imgui.table_next_column()
            archived = bool(row.get("archived"))
            imgui.text_colored(
                status_badge_color("warning" if archived else "ok"),
                "yes" if archived else "no",
            )
            imgui.table_next_column()
            draw_clipped_text(row.get("raw_file_name"), 34)
            imgui.table_next_column()
            with disabled_when(is_selected):
                if imgui.button("View", (70, 0)):
                    state.antidengue_history_selected_run_key = run_key
                    selected = row
            imgui.same_line()
            archive_label = "Restore" if archived else "Archive"
            if imgui.button(archive_label, (82, 0)):
                success, message = set_antidengue_run_archived(
                    state,
                    run_key,
                    archived=not archived,
                )
                _record_history_action_result(state, success, message)
                if success and state.antidengue_history_filter != "all":
                    state.antidengue_history_selected_run_key = ""
            imgui.pop_id()
        imgui.end_table()

    details, detail_error = load_antidengue_run_details(
        state,
        state.antidengue_history_selected_run_key,
    )
    if detail_error:
        imgui.text_colored(status_badge_color("warning"), detail_error)
        return

    summary = selected.get("summary_dict") or {}
    quality_gate = summary.get("quality_gate") or {}
    whatsapp = summary.get("whatsapp") or {}
    outputs = summary.get("outputs") or {}

    imgui.separator_text("Selected Run")
    if imgui.begin_table("antidengue_history_selected_metrics", 6):
        for label in ("Started", "Status", "Mode", "Raw rows", "Dormant", "Quality"):
            imgui.table_setup_column(label)
        imgui.table_headers_row()
        imgui.table_next_row()
        imgui.table_next_column()
        imgui.text(_format_history_time(selected.get("started_at")))
        imgui.table_next_column()
        imgui.text_colored(status_badge_color(str(selected.get("status") or "")), str(selected.get("status") or "-"))
        imgui.table_next_column()
        imgui.text_colored(
            status_badge_color("warning" if _run_is_dry(selected) else "ok"),
            _run_mode_label(selected),
        )
        imgui.table_next_column()
        filter_info = summary.get("filter") or {}
        imgui.text(str(filter_info.get("raw_rows") or "-"))
        imgui.table_next_column()
        imgui.text(str(selected.get("dormant_count") or 0))
        imgui.table_next_column()
        imgui.text_colored(
            status_badge_color("passed" if quality_gate.get("passed") else "failed"),
            "passed" if quality_gate.get("passed") else "blocked",
        )
        imgui.end_table()

    selected_run_key = str(selected.get("run_key") or "")
    selected_archived = bool(selected.get("archived"))
    if imgui.button("Restore Run" if selected_archived else "Archive Run", (125, 0)):
        success, message = set_antidengue_run_archived(
            state,
            selected_run_key,
            archived=not selected_archived,
        )
        _record_history_action_result(state, success, message)
        if success and state.antidengue_history_filter != "all":
            state.antidengue_history_selected_run_key = ""
            return
    imgui.same_line()
    if state.antidengue_history_pending_delete_run_key == selected_run_key:
        if imgui.button("Confirm Delete", (135, 0)):
            _record_history_action_result(
                state,
                *delete_antidengue_run_history(state, selected_run_key),
            )
            return
        imgui.same_line()
        if imgui.button("Cancel Delete", (125, 0)):
            state.antidengue_history_pending_delete_run_key = ""
    else:
        if imgui.button("Delete Run", (110, 0)):
            state.antidengue_history_pending_delete_run_key = selected_run_key

    output_dir = _history_summary_output_dir(selected)
    if output_dir:
        with disabled_when(not output_dir.exists()):
            if imgui.button("Open Output Folder", (160, 0)):
                open_desktop_path(state, output_dir, "AntiDengue history output folder")
        imgui.same_line()
    delivery_audit_value = str(outputs.get("officer_delivery_audit") or "").strip()
    delivery_audit = Path(delivery_audit_value) if delivery_audit_value else None
    with disabled_when(delivery_audit is None or not delivery_audit.exists()):
        if imgui.button("Open Delivery Audit", (160, 0)):
            assert delivery_audit is not None
            open_desktop_path(state, delivery_audit, "AntiDengue history delivery audit")

    lifecycle_rows = details.get("lifecycle") or []
    imgui.separator_text("Run Lifecycle")
    if not lifecycle_rows:
        imgui.text_disabled("No lifecycle events saved for this run.")
    elif imgui.begin_table("antidengue_history_lifecycle", 4):
        for label in ("Stage", "Status", "Message", "Time"):
            imgui.table_setup_column(label)
        imgui.table_headers_row()
        for index, row in enumerate(lifecycle_rows):
            imgui.push_id(f"history_lifecycle_{index}")
            status = str(row.get("status") or "")
            imgui.table_next_row()
            imgui.table_next_column()
            imgui.text(str(row.get("stage") or "-").replace("_", " "))
            imgui.table_next_column()
            imgui.text_colored(status_badge_color(status), status.replace("_", " ") or "-")
            imgui.table_next_column()
            draw_clipped_text(row.get("message"), 96)
            imgui.table_next_column()
            draw_clipped_text(str(row.get("created_at") or ""), 24)
            imgui.pop_id()
        imgui.end_table()
        lifecycle_total = int(selected.get("stage_count") or len(lifecycle_rows))
        if lifecycle_total > len(lifecycle_rows):
            imgui.text_disabled(
                f"Showing first {len(lifecycle_rows)} of {lifecycle_total} lifecycle events."
            )

    dispatch_plan_rows = details.get("dispatch_plan") or []
    imgui.separator_text("Dispatch Plan")
    if not dispatch_plan_rows:
        imgui.text_disabled("No dispatch plan saved for this run.")
    elif imgui.begin_table("antidengue_history_dispatch_plan", 9):
        for label in (
            "Channel",
            "Recipient",
            "Target",
            "Route",
            "Message",
            "Rows",
            "Status",
            "Copy",
            "Cause",
        ):
            imgui.table_setup_column(label)
        imgui.table_headers_row()
        for index, row in enumerate(dispatch_plan_rows):
            imgui.push_id(f"history_dispatch_plan_{index}")
            status = str(row.get("status") or "")
            payload = row.get("payload_dict") or {}
            planned_text = str(payload.get("text") or "")
            imgui.table_next_row()
            imgui.table_next_column()
            imgui.text(str(row.get("channel") or "-"))
            imgui.table_next_column()
            draw_clipped_text(row.get("recipient_name"), 30)
            imgui.table_next_column()
            draw_clipped_text(row.get("target"), 34)
            imgui.table_next_column()
            route_kind = str(row.get("route_kind") or "-").replace("_", " ")
            imgui.text(route_kind)
            imgui.table_next_column()
            message_mode = str(row.get("message_mode") or "-").replace("_", " ")
            draw_clipped_text(message_mode, 28)
            imgui.table_next_column()
            imgui.text(str(row.get("row_count") or 0))
            imgui.table_next_column()
            imgui.text_colored(status_badge_color(status), status.replace("_", " ") or "-")
            imgui.table_next_column()
            with disabled_when(not planned_text):
                if imgui.button("Copy", (70, 0)):
                    imgui.set_clipboard_text(planned_text)
                    state.runner.status_message = (
                        "Copied planned dispatch message for "
                        f"{row.get('recipient_name') or row.get('target') or 'recipient'}"
                    )
                    state.runner.append_log(state.runner.status_message)
            imgui.table_next_column()
            draw_clipped_text(row.get("cause"), 46)
            imgui.pop_id()
        imgui.end_table()
        dispatch_plan_total = int(selected.get("dispatch_plan_count") or len(dispatch_plan_rows))
        if dispatch_plan_total > len(dispatch_plan_rows):
            imgui.text_disabled(
                f"Showing first {len(dispatch_plan_rows)} of {dispatch_plan_total} dispatch plan items."
            )

    quality_rows = details.get("quality") or []
    if quality_rows:
        imgui.separator_text("Quality Issues")
        if imgui.begin_table("antidengue_history_quality", 3):
            imgui.table_setup_column("Severity")
            imgui.table_setup_column("Category")
            imgui.table_setup_column("Message")
            imgui.table_headers_row()
            for index, row in enumerate(quality_rows):
                imgui.push_id(f"history_quality_{index}")
                imgui.table_next_row()
                imgui.table_next_column()
                severity = str(row.get("severity") or "")
                imgui.text_colored(status_badge_color(severity), severity)
                imgui.table_next_column()
                imgui.text(str(row.get("category") or "-"))
                imgui.table_next_column()
                draw_clipped_text(row.get("message"), 90)
                imgui.pop_id()
            imgui.end_table()

    delivery_rows = details.get("delivery") or []
    imgui.separator_text("Delivery Events")
    if not delivery_rows:
        imgui.text_disabled("No delivery events saved for this run.")
    elif imgui.begin_table("antidengue_history_delivery", 7):
        for label in ("Recipient", "Role", "Target", "Status", "Cause", "Manual", "Job"):
            imgui.table_setup_column(label)
        imgui.table_headers_row()
        for index, row in enumerate(delivery_rows):
            imgui.push_id(f"history_delivery_{index}")
            status = str(row.get("status") or "")
            payload = row.get("payload_dict") or {}
            imgui.table_next_row()
            imgui.table_next_column()
            draw_clipped_text(row.get("recipient_name"), 32)
            imgui.table_next_column()
            imgui.text(str(row.get("role") or "-"))
            imgui.table_next_column()
            draw_clipped_text(row.get("target"), 34)
            imgui.table_next_column()
            imgui.text_colored(status_badge_color(status), status or "-")
            imgui.table_next_column()
            draw_clipped_text(row.get("cause"), 46)
            imgui.table_next_column()
            can_copy = (
                status.lower() != "delivered"
                and str(row.get("role") or "").upper() in {"DDEO", "AEO"}
                and bool(payload)
            )
            with disabled_when(not can_copy):
                if imgui.button("Copy", (70, 0)):
                    imgui.set_clipboard_text(build_antidengue_manual_message(payload))
                    state.runner.status_message = (
                        "Copied manual WhatsApp message for "
                        f"{row.get('recipient_name') or row.get('target') or 'recipient'}"
                    )
                    state.runner.append_log(state.runner.status_message)
            imgui.table_next_column()
            imgui.text(str(row.get("job_id") or "-")[:12])
            imgui.pop_id()
        imgui.end_table()
        delivery_total = int(selected.get("delivery_count") or len(delivery_rows))
        if delivery_total > len(delivery_rows):
            imgui.text_disabled(
                f"Showing first {len(delivery_rows)} of {delivery_total} delivery events. Open the delivery audit for the full list."
            )

    dormant_rows = details.get("dormant") or []
    imgui.separator_text("Dormant Schools")
    if not dormant_rows:
        imgui.text_disabled("No dormant school rows saved for this run.")
    elif imgui.begin_table("antidengue_history_dormant", 4):
        for label in ("EMIS", "School", "Tehsil", "Markaz"):
            imgui.table_setup_column(label)
        imgui.table_headers_row()
        for index, row in enumerate(dormant_rows):
            imgui.push_id(f"history_dormant_{index}")
            imgui.table_next_row()
            imgui.table_next_column()
            imgui.text(str(row.get("school_emis") or "-"))
            imgui.table_next_column()
            draw_clipped_text(row.get("school_name"), 48)
            imgui.table_next_column()
            imgui.text(str(row.get("tehsil") or "-"))
            imgui.table_next_column()
            draw_clipped_text(row.get("markaz"), 42)
            imgui.pop_id()
        imgui.end_table()
        dormant_total = int(selected.get("dormant_count") or len(dormant_rows))
        if dormant_total > len(dormant_rows):
            imgui.text_disabled(
                f"Showing first {len(dormant_rows)} of {dormant_total} dormant schools. Open the output folder for the full report."
            )


ANTIDENGUE_VIEWS = (
    ("dispatch", "Dispatch Center"),
    ("health", "Health Stats"),
    ("history", "Run History"),
    ("failed", "Failed Officers"),
    ("groups", "Fixed / Group Delivery"),
    ("manual", "Manual Report"),
    ("master", "Master Data"),
    ("database", "Database"),
    ("files", "Files"),
    ("run", "Run / Schedule"),
)


def draw_antidengue_view_button(
    state: GuiState,
    view_id: str,
    label: str,
    current_view: str,
    width: int,
) -> str:
    selected = current_view == view_id
    button_label = f"{label}{' *' if selected else ''}"
    with disabled_when(selected):
        if imgui.button(button_label, (width, 0)):
            state.antidengue_view = view_id
            return view_id
    return current_view


def draw_antidengue_run_status_screen(state: GuiState) -> None:
    summary_path, summary, error = latest_antidengue_summary(state)

    imgui.separator_text("AntiDengue Operations")

    run_dir = summary_path.parent if summary_path else None
    whatsapp = summary.get("whatsapp") or {}
    delivery = whatsapp.get("delivery") or {}
    quality_gate = summary.get("quality_gate") or {}
    filter_info = summary.get("filter") or {}
    outputs = summary.get("outputs") or {}
    audit_rows = whatsapp.get("officer_delivery_audit") or []
    fallback = whatsapp.get("fallback_delivery") or {}
    fallback_delivery = fallback.get("delivery") or {}

    current_view = state.antidengue_view if state.antidengue_view in dict(ANTIDENGUE_VIEWS) else "dispatch"
    failed_count = len(failed_officer_rows(audit_rows))
    group_failure_count = len(
        [
            item
            for item in failed_delivery_statuses(delivery)
            if str(item.get("target") or "").endswith("@g.us")
        ]
    )
    fallback_group_failure_count = len(
        [
            item
            for item in ((fallback.get("delivery") or {}).get("statuses") or {}).values()
            if str(item.get("target") or "").endswith("@g.us")
            and str(item.get("status") or "").lower() != "delivered"
        ]
    )

    draw_antidengue_worker_controls(state)
    imgui.separator()

    nav_items = (
        ("dispatch", "Dispatch Center", 170),
        ("health", "Health Stats", 135),
        ("history", "Run History", 135),
        ("failed", f"Failed Officers ({failed_count})", 180),
        ("groups", f"Fixed / Group ({group_failure_count + fallback_group_failure_count})", 175),
        ("manual", "Manual Report", 150),
        ("master", "Master Data", 135),
        ("database", "Database", 120),
        ("files", "Files", 90),
        ("run", "Run / Schedule", 150),
    )
    for index, (view_id, label, width) in enumerate(nav_items):
        if index:
            imgui.same_line()
        current_view = draw_antidengue_view_button(
            state,
            view_id,
            label,
            current_view,
            width,
        )

    imgui.separator()
    if current_view == "dispatch":
        draw_antidengue_dispatch_center_section(state)
    elif current_view == "health":
        if error:
            imgui.text_colored(status_badge_color("warning"), error)
            imgui.text_disabled("Run History, Database, Manual Report, and Run / Schedule remain available.")
        else:
            draw_antidengue_health_section(
                summary,
                whatsapp,
                delivery,
                quality_gate,
                filter_info,
                fallback,
                fallback_delivery,
            )
    elif current_view == "history":
        draw_antidengue_run_history_section(state)
    elif current_view == "failed":
        draw_antidengue_failed_officers_section(state, audit_rows, delivery)
    elif current_view == "groups":
        draw_antidengue_group_delivery_section(delivery, fallback)
    elif current_view == "manual":
        draw_antidengue_manual_report_section(state)
    elif current_view == "master":
        draw_antidengue_master_data_section(state)
    elif current_view == "database":
        draw_antidengue_database_section(state)
    elif current_view == "files":
        draw_antidengue_files_section(state, summary_path, run_dir, outputs)
    elif current_view == "run":
        draw_antidengue_run_controls(state)


def draw_antidengue_run_controls(state: GuiState) -> None:
    command = command_by_id("antidengue_run")
    values = state.command_values(command)
    values.setdefault("PORTAL_LOGIN_MODE", "manual")

    imgui.separator_text("Run")
    imgui.text_wrapped(command.summary)
    imgui.text("Folder:")
    imgui.same_line()
    imgui.text_disabled(str(command.working_dir(state.project_root)))
    imgui.text("Command:")
    imgui.same_line()
    imgui.text_disabled(" ".join(command.build(state.project_root, values)))

    login_mode = str(values.get("PORTAL_LOGIN_MODE") or "manual")
    if login_mode not in {value for value, _label in PORTAL_LOGIN_MODES}:
        login_mode = "manual"
        values["PORTAL_LOGIN_MODE"] = login_mode

    auto_click = login_mode == "auto"
    changed, auto_click = imgui.checkbox("Auto click Sign In", auto_click)
    if changed:
        values["PORTAL_LOGIN_MODE"] = "auto" if auto_click else "manual"
        state.persist_command_values(command)
        login_mode = values["PORTAL_LOGIN_MODE"]
    imgui.same_line()

    mode_index = _option_index(PORTAL_LOGIN_MODES, login_mode)
    imgui.set_next_item_width(190)
    changed, mode_index = imgui.combo(
        "Login mode",
        mode_index,
        [label for _value, label in PORTAL_LOGIN_MODES],
    )
    if changed:
        values["PORTAL_LOGIN_MODE"] = PORTAL_LOGIN_MODES[mode_index][0]
        state.persist_command_values(command)

    with disabled_when(state.runner.is_running or state.command_start_in_progress):
        if imgui.button("Run Now", (120, 0)):
            start_command_with_services(state, command, values)
    imgui.same_line()
    with disabled_when(not state.runner.is_running):
        if imgui.button("Stop", (90, 0)):
            state.runner.stop()

    imgui.separator_text("Schedule")
    form = state.antidengue_form
    mode_values = ["once", "daily", "weekly"]
    mode_labels = ["Once", "Daily", "Weekly"]
    try:
        mode_index = mode_values.index(form.get("mode", "daily"))
    except ValueError:
        mode_index = 1
    imgui.set_next_item_width(130)
    changed, mode_index = imgui.combo("Mode", mode_index, mode_labels)
    if changed:
        form["mode"] = mode_values[mode_index]
        state.persist_antidengue_form()

    if form.get("mode") == "once":
        imgui.same_line()
        imgui.set_next_item_width(140)
        changed, date_value = imgui.input_text("Date", form.get("date", ""))
        if changed:
            form["date"] = date_value
            state.persist_antidengue_form()
        imgui.same_line()
        if imgui.button("Today", (70, 0)):
            form["date"] = date.today().isoformat()
            state.persist_antidengue_form()

    if form.get("mode") == "weekly":
        selected_days = parse_weekdays(form.get("weekdays", ""))
        for index, label in enumerate(WEEKDAYS):
            if index:
                imgui.same_line()
            checked = index in selected_days
            changed, checked = imgui.checkbox(label, checked)
            if changed:
                if checked:
                    selected_days.add(index)
                else:
                    selected_days.discard(index)
                form["weekdays"] = ",".join(str(day) for day in sorted(selected_days))
                state.persist_antidengue_form()

    imgui.set_next_item_width(360)
    changed, times_value = imgui.input_text(
        "Times", form.get("times", "09:00, 13:00")
    )
    if changed:
        form["times"] = times_value
        state.persist_antidengue_form()
    imgui.same_line()
    if imgui.button("+ Now", (70, 0)):
        current = parse_schedule_times(form.get("times", ""))
        now_time = datetime.now().strftime("%H:%M")
        if now_time not in current:
            current.append(now_time)
        form["times"] = ", ".join(sorted(current))
        state.persist_antidengue_form()
    imgui.same_line()
    with disabled_when(not parse_schedule_times(form.get("times", ""))):
        if imgui.button("Add Schedule", (130, 0)):
            state.add_antidengue_schedule()

    imgui.text_disabled("Use 24-hour HH:MM times separated by commas.")

    schedules = state.antidengue_schedules()
    if imgui.begin_table("antidengue_schedules", 6):
        imgui.table_setup_column("Enabled")
        imgui.table_setup_column("Mode")
        imgui.table_setup_column("Date / Days")
        imgui.table_setup_column("Times")
        imgui.table_setup_column("Next")
        imgui.table_setup_column("Actions")
        imgui.table_headers_row()
        now = datetime.now()
        for schedule in list(schedules):
            imgui.push_id(str(schedule.get("id", "")))
            imgui.table_next_row()

            imgui.table_next_column()
            enabled = bool(schedule.get("enabled", True))
            changed, enabled = imgui.checkbox("##enabled", enabled)
            if changed:
                schedule["enabled"] = enabled
                state.update_antidengue_schedule(schedule)

            imgui.table_next_column()
            imgui.text(str(schedule.get("mode", "")))

            imgui.table_next_column()
            if schedule.get("mode") == "weekly":
                imgui.text(format_weekdays(str(schedule.get("weekdays", ""))))
            elif schedule.get("mode") == "once":
                imgui.text(str(schedule.get("date", "")))
            else:
                imgui.text("Every day")

            imgui.table_next_column()
            imgui.text(", ".join(parse_schedule_times(str(schedule.get("times", "")))))

            imgui.table_next_column()
            imgui.text(next_schedule_run(schedule, now) if enabled else "-")

            imgui.table_next_column()
            with disabled_when(state.runner.is_running or state.command_start_in_progress):
                if imgui.button("Run Now", (80, 0)):
                    start_command_with_services(state, command, values)
            imgui.same_line()
            if imgui.button("Delete", (75, 0)):
                state.remove_antidengue_schedule(str(schedule.get("id", "")))
            imgui.pop_id()
        imgui.end_table()


def draw_antidengue_dashboard(state: GuiState) -> None:
    draw_antidengue_run_status_screen(state)


def draw_sms_campaign_list_dashboard(state: GuiState, command: CommandSpec) -> None:
    imgui.text(command.title)
    imgui.text_wrapped("Manage saved SMS campaigns. Select a row action to edit, send, inspect failures, or archive a campaign while preserving audit history.")

    campaigns = discovered_sms_campaigns(state, command)
    total = len(campaigns)
    needs_attention = sum(1 for item in campaigns if item.get("status") == "needs_attention")
    running = sum(1 for item in campaigns if item.get("status") in {"running", "submitted"})
    verified = sum(1 for item in campaigns if item.get("status") == "verified")

    if imgui.begin_table("sms_campaign_overview_metrics", 4):
        for label in ("Campaigns", "Needs attention", "Active/submitted", "Verified"):
            imgui.table_setup_column(label)
        imgui.table_headers_row()
        imgui.table_next_row()
        for value, color in (
            (str(total), (0.82, 0.82, 0.82, 1.0)),
            (str(needs_attention), sms_campaign_status_color("needs_attention")),
            (str(running), sms_campaign_status_color("submitted")),
            (str(verified), sms_campaign_status_color("verified")),
        ):
            imgui.table_next_column()
            imgui.text_colored(color, value)
        imgui.end_table()

    if imgui.button("Refresh Campaigns", (150, 0)):
        campaigns = discovered_sms_campaigns(state, command, force=True)
        state.runner.append_log(f"Refreshed SMS campaigns: {len(campaigns)} saved")
    imgui.same_line()
    if imgui.button("New Campaign", (130, 0)):
        create_command = command_by_id("sms_campaign_create")
        values = state.command_values(create_command)
        for key in ("campaign_id", "description", "message", "sheet", "column"):
            values[key] = ""
        state.persist_command_values(create_command)
        state.selected_command_id = create_command.id

    imgui.separator_text("Saved Campaigns")
    if not campaigns:
        imgui.text_disabled("No saved SMS campaigns yet. Click New Campaign or use Create SMS Campaign.")
        return

    if imgui.begin_table("sms_campaigns_table", 10):
        for label in ("Campaign", "Status", "Recipients", "Pending", "Sent", "Failed", "Unknown", "Updated", "Source", "Actions"):
            imgui.table_setup_column(label)
        imgui.table_headers_row()

        for campaign in campaigns:
            name = campaign.get("campaign_name", "")
            imgui.push_id(name)
            imgui.table_next_row()

            imgui.table_next_column()
            imgui.text_wrapped(name)
            description = campaign.get("description", "")
            if description:
                imgui.text_disabled(description[:120])

            imgui.table_next_column()
            status = campaign.get("status", "-")
            imgui.text_colored(sms_campaign_status_color(status), status)

            imgui.table_next_column()
            imgui.text(campaign.get("recipients", "0"))

            imgui.table_next_column()
            imgui.text(campaign.get("pending", "0"))

            imgui.table_next_column()
            delivered = int(campaign.get("delivered") or 0)
            sent = int(campaign.get("sent") or 0)
            imgui.text(str(sent + delivered))

            imgui.table_next_column()
            failed = campaign.get("failed", "0")
            failed_color = sms_campaign_status_color("needs_attention") if failed not in {"", "0"} else (0.82, 0.82, 0.82, 1.0)
            imgui.text_colored(failed_color, failed or "0")

            imgui.table_next_column()
            imgui.text(campaign.get("unknown", "0") or "0")

            imgui.table_next_column()
            imgui.text_disabled(campaign.get("updated_at", ""))

            imgui.table_next_column()
            source = Path(campaign.get("source_file", "")).name
            sheet = campaign.get("sheet", "")
            column = campaign.get("number_column", "")
            imgui.text_wrapped(source)
            imgui.text_disabled(f"{sheet} / {column}")

            imgui.table_next_column()
            if imgui.button("Edit", (52, 0)):
                apply_sms_campaign_to_command(state, "sms_campaign_update", campaign)
                state.selected_command_id = "sms_campaign_update"
            imgui.same_line()
            if imgui.button("Send", (55, 0)):
                apply_sms_campaign_to_command(state, "sms_campaign_send", campaign)
                state.selected_command_id = "sms_campaign_send"
            imgui.same_line()
            if imgui.button("Summary", (82, 0)):
                apply_sms_campaign_to_command(state, "sms_campaign_show", campaign)
                state.selected_command_id = "sms_campaign_show"
            imgui.same_line()
            if imgui.button("Failed", (65, 0)):
                apply_sms_campaign_to_command(state, "sms_campaign_failed", campaign)
                state.selected_command_id = "sms_campaign_failed"
            imgui.same_line()
            if imgui.button("Archive", (76, 0)):
                apply_sms_campaign_to_command(state, "sms_campaign_delete", campaign)
                state.show_confirm_for = "sms_campaign_delete"
            imgui.pop_id()

        imgui.end_table()


def whatsapp_service_id(profile_id: str) -> str:
    return "whatsapp" if profile_id == "default" else f"whatsapp_{profile_id}"


def whatsapp_service_for_account(state: GuiState, profile_id: str):
    return state.services.services.get(whatsapp_service_id(profile_id))


def whatsapp_qr_path(service) -> Path | None:
    if service is None:
        return None
    service_env = dict(service.spec.env)
    return service.spec.cwd / service_env.get("WA_QR_IMAGE_PATH", "data/whatsapp-login-qr.png")


def draw_whatsapp_accounts_dashboard(state: GuiState, command: CommandSpec) -> None:
    imgui.text(command.title)
    imgui.text_wrapped(
        "Manage isolated WhatsApp accounts. Each profile has its own auth folder, lock file, queue subjects, QR login, and worker service."
    )

    accounts = state.services.whatsapp_accounts()
    active_accounts = [account for account in accounts if bool(account.get("enabled", True))]
    running_count = sum(
        1
        for account in accounts
        if (service := whatsapp_service_for_account(state, str(account.get("id", "")))) and service.is_running
    )

    if imgui.begin_table("whatsapp_accounts_metrics", 3):
        for label in ("Accounts", "Enabled", "Running"):
            imgui.table_setup_column(label)
        imgui.table_headers_row()
        imgui.table_next_row()
        for value in (len(accounts), len(active_accounts), running_count):
            imgui.table_next_column()
            imgui.text(str(value))
        imgui.end_table()

    imgui.separator_text("Accounts")
    if imgui.button("Add New Account", (170, 0)):
        state.new_whatsapp_account_id = ""
        state.new_whatsapp_account_name = ""
        state.whatsapp_account_setup_service_id = ""
        state.selected_command_id = "whatsapp_account_add"
    imgui.same_line()
    if imgui.button("Refresh Accounts", (150, 0)):
        state.services.reload_specs()
        state.runner.append_log("WhatsApp account registry refreshed.")
    imgui.text_disabled("Each account is isolated. Removing a profile preserves its credential folder on disk.")

    for account in accounts:
        draw_whatsapp_account_card(state, account)

    draw_whatsapp_reset_modal(state)


def draw_whatsapp_account_card(state: GuiState, account: dict) -> None:
    profile_id = str(account.get("id", "")).strip()
    if not profile_id:
        return
    service = whatsapp_service_for_account(state, profile_id)
    enabled = bool(account.get("enabled", True))
    auth_dir = str(account.get("auth_dir") or f"auth_info_baileys_{profile_id}")
    service_env = dict(service.spec.env) if service else {}
    qr_path = whatsapp_qr_path(service)
    title = str(account.get("name") or profile_id)

    imgui.push_id(profile_id)
    imgui.separator_text(title)
    imgui.text_disabled(f"Account ID: {profile_id}")

    if imgui.begin_table("account_card_details", 3):
        for label in ("Service", "Login", "Queue subjects"):
            imgui.table_setup_column(label)
        imgui.table_headers_row()
        imgui.table_next_row()

        imgui.table_next_column()
        if service:
            imgui.text_colored(service_status_color(service.is_running, service.last_exit_code), service.status_message)
            imgui.text_disabled(f"Uptime: {format_elapsed(service.elapsed_seconds()) if service.is_running else '-'}")
        else:
            imgui.text_colored((0.95, 0.35, 0.28, 1.0), "No service registered")

        imgui.table_next_column()
        imgui.text_disabled(auth_dir)
        if qr_path and qr_path.is_file():
            imgui.text_colored((0.28, 0.85, 0.48, 1.0), "QR ready to scan")
        else:
            imgui.text_disabled("No QR pending")

        imgui.table_next_column()
        imgui.text_disabled(service_env.get("NATS_HEALTH_SUBJECT", "whatsapp.worker.health") if service else "-")
        imgui.text_disabled(service_env.get("NATS_GROUP_MEMBERS_SUBJECT", "whatsapp.worker.group-members") if service else "-")
        imgui.end_table()

    changed, new_enabled = imgui.checkbox("Enabled", enabled)
    if changed:
        ok, message = state.services.update_whatsapp_account(profile_id, enabled=new_enabled)
        state.runner.append_log(message)

    imgui.separator_text("Actions")
    with disabled_when(not service or service.is_running or not enabled):
        if imgui.button("Start Worker", (125, 0)) and service:
            state.services.start(service.spec.id)
    imgui.same_line()
    with disabled_when(not service or not service.is_running):
        if imgui.button("Stop Worker", (120, 0)) and service:
            state.services.stop(service.spec.id)
    imgui.same_line()
    with disabled_when(not service):
        if imgui.button("Restart", (90, 0)) and service:
            state.services.restart(service.spec.id)
    imgui.same_line()
    with disabled_when(not service):
        if imgui.button("Service Details", (135, 0)) and service:
            state.current_service_id = service.spec.id

    with disabled_when(not qr_path or not qr_path.is_file()):
        if imgui.button("Open QR Image", (135, 0)) and qr_path and service:
            open_external_file(state, qr_path, "WhatsApp QR code", service.spec.id)
    imgui.same_line()
    with disabled_when(not service or service.is_running):
        if imgui.button("Reset Login / New QR", (165, 0)) and service:
            state.whatsapp_auth_reset_service_id = service.spec.id
            state.show_whatsapp_auth_reset_confirm = True
    imgui.same_line()
    with disabled_when(profile_id == "default" or (service is not None and service.is_running)):
        if imgui.button("Remove Profile", (135, 0)):
            ok, message = state.services.remove_whatsapp_account(profile_id)
            state.runner.append_log(message)

    if state.edit_whatsapp_account_id == profile_id:
        imgui.set_next_item_width(320)
        changed, edit_name = imgui.input_text("Display name", state.edit_whatsapp_account_name)
        if changed:
            state.edit_whatsapp_account_name = edit_name
        if imgui.button("Save Name", (105, 0)):
            ok, message = state.services.update_whatsapp_account(profile_id, name=state.edit_whatsapp_account_name)
            state.runner.append_log(message)
            state.edit_whatsapp_account_id = ""
            state.edit_whatsapp_account_name = ""
        imgui.same_line()
        if imgui.button("Cancel Rename", (130, 0)):
            state.edit_whatsapp_account_id = ""
            state.edit_whatsapp_account_name = ""
    else:
        if imgui.button("Rename Account", (140, 0)):
            state.edit_whatsapp_account_id = profile_id
            state.edit_whatsapp_account_name = title

    imgui.pop_id()


def draw_whatsapp_account_add_screen(state: GuiState, command: CommandSpec) -> None:
    imgui.text(command.title)
    imgui.text_wrapped(
        "Create a new isolated WhatsApp profile, start its worker, then scan the generated QR code from WhatsApp on your phone. "
        "Once linked, the credentials are stored in that account's own auth folder and reused on restart."
    )

    imgui.separator_text("New account")
    imgui.set_next_item_width(260)
    changed, account_id = imgui.input_text("Account ID", state.new_whatsapp_account_id)
    if changed:
        state.new_whatsapp_account_id = account_id
    imgui.same_line()
    normalized = normalize_whatsapp_account_id(state.new_whatsapp_account_id)
    imgui.text_disabled(f"Will save as: {normalized or '-'}")

    imgui.set_next_item_width(420)
    changed, account_name = imgui.input_text("Display name", state.new_whatsapp_account_name)
    if changed:
        state.new_whatsapp_account_name = account_name
    imgui.text_disabled("Use a short stable ID like office-2, deo-main, ops-heads. Do not use spaces; they are converted to dashes.")

    existing_ids = {str(account.get("id")) for account in state.services.whatsapp_accounts()}
    cannot_create = not normalized or normalized == "default" or normalized in existing_ids
    if normalized in existing_ids:
        imgui.text_colored((0.95, 0.45, 0.25, 1.0), "That account ID already exists.")

    with disabled_when(cannot_create):
        if imgui.button("Create Account & Start QR Login", (260, 0)):
            added, message = state.services.add_whatsapp_account(state.new_whatsapp_account_id, state.new_whatsapp_account_name)
            state.runner.append_log(message)
            if added:
                service_id = whatsapp_service_id(normalized)
                state.whatsapp_account_setup_service_id = service_id
                state.services.start(service_id)
                state.new_whatsapp_account_id = ""
                state.new_whatsapp_account_name = ""

    imgui.same_line()
    if imgui.button("Back to Accounts", (150, 0)):
        state.selected_command_id = "whatsapp_accounts"

    service = state.services.services.get(state.whatsapp_account_setup_service_id)
    if not service:
        return

    imgui.separator_text("QR login setup")
    imgui.text(service.spec.title)
    imgui.text_colored(service_status_color(service.is_running, service.last_exit_code), service.status_message)
    imgui.text_disabled("Wait for the worker log to say the QR code is ready, then open and scan it.")
    qr_path = whatsapp_qr_path(service)
    if qr_path and qr_path.is_file():
        imgui.text_colored((0.28, 0.85, 0.48, 1.0), f"QR image ready: {qr_path}")
        if imgui.button("Open QR Image", (140, 0)):
            open_external_file(state, qr_path, "WhatsApp QR code", service.spec.id)
    else:
        imgui.text_disabled("QR image is not ready yet. Keep this screen open for a few seconds, or open Service Details to watch logs.")
    imgui.same_line()
    if imgui.button("Service Details", (135, 0)):
        state.current_service_id = service.spec.id
    imgui.same_line()
    with disabled_when(not service.is_running):
        if imgui.button("Stop Worker", (110, 0)):
            state.services.stop(service.spec.id)


def draw_whatsapp_reset_modal(state: GuiState) -> None:
    if state.show_whatsapp_auth_reset_confirm:
        imgui.open_popup("Confirm WhatsApp login reset")
    opened = imgui.begin_popup_modal("Confirm WhatsApp login reset", True)
    if not opened[0]:
        return
    service_id = state.whatsapp_auth_reset_service_id or "whatsapp"
    service = state.services.services.get(service_id)
    imgui.text("Reset WhatsApp Login")
    if service:
        imgui.text_disabled(service.spec.title)
    imgui.text_wrapped(
        "Use this only when WhatsApp rejects the preserved session. "
        "The current credentials will be archived and a new QR login will start."
    )
    with disabled_when(service is None or service.is_running):
        if imgui.button("Archive and Reset", (150, 0)):
            state.show_whatsapp_auth_reset_confirm = False
            reset_whatsapp_authentication(state, service_id)
            imgui.close_current_popup()
    imgui.same_line()
    if imgui.button("Cancel", (100, 0)):
        state.show_whatsapp_auth_reset_confirm = False
        state.whatsapp_auth_reset_service_id = ""
        imgui.close_current_popup()
    imgui.end_popup()


def draw_command_detail(state: GuiState, command: CommandSpec) -> None:
    if command.id == "whatsapp_accounts":
        draw_whatsapp_accounts_dashboard(state, command)
        return
    if command.id == "whatsapp_account_add":
        draw_whatsapp_account_add_screen(state, command)
        return
    if command.id == "sms_campaign_list":
        draw_sms_campaign_list_dashboard(state, command)
        return

    values = state.command_values(command)
    validation_error = command_validation_error(state, command, values)
    imgui.text(command.title)
    imgui.text_wrapped(command.summary)
    if command.aliases:
        imgui.text_disabled("Aliases: " + ", ".join(command.aliases))

    if command.env_fields:
        imgui.separator_text("Limits and Environment")
        for field in command.env_fields:
            draw_env_field(
                state,
                command,
                values,
                field.key,
                field.label,
                field.kind,
            )

    if command.arg_fields:
        visible_fields = [
            field
            for field in command.arg_fields
            if not field.kind.startswith("hidden_env:")
        ]
        if visible_fields:
            imgui.separator_text("Options")
            for field in visible_fields:
                draw_field(state, command, values, field)

    imgui.separator_text("Command")
    if validation_error:
        imgui.text_colored((0.95, 0.35, 0.28, 1.0), validation_error)
    try:
        command_line = shlex.join(command.build(state.project_root, values))
        imgui.text_wrapped(command_line)
    except Exception as exc:
        imgui.text_colored((0.95, 0.35, 0.28, 1.0), str(exc))

    command_busy = state.runner.is_running or state.command_start_in_progress
    with disabled_when(command_busy or bool(validation_error)):
        if imgui.button("Run", (100, 0)):
            if command.confirm_label:
                state.show_confirm_for = command.id
            else:
                start_command_with_services(state, command, values)
    imgui.same_line()
    with disabled_when(not state.runner.is_running):
        if imgui.button("Stop", (100, 0)):
            state.runner.stop()
    if command.id in {"crm_filter_excel_duplicates", "crm_filter_pdf_duplicates"}:
        imgui.same_line()
        with disabled_when(running or bool(validation_error)):
            if imgui.button("Clear Results", (130, 0)):
                clear_duplicate_filter_results(state, command, values)
    if command.id == "crm_filter_excel_duplicates":
        imgui.same_line()
        with disabled_when(running or bool(validation_error)):
            if imgui.button("Prepare Submitted", (155, 0)):
                notice_command = command_by_id("crm_prepare_duplicate_notices")
                notice_values = copy_excel_filter_values_to_notice_command(
                    state, values, notice_command, "submitted"
                )
                start_command_with_services(
                    state,
                    notice_command,
                    notice_values,
                    source="Preparing submitted notice list from the current filter output.",
                )
        imgui.same_line()
        with disabled_when(running or bool(validation_error)):
            if imgui.button("Send Submitted", (140, 0)):
                notice_command = command_by_id("crm_send_duplicate_notices")
                copy_excel_filter_values_to_notice_command(
                    state, values, notice_command, "submitted"
                )
                state.selected_command_id = notice_command.id
                state.show_confirm_for = notice_command.id
        imgui.same_line()
        with disabled_when(running or bool(validation_error)):
            if imgui.button("Prepare Not Relevant", (180, 0)):
                notice_command = command_by_id("crm_prepare_duplicate_notices")
                notice_values = copy_excel_filter_values_to_notice_command(
                    state, values, notice_command, "not-relevant"
                )
                start_command_with_services(
                    state,
                    notice_command,
                    notice_values,
                    source="Preparing not-relevant notice list from the current filter output.",
                )
        imgui.same_line()
        with disabled_when(running or bool(validation_error)):
            if imgui.button("Send Not Relevant", (165, 0)):
                notice_command = command_by_id("crm_send_duplicate_notices")
                copy_excel_filter_values_to_notice_command(
                    state, values, notice_command, "not-relevant"
                )
                state.selected_command_id = notice_command.id
                state.show_confirm_for = notice_command.id


def draw_confirm_modal(state: GuiState) -> None:
    if state.show_confirm_for is None:
        return

    command = command_by_id(state.show_confirm_for)
    imgui.open_popup("Confirm command")
    opened = imgui.begin_popup_modal("Confirm command", True)
    if not opened[0]:
        return

    imgui.text(command.title)
    imgui.text_wrapped(command.confirm_label)
    validation_error = command_validation_error(state, command, state.command_values(command))
    if validation_error:
        imgui.text_colored((0.95, 0.35, 0.28, 1.0), validation_error)
    imgui.spacing()
    with disabled_when(bool(validation_error)):
        if imgui.button("Run Command", (130, 0)):
            start_command_with_services(state, command, state.command_values(command))
            state.show_confirm_for = None
            imgui.close_current_popup()
    imgui.same_line()
    if imgui.button("Cancel", (100, 0)):
        state.show_confirm_for = None
        imgui.close_current_popup()
    imgui.end_popup()


def draw_logs(state: GuiState) -> None:
    imgui.separator_text("Activity")
    imgui.text_colored(status_color(state), state.runner.status_message)
    if state.runner.is_running:
        elapsed = state.runner.elapsed_seconds()
        imgui.same_line()
        imgui.text(f"{elapsed // 60:02d}:{elapsed % 60:02d}")

    if imgui.button("Clear", (90, 0)):
        state.runner.clear_log()
    imgui.same_line()
    if imgui.button("Copy", (90, 0)):
        imgui.set_clipboard_text("\n".join(state.runner.log_lines))
    imgui.same_line()
    changed, state.auto_scroll = imgui.checkbox("Auto-scroll", state.auto_scroll)
    imgui.same_line()
    imgui.text_disabled(f"{len(state.runner.log_lines)} lines")

    available = imgui.get_content_region_avail()
    imgui.begin_child("activity_log", (0, max(180, available.y - 8)), True)
    for line in state.runner.log_lines:
        imgui.text_unformatted(line)
    if state.auto_scroll and state.runner.is_running:
        imgui.set_scroll_here_y(1.0)
    imgui.end_child()


def draw_service_detail(state: GuiState, service_id: str) -> None:
    service = state.services.services[service_id]
    imgui.spacing()
    imgui.text(service.spec.title)
    imgui.text_wrapped(service.spec.summary)
    imgui.spacing()

    imgui.text("Status:")
    imgui.same_line()
    imgui.text_colored(
        service_status_color(service.is_running, service.last_exit_code),
        service.status_message,
    )
    if service.is_running:
        imgui.same_line()
        imgui.text_disabled(format_elapsed(service.elapsed_seconds()))

    imgui.text("Folder:")
    imgui.same_line()
    imgui.text_wrapped(str(service.spec.cwd))
    imgui.text("Command:")
    imgui.same_line()
    imgui.text_wrapped(" ".join(service.spec.command))

    with disabled_when(service.is_running):
        if imgui.button("Start", (90, 0)):
            state.services.start(service_id)
    imgui.same_line()
    with disabled_when(not service.is_running):
        if imgui.button("Stop", (90, 0)):
            state.services.stop(service_id)
    imgui.same_line()
    if imgui.button("Restart", (100, 0)):
        state.services.restart(service_id)
    imgui.same_line()
    if imgui.button("Clear Log", (100, 0)):
        state.services.clear_log(service_id)
    imgui.same_line()
    changed, state.service_auto_scroll = imgui.checkbox(
        "Auto-scroll", state.service_auto_scroll
    )

    if service.spec.kind == "whatsapp":
        service_env = dict(service.spec.env)
        imgui.text("Profile ID:")
        imgui.same_line()
        imgui.text_disabled(service.spec.profile_id)
        imgui.text("Auth folder:")
        imgui.same_line()
        imgui.text_disabled(service_env.get("WA_AUTH_DIR", "auth_info_baileys"))
        imgui.text("Health subject:")
        imgui.same_line()
        imgui.text_disabled(
            service_env.get("NATS_HEALTH_SUBJECT", "whatsapp.worker.health")
        )
        qr_path = service.spec.cwd / service_env.get(
            "WA_QR_IMAGE_PATH", "data/whatsapp-login-qr.png"
        )
        imgui.separator_text("WhatsApp Login")
        if qr_path.is_file():
            imgui.text_colored(
                (0.28, 0.85, 0.48, 1.0),
                "QR code ready - open it and scan with WhatsApp Linked Devices.",
            )
            if imgui.button("Open QR Code", (140, 0)):
                open_external_file(state, qr_path, "WhatsApp QR code", service_id)
            imgui.same_line()
            imgui.text_disabled(str(qr_path))
        else:
            imgui.text_disabled(
                "No QR image is pending. It appears here when WhatsApp needs authentication."
            )
        with disabled_when(service.is_running):
            if imgui.button("Reset Login & Create QR", (190, 0)):
                state.whatsapp_auth_reset_service_id = service_id
                state.show_whatsapp_auth_reset_confirm = True
        imgui.same_line()
        imgui.text_disabled("Manual only; existing credentials are archived, never deleted.")

        draw_whatsapp_reset_modal(state)

    imgui.separator_text("Service Log")
    available = imgui.get_content_region_avail()
    imgui.begin_child(f"{service_id}_service_log", (0, max(260, available.y - 8)), True)
    for line in service.log_lines:
        imgui.text_unformatted(line)
    if state.service_auto_scroll and service.is_running:
        imgui.set_scroll_here_y(1.0)
    imgui.end_child()
