#!/usr/bin/env python3
"""Build a compact, unfiltered AI context bundle for Deomee projects.

This edition keeps the proven v4 size-optimized selection strategy while
removing privacy and secret redaction. It includes source selected with real
.gitignore semantics, plus explicitly discovered ignored credential/config
files such as .env, private keys, credentials JSON, Frappe configuration and
small authentication state files.

It deliberately remains compact:
- dependency/build/cache directories stay excluded
- raw spreadsheets are represented by unfiltered bounded previews by default
- raw SQLite databases are represented by unfiltered read-only diagnostics
- logs are represented by recent unfiltered tails
- general binary files remain excluded unless --include-binaries is supplied
- selected secret/config files are included even when ignored or binary

The resulting ZIP may contain live passwords, tokens, private keys, personal
data and session credentials. Treat it as a confidential backup.
"""

from __future__ import annotations

import argparse
import ast
import base64
import csv
import datetime as dt
import hashlib
import json
import mimetypes
import os
import platform
import re
import shutil
import shlex
import socket
import sqlite3
import subprocess
import sys
import tempfile
import textwrap
import traceback
import zipfile
import zlib
from collections import Counter, defaultdict
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Any, Iterable, Iterator, Mapping, Sequence

TOOL_VERSION = "2026.07.20.1-compact-unfiltered"
DB_COMPONENT_VERSION = "2026.07.19.3"

# Embedded, compressed copy of the user's proven DB diagnostic collector.
DB_COMPONENT_B85 = 'c-rlKTYKY1j_A971(!M57CVyc?p$_8iL>jporyQyZ6Dj2*~HOsX^FNOOQb?lzQoplKLF}}Aw_q4X7}Ob`_iV!DgcE-p-?Ck>XU!jT2<xNd6sXb`Ax9Audj;y^IjN+Z^|^O(;x}p?`&A)^ZQ_yC6{?o)!8&SU*)rT8e9}*P+z6NURo^DG}v9$#UiP*A`gz{Ney2Y!QOYHUhg=a7K=ri&(c}2s*+2nSr!Ybn#elCrKmH@%5+*6<^8zV3*f)iO;E0KPF=9IOzNwxy4X@xV^KApoyUsI(egg{l^_JcaJVefn=HK@UZ?lL?%weA_Uo^<|LxVEhueSp68;-*?|=UJmv4H#Qy{9Emf5lnro@8^nDEQNQ7|p?I?3`Zza005!7Qnh^Q6MESLq@d;OD9?p+H>?f()8^hc)wAP!_lFaRu+8T9xEko&5oPfELx+B1OJuWsw7S>R^%92?C~qi{%_35TnwYw7d@h;XH*;z)Sc9odYN@vddM;SWcf2M-t4^bO{udMO6*wS$>V^(q&QB)MHgT2ll43tvR-?4*qtP)Ya~C8E_6&gagPMVw<w6PV?zKH6^R66|{J@SR`czxOxY5kRx%yZJ8mb>Vi5|G6yTM0_h#%uL8Zfz~34C&60UF0={Qeu!087>lEGx=`5pWW%+czngQ%(T2@7#%!7<7U1ack0F41CHRH5o%uO<1rPU~azU-xU$rPwe7Acfof&isKzFM49<^Md`V>RI;iRF1Q%dp$ZG`LLjRhFk9_jw9r&PF&YdOhloIKEibt1^w_AX^Y$lRPhK>IdM8cvoI7ld?+1Z{)~VUx;7R>P9>QyHb>eAgO{`E#IcoY>~{x&#R=mnrG+Y=WoCZ@my5mNsI&WyOhYQt5uyzU{x`_P9>&lb-pZ%DeyqPxtEW1x>!P|NI)Rt)AT%<UW*s2RW=jPe`KsIacNr2QP7Ylf=zEwtkQA-jk`!z^Lmy|YgQNsH8f8YJc6I>Lw&!5At&DN=J$i(06?Iy4}y2e62A_Dll1o$h$+XsD(3)eMBEoOe_s`K+UuRXIX*Z#jsLoTd~$I3KG+FDR0N~#e;d8}^XPNv;FJ9~$NQ)8{`<ei$NTWTOh+hJ(5gPEk}~Z7`}k=0<m7LM$9s?Tu=jX+_~ZWjM^^Q5_syIAlau(z{huFqj}GYX(ec4wcTe}}`OV?sj|cmYyFZ=&aCm(1%kC-A@dy?7_TQf#?7oe{0mGAy4*v4~@Ob~t?#ce37YV}N9KQYO-TM~WmPu9J7Uk@b9%hehgghow?BKI_IQe(bfRD*Clh07;?~id*7uRVH&;JLWA3?cZQ2~1Uc%H(TC?C~`eVi7>b(S_#djt$Ve82m4Grfo!{w#`fc>hQv^sxpN#h!l4W{)+DBxnJs*$nXJ+4M1)%}QuZ#Hno};vMwE?qBx1)4<fgBdCFUpy9551jR#h0SNIUjPSI2Ov|z;;XTX+^Xd_(05Jwum!M!kEUU*VPnOkHQ9qKh+(_aciUp_d=y?D8{p0=jZ;(z#UC8F3S)9}yLbzJatTz`hD^*v<s|t8#6)KBA(p-B117>BEfU#5-H^wWWQT2E8q^jd8HDG145F68bbNKG)?LM?<|G$n7k58#Z2miHyg6)B(EaGqgr9MH$3?q}xSfp*!4Cak8cys&?STqIctpb>-%Jc%}0vOgbqrlJKR~gO?FwKAg0^F`ac>_2i6Q|x_3aq&-%Ipu)#RHgpLG7V}m=)6%O*+)ZJVBtS`HPE-Yzmlhc83lh^}he<?OQ?q>F#%LiTDT3p?SJE&!#a<Xcd5$T!My465=VS`O5!E5V7GifrxS9HVVqmKqn|+QmlSJFgMO;adBRy<xSF34y-4zUW|{+1m;=jl`3AY&ga>*tvc8g^GsD=fMBF$Om@<=wqIip#>on1b@>HMr=pl!KjSMfIOV%rv`s)TV^CjYL%AQ8ONUJPky$-4%ps}o<1|}l$k1w9_zEDkD1FUe;zdfPM00_vSd~5qIIWWX<$RfC)e`$921&)q)Gbk^H43$_DmR?tOA!2JTLHV5WD#av)%S5S&vZ$a1qgIw!y6EDbf!>4OY1Vzh?u3INNZz+-0BUj)2k|*Ym!x7Cx0js`B=7p$G@%0tU^;mmk};)q6{d7vV^wF?+I`LSaG>Z;~7XK5T3@%0z@tImceDfY&G|O!?AjUEG-)fR2iz-rgCbT1Jn97(8`w%SYi!2_xN4>_Tc@G`+JVQQk|Ng!|0jCe1h}zIxmpa&`non57=cY5TF#*%WvtlrP#TtT>Rf-ST-!0%W$(4WB?o8sa$7(uq??pk#80H4OFW39M20Sn|`i<IA<67IAi)CHU_r#PBatx^$<)bG7iCrqK<`Les^%HbuW=K4nE)B-nNM{-quNV9m7O;UfjWlSL%aDeqnM+rz<v6HkW1GF_8K67+Y3Cxi2>tixtbd!G+;D&tUZ8xMUPJS79a&tT)ZJse(EvA9687<!9)q9xP^pbXmfb*nbJ8YBk1XN)!yg4c>zZJ0=j?=Rtryy#s5kQB{K#Q$`#H*l{GGC_!;q6m_4>xG}AU2p&E#cvCQ5vKhH2c7ii1Hv(g{nBS!RXb`-#-r!mY`)Wd!(F-5}zzSP~Hs=14AjmF)zACg8gv)zQ+XxiWJd8$JMKaru$da-O!wQi!4HbiVrynWM%o2tbw1%0r>MEPhpg|V(2+-pw3VV`Hs@F=R!Y(jlou2l>LdrXZ7Lj0k)oFgy$5pD<9XO`|?_oU)|4bML=6e`h-*C{b1-Y=0p!IGGaSV#=tTMU9ZlX%)kc-h)?Sp|)%s`dxgsb{u_@{6X(Bh=p3DGixWjeIK0_&*8$*a<_;0s&}CATA-f|t-*t=o}4Dgl`B?}#{4-GYtY4?hVbd>`(FR1T=ZB4*#>Y@g+;RKA2&=76;curiQHS0A7Zf>)7V4hq1(4q2iy{q+&e3ZNj;)7ka}X?u7Pz8D1Ii*T)d8onAshwR9LCh-QQvsdHctBFZPbp`}ZM2Vi;YfNysAjQHN7O%>ApQ#7o85+Guw-gMO6jv&RKGN$ISOm6gK{FA+NC)YNj#C2KKOBk|LwF&_O;L@wKW+pI?0yJ~(cSN0VH*E*{5EujQWZZli%tNA=GbHK;YG|C0<1Dd5jGVED?$q+L5wQji5OMfB?dU=L&BX-K+Z*2{SNY%j<>e99z@X>$fFnd;L$as+4+k|DTj5a_d5p+T;SSAya~B1RLMn(zG)Nz>@%d0x*7hmmeuI(7EL547EqYX(YbsNO}ND2$URw$<eK`mC(9FeQE!S*h$ON(#!5jOaOoT;Rg9x3zwBSZydDd$(avePN?FzPZkjIZV4wZ~FPttYVIZ!~`{+UWW(Qaq_0Rr3-ujQpXVDm*@b7=X@Giq&UyWbBe2LFQ7GnuIFj)uo>0dQz!e1%wY%pdO&7`PzGO*on*+1pg20@<&3CylAug;4E{+xqhA2sMy0oVg)R*!@yb`bC%^nh=g;DHV=P(G5nAGN?nC+X;?pAPm0G{gtOyF1!re?8z)h^cxD-1B>wvfZi=VI0$^nE9OJ80;1HI6D983z|1LjI;3<+kgIQ5{-Z$d>JOyG|NITA;LL`U3~>L;ek`KZYG7hVm-uQL$0H-(c=+KYu7L-pqA#GM9aP3hNC}+72|;EMs)pSmy=ry9z7g9b1^D(lb)Z~e87@#Dh7+5Ia+KPmE#s?<{i7QX4xfftn3J<{iwQ1UVr&jALm=n{N^b;8eOG#3_?FT8^8LBQCh8FrFzF4m~7Pjx(C%m9AU|AbT<5AY&7X2{Dx5vih-dYKQO54hZg}l@xB<p2tE^Vf#3WoomXk_;>G&gP)=kBCI%A%U}=m1D{~;WXW|MvAp_eOKH=~i;3OVj!Kfr%hko|Z8QgANy9aG05qb53Qw?7}MN(mHq#JBNqh_!P8;ooilu<`wP5S{G&YDbupWa8tXCTS;<d3Q}DW_L`Uekid2|VLpk+Sd`=F|_0MOH(nd>a}yU=ec>xDNOyqj3fSwl^ZCA{)5**GFy-sDDXx@S97H<BESWHKrIM7LWApd=HDeR7u~C0B=!LrT;8^!ggBK&!^X{5!sb4eDJHmjq*ZOJyVoVOGdrc5iBeswh<2N+*>ffZpUI+sv+2>U##U97U%kyHe{MdGcA3K`D&4i1qpBh@VjFfw~YB9W)++@l0zG<F`7m<!gJQZNbDJY=`gIFRaGk)&HcHVkcNI$inn3tD+}wxXh!^_d2ySTu;$0rTgV-jIMLyzT(~yr#$_<gEv~Id1X=TPLvtS-RS=4)6RaOJhzD*7!#YAbB2OVU@j@ZX5TLE5r&$6f^mRLztp<8#aG%a_-e~%Rgr;X0|A>N*W(xm^aja`Dk^$}7d7WYC$6y|*(c&7NL)<+@s}JDB9q7ZjxYiaURE41--v{Cv%~p$L)z?%EB8*u04bc1=AmtUtX^>GXbYg{DLeMaVEy6;kA0;wcHLY%pMuD<`CA<hYICIq2<l`I52MJ-)7I`KI(DHLBH#Ia<dAJ)0B;zD=#wWKkZd9geKE?@0yoX1=jv=>Gam&@CMhCZr_=mA=+~6m!9tgTKDDdqlnay0&ND(^%rou8+6+zz<<3Ir)`r)mo@q-{wZ_&B5g9<6BV9;TdnxcT1hlUw+>1Yooh`-T`J}Wn1xG7R-hmjFkRT=wngAz5A!4WU3?3y^&JzTSu5EtGXpNvMfnSiTPpm#uh7zu4_tf90yY%K)1T;;f3yMXp_&yy5e?C(J^y`2fM#zijK!MKgS6TI4PoMH`E1#YOI+{h535g_dw96*{0y`{qd`40R)a5Vakt8%M)2E^@XXpRn!_MI}ILfgymR_^e>7C~3(^qOl4`N6T|%-JI3{0+yA?F?B(RQ6B`X&b3GR#DbYQH=c(-cGkDG*>GMzucm0uPX;f2fXt*4!q3?LYkZ4<wf{_o9TV19F5`_2TL5U$H4=<SchUW<T9;eW<t<t>VqIJY2TTz8)sR9kzvXWMyg4=JcGx885Ss{T<BPsrKZJrA!)yInCRCHmqk@y!UXmEJXB5O5NNxE#X;h*k))y1F=k~r)ieq#q0An_EN6}$+=*ocxmOuWRECp$c^QY%L^pau7ETe8UgT{trB{(B%7Ou`lPXFORsXUhK~5MVb5Wy#Q*08(aT~kU<UoiCjF)H%{>4sq*{JtDtGC1cc+F0s?E4~=ma0e1*2eD$xxXwiS|AlN#VRvzgJ=)|@;F;57I4PMMCUYzG@|#hiZAUmN`#1U6!xHjG0_Zxw|hN$(2<(%WBrR*9x*`_!DP+|VjNy3S*a`ahAwT?<e*%Zs||}Kqf@Y}b=20z4RC@LhvxEjD!#Evi#1ZJV|^-76p0NMk73RS-dND(mLLpJfc|>Ev&9^4@SZsz@xKS}TdIn32lW5~+oGV~9=l)B0=hCBi!~&xMDL=uTKQ@|=Z#RP#JjeNV%NdKC|*aT(cIU#W<njN)=eZl&3G^bY{gbbn(-|G%g?Q?6oEz!K;-8pAZzbizP@*L67P04q=mW=YRLgL@d(bc%oa&`k4uMeY(ZaM$LaupABP7b5SwO~IW0yUkow2Y5c$Ca2N-|f8O|!tK&NPJazVO9IzuNzs};CYq6=&#5gY9tkFIf4i>*q!TS2DZ50Hoej}(JY7Hq{>fvo4bn7w{yn}UOYLWwJnJB;geAQk^8v&mH6T(es{h=B~|zpHyXc~l$yZ;`%=LBd8X5GW;IMIS&#f;QZSdss&!*^X$G%;$Zucf{P4%#-rtTPL>)x;Ca8*_=GFM#3<_DvgVBMzO1?B)Gqi88B*jENlo)>R3meBR2A^5g+LR7}*frjq7_}LPrIpQsX&gxIkn^0b5TghVqCa-Tu*xWWxqxD>i8d4LBsyg{@u;9<N^9#uRu<>%xJ)jULalMW(;MfGLWbev9NTW_}X+wP~ZMr-eKC<arQy1)La$!<J^cGd2}Zm=sSHW2-f9FpFc4F@&7FU{F<BG@g)eVcY6feN%9|v=3b@x~<4xF}BfO{as(O*NB1vP##_e-v--bYhleHWqIF%Kkz|N7~BmI0n36T%Pe56ik?={Ml$Rq`N20k90HjSvQ&(L5A}UKMH7D(de$lbpD%(hdiy!8yzNpqq3g<{u9))RD;qp1h?e-t{@eXGr@_nM`{ToR!GpPG8klCi2xq~NAOTphfx-L?fd+@id;7=1cRvR$OCAX#S_gakCvU=@0VND5jx`ulBSWEPa4>u3#{sK2fccq~A2UX+S$oxaN?g`kdrF;r+>!0RV~bxrifODsLfNT6f|5`5ot<z50YM;qUqF(lSt{XzIf533_a}6f2bur&ORa`SlF>GC-)eOX9%}DdpPi#&E4xqoT$igHkJilew!^^lwb3wt<i`PV)QG7>t3Ru*+QNT)Q97C*v^~V<V`2o%9cl8aElJ24;}1m{ea7gA54dx67Dkri!&&!QJEo&~k<6;TfNAYsGvk)Z{pZQy`#p->*+<*75vV^@k#`53RAYUbV&ELk+v1F+&oW?+av9jc4!W`M4t~J%%nV;NJbCOK&$zuPz~$dJZAUQdGQ>~_6}rQ59Z94{Mo($I%_?(oGMRTl!DbcPAZF!U?uy+^3!*Q^q;<be?Tujcd413Koe3CQ4vQ~5`gWCH=f!Ovx+~Wyn~<#P<b2*#Rh)u$YKvW12EdCH`}C|Ys2mU(0|7CbpyGdJ+Uj&7p060#iE_}~jM4p<>=;q27;abr)kU4mKSat1HFi3uK{=}s;)or;?3-hTikcov<y?ADoi`2?heEy~z&Is&JQhrYcg1wNDob_*ZOm9^#w0Z|e85VR<CbQnKBJ!R5Yfi&&OBM1&ywJ79Q22GXRlyj-<@qwcr(Omj!h-BNy!)C_PWUNv!Y#Ztd1<-A=MXMPss>i<^c=BSP@(9U2O@&O?&<$o9XKq7B}LqgSxGn#c-ZgZp;lc3utHls+i;9Q`CLXNu%Eiurn{h@y$cF4w>E{HI?L-X&+}ufEPt#8=ID*U|3v)RDR5-DXON0q@?+30iutR2DObj)v}U$npVlxoc&Y_F!<XKusUtC>v$B*IyiVA^dEFiti|zW)kOANE@jhnDbMS$Lz!LIw0@Bjr3oN<$n336W>Qmv3wZ#k%W9OO{9#C4+_N)U#wyk4L&v?HmNQRc`$%twyX%(RB$;se@+gI{OcR*P;6)2~KC%$U1)f^N9XL*u!5q~&E%JNaB)>YSNSjM_VZQDr%m%2nwu{U~!@CZ!V_kC7?Pp7R3%i(%Yee{DSuJwZ;d{1_G4bkxu6ByeCFDmntb4nfC^nZ`CUk8<We$_I-XyMgD0$1+t2)gMLk@&aD`{<aIg-uT&Kv#^9)HcUP$%(ZE(F=%Ell(?@+r*J&H!5CLKKg+Uy|z;S3QHkN!MU<Tveuhd+}p!mURM#o95s_$77Lh^sDGH6P-MI`OoILLfn<*iqQk|1^$n-TOC{N%x!TP^`FIUg|I7^-IO5zd<LsCyH8~-nv|%WN~K@eIuf;$Vg@?wv=Z88=)}=ld{EuDg4Q%h7uZ->8?BqP>{EECsQ}xpY2X~CcQuA!xYoEb#2Ww)MB*6YH~iGOA4L{c%0;#ld>9^2CxioU*O+K&y~}%y_6xKw(Ynpz$+`=}O1(NYj=6$N6x;@sbVTat9zXk!2lw3>(fFu|KYJl0g~x218KoyghymOlFdT|(3d1RowT~20N^_mdya69Cz#H-sShB3XY_PCeRA(4E?@VplPLfl(UBUXCLWjqWER2m84V47cr`3XW&#ZG5&!p?);1ec@L14-D*lc|31cHJoSx?IgC3-zyJP40gw#SpEQPdlaF=2!E4$mqJj5HcBN_pHE=pu$0k8+b$De%A<({w(*s>#hg!i{eV|G1-e++iz3IiTWBpHCtKRJB>jvScPKcb9+~Z5WtTC-Q0%OgqvEDbJb`gp%pqk+Zrk0a}KW%{1ugP(FYRy}C<@?A}%jK~q_#h2R6kvaC2p;)YhX<ZHo57L7~{rPrx`in-Q;Ws|k0hmce18+ZP*Q(>92pvaH)6nnzWt;GkxnL;s%K41~R!MOq<LO7ir3Ztb)Dpwiysn4(CoyH6QnGScCLsV_rmXHw9IgH%o<r!(JZMIx!C=3wZ{S>*U!cWi|H5wJIzOzCnFhvVkK3aIC#!j;e_VG^WW7_HfIA;iBV(t|A>Jj1(6b{>NsmmmORw?!o?6HNV9bDWxmT+;a^^Ro(i?`Bbgu+@St3kpNHj(JIu%-mGG=7%mG{5LtRL9g%c%u1QWhy`mf%#<v0=uw6AXENX<g3@4XxB?>a9QR<8?1c039435n=pC$K?|z(dIMOaq^9jKlcK|4BhO+N9=jD3id)L^W?<m2Hh`6-Tv*l`$OF&2rkCY3K`v8d%S;cd_brF41OrryRR1!aQqzoqX9zY8_lEqJTn|{I$%>OTq}YUJe;{bq&~cz?c{2UkGX+)J2yubP1fSllp8|v$3=RY>rKUg8CWoQRzK1h*Qqb0np+fHpv3KC{{FQ-vpbA3pwDJ>UXLp{Qr}Ke`JQgR`$?(rDX|d1Lko=tyfXT*L#X{g&d$8Thc9V$hAxhiF)rmY^=u?f@mV~Q;p)W1r`VfbsgQBauQE)tTARzngmuiQ@Ox9|tH216FlCk;vnO{NqpiDdzi1e&YG^p<%M}-b&qHkQw<ggYA)~Agc*(pD9yUEFtJ9(1rWvAHXBihrQPgqA-K;318hkBX9_rvH{8GxM@XyWhzZIFOlT4Ck`eo6zZOv0RYTZB53ZG!WtFrL>H79sVjA~Rit3~6p&_SFus^2w^Z+4hm3<-Sv7hQ!4zsER_#?C`2HzV^i6dP&*92TH^ju}G)N37E21bp~qdfU61Wj4irkXw-ya#Mnzc?n&#4*@{^;EbzZ)pC(%@3%aUA9H%<-_#-JK0lrZ(r%;6RT9H*w)+ZTD<>?;_v9Y7jun}Ze5K&pa@In!3`>NM+JS}85By4K5lhy?tm(3YjXec}%qZp0xZ`Q7hmq$j~k7CbSi4QG^>0nu>EefPij0j_(S<aKBzKTW8P*H|^PVkfR<@7v}qpJ3}*C|RAD*z(Q*+9%y0Dw{H-7?8%pi4@NThwabIbYY0FVzzB$^#0sQJMc(r#KsH7H&EqORVAG@zGa=>22ILFjD2V<-x(Qr-r<`N}VUn6$i!e*&|lOEU@OL^(vXi73xh&rWo?nNSKg;UN_aGD&lShyAmcSz%!yIqckj}>9@VoZUm-zy<M1Rh*F$b35OzYX*;V*dG1e~;yq-01(Ok}d>rR2lyuB%QjX&OtL@ic1TO>p6Agm%u+2lKXcqyt1m91#)j=%B(u6k}AZS^W+Uhe^uv!>h<$}NtbBBSf%+Y&8qcx2q+>P`ra?A@ky;{?n+|(8t*z}K^U){VT$*iHK{>iR(o0NFVnZKg>(XM(v$BU(A{;K9jC!NVP9d+<mH9y)_rONWwwZ1yJQJ9a7M5%KxU{s`L(|P&sRI)5dts}sI-EQiIZeVz6A0WZ@Zh&}cA0W{Q-2m~@K0r=?Av=}v*}w=b;$DlR=Y7_h3mu^NS+MRIhaLE^-Xtqhw~Wn^erC%QB_ub*&@IwDyak%r5WUQYS0DL4q+a2jw5bS*+~_F}YEyUkko)@jX}H^(-W~sj<{O$!l3<9K2feBPeCT~b`ZU~a&F_wXL-Wn?+brVNIQL<+i+Bxp9)2vQ&|Z7C77^5e^ymjd`Ri_nLVdX$^~Twa^^`YSik7kd+nbk0sTHx*;y*p5Wfw86tI$wF2TLRSnq#>iEC7yk!kcVFU+iNF5B;}~X1RVxH1bDt>zEi@=X#%Vs@Gg#PFmRT42UGCHaxLr6XI!I^~+aDyRPPAi4&lF#96_OACJuf=+CCZ_pU!0nSkWO!X1vW5nll){Yjp#)BBI6M$4+)=q6E#hZ%W|9*(uAN&MEQk0cYJO#INr4N@PEfgnx1IlA;4sXh{s2FyG1REUysx79~tRj|4uRo7)b7itc1(&Q82>r6TtMqW`X09wqH14@jezj9oed=6%?7a6-kjf2EUMm^9{KjdYlgjJ*%mlbGf7WAWVW**Q`<>1gUh*Xuv(K0>V0w>n0qqjjTElTN26W;d7Erx*jSCRa{NcD7km6mCo<xfBCh?hLt=N<iNe6$zh?kRa{nvTQvHk^75!|;2gmSy1}w`V|i;Ni^?n2j9cUC9ZmF5V{B5sbNj-64Y>t+ZSgY+HIH!l}g`jC%-@n%8i~mM0dIcL#a)q}4G)sY-eA<`m_Tkz3h7(nLu^;;Hj2pV8sJftG^`4W*QBnBAeb{q84mwPcmiItaCjPdbx3W{qL>qmiVt>dVJ|$je~JC;EypeKO^pj-HqmOg48QHW+X$7+h^h5yrMi&+ng>Y{EK&>kX@6T}kk}1>tBnk<&E{8`4b(PrsR8!o7&h5kJ%S$q77D3F;?ec_VpmbP2uXwZy4vO6vDyE>g;lS!glSs~OzO6x<yXEmi10l3_ffy~Z!>JHe{PRIzLo#pLm$gbfs2L;eyTecLfz$p&_?JgVV0>0(*mf2&_>;NG_2x!iy+t0|?SopFPh#r<51qN*t~c#-!b<B&ey@cf|%^nG(y>p^=fO<dy{dzeqYTj33liOh(EM{SF$8uC+D@>1$8yglK5OJ4WQ@Ac-hw49l+nf&_k9ho76rYa+?vDr2`pBcHH=~{C&xp;2_?>GT&N(A5mEofy2L7;moynv58Ps-%>32}%#$rIl6)|wGmaz`<7Y)W9jwz#xJkOMUfzB&Bq{b~PY)MYYsHp4`cG``V%Aw>u?t!=Y$fob1(@;Eoc8)dhmD?^%-5bHS^Wr{(1XXP2BcXDdv8OV)nv#jmBHI3;>ThXg48I**{L*{l0Nh`xCZ*0fJZ7&qy7e?;3CuNp1I`Dx=SbI{warr!>f|Jh4-;C^d>qx}_I7fn{G%FMjMiMtJBi_VqiZaZ8uctrLQ=e(2_=sa;lHIJz_6a@OU;uRTi{iMNYN;ZNXw@iKX>%1qCj+HKWs1_Dit~~N2w71pqtz_VtF*jHyj2x;n$cpttOkPnikoW-XR@lovfTTzE|X<ig8r#uy_8lNXM?N_A1}bGP@PW#wj3(*`F&6<Q|2^E=3DIeAX_FeZ!Q!l(~Fcsm1e>DeemY^9lV)>)K{aPxF18inOvMgmWLwvkwmvxXxQFT3YZd6*R(eQG)lC>;wht=0IAQdQ0yDn<$#t;U~lBHoC^n?ZguC;ejcVqP23jo1jaEG>!^!?CGv3C0n)-=KRn}lc$e4K*LH|J8`LUy;IMa_YFiC_mcG)!sbGDBhmkF94pW;%O@rOQOb1juLmhw|^RsY5z>;jxn(Rm^bl7tb6$Gh{ZmhH;o#%9D%$6DPwoz&j;93XZ(__Y`F5;^yJPQLfaI*0xxfTx`k@53EgZn(P=KwsHQ{wY5A$1rq^O6262u+CUIE>9BNABgB=I4}rb-bDDL9hHjA8H*Q-B2hb$l$zKiN!MeiRSIzJm-y_@iS<tuB+j(uIJh+NsT3QPqr4fb<NmT7HyUFaXQuq!^TRftk6!LC6i-I>^rth?$tQ&Hn&dJ)ql$U0dDuRlMb{}=3YE$QQ3An7M;;GYtnD38R!F^JK8k;<w<!m{}BQ%26Pbg!Kh5j1zAy&<Etz`wgAwV06rM{_ZBeRLia)7pIR#G%K{$+{d)@-8D`@`;GZ^Dp0!sNg)H)w@HyhL(Qb8WG?_haNudTJX=%2h$1xw-27*(;Kk+q`OU)_EDP)f6hUtZ>;lUKW!=2$wp$N|6s~IA+b`wBhhC8#4k0L`!1hygN1wp?z!@%m&D-m-~Xs0ZWB&X<FLVtzk++)HiZbgP)+L?f!ndJ*rQPpYdKp>SqTYYsYMhzk2_-U;^_GA_#oU}(E>TC4Th5>(V9<|L_AiJR{H{g@B?kHVORS-D>Qa53F6m#U91reX!2>GYBtY?$hx#NBPpkW{JU6ck+D7o-ZXlIe&BZl(5m>Nzp(S*|CCb6I^C5Pv6EN$2h-CEQ~-V-|-e9Q!*C~~{58z~Bc*R^0#p>nhv%DcT``oh}ZDghRoqzD+dq=@&US9RSby@i?SQ8~9X@<S+7b_QJ&9(hxp0aocxfNsx6jYur5`xIOfjfnxsw%3&mDz@7Xj!kMjRk@S$jC&t(BTx~hJ6ZGvZ{)qJfP=tJW{*t^R{udANuNF|%LcbYjH0F#Kdfx?(V+7(O|JbaM`$*&b?ylU7G>>dVDbKg!+|y5`uI7no?$+HJ7-SQ1cIT|y5^-W^;eV@B~{|TCf);g%IW_sO>$N+M4Co6O{_jrm@GBMY1$T>8|tmbM<=2i2D|^_NlioBi1oqP6|pu;Y(lK=h%m6q3nRLs6k2=}KGVjx(ZZ|lC=IRGfKI6|8<qOf9iui8bVBUK`9yRZF&T={#>hT)N6Gc^6MC^J0`)^icjI!aagt?9$TQZ$ut^iJxX)GokEdMB*T&o}DVE~9fmF?ocutCaYe-40!f6q0kAP0J|0Bt!-zMEYk~dLnMa8eR<Fv*~*KZck%xw=neD@)yKq&T)pM_tRtIJt~+6S>zY{1wximbApZTV-R+2`fkNO;4({qu}&kIQc(LB*=REKnh+!)(t$isd%kNCiME7XKhBW?iUQ>kcD5VR~btNb(S^rK9#fZmjCHNL}G)u#8kMBDAt^z~gJ%>b^pQA&Le~25`Q@>{PV;vI*yxbvD-MsT=0HxXpCXO<psTl|5@6-a?GILSjTo4JR%WW6Kh8w!r~1-~}kR2Y*^SL>Oe2=^`nwE1BRx+<o{_<e-A`5@Wd0Inm=bu9(zod$1roA+2~qQV11{fRL6+CYzJI$U)HNS}2jR((2Ysi6S-{tYEU_;{OxP71GBeP|ZR5L9I(}6Hqd^CO}xv+Aemd9S!V3z`p5M*(t3cPRxyazGOh%nkAvdT$T8Si>YNqntFfx*n^WU$&CnSnr>`eAG$$R@(z4j*A?aRD#=w(vF?|teN1KK-NPDhCe3Z3J_nCmB{svT=ZhqpQ^JKX#am3nb-+NeqVuGxWA>rzaP*lM9K?Xjvq}=&#ER|=KsllIjGYF;CyoBGzg1$|mDlaCZXz>|ZuT%0viK%-vPZZanetQ<s{8|v;)&FHbo32At7u!TkUAOevA~TJb~I!9OzV65Al}@QZ-u~NFn;V-k><$1TB@%R2$+IxaE2S;Xm+@kD@JbZH<hNug?bh82inCokD6Xq^V~(7%zmSyZoJzy$!s)AbE`D!V$2&Ab#>q*gK=d(rn^ex=~aTQ!bGZL%wb4(5wZg>=FvajuyK;t)I9HiaUzK1PL5uu4(a;lN~Qto$WxtO7W_E9w>~d%@8VsZ$D6BbrCbnodR1lf#=7jYQ)`B?PqkQ4dTXj8VWTXU#oQ%Zj~KxMp?|pFKzH-Cpm}ynhI;T4Mj|Io>B#IW{BGx-F~TchqqwW&kneT!Ze<5s9NmTSVB_b#YF^*uLGS|>VNO?Yil{LsL%>>{H>PuL<w7Lhh|yl36}RHt@;Gn~y3s)^_+{LB%TGwLJxHnf-8<I9PbkX$4P)T1`LFp;xhxNpZ}1tLcr<YQ1QQrBp9SHsIXeKN@{{UC#hgJ$(nSDMYGy>26!4ii*b0tOcmz$_O+@;wMB(XGR&iu_VCpi>=~B`e&wL92=)iA%l~MwWAX(MLg0$8U6NW?oEP`1Af`O}O9k8MV(GLPCte_YxJUU<;ou<8-bPDW*=Q3r7wo1@nqXxn+!l7y`EJPnTlZaz4)MWYsGghswipSj8v}$QHoVS|7$Fc;LelXstcN0LdfY=Dsnu^=us+bWQo;f9Zp&W@h%rspsKrJ(!7wQ3s<0a-7+nU1sRN|J4a3(yc!~zU9`|yH1G6o<+gg*h%S2KfKNckenL?eB@9s1lvwjgGIa28ZF?u2cPt$6^DYoMHfu0w5fit7{3&}WM@p41}aemIFbk;Lcc=Ow3+hr0&><~c(dqyuVA?0z*0{u@;57O0!}<fnJ<c8`A^EoT3%7kSN#u)~mdMP4lKM|j7NQR;K86wO0y(C}HvE&qj9J`+;JGmE*UzZ?a}$*nvaOWKj2ZpAbXW$@qZ7lC_e)Snn<r%^ex`=E)DQeU%4W|3UW^iF*qKv*ZJ5j}Ivz&I@3@EMOBH-4nWk$$na=fHN)oXR|N#oqqmyZ!xO_ovgtce|$thwp=<x4Wm`9|DVp!QSra?svN<`@!D9?qA*?o}3=&mpq3%Vem5e@=yAp{)?D2K=MTE*#MT)2|GS%sDhIzEVSx?n=Ve(iarb5$F97`ryQr6n2TPh{4Fn5<=m;y&ZTy$&d;#gAWoSu&9aIV8MU5PMby1k;Wxz`Vc_L_q6?`o`{C7X6Jdy;K(v@3mJ-;5((=Y>c}ewy1oFbj=E7C9vZ1(}j*>tKwBBqnq?OF;wRkxv3Q>_4zT@C!+8RkJyRaTDC3OM&B3brHusLKsVn)xfu1J+RXgu+W0@IFmN0yQ8$3r*Qn$C9HnSZ-5^lu|xghw3rgPw-<g~m`*y3EP}%pJ6=%@WUzEhy}5CjdhAJV*|_FjDrl1KFkXojxE&9=~=Vv~dxLl;au<J-J^yhyzslEYMesBL&7txaTn)+^c2$w`md_p37NmA{)8=2=v+b_4dRYpqlz~GwSUKK-)eLP~H_h^C})R-!i_S6A$0s=`Xb#0pQ!%sKBFHkdF>}ew1QfrB6{JFl%Zn8cm<Dh0Ke{Yi{)X1I8XG;rHxbsvH&?unPI*;Am`@02$`SMtxT|=F9$JwMg<I+A)-n0TnYF-!?F*_Hp8G0Co%lnJsv1AYa_hg1zr#Wfv@7__$~ViX(vXAPj<|AK{<x@gIuePuVgE>>tp1v?_4nGan<*Ski!HpAu8yl3H98<%hyNm9R9<vkTD07PzCd9I~tytQ<=-QEe%&W&_-6i&<0ZQb2q&t!@N$EVBSlnc-)nJVc}}vQEWCT!PLwdae^8t1wqz=0%lx8tvMwzOFG+rUY_{(kw_DELaULTA}^~EdkGg&;;ACp&gp#!etOttMjS;lNJzMs#r^!k$6pNQBKU;<bmB1{&WP49XR|Qv_I4gE;34=<!g#GfS7~dC%s_426>SWF976RtHDhs8>s#JBn5qlROCgSdi2MBev_3&PEm!+q^fQ)Bf(Z(T&H=pr8^%)@(W7Nx0=sKF07*_Cx!?kJTGR%p$*{JEcB%H27Dj`KNGR&bcJe3H*H+f^Jurcq{aOaed*6qW=sPscVgVijpG5%af_s`(=wjUff+lpGOU{RRD~bX`SN>FG}7>mk{MX_98o_U4wq$t0er)<C~zMeZ^qlfD>(+i6~NdDeV!vM`!)HLhBFu*Q{)o)o7l`&tNWM2R@}mDAA+eL3WO-}CG?M6(ZoL;zx|rJK&-ec;i8;{8`?(dWr!Id@xnb~Wyq#Wyd{d|OWQe{2;7=g8`OnROX=4*nWjli7RCaqWS|xTW>2@_r_1{4p=R9@doe?lME-hPAiupza~K3H^<)J~0dGFUAWGDtb4*8xcjVP<t_AgfA29fVgx^ZcBDot<S{N@CUv9&Hf)dnl(B<bDX$PE=Sk9VO_}D>g>S;%1U?J_wXR2wk{5sf%p_t^D0t$b(W8^|GlwN7o$#PhHR}^yqc(`PVPj>0EB8K(?IXxU)kbt16tAm)+{_7ykv1ybUg;#DTj$Lpv6qnix;Ul{}hS@lk0*^^n7Rwbd;^c?j;p;EIVn?}(Dd;I7kn;8({r=(x%*Cn(Xuh$`zM!2(SEYO{T4K78ac90^90Qt4-2PmS*d@^Tt}o7p7j!!`&L=TfEHS?8W1*4$2A82o+Gu(^>+|)YPzN!ngMs?Rk9weUy(;FIjae&4kOzdA0YTFbCeXza*TVG(^BLf)QI01i>JdKwQRL}pRZnTx_yXVb!%u(ybn)rz^wS?cefR0fFXGrQ<JV5W=;uET%E98l@af3m?0m@8N1aWF50b3)=#LEKdc+$eIA>(;YDm_}a$3b{ewpPI;z{~|XYSOJaUD?ceVlz2%yd~|BC~M+`1tU69Gv{$x4Qy5OSmSN3DTZFy05Qr8dnqWNOL&|fM9OZJ_DP38124E18uCaTOU_cs~d_+UNU?SbR$&E(u>C$r`Ca?(t~e~j`#n1u>ZIC$Nitb4I_JZ=XKBUy|u@@(ar!~r;1odr+BYf3CqPjF#s?t*hO|{%!g(dv-^e+1YRJh7=OOPdxXSASK>PgPN-^!R`Jo|8VdC>e*-K9SYA_=4~(pHgn|dQ82Bw7$WVozVjP+lT+Cz$v(M==)i+1r5_Bh;^6^sUR#grs=`P)H6IW<@1Va$LH^&D@r}1C+k53K`-&;1C?s~*f;7ubqs|sck+1F`>SHM{Bu^alr@A#_lmI&Im5HI`QJHEIOV-f}aI1&&33M9&i;JzC1rhCsRVG3Fvmx4vr3o9_(4pd;O)Zacb<wUFx=tU3W%M7=pZZh5_cq{>PXePAt0Uud_)y<GpJpTA$e|ImOa7Ehy@VrcNI(*^<$G^hUa9CXxw?jV3Y=X*fjckC*zY3^j6F8fJLel*lM0*s&>MvljIj3nx9~nMM;dLyJdRbFfJ0-<ZM2t6I03tGd^J!GR^Id>IHesw~^`=VJYfnXyU{cf&zdc3K)MHTvv0cma$#>m`9Zw*tJ0Coi^{eAVJ^fufyS3R2S-$pO(*lK5((Z?sFPWt?!>cZDXiTBL9l;NbihKEz^3fL6h{M*>V+u8SX5f_xe|frAg-&c1>esL{=TvhdQ+;#F$nqOJ20WwhgPyR{a38gUWt&>9h-H#pf4Sz=q%`BQd1r(Pi$I<AX$qx0y-v6D?Ho#&s%0oemF+p&Ex0a+`2?>UrYfBK^xU;(iMM}HMFTl|-w_A}MwD_MobVuen0jE5&Qi+?z^K=Z(s7nw6d_AU8hP(il<dK<X_(3`!D{*)l$h6x?ev9pSEtI2PU+vnll{}+^mzCE$?hB4c07j3AUJ&g_UEwSEVP{!7P}SRH6!KDB^mw-YMl;;0v~F+2`t7GmMW}v3|998u5bDPHX6k&xmAqrba=52*U{4(?(a4}OU3FPcATCLSY(m4OZKBM`aooC$Ztl)%6$MLAtI(Y`cUx#jX0jKtL3c8oOAWZb$2?qMg<Iq4C9)gOE4I`%?vcl{jcYvt6bPJvnVNP5j|52-DZe{bGDjP&em8IaqA{oiIXj6oYw9)?@R0h(U~q@E44YyJ1YNqJWxr#35R>-GaE$IefS`+Io5GgRS$D0n+2=J;vG$7{@x=S35!z>>Ibd<dc;UyQ@VWZ13O1b-#$d6e`F$aw}qa3=3cDTnQZJt<T`&I@%Mqh5Z;Y88GrhKV)5StzQV>qB#?gWe8lHl_ipTt{YpEMYFrkjk+GevBYE%bXr0(h4B6{^K%v@0Q)#tWvLV@Tz4Mqvs<ng^sMb=tu(cNSER!2Jw5bp#Pj4-PsoGx+%ym)JUZwkL$f?y~5#$emFxZHTS6mNqMN_7uYIWW(!?VBd4u468e{2u`9FK;R&rq;q_(uwLGtAo(XENLY>}x)#EmC4Sx;5<*hAjGf6S!%0vk4j$*$B*9*O-by3G_c=?rrVSht{v+KnGJa`UZbE*j&-oN1N+((^p-Pm&r6ePo~#&W{N+Lm@^LAFzs&`AnO*P?se9KYW>M$-<)e-p02Al>nSUtZeuy~sM*J<2#<;T6tOiJwc&P_4RcxVY38WeDY`L=>tDn%=ps$)r-%h26#^mrW&c}*VzU#^H|<ysyl4abC(I!m9B9uZf4oQj>5^ym{>Mn)lbXGGb;f;O6p__gEkBV|YtmMGMMfMQCH7~(J5=?D1mQJ&EGXK5pw-jj0K;g+K%2J<sz7(nqd@T}U<a#5!AY7nL?w;HaV#@n^F)RK3f2cd9lyo%Fir0L^{=^S0!;I1L6;1MtNLR2C#!@O;hg1on~Eh4%+EQ?s%257`cB*jc8|B37RDF~yMIF!Co(bQNeZe?6@9^o>_$fY^#)KI!v;eZ0}~luXM6h--i3An&o(L&mw_TJuB{zSjWhv%i{De`j}b&p$!SfTMv|8zbpdQ1(-fEOG{E+r>7)SoE~t!HL=SZ1uwoa0-nc^#V19_TGf!CE*d757bfLAnEY&=Kr6J`TDAr}Y53+0gd|YcvQ>Vy`JT>)}_n@kx1m+c6Wo#G+oi%G0t!kWb65JGHn~f(4*hi_lW{#p>Mr;f{Wk5j>@VZpMNw>^Du5#=V?L07T4E>R@NQ_^yB?g)ISz)}of)vyOPnx1Xbw?IFIEwf7zkj=Xy1%DN%+s56zVj9BD<$Rh3U6(6&J19E6cWB0$0IPx%Xx9$4_}6nKLagc%_=gP`VDDjIZD<qeJnX(_2{;Ayji|3I6krCC!%8BunX)06(hoR9n4&e8Ecu^oz*yjNlt#mdmZaOMmWV4ya!R~zuJENMes7fKhb#f`eGft%f1WcZI3)orl$;EP6!Xvo3V~D&Of!p{r0Qq>!50$tPSGBlPe6l#V{fjhXy@5pz(VTLp8w7Rn}8<t8<y$s`(sURHM)c$KCGrdeA6gLD3FC9Ag-69P>SZGRdIs3FSQ4zsqpQWORq<{{zn_kaG'

HARD_EXCLUDED_DIRS = {
    ".git", ".hg", ".svn", ".venv", "venv", "env", "node_modules",
    "__pycache__", ".pytest_cache", ".mypy_cache", ".ruff_cache",
    ".tox", ".nox", ".coverage", "coverage", "htmlcov", ".astro",
    ".next", ".nuxt", ".svelte-kit", "dist", "build", "target",
    ".turbo", ".nx", ".parcel-cache", ".vite", ".cache",
    ".update-backups", "update-backups", "backups", ".idea", ".vscode",
    "playwright-report", "test-results", "blob-report",
    "whatsapp-session", "whatsapp-sessions", ".wwebjs_auth", ".wwebjs_cache",
    "browser-profile", "browser-profiles", "chrome-profile", "chromium-profile",
}

SECRET_FILE_PATTERNS = (
    re.compile(r"^\.env(?:\..+)?$", re.I),
    re.compile(r".*\.(?:pem|key|p12|pfx|jks|keystore)$", re.I),
    re.compile(r"^(?:id_rsa|id_ed25519|credentials\.json|service-account.*\.json)$", re.I),
    re.compile(r".*(?:cookie|session|auth_state|creds).*\.(?:json|db|sqlite|bin)$", re.I),
)

SENSITIVE_INCLUDE_PATTERNS = (
    re.compile(r"^\.env(?:\..+)?$", re.I),
    re.compile(r".*\.(?:pem|key|p12|pfx|jks|keystore)$", re.I),
    re.compile(r"^(?:id_rsa|id_ed25519|credentials\.json|service-account.*\.json)$", re.I),
    re.compile(r"^(?:site_config|common_site_config|auth_state|creds|cookies?|tokens?|secrets?)\.(?:json|ya?ml|toml|ini|cfg|conf|txt|db|sqlite|sqlite3)$", re.I),
    re.compile(r"^(?:helpdesk|paperless|docker|compose|database|whatsapp).*\.env$", re.I),
)

# These directories remain excluded from normal source scanning, but a few
# small credential files inside them may still be collected explicitly.
SENSITIVE_SESSION_DIRS = {
    "whatsapp-session", "whatsapp-sessions", ".wwebjs_auth",
    "browser-profile", "browser-profiles", "chrome-profile", "chromium-profile",
}
SENSITIVE_FILE_MAX_BYTES = 10 * 1024 * 1024


RAW_DATABASE_EXTENSIONS = {".db", ".sqlite", ".sqlite3", ".duckdb", ".mdb"}
LOG_EXTENSIONS = {".log", ".out", ".err"}
BINARY_SKIP_EXTENSIONS = {
    ".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp", ".ico", ".svgz",
    ".mp3", ".wav", ".ogg", ".mp4", ".mkv", ".avi", ".mov", ".webm",
    ".zip", ".tar", ".gz", ".bz2", ".xz", ".7z", ".rar", ".zst",
    ".exe", ".dll", ".so", ".dylib", ".a", ".o", ".class", ".jar",
    ".pyc", ".pyo", ".wasm", ".woff", ".woff2", ".ttf", ".otf",
    ".pdf", ".odt", ".doc", ".docx", ".ppt", ".pptx",
}
TEXT_EXTENSIONS = {
    ".py", ".pyi", ".js", ".jsx", ".ts", ".tsx", ".mjs", ".cjs",
    ".astro", ".vue", ".svelte", ".html", ".htm", ".css", ".scss",
    ".less", ".md", ".mdx", ".rst", ".txt", ".toml", ".yaml", ".yml",
    ".json", ".jsonl", ".ndjson", ".xml", ".ini", ".cfg", ".conf",
    ".properties", ".env.example", ".sh", ".bash", ".zsh", ".fish",
    ".sql", ".graphql", ".gql", ".proto", ".csv", ".tsv", ".go",
    ".rs", ".java", ".kt", ".kts", ".gradle", ".dockerfile",
}

PERSONAL_VALUE_RE = re.compile(
    r"(?P<email>[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,})|"
    r"(?P<jid>\b\d{7,20}(?:@(?:s\.whatsapp\.net|g\.us))\b)|"
    r"(?P<phone>(?<!\w)(?:\+?\d[\d ()-]{7,}\d)(?!\w))",
    re.I,
)
SECRET_NAME_RE = re.compile(
    r"(?:password|passwd|secret|api[_-]?key|access[_-]?key|private[_-]?key|"
    r"token|auth[_-]?token|bearer[_-]?token|refresh[_-]?token|"
    r"client[_-]?secret|cookie)",
    re.I,
)
SECRET_ASSIGNMENT_RE = re.compile(
    r"(?im)^(?P<prefix>\s*(?:[\"']?[A-Za-z0-9_.-]+[\"']?\s*[:=]\s*))"
    r"(?P<quote>[\"']?)(?P<value>[^\r\n,}\]]{4,})(?P=quote)\s*(?P<suffix>[,]?)$"
)
BEARER_RE = re.compile(r"(?i)(authorization\s*[:=]\s*[\"']?bearer\s+)[A-Za-z0-9._~+/=-]+")
URL_CREDENTIAL_RE = re.compile(r"(?P<scheme>[a-z][a-z0-9+.-]*://)(?P<user>[^:/\s]+):(?P<password>[^@/\s]+)@", re.I)
PRIVATE_KEY_RE = re.compile(
    r"-----BEGIN [A-Z0-9 ]*PRIVATE KEY-----.*?-----END [A-Z0-9 ]*PRIVATE KEY-----",
    re.S,
)
HIGH_CONFIDENCE_TOKEN_RE = re.compile(
    r"\b(?:AKIA[0-9A-Z]{16}|gh[pousr]_[A-Za-z0-9_]{30,}|sk-[A-Za-z0-9_-]{20,})\b"
)

LANGUAGE_BY_SUFFIX = {
    ".py": "Python", ".pyi": "Python", ".ts": "TypeScript", ".tsx": "TypeScript",
    ".js": "JavaScript", ".jsx": "JavaScript", ".astro": "Astro",
    ".html": "HTML", ".css": "CSS", ".scss": "SCSS", ".sql": "SQL",
    ".rs": "Rust", ".go": "Go", ".java": "Java", ".kt": "Kotlin",
    ".kts": "Kotlin", ".json": "JSON", ".yaml": "YAML", ".yml": "YAML",
    ".toml": "TOML", ".md": "Markdown", ".sh": "Shell", ".csv": "CSV",
}


@dataclass
class Candidate:
    source: Path
    archive_name: str
    project: str
    relative_path: str
    size: int
    category: str
    priority: int


@dataclass
class ManifestRecord:
    archive_name: str
    source_path: str
    project: str
    relative_path: str
    size_original: int
    size_bundled: int
    sha256_original: str
    sha256_bundled: str
    lines: int | None
    language: str | None
    category: str
    redacted: bool
    binary: bool


def eprint(*args: Any) -> None:
    print(*args, file=sys.stderr)


def human_size(size: int) -> str:
    units = ("B", "KiB", "MiB", "GiB", "TiB")
    value = float(size)
    for unit in units:
        if value < 1024.0 or unit == units[-1]:
            return f"{value:.2f} {unit}"
        value /= 1024.0
    return f"{size} B"


def sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def sha256_file(path: Path) -> str:
    hasher = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            hasher.update(chunk)
    return hasher.hexdigest()


def write_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, indent=2, ensure_ascii=False, sort_keys=True, default=str), encoding="utf-8")


def write_text(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text, encoding="utf-8")


def run_command(
    command: Sequence[str],
    cwd: Path,
    timeout: int = 30,
    env: Mapping[str, str] | None = None,
    input_text: str | None = None,
) -> dict[str, Any]:
    started = dt.datetime.now(dt.timezone.utc)
    try:
        proc = subprocess.run(
            list(command), cwd=cwd, env=dict(env) if env else None,
            text=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE,
            input=input_text,
            timeout=timeout, check=False,
        )
        return {
            "command": list(command), "cwd": str(cwd), "returncode": proc.returncode,
            "started_at": started.isoformat(),
            "duration_seconds": (dt.datetime.now(dt.timezone.utc) - started).total_seconds(),
            "stdout": proc.stdout, "stderr": proc.stderr,
        }
    except subprocess.TimeoutExpired as exc:
        return {
            "command": list(command), "cwd": str(cwd), "timeout": timeout,
            "started_at": started.isoformat(),
            "duration_seconds": (dt.datetime.now(dt.timezone.utc) - started).total_seconds(),
            "error": "timeout", "stdout": exc.stdout or "", "stderr": exc.stderr or "",
        }
    except Exception as exc:
        return {
            "command": list(command), "cwd": str(cwd), "started_at": started.isoformat(),
            "duration_seconds": (dt.datetime.now(dt.timezone.utc) - started).total_seconds(),
            "error": f"{type(exc).__name__}: {exc}",
        }


def safe_environment() -> dict[str, str]:
    """Return the complete process environment without redaction."""
    return dict(sorted(os.environ.items()))


def stable_redaction(value: str) -> str:
    digest = hashlib.sha256(value.encode("utf-8", errors="replace")).hexdigest()[:16]
    digits = re.sub(r"\D", "", value)
    suffix = f" last4:{digits[-4:]}" if digits else ""
    return f"<redacted sha256:{digest}{suffix}>"


def redact_personal_values(text: str) -> tuple[str, int]:
    """Compatibility no-op: this bundle does not redact personal values."""
    return text, 0


def redact_text(text: str, include_sensitive: bool = True) -> tuple[str, dict[str, int]]:
    """Return text unchanged; privacy and secret filtering are disabled."""
    return text, {}


def is_secret_filename(path: Path) -> bool:
    return any(pattern.match(path.name) for pattern in SECRET_FILE_PATTERNS)


def is_probably_text(path: Path) -> bool:
    if path.suffix.lower() in TEXT_EXTENSIONS or path.name in {
        "Dockerfile", "Makefile", "Procfile", "Justfile", "README", "LICENSE",
    }:
        return True
    try:
        sample = path.read_bytes()[:8192]
    except OSError:
        return False
    if b"\0" in sample:
        return False
    if not sample:
        return True
    printable = sum(byte in b"\t\n\r" or 32 <= byte <= 126 or byte >= 128 for byte in sample)
    return printable / len(sample) > 0.92


def hard_excluded(relative: Path) -> bool:
    return any(part in HARD_EXCLUDED_DIRS for part in relative.parts)


def create_temporary_git_dir() -> tempfile.TemporaryDirectory[str]:
    temp = tempfile.TemporaryDirectory(prefix="deomee-ai-git-")
    subprocess.run(["git", "init", "--quiet", "--bare", temp.name], check=True)
    return temp


def gitignore_paths(project: Path, temp_git_dir: Path) -> tuple[list[Path], list[str]]:
    command = [
        "git", f"--git-dir={temp_git_dir}", f"--work-tree={project}",
        "ls-files", "--others", "--exclude-per-directory=.gitignore", "-z", "--", ".",
    ]
    result = subprocess.run(command, stdout=subprocess.PIPE, stderr=subprocess.PIPE, check=True)
    paths: list[Path] = []
    warnings: list[str] = []
    for raw in result.stdout.split(b"\0"):
        if not raw:
            continue
        relative_text = os.fsdecode(raw)
        relative = Path(relative_text)
        full = project / relative
        if relative_text.endswith("/") or full.is_dir():
            warnings.append(f"Nested repository or directory entry skipped: {project.name}/{relative_text.rstrip('/')}")
            continue
        if full.is_symlink():
            warnings.append(f"Symbolic link skipped: {project.name}/{relative_text}")
            continue
        if full.is_file():
            paths.append(relative)
    return paths, warnings


def select_projects(workspace: Path, requested: Sequence[str], skipped: set[str]) -> list[Path]:
    if requested:
        result: list[Path] = []
        for name in requested:
            candidate = (workspace / name).resolve()
            try:
                candidate.relative_to(workspace)
            except ValueError as exc:
                raise SystemExit(f"Project must be inside workspace: {name}") from exc
            if not candidate.is_dir():
                raise SystemExit(f"Project does not exist: {candidate}")
            result.append(candidate)
        return result
    return [
        child.resolve() for child in sorted(workspace.iterdir(), key=lambda p: p.name.casefold())
        if child.is_dir() and not child.name.startswith(".") and child.name not in skipped
    ]


def scan_operational_files(
    workspace: Path,
    projects: Sequence[Path],
    max_log_files: int,
    max_local_dbs: int,
    max_local_db_bytes: int,
) -> tuple[list[Candidate], list[dict[str, Any]]]:
    """Find logs and local databases even when .gitignore excludes them."""
    log_items: list[tuple[float, Candidate]] = []
    db_items: list[tuple[float, Candidate]] = []
    skipped: list[dict[str, Any]] = []
    roots = [(project.name, project) for project in projects]
    roots.append(("_workspace_root", workspace))
    seen: set[Path] = set()
    for project_name, root in roots:
        try:
            iterator = root.rglob("*")
        except OSError:
            continue
        for source in iterator:
            try:
                if not source.is_file() or source.is_symlink():
                    continue
                relative = source.relative_to(root)
                if hard_excluded(relative):
                    continue
                resolved = source.resolve()
                if resolved in seen:
                    continue
                seen.add(resolved)
                stat = source.stat()
            except (OSError, ValueError):
                continue
            suffix = source.suffix.lower()
            name_lower = source.name.lower()
            if suffix in LOG_EXTENSIONS or name_lower.endswith((".log.1", ".log.txt")):
                candidate = Candidate(
                    source=source,
                    archive_name=(Path("source") / project_name / relative).as_posix(),
                    project=project_name, relative_path=relative.as_posix(),
                    size=stat.st_size, category="log", priority=4,
                )
                log_items.append((stat.st_mtime, candidate))
            elif suffix in RAW_DATABASE_EXTENSIONS:
                if stat.st_size > max_local_db_bytes:
                    skipped.append({
                        "path": str(source), "reason": "local database exceeds max-local-db-mb",
                        "size": stat.st_size,
                    })
                    continue
                candidate = Candidate(
                    source=source,
                    archive_name=(Path("source") / project_name / relative).as_posix(),
                    project=project_name, relative_path=relative.as_posix(),
                    size=stat.st_size, category="database", priority=3,
                )
                db_items.append((stat.st_mtime, candidate))
    log_items.sort(key=lambda item: item[0], reverse=True)
    db_items.sort(key=lambda item: item[0], reverse=True)
    if len(log_items) > max_log_files:
        for _, candidate in log_items[max_log_files:]:
            skipped.append({"path": str(candidate.source), "reason": "older log beyond max-log-files"})
    if len(db_items) > max_local_dbs:
        for _, candidate in db_items[max_local_dbs:]:
            skipped.append({"path": str(candidate.source), "reason": "older database beyond max-local-dbs"})
    return [item[1] for item in log_items[:max_log_files]] + [item[1] for item in db_items[:max_local_dbs]], skipped


def is_sensitive_include_file(path: Path) -> bool:
    return any(pattern.match(path.name) for pattern in SENSITIVE_INCLUDE_PATTERNS)


def scan_sensitive_files(
    workspace: Path,
    projects: Sequence[Path],
    max_file_bytes: int,
) -> tuple[list[Candidate], list[dict[str, Any]]]:
    """Find ignored credential/config files without enabling a broad raw-file scan.

    This preserves v4's compact source selection while ensuring .env files,
    private keys, credentials JSON, Frappe site configs and small auth-state
    files are available to the AI exactly as stored.
    """
    candidates: list[Candidate] = []
    skipped: list[dict[str, Any]] = []
    roots = [(project.name, project) for project in projects]
    roots.append(("_workspace_root", workspace))
    seen: set[Path] = set()

    for project_name, root in roots:
        try:
            iterator = root.rglob("*")
        except OSError:
            continue
        for source in iterator:
            try:
                if not source.is_file() or source.is_symlink() or not is_sensitive_include_file(source):
                    continue
                relative = source.relative_to(root)
                excluded_parts = {part for part in relative.parts if part in HARD_EXCLUDED_DIRS}
                if excluded_parts and not excluded_parts.issubset(SENSITIVE_SESSION_DIRS):
                    continue
                resolved = source.resolve()
                if resolved in seen:
                    continue
                seen.add(resolved)
                stat = source.stat()
                limit = min(
                    SENSITIVE_FILE_MAX_BYTES,
                    max_file_bytes if max_file_bytes > 0 else SENSITIVE_FILE_MAX_BYTES,
                )
                if stat.st_size > limit:
                    skipped.append({
                        "path": str(source),
                        "reason": "sensitive/config file exceeds compact safety limit",
                        "size": stat.st_size,
                    })
                    continue
                candidates.append(Candidate(
                    source=source,
                    archive_name=(Path("source") / project_name / relative).as_posix(),
                    project=project_name,
                    relative_path=relative.as_posix(),
                    size=stat.st_size,
                    category="sensitive",
                    priority=0,
                ))
            except (OSError, ValueError) as exc:
                skipped.append({"path": str(source), "reason": f"sensitive-file scan failed: {exc}"})
    return candidates, skipped



def classify_candidate(path: Path) -> tuple[str, int]:
    suffix = path.suffix.lower()
    name = path.name.lower()
    if suffix in LOG_EXTENSIONS or name.endswith((".log.1", ".log.txt")):
        return "log", 4
    if suffix in RAW_DATABASE_EXTENSIONS:
        return "database", 3
    if suffix in {".xlsx", ".xlsm", ".xls"}:
        return "spreadsheet", 2
    if is_probably_text(path):
        if name in {"pyproject.toml", "package.json", "uv.lock", "package-lock.json", "pnpm-lock.yaml", "cargo.toml", "go.mod", "dockerfile", "docker-compose.yml", "docker-compose.yaml"}:
            return "manifest", 0
        return "source", 1
    return "binary", 5


def scan_workspace(
    workspace: Path,
    projects: Sequence[Path],
    output: Path,
    max_file_bytes: int,
) -> tuple[list[Candidate], list[dict[str, Any]], list[str]]:
    candidates: list[Candidate] = []
    skipped: list[dict[str, Any]] = []
    warnings: list[str] = []
    output_paths = {output.resolve(), output.with_suffix(output.suffix + ".part").resolve()}
    with create_temporary_git_dir() as temp_name:
        temp_git = Path(temp_name)
        for project in projects:
            print(f"Scanning {project.name} ...", flush=True)
            try:
                relative_paths, project_warnings = gitignore_paths(project, temp_git)
                warnings.extend(project_warnings)
            except Exception as exc:
                warnings.append(f"Gitignore scan failed for {project}: {exc}; using filesystem fallback")
                relative_paths = [p.relative_to(project) for p in project.rglob("*") if p.is_file()]
            for relative in relative_paths:
                source = project / relative
                if hard_excluded(relative):
                    skipped.append({"path": str(source), "reason": "hard-excluded directory"})
                    continue
                try:
                    resolved = source.resolve()
                    stat = source.stat()
                except OSError as exc:
                    skipped.append({"path": str(source), "reason": f"stat failed: {exc}"})
                    continue
                if resolved in output_paths:
                    continue
                if stat.st_size > max_file_bytes:
                    skipped.append({"path": str(source), "reason": f"larger than max-file-size ({human_size(stat.st_size)})"})
                    continue
                category, priority = classify_candidate(source)
                candidates.append(Candidate(
                    source=source,
                    archive_name=(Path("source") / project.name / relative).as_posix(),
                    project=project.name,
                    relative_path=relative.as_posix(),
                    size=stat.st_size,
                    category=category,
                    priority=priority,
                ))

    # Include useful workspace-root files too.
    for source in sorted(workspace.iterdir(), key=lambda p: p.name.casefold()):
        if not source.is_file() or source.is_symlink():
            continue
        if source.resolve() in output_paths:
            continue
        if source.stat().st_size > max_file_bytes:
            skipped.append({"path": str(source), "reason": "workspace root file larger than max-file-size"})
            continue
        category, priority = classify_candidate(source)
        candidates.append(Candidate(
            source=source,
            archive_name=(Path("source") / "_workspace_root" / source.name).as_posix(),
            project="_workspace_root",
            relative_path=source.name,
            size=source.stat().st_size,
            category=category,
            priority=priority,
        ))
    return candidates, skipped, warnings


def python_symbols(path: Path, text: str) -> dict[str, Any]:
    result: dict[str, Any] = {"path": str(path), "language": "Python", "imports": [], "symbols": [], "routes": [], "tasks": [], "models": []}
    try:
        tree = ast.parse(text)
    except SyntaxError as exc:
        result["parse_error"] = f"{exc.msg} at line {exc.lineno}"
        return result
    doc = ast.get_docstring(tree)
    if doc:
        result["module_doc"] = doc.splitlines()[0][:300]
    for node in ast.walk(tree):
        if isinstance(node, ast.Import):
            result["imports"].extend(alias.name for alias in node.names)
        elif isinstance(node, ast.ImportFrom):
            module = node.module or ""
            result["imports"].append(module + ":" + ",".join(alias.name for alias in node.names))
        elif isinstance(node, (ast.FunctionDef, ast.AsyncFunctionDef)):
            decorators = [ast.unparse(d) if hasattr(ast, "unparse") else "" for d in node.decorator_list]
            item = {"kind": "async_function" if isinstance(node, ast.AsyncFunctionDef) else "function", "name": node.name, "line": node.lineno, "decorators": decorators}
            result["symbols"].append(item)
            for decorator in decorators:
                if re.search(r"\.(?:get|post|put|patch|delete|options|head)\(", decorator):
                    result["routes"].append(item | {"decorator": decorator})
                if "task" in decorator.lower() or "shared_task" in decorator.lower():
                    result["tasks"].append(item | {"decorator": decorator})
        elif isinstance(node, ast.ClassDef):
            bases = [ast.unparse(base) if hasattr(ast, "unparse") else "" for base in node.bases]
            item = {"kind": "class", "name": node.name, "line": node.lineno, "bases": bases}
            result["symbols"].append(item)
            tablename = None
            for child in node.body:
                if isinstance(child, ast.Assign):
                    for target in child.targets:
                        if isinstance(target, ast.Name) and target.id == "__tablename__":
                            try:
                                tablename = ast.literal_eval(child.value)
                            except Exception:
                                tablename = None
            if tablename or any(base.endswith(("Base", "SQLModel")) for base in bases):
                result["models"].append(item | {"table": tablename})
    result["imports"] = sorted(set(result["imports"]))
    return result


def web_symbols(path: Path, text: str) -> dict[str, Any]:
    language = LANGUAGE_BY_SUFFIX.get(path.suffix.lower(), "Web")
    imports = re.findall(r"(?m)^\s*import\s+(?:.+?\s+from\s+)?[\"']([^\"']+)[\"']", text)
    exports = re.findall(r"(?m)^\s*export\s+(?:default\s+)?(?:async\s+)?(?:function|class|const|let|var|interface|type)\s+([A-Za-z_$][\w$]*)", text)
    actions = re.findall(r"(?m)\b(?:GET|POST|PUT|PATCH|DELETE)\s*[:=]\s*(?:async\s*)?", text)
    return {
        "path": str(path), "language": language,
        "imports": sorted(set(imports)), "exports": sorted(set(exports)),
        "http_action_markers": len(actions),
    }


def analyze_text_file(archive_name: str, text: str) -> dict[str, Any] | None:
    suffix = Path(archive_name).suffix.lower()
    if suffix in {".py", ".pyi"}:
        return python_symbols(Path(archive_name), text)
    if suffix in {".ts", ".tsx", ".js", ".jsx", ".mjs", ".cjs", ".astro", ".vue", ".svelte"}:
        return web_symbols(Path(archive_name), text)
    return None


def xlsx_preview(path: Path, include_sensitive: bool, max_rows: int = 25, max_cols: int = 30) -> dict[str, Any]:
    try:
        import openpyxl  # type: ignore
    except Exception as exc:
        return {"available": False, "error": f"openpyxl unavailable: {exc}"}
    result: dict[str, Any] = {"available": True, "workbook": path.name, "sheets": []}
    try:
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=False)
        for sheet in workbook.worksheets[:30]:
            rows: list[list[Any]] = []
            formulas = 0
            for row_index, row in enumerate(sheet.iter_rows(max_row=max_rows, max_col=max_cols, values_only=False), start=1):
                values: list[Any] = []
                for cell in row:
                    value = cell.value
                    if isinstance(value, str) and value.startswith("="):
                        formulas += 1
                    if isinstance(value, (dt.datetime, dt.date, dt.time)):
                        value = value.isoformat()
                    elif value is not None and not isinstance(value, (str, int, float, bool)):
                        value = str(value)
                    if isinstance(value, str):
                        value, _ = redact_text(value, include_sensitive)
                    values.append(value)
                rows.append(values)
            result["sheets"].append({
                "name": sheet.title, "max_row": sheet.max_row, "max_column": sheet.max_column,
                "sample_rows": rows, "sample_formula_cells": formulas,
                "sample_limits": {"rows": max_rows, "columns": max_cols},
            })
        workbook.close()
    except Exception as exc:
        return {"available": False, "error": f"{type(exc).__name__}: {exc}", "traceback": traceback.format_exc()}
    return result


def sanitize_sqlite_value(column: str, value: Any, include_sensitive: bool) -> Any:
    """Convert SQLite values to JSON-safe values without redaction."""
    if value is None or isinstance(value, (int, float, bool)):
        return value
    if isinstance(value, bytes):
        return {
            "binary_length": len(value),
            "sha256": sha256_bytes(value),
            "base64": base64.b64encode(value).decode("ascii") if len(value) <= 1024 * 1024 else None,
        }
    text = str(value)
    if len(text) > 4000:
        return {"truncated": text[:4000], "length": len(text), "sha256": sha256_bytes(text.encode())}
    return text


def sqlite_diagnostic(
    path: Path,
    include_sensitive: bool,
    sample_rows: int = 50,
    full_table_max_rows: int = 100000,
) -> dict[str, Any]:
    report: dict[str, Any] = {"path": str(path), "size_bytes": path.stat().st_size, "sha256": sha256_file(path), "tables": {}}
    crm_name = any(token in path.name.casefold() for token in ("crm", "complaint", "reply", "paperless"))
    try:
        uri = f"file:{path.resolve().as_posix()}?mode=ro"
        connection = sqlite3.connect(uri, uri=True)
        connection.row_factory = sqlite3.Row
        tables = connection.execute("SELECT name, sql FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%' ORDER BY name").fetchall()
        for table_row in tables:
            table = table_row["name"]
            quoted = '"' + table.replace('"', '""') + '"'
            count = int(connection.execute(f"SELECT COUNT(*) FROM {quoted}").fetchone()[0])
            table_lower = table.casefold()
            complete_requested = table_lower.startswith("crm_") or any(
                token in table_lower for token in ("complaint", "reply", "paperless", "case")
            ) or crm_name
            if complete_requested:
                limit = count if full_table_max_rows <= 0 else min(count, full_table_max_rows)
                mode = "complete" if count <= limit else "complete_safety_capped"
            else:
                limit = min(count, sample_rows)
                mode = "sample"
            rows = connection.execute(f"SELECT * FROM {quoted} LIMIT ?", (max(limit, 1),)).fetchall() if count else []
            sanitized = [
                {key: sanitize_sqlite_value(key, row[key], include_sensitive) for key in row.keys()}
                for row in rows
            ]
            report["tables"][table] = {
                "row_count": count,
                "exported_rows": len(sanitized),
                "truncated": count > len(sanitized),
                "mode": mode,
                "schema_sql": table_row["sql"],
                "rows": sanitized,
            }
        connection.close()
    except Exception as exc:
        report["error"] = f"{type(exc).__name__}: {exc}"
        report["traceback"] = traceback.format_exc()
    return report


def copy_candidates(
    candidates: Sequence[Candidate],
    bundle: Path,
    include_sensitive: bool,
    include_binaries: bool,
    max_total_bytes: int,
    log_tail_bytes: int,
    full_table_max_rows: int,
) -> tuple[list[ManifestRecord], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    manifest: list[ManifestRecord] = []
    skipped: list[dict[str, Any]] = []
    redactions: list[dict[str, Any]] = []
    symbols: list[dict[str, Any]] = []
    local_data_reports: list[dict[str, Any]] = []
    used = 0

    for candidate in sorted(candidates, key=lambda c: (c.priority, c.archive_name.casefold())):
        if candidate.category not in {"database", "log"} and used + candidate.size > max_total_bytes:
            skipped.append({"path": str(candidate.source), "reason": "max total source size reached", "size": candidate.size})
            continue
        original_hash = sha256_file(candidate.source)

        if candidate.category == "database":
            report = sqlite_diagnostic(candidate.source, include_sensitive, full_table_max_rows=full_table_max_rows)
            report_name = re.sub(r"[^A-Za-z0-9_.-]+", "_", f"{candidate.project}_{candidate.relative_path}") + ".json"
            write_json(bundle / "local_databases" / report_name, report)
            local_data_reports.append({"source": str(candidate.source), "report": f"local_databases/{report_name}"})
            skipped.append({"path": str(candidate.source), "reason": "raw database excluded; read-only diagnostic included"})
            continue

        if candidate.category == "spreadsheet":
            preview = xlsx_preview(candidate.source, include_sensitive)
            preview_name = re.sub(r"[^A-Za-z0-9_.-]+", "_", f"{candidate.project}_{candidate.relative_path}") + ".json"
            write_json(bundle / "data_previews" / preview_name, preview)
            local_data_reports.append({"source": str(candidate.source), "report": f"data_previews/{preview_name}"})
            if not include_binaries:
                skipped.append({"path": str(candidate.source), "reason": "raw spreadsheet excluded; workbook preview included"})
                continue

        if candidate.category == "log":
            try:
                data = candidate.source.read_bytes()
                tail = data[-log_tail_bytes:]
                text = tail.decode("utf-8", errors="replace")
                text, stats = redact_text(text, include_sensitive)
                target = bundle / "logs" / candidate.project / (candidate.relative_path + ".tail.txt")
                write_text(target, text)
                redactions.append({"path": str(candidate.source), "kind": "log-tail", "stats": stats})
            except Exception as exc:
                skipped.append({"path": str(candidate.source), "reason": f"log tail failed: {exc}"})
            continue

        binary = not is_probably_text(candidate.source)
        if binary and not include_binaries and candidate.category != "sensitive":
            skipped.append({"path": str(candidate.source), "reason": "binary excluded by default", "size": candidate.size})
            continue

        target = bundle / candidate.archive_name
        target.parent.mkdir(parents=True, exist_ok=True)
        redacted = False
        line_count: int | None = None
        language = LANGUAGE_BY_SUFFIX.get(candidate.source.suffix.lower())

        if binary:
            shutil.copy2(candidate.source, target)
            bundled_data = target.read_bytes()
        else:
            raw = candidate.source.read_bytes()
            text = raw.decode("utf-8", errors="replace")
            line_count = len(text.splitlines())
            bundled_text, stats = redact_text(text, include_sensitive)
            redacted = any(stats.values())
            if redacted:
                redactions.append({"path": str(candidate.source), "kind": "source", "stats": stats})
            write_text(target, bundled_text)
            bundled_data = bundled_text.encode("utf-8")
            analysis = analyze_text_file(candidate.archive_name, bundled_text)
            if analysis:
                symbols.append(analysis)

        used += len(bundled_data)
        manifest.append(ManifestRecord(
            archive_name=candidate.archive_name,
            source_path=str(candidate.source), project=candidate.project,
            relative_path=candidate.relative_path, size_original=candidate.size,
            size_bundled=len(bundled_data), sha256_original=original_hash,
            sha256_bundled=sha256_bytes(bundled_data), lines=line_count,
            language=language, category=candidate.category, redacted=redacted, binary=binary,
        ))
    return manifest, skipped, redactions, symbols, local_data_reports


def find_git_repositories(workspace: Path, projects: Sequence[Path]) -> list[Path]:
    roots: set[Path] = set()
    for candidate in [workspace, *projects]:
        result = run_command(["git", "rev-parse", "--show-toplevel"], candidate, timeout=10)
        if result.get("returncode") == 0 and result.get("stdout", "").strip():
            roots.add(Path(result["stdout"].strip()).resolve())
    return sorted(roots)


def collect_git(bundle: Path, repositories: Sequence[Path], include_sensitive: bool) -> list[dict[str, Any]]:
    reports: list[dict[str, Any]] = []
    commands = {
        "revision": ["git", "rev-parse", "HEAD"],
        "branch": ["git", "branch", "--show-current"],
        "status": ["git", "status", "--short", "--branch"],
        "recent_log": ["git", "log", "-30", "--date=iso-strict", "--pretty=format:%H%x09%ad%x09%an%x09%s"],
        "diff_stat": ["git", "diff", "--stat", "--"],
        "diff": ["git", "diff", "--no-ext-diff", "--"],
        "staged_diff": ["git", "diff", "--cached", "--no-ext-diff", "--"],
        "submodules": ["git", "submodule", "status", "--recursive"],
    }
    for index, repo in enumerate(repositories, start=1):
        repo_id = f"repo-{index}-{repo.name}"
        report: dict[str, Any] = {"path": str(repo), "id": repo_id, "commands": {}}
        for label, command in commands.items():
            result = run_command(command, repo, timeout=60)
            for stream in ("stdout", "stderr"):
                if isinstance(result.get(stream), str):
                    result[stream], _ = redact_text(result[stream], include_sensitive)
            output = (result.get("stdout") or "") + (("\n" + result.get("stderr")) if result.get("stderr") else "")
            if len(output.encode("utf-8", errors="replace")) > 10 * 1024 * 1024:
                output = output[:10 * 1024 * 1024] + "\n<output truncated at 10 MiB>\n"
            output_path = bundle / "git" / repo_id / f"{label}.txt"
            write_text(output_path, output)
            report["commands"][label] = {
                key: value for key, value in result.items() if key not in {"stdout", "stderr"}
            } | {
                "output_file": output_path.relative_to(bundle).as_posix(),
                "output_bytes": len(output.encode("utf-8", errors="replace")),
                "output_sha256": sha256_bytes(output.encode("utf-8", errors="replace")),
            }
        reports.append(report)
    write_json(bundle / "git" / "repositories.json", reports)
    return reports


def collect_runtime(bundle: Path, workspace: Path, platform_root: Path, include_sensitive: bool) -> dict[str, Any]:
    metadata: dict[str, Any] = {
        "generated_at": dt.datetime.now(dt.timezone.utc).isoformat(),
        "tool_version": TOOL_VERSION,
        "host": {
            "hostname": socket.gethostname(), "platform": platform.platform(),
            "python": sys.version, "executable": sys.executable,
            "cwd": str(Path.cwd()), "workspace": str(workspace), "platform_root": str(platform_root),
        },
        "environment": safe_environment(),
        "commands": {},
    }
    command_specs = {
        "python_version": ([sys.executable, "--version"], workspace),
        "uv_version": (["uv", "--version"], platform_root),
        "node_version": (["node", "--version"], platform_root),
        "npm_version": (["npm", "--version"], platform_root),
        "docker_version": (["docker", "--version"], platform_root),
        "docker_ps": (["docker", "ps", "--format", "{{.Names}}\t{{.Image}}\t{{.Status}}\t{{.Ports}}"], platform_root),
        "docker_compose_ps": (["docker", "compose", "ps"], platform_root),
        "listening_ports": (["ss", "-ltnp"], platform_root),
        "uv_pip_list": (["uv", "pip", "list"], platform_root),
    }
    for label, (command, cwd) in command_specs.items():
        result = run_command(command, cwd, timeout=30)
        for stream in ("stdout", "stderr"):
            if isinstance(result.get(stream), str):
                result[stream], _ = redact_text(result[stream], include_sensitive)
        output = (result.get("stdout") or "") + (("\n" + result.get("stderr")) if result.get("stderr") else "")
        output_path = bundle / "runtime" / f"{label}.txt"
        write_text(output_path, output)
        metadata["commands"][label] = {
            key: value for key, value in result.items() if key not in {"stdout", "stderr"}
        } | {
            "output_file": output_path.relative_to(bundle).as_posix(),
            "output_bytes": len(output.encode("utf-8", errors="replace")),
            "output_sha256": sha256_bytes(output.encode("utf-8", errors="replace")),
        }
    write_json(bundle / "runtime" / "metadata.json", metadata)
    return metadata


def run_checks(
    bundle: Path,
    platform_root: Path,
    mode: str,
    custom_commands: Sequence[str],
    timeout: int,
    include_sensitive: bool,
) -> list[dict[str, Any]]:
    specs: list[tuple[str, list[str], Path]] = []
    if mode in {"quick", "full"}:
        specs.extend([
            ("git_diff_check", ["git", "diff", "--check"], platform_root),
            ("python_compile", ["uv", "run", "python", "-m", "compileall", "-q", "automation_api", "automation_worker", "packages"], platform_root),
            ("alembic_heads", ["uv", "run", "alembic", "heads"], platform_root),
        ])
    if mode == "full":
        specs.append(("pytest_full", ["uv", "run", "pytest", "-q"], platform_root))
        web_root = platform_root / "apps" / "web"
        if web_root.is_dir():
            specs.append(("astro_build", ["npm", "run", "build"], web_root))
    for index, command_text in enumerate(custom_commands, start=1):
        specs.append((f"custom_{index}", ["bash", "-lc", command_text], platform_root))

    results: list[dict[str, Any]] = []
    for label, command, cwd in specs:
        print(f"Running check: {label} ...", flush=True)
        result = run_command(command, cwd, timeout=timeout)
        for stream in ("stdout", "stderr"):
            if isinstance(result.get(stream), str):
                result[stream], _ = redact_text(result[stream], include_sensitive)
        log = f"$ {' '.join(command)}\n\nSTDOUT\n{'-'*72}\n{result.get('stdout','')}\n\nSTDERR\n{'-'*72}\n{result.get('stderr','')}\n"
        if result.get("error"):
            log += f"\nERROR: {result['error']}\n"
        output_path = bundle / "checks" / f"{label}.log"
        write_text(output_path, log)
        summary = {key: value for key, value in result.items() if key not in {"stdout", "stderr"}}
        summary.update({
            "label": label,
            "output_file": output_path.relative_to(bundle).as_posix(),
            "output_bytes": len(log.encode("utf-8", errors="replace")),
            "output_sha256": sha256_bytes(log.encode("utf-8", errors="replace")),
        })
        results.append(summary)
    write_json(bundle / "checks" / "results.json", results)
    return results



def collect_crm_source_focus(bundle: Path, manifest: Sequence[ManifestRecord], symbols: Sequence[Mapping[str, Any]]) -> dict[str, Any]:
    """Create a compact index of all source files relevant to CRM/complaints/replies/Paperless."""
    keywords = (
        "crm", "complaint", "reply", "replies", "paperless", "inbound", "case",
        "formal_letter", "formal-letter", "odt", "ticket", "helpdesk",
    )
    selected: list[dict[str, Any]] = []
    for record in manifest:
        haystack = f"{record.project}/{record.relative_path}".casefold()
        score = sum(1 for keyword in keywords if keyword in haystack)
        if score:
            selected.append(asdict(record) | {"focus_score": score})
    selected.sort(key=lambda row: (-int(row["focus_score"]), str(row["archive_name"]).casefold()))

    symbol_items = []
    for item in symbols:
        path = str(item.get("path", ""))
        lowered = path.casefold()
        route_text = json.dumps(item.get("routes") or [], ensure_ascii=False).casefold()
        if any(keyword in lowered or keyword in route_text for keyword in keywords):
            symbol_items.append(item)

    table_names: set[str] = set()
    endpoints: list[dict[str, Any]] = []
    tests: list[str] = []
    migrations: list[str] = []
    ui_files: list[str] = []
    for row in selected:
        archive_name = str(row["archive_name"])
        source_path = bundle / archive_name
        lowered = archive_name.casefold()
        if "test" in lowered:
            tests.append(archive_name)
        if "migration" in lowered or "/versions/" in lowered:
            migrations.append(archive_name)
        if "/apps/web/" in lowered or lowered.endswith((".astro", ".tsx", ".ts", ".js")):
            ui_files.append(archive_name)
        if not source_path.is_file() or row.get("binary"):
            continue
        try:
            content = source_path.read_text(encoding="utf-8", errors="replace")
        except OSError:
            continue
        table_names.update(re.findall(r"__tablename__\s*=\s*[\"']([^\"']+)[\"']", content))
        for method, route in re.findall(r"@(?:router|app)\.(get|post|put|patch|delete)\(\s*[\"']([^\"']+)", content, re.I):
            if any(keyword in route.casefold() for keyword in keywords):
                endpoints.append({"method": method.upper(), "route": route, "file": archive_name})

    summary = {
        "file_count": len(selected),
        "table_names": sorted(table_names),
        "endpoints": endpoints,
        "tests": sorted(set(tests)),
        "migrations": sorted(set(migrations)),
        "ui_files": sorted(set(ui_files)),
        "files": selected,
        "symbols": symbol_items,
    }
    write_json(bundle / "analysis" / "crm_source_index.json", summary)

    lines = [
        "# CRM / Complaint Architecture Source Index",
        "",
        f"Focused files: {len(selected)}",
        "",
        "## Database tables discovered in focused source",
        "",
    ]
    lines.extend(f"- `{name}`" for name in sorted(table_names))
    lines.extend(["", "## API endpoints discovered", ""])
    lines.extend(f"- `{item['method']} {item['route']}` — `{item['file']}`" for item in endpoints)
    lines.extend(["", "## Primary UI files", ""])
    lines.extend(f"- `{path}`" for path in sorted(set(ui_files))[:200])
    lines.extend(["", "## Tests", ""])
    lines.extend(f"- `{path}`" for path in sorted(set(tests))[:300])
    lines.extend(["", "## Migrations", ""])
    lines.extend(f"- `{path}`" for path in sorted(set(migrations))[:200])
    write_text(bundle / "analysis" / "CRM_SOURCE_MAP.md", "\n".join(lines) + "\n")
    return summary


def resolve_frappe_root(requested: Path | None) -> Path | None:
    candidates: list[Path] = []
    if requested:
        candidates.append(requested.expanduser())
    candidates.extend([
        Path.home() / "frappe-helpdesk-docker",
        Path.home() / "frappe_docker",
    ])
    for candidate in candidates:
        candidate = candidate.resolve()
        repo = candidate / "frappe_docker" if (candidate / "frappe_docker" / "compose.yaml").is_file() else candidate
        if (repo / "compose.yaml").is_file():
            return candidate
    return None


def frappe_compose_command(root: Path) -> tuple[Path, list[str]]:
    repo = root / "frappe_docker" if (root / "frappe_docker" / "compose.yaml").is_file() else root
    env_file = root / "helpdesk.env"
    env_values: dict[str, str] = {}
    if env_file.is_file():
        for raw in env_file.read_text(encoding="utf-8", errors="replace").splitlines():
            line = raw.strip()
            if line and not line.startswith("#") and "=" in line:
                key, value = line.split("=", 1)
                env_values[key.strip()] = value.strip().strip("\"'")
    project_name = env_values.get("COMPOSE_PROJECT_NAME") or "frappe-helpdesk"
    command = ["docker", "compose", "--project-name", project_name]
    if env_file.is_file():
        command.extend(["--env-file", str(env_file)])
    compose_files = [
        repo / "compose.yaml",
        repo / "overrides" / "compose.mariadb.yaml",
        repo / "overrides" / "compose.redis.yaml",
        repo / "overrides" / "compose.noproxy.yaml",
    ]
    for compose_file in compose_files:
        if compose_file.is_file():
            command.extend(["-f", str(compose_file)])
    return repo, command



def build_frappe_helpdesk_context(parsed: Mapping[str, Any]) -> dict[str, Any]:
    """Cross-link Frappe HD Ticket rows with replies, notes, assignments and files."""
    tables = parsed.get("tables") if isinstance(parsed, Mapping) else None
    if not isinstance(tables, Mapping):
        return {"available": False, "reason": "Frappe table export is unavailable"}

    def rows(table: str) -> list[dict[str, Any]]:
        payload = tables.get(table)
        if not isinstance(payload, Mapping):
            return []
        values = payload.get("rows")
        return [dict(value) for value in values if isinstance(value, Mapping)] if isinstance(values, list) else []

    tickets = rows("tabHD Ticket")
    if not tickets:
        return {
            "available": True,
            "ticket_count": 0,
            "cases": [],
            "taxonomy": {},
            "note": "No HD Ticket rows were present in the exported Frappe site.",
        }

    communications = rows("tabCommunication")
    comments = rows("tabComment")
    assignments = rows("tabToDo")
    files = rows("tabFile")
    versions = rows("tabVersion")
    hd_tables = {
        name: rows(str(name))
        for name in tables
        if str(name).startswith("tabHD ") and str(name) != "tabHD Ticket"
    }

    def matching(source: Iterable[Mapping[str, Any]], ticket_name: str, fields: Sequence[str]) -> list[dict[str, Any]]:
        return [
            dict(row) for row in source
            if any(str(row.get(field) or "") == ticket_name for field in fields)
        ]

    cases: list[dict[str, Any]] = []
    for ticket in tickets:
        ticket_name = str(ticket.get("name") or ticket.get("id") or "")
        hd_related: dict[str, list[dict[str, Any]]] = {}
        for table_name, table_rows in hd_tables.items():
            related = matching(
                table_rows,
                ticket_name,
                ("ticket", "ticket_id", "parent", "reference_name", "reference", "hd_ticket"),
            )
            if related:
                hd_related[table_name] = related
        cases.append({
            "ticket": ticket,
            "communications_and_replies": matching(communications, ticket_name, ("reference_name", "reference", "parent")),
            "internal_comments": matching(comments, ticket_name, ("reference_name", "reference", "parent")),
            "assignments": matching(assignments, ticket_name, ("reference_name", "reference", "parent")),
            "files": matching(files, ticket_name, ("attached_to_name", "reference_name", "parent")),
            "versions": matching(versions, ticket_name, ("docname", "reference_name", "parent")),
            "helpdesk_related_records": hd_related,
        })

    def values(field: str) -> dict[str, int]:
        counts: Counter[str] = Counter()
        for ticket in tickets:
            value = ticket.get(field)
            counts["<empty>" if value in (None, "") else str(value)] += 1
        return dict(counts.most_common())

    taxonomy_fields = (
        "status", "ticket_type", "priority", "team", "agent_group", "customer",
        "contact", "raised_by", "via_customer_portal", "agreement_status",
    )
    taxonomy = {field: values(field) for field in taxonomy_fields if any(field in ticket for ticket in tickets)}
    return {
        "available": True,
        "ticket_count": len(tickets),
        "communication_count": len(communications),
        "comment_count": len(comments),
        "assignment_count": len(assignments),
        "file_count": len(files),
        "version_count": len(versions),
        "taxonomy": taxonomy,
        "cases": cases,
    }


def frappe_summary_markdown(context: Mapping[str, Any]) -> str:
    if not context.get("available"):
        return "# Frappe Helpdesk Context\n\nUnavailable: " + str(context.get("reason")) + "\n"
    lines = [
        "# Frappe Helpdesk Complaint Context",
        "",
        f"- Tickets: {context.get('ticket_count', 0)}",
        f"- Communications/replies: {context.get('communication_count', 0)}",
        f"- Internal comments: {context.get('comment_count', 0)}",
        f"- Assignments: {context.get('assignment_count', 0)}",
        f"- Files: {context.get('file_count', 0)}",
        f"- Version/audit rows: {context.get('version_count', 0)}",
        "",
        "## AI reading order",
        "",
        "1. `frappe/HELPDESK_SUMMARY.md`",
        "2. `frappe/helpdesk_taxonomy.json`",
        "3. `frappe/helpdesk_cases_with_context.json`",
        "4. `frappe/helpdesk_database.json` for raw unfiltered table exports",
    ]
    return "\n".join(lines) + "\n"


def collect_frappe_helpdesk(
    bundle: Path,
    requested_root: Path | None,
    requested_site: str | None,
    max_rows: int,
    timeout: int,
    include_sensitive: bool,
    skip: bool,
) -> dict[str, Any]:
    """Best-effort, read-only Frappe Helpdesk runtime and database context collection."""
    if skip:
        result = {"status": "skipped"}
        write_json(bundle / "frappe" / "collector_result.json", result)
        return result
    root = resolve_frappe_root(requested_root)
    if root is None:
        result = {"status": "not_found", "reason": "No Frappe Helpdesk Docker installation auto-detected"}
        write_json(bundle / "frappe" / "collector_result.json", result)
        return result
    repo, compose = frappe_compose_command(root)
    frappe_dir = bundle / "frappe"
    frappe_dir.mkdir(parents=True, exist_ok=True)

    # Copy small text configuration files exactly as stored.
    copied_configs: list[str] = []
    config_paths = [root / "helpdesk.env", repo / "compose.yaml"]
    if (repo / "overrides").is_dir():
        config_paths.extend(sorted((repo / "overrides").glob("compose*.yaml")))
    for path in config_paths:
        if not path.is_file() or path.stat().st_size > 2 * 1024 * 1024:
            continue
        content, _ = redact_text(path.read_text(encoding="utf-8", errors="replace"), include_sensitive=False)
        target = frappe_dir / "config" / path.name
        write_text(target, content)
        copied_configs.append(target.relative_to(bundle).as_posix())

    command_reports: dict[str, Any] = {}
    for label, suffix in {
        "services": ["config", "--services"],
        "status": ["ps"],
        "images": ["images"],
        "logs": ["logs", "--tail=500"],
    }.items():
        result = run_command([*compose, *suffix], repo, timeout=timeout)
        output = (result.get("stdout") or "") + (("\n" + result.get("stderr")) if result.get("stderr") else "")
        output, _ = redact_text(output, include_sensitive)
        output_file = frappe_dir / f"docker_{label}.txt"
        write_text(output_file, output)
        command_reports[label] = {key: value for key, value in result.items() if key not in {"stdout", "stderr"}} | {"output_file": output_file.relative_to(bundle).as_posix()}

    in_container = r"""
import datetime as dt, decimal, json, sys, uuid
from pathlib import Path
import frappe

def clean(value):
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, (dt.datetime, dt.date, dt.time, decimal.Decimal, uuid.UUID)):
        return str(value)
    if isinstance(value, bytes):
        return {"binary_length": len(value)}
    if isinstance(value, dict):
        return {str(k): clean(v) for k, v in value.items()}
    if isinstance(value, (list, tuple)):
        return [clean(v) for v in value]
    return str(value)

requested = sys.argv[1].strip() if len(sys.argv) > 1 else ""
max_rows = int(sys.argv[2]) if len(sys.argv) > 2 else 5000
sites_root = Path("/home/frappe/frappe-bench/sites").resolve()
sites = sorted(
    p.name
    for p in sites_root.iterdir()
    if p.is_dir() and (p / "site_config.json").is_file()
)
site = requested or (sites[0] if sites else "")
if not site:
    print(json.dumps({"status": "no_site", "sites": sites, "sites_path": str(sites_root)}))
    raise SystemExit(0)
if site not in sites:
    print(json.dumps({
        "status": "site_not_found",
        "site_requested": site,
        "sites": sites,
        "sites_path": str(sites_root),
    }))
    raise SystemExit(0)
# Direct execution via the virtualenv Python does not receive Bench CLI's
# sites-path context. Pass it explicitly. The outer command starts Python from
# the sites directory and enables stream-only logging, preserving Frappe's
# expected relative paths without creating collector log files.
frappe.init(site=site, sites_path=str(sites_root))
frappe.connect()
try:
    all_tables = [row[0] for row in frappe.db.sql("SHOW TABLES")]
    hd_tables = sorted(table for table in all_tables if table.startswith("tabHD "))
    exact = {
        "tabToDo", "tabCommunication", "tabComment", "tabVersion", "tabFile",
        "tabCustom Field", "tabProperty Setter", "tabAssignment Rule",
        "tabWorkflow", "tabWorkflow State", "tabWorkflow Transition",
        "tabUser", "tabRole", "tabHas Role", "tabDocPerm",
    }
    selected = hd_tables + sorted(table for table in exact if table in all_tables)
    filters = {
        "tabToDo": "reference_type='HD Ticket'",
        "tabCommunication": "reference_doctype='HD Ticket'",
        "tabComment": "reference_doctype='HD Ticket'",
        "tabVersion": "ref_doctype='HD Ticket'",
        "tabFile": "attached_to_doctype='HD Ticket'",
        "tabCustom Field": "dt LIKE 'HD %'",
        "tabProperty Setter": "doc_type LIKE 'HD %'",
    }
    site_config_path = sites_root / site / "site_config.json"
    common_config_path = sites_root / "common_site_config.json"
    output = {
        "status": "ok",
        "site": site,
        "sites": sites,
        "installed_apps": frappe.get_installed_apps(),
        "site_config": json.loads(site_config_path.read_text(encoding="utf-8")) if site_config_path.is_file() else None,
        "common_site_config": json.loads(common_config_path.read_text(encoding="utf-8")) if common_config_path.is_file() else None,
        "tables": {},
    }
    for table in selected:
        where = filters.get(table)
        count_sql = f"SELECT COUNT(*) FROM `{table}`" + (f" WHERE {where}" if where else "")
        count = int(frappe.db.sql(count_sql)[0][0])
        limit = min(count, max_rows) if max_rows > 0 else count
        query = f"SELECT * FROM `{table}`" + (f" WHERE {where}" if where else "")
        columns = [row[0] for row in frappe.db.sql(f"SHOW COLUMNS FROM `{table}`")]
        if "modified" in columns:
            query += " ORDER BY modified DESC"
        query += f" LIMIT {max(limit, 1)}"
        rows = frappe.db.sql(query, as_dict=True) if count else []
        output["tables"][table] = {"row_count": count, "exported_rows": len(rows), "truncated": count > len(rows), "rows": clean(rows)}
    print(json.dumps(output, ensure_ascii=False, default=clean))
finally:
    frappe.destroy()
"""
    site_arg = shlex.quote(requested_site or "")
    shell = "cd /home/frappe/frappe-bench/sites && FRAPPE_STREAM_LOGGING=1 ../env/bin/python - " + site_arg + " " + str(max_rows)
    db_result = run_command(
        [*compose, "exec", "-T", "backend", "bash", "-lc", shell],
        repo,
        timeout=timeout,
        input_text=in_container,
    )
    raw_output = str(db_result.get("stdout") or "").strip()
    parsed: Any
    try:
        # Bench/container startup text can precede JSON; prefer the final JSON-looking line.
        json_line = next((line for line in reversed(raw_output.splitlines()) if line.lstrip().startswith("{")), raw_output)
        parsed = json.loads(json_line)
        serialized = json.dumps(parsed, indent=2, ensure_ascii=False, default=str)
        serialized, _ = redact_text(serialized, include_sensitive)
        parsed = json.loads(serialized)
        write_json(frappe_dir / "helpdesk_database.json", parsed)
        helpdesk_context = build_frappe_helpdesk_context(parsed)
        write_json(frappe_dir / "helpdesk_cases_with_context.json", helpdesk_context.get("cases", []))
        write_json(frappe_dir / "helpdesk_taxonomy.json", helpdesk_context.get("taxonomy", {}))
        write_json(frappe_dir / "helpdesk_context_metadata.json", {key: value for key, value in helpdesk_context.items() if key not in {"cases", "taxonomy"}})
        write_text(frappe_dir / "HELPDESK_SUMMARY.md", frappe_summary_markdown(helpdesk_context))
        status = "collected" if parsed.get("status") == "ok" else str(parsed.get("status") or "unknown")
    except Exception as exc:
        sanitized, _ = redact_text(raw_output + "\n" + str(db_result.get("stderr") or ""), include_sensitive)
        write_text(frappe_dir / "database_collector.log", sanitized)
        parsed = {"parse_error": f"{type(exc).__name__}: {exc}"}
        status = "failed"

    result = {
        "status": status,
        "root": str(root),
        "repo": str(repo),
        "site_requested": requested_site,
        "config_files": copied_configs,
        "commands": command_reports,
        "database_command": {key: value for key, value in db_result.items() if key not in {"stdout", "stderr"}},
        "database_summary": {
            "site": parsed.get("site") if isinstance(parsed, dict) else None,
            "installed_apps": parsed.get("installed_apps") if isinstance(parsed, dict) else None,
            "table_count": len(parsed.get("tables", {})) if isinstance(parsed, dict) else 0,
        },
    }
    write_json(frappe_dir / "collector_result.json", result)
    return result

def extract_and_run_database_component(
    bundle: Path,
    temp_root: Path,
    platform_root: Path,
    args: argparse.Namespace,
) -> dict[str, Any]:
    db_source = zlib.decompress(base64.b85decode(DB_COMPONENT_B85.encode("ascii")))
    tooling = bundle / "tooling"
    tooling.mkdir(parents=True, exist_ok=True)
    component_path = tooling / "embedded_database_diagnostic.py"
    component_path.write_bytes(db_source)

    if args.skip_db:
        return {"status": "skipped"}

    temp_script = temp_root / "embedded_database_diagnostic.py"
    temp_script.write_bytes(db_source)
    db_zip = temp_root / "database-diagnostic.zip"
    component_args = [
        str(temp_script), "--project-root", str(platform_root),
        "--latest-previews", str(args.latest_previews), "--max-rows", str(args.max_rows),
        "--full-table-max-rows", str(args.full_table_max_rows),
        "--output", str(db_zip),
    ]
    if args.database_url:
        component_args.extend(["--database-url", args.database_url])
    for key in args.preview_key:
        component_args.extend(["--preview-key", key])
    if args.include_sensitive:
        component_args.append("--include-sensitive")
    if args.verify_artifacts:
        component_args.append("--verify-files")

    interpreters: list[tuple[str, list[str]]] = []
    uv = shutil.which("uv")
    if uv:
        interpreters.append(("uv-run", [uv, "run", "python"]))
    venv_python = platform_root / ".venv" / "bin" / "python"
    if venv_python.is_file():
        interpreters.append(("project-venv", [str(venv_python)]))
    interpreters.append(("current-python", [sys.executable]))

    attempts: list[dict[str, Any]] = []
    combined_logs: list[str] = []
    final_result: dict[str, Any] = {}
    for label, prefix in interpreters:
        if db_zip.exists():
            db_zip.unlink()
        command = [*prefix, *component_args]
        result = run_command(command, platform_root, timeout=args.db_timeout)
        stdout = str(result.get("stdout") or "")
        stderr = str(result.get("stderr") or "")
        combined_logs.append(f"=== Attempt: {label} ===\n$ {' '.join(command)}\n\n{stdout}\n{stderr}\n")
        attempt = {key: value for key, value in result.items() if key not in {"stdout", "stderr"}} | {"interpreter": label}
        attempts.append(attempt)
        final_result = result
        if db_zip.is_file():
            break
        # Retry only interpreter/dependency failures. A genuine DB connection failure
        # already creates a diagnostic ZIP and therefore exits this loop above.
        output_lower = (stdout + "\n" + stderr).casefold()
        if not any(token in output_lower for token in ("sqlalchemy is not available", "no module named", "failed to spawn", "not found")):
            break

    database_dir = bundle / "database"
    database_dir.mkdir(parents=True, exist_ok=True)
    if db_zip.is_file():
        with zipfile.ZipFile(db_zip) as archive:
            archive.extractall(database_dir)
        status = "collected"
    else:
        status = "failed"

    output = "\n".join(combined_logs)
    output, _ = redact_text(output, args.include_sensitive)
    output_path = database_dir / "collector.log"
    write_text(output_path, output)
    result = {
        "status": status,
        "component_version": DB_COMPONENT_VERSION,
        "attempts": attempts,
        "final_returncode": final_result.get("returncode"),
        "output_file": output_path.relative_to(bundle).as_posix(),
        "output_bytes": len(output.encode("utf-8", errors="replace")),
        "output_sha256": sha256_bytes(output.encode("utf-8", errors="replace")),
    }
    write_json(database_dir / "collector_result.json", result)
    return result


def build_tree(records: Sequence[ManifestRecord]) -> str:
    grouped: dict[str, list[str]] = defaultdict(list)
    for record in records:
        grouped[record.project].append(record.relative_path)
    lines: list[str] = []
    for project in sorted(grouped):
        lines.append(project + "/")
        for path in sorted(grouped[project], key=str.casefold):
            lines.append("  " + path)
        lines.append("")
    return "\n".join(lines)


def write_manifest_csv(path: Path, records: Sequence[ManifestRecord]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not records:
        path.write_text("", encoding="utf-8")
        return
    fieldnames = list(asdict(records[0]).keys())
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for record in records:
            writer.writerow(asdict(record))


def generate_code_map(symbols: Sequence[Mapping[str, Any]]) -> str:
    lines = ["# Code Map", "", "Automatically extracted symbols, routes, tasks and models.", ""]
    for item in sorted(symbols, key=lambda x: str(x.get("path", "")).casefold()):
        lines.append(f"## `{item.get('path')}`")
        if item.get("module_doc"):
            lines.append(str(item["module_doc"]))
        if item.get("parse_error"):
            lines.append(f"Parse error: {item['parse_error']}")
        routes = item.get("routes") or []
        tasks = item.get("tasks") or []
        models = item.get("models") or []
        exports = item.get("exports") or []
        symbols_list = item.get("symbols") or []
        if routes:
            lines.append("- Routes: " + ", ".join(f"{r.get('name')}@{r.get('line')}" for r in routes))
        if tasks:
            lines.append("- Tasks: " + ", ".join(f"{r.get('name')}@{r.get('line')}" for r in tasks))
        if models:
            lines.append("- Models: " + ", ".join(f"{r.get('name')}[{r.get('table') or '?'}]" for r in models))
        if exports:
            lines.append("- Exports: " + ", ".join(map(str, exports[:30])))
        if symbols_list and not (routes or tasks or models):
            lines.append("- Symbols: " + ", ".join(f"{r.get('name')}@{r.get('line')}" for r in symbols_list[:40]))
        lines.append("")
    return "\n".join(lines)


def generate_start_here(
    args: argparse.Namespace,
    workspace: Path,
    platform_root: Path,
    projects: Sequence[Path],
    records: Sequence[ManifestRecord],
    skipped: Sequence[Mapping[str, Any]],
    db_result: Mapping[str, Any],
    checks: Sequence[Mapping[str, Any]],
    git_reports: Sequence[Mapping[str, Any]],
    frappe_result: Mapping[str, Any],
) -> str:
    by_language = Counter(record.language or "Other" for record in records)
    by_project = Counter(record.project for record in records)
    check_summary = [f"- {r.get('label')}: return code {r.get('returncode', r.get('error', 'unknown'))}" for r in checks]
    notes = list(args.context_note)
    if args.context_file:
        try:
            notes.append(Path(args.context_file).read_text(encoding="utf-8", errors="replace"))
        except Exception as exc:
            notes.append(f"Unable to read context file: {exc}")
    return textwrap.dedent(f"""
    # START HERE FOR AI

    This ZIP is a machine-generated, read-only context bundle for the Deomee workspace.
    It is intended to minimize guessing in a new chat.

    ## User-provided task context

    {chr(10).join(notes) if notes else 'No explicit context note was supplied. Ask the user what change or investigation is required.'}

    ## Workspace

    - Generated: {dt.datetime.now(dt.timezone.utc).isoformat()}
    - Tool version: {TOOL_VERSION}
    - Workspace root on source machine: `{workspace}`
    - Automation platform root: `{platform_root}`
    - Selected projects: {', '.join(project.name for project in projects)}
    - Bundled source files: {len(records):,}
    - Skipped/excluded items: {len(skipped):,}
    - Database collector: {db_result.get('status', 'unknown')}
    - Frappe Helpdesk collector: {frappe_result.get('status', 'unknown')}
    - Git repositories: {len(git_reports)}

    ## Source distribution

    By project: {json.dumps(dict(by_project), ensure_ascii=False)}

    By language: {json.dumps(dict(by_language), ensure_ascii=False)}

    ## Suggested reading order

    1. `START_HERE_FOR_AI.md` (this file)
    2. `NEW_CHAT_PROMPT.md`
    3. `analysis/PROJECT_TREE.txt`
    4. `analysis/CODE_MAP.md` and `analysis/code_symbols.json`
    5. `manifest/files.json` for hashes, sizes and exact source paths
    6. `git/repositories.json` plus each repository's status/diff
    7. `checks/results.json` and `checks/*.log`
    8. `database/crm/CRM_SUMMARY.md` and `database/crm/crm_cases_with_context.json`
    9. `database/SUMMARY.txt`, `database/schema.json`, and `database/whatsapp_preview_deep_report.json`
    10. `frappe/collector_result.json` and `frappe/helpdesk_database.json` when available
    11. Relevant files under `source/`

    ## Test/build checks

    {chr(10).join(check_summary) if check_summary else '- Checks were not run. Use --run-checks quick or full next time.'}

    ## Data safety

    - No privacy or secret redaction is performed.
    - Ignored `.env`, private-key, credential, Frappe config and selected auth-state files are included explicitly.
    - Dependency/build/cache directories remain excluded to keep the bundle compact.
    - Raw SQLite databases are replaced by unfiltered read-only schema/count/sample reports.
    - Workbooks are represented by unfiltered bounded previews unless binaries were explicitly included.
    - Treat this ZIP as a confidential credential-bearing backup.

    ## Important interpretation rules

    - Treat files under `source/` as the current uploaded source snapshot, not necessarily committed Git state.
    - Use Git status and diffs to distinguish committed code from local changes.
    - Database exports are unfiltered and may be sampled; consult `database/table_export_manifest.json`.
    - Existing frozen WhatsApp previews are historical snapshots and may intentionally differ from current configuration.
    """).strip() + "\n"


def generate_new_chat_prompt(args: argparse.Namespace) -> str:
    context = "\n".join(args.context_note).strip()
    return textwrap.dedent(f"""
    I have uploaded a Deomee AI context bundle generated from my current workspace and database.

    Start by reading `START_HERE_FOR_AI.md`, then inspect the project tree, code map, Git state,
    checks, database summary, schema and any selected WhatsApp preview reports. Base conclusions on
    the bundled source and diagnostics rather than assumptions. Preserve existing behavior and tests.

    Current task/context:
    {context or '[Describe the issue or requested change here]'}

    Before proposing a patch, identify the root cause across data model, backend, worker, API and UI.
    State any missing evidence clearly. For code changes, include regression, integration and contract tests.
    """).strip() + "\n"


def zip_bundle(bundle: Path, output: Path, level: int) -> None:
    part = output.with_suffix(output.suffix + ".part")
    output.parent.mkdir(parents=True, exist_ok=True)
    if part.exists():
        part.unlink()
    try:
        with zipfile.ZipFile(part, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=level, allowZip64=True) as archive:
            archive.comment = f"Deomee AI context bundle {TOOL_VERSION}".encode("ascii")
            for path in sorted(bundle.rglob("*")):
                if path.is_file():
                    archive.write(path, path.relative_to(bundle))
        os.replace(part, output)
    except BaseException:
        if part.exists():
            part.unlink()
        raise


def resolve_platform_root(workspace: Path, requested: Path | None) -> Path:
    if requested:
        candidate = requested.expanduser()
        if not candidate.is_absolute():
            candidate = workspace / candidate
        candidate = candidate.resolve()
        if not (candidate / "pyproject.toml").is_file():
            raise SystemExit(f"Platform root does not look valid: {candidate}")
        return candidate
    candidates = [workspace, workspace / "automation-platform"]
    candidates.extend(child for child in workspace.iterdir() if child.is_dir())
    for candidate in candidates:
        if (candidate / "pyproject.toml").is_file() and (candidate / "packages").is_dir():
            return candidate.resolve()
    raise SystemExit("Could not auto-detect automation-platform; pass --platform-root")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("workspace", nargs="?", default=".", help="Workspace containing project folders")
    parser.add_argument("--platform-root", type=Path, help="Automation-platform root, absolute or relative to workspace")
    parser.add_argument("--project", action="append", default=[], help="Immediate child project to include; repeatable")
    parser.add_argument("--skip-project", action="append", default=[], help="Immediate child project to skip; repeatable")
    parser.add_argument("--output", type=Path, help="Output ZIP path")
    parser.add_argument("--preview-key", action="append", default=[], help="Deeply inspect this preview key; repeatable")
    parser.add_argument("--latest-previews", type=int, default=20)
    parser.add_argument("--max-rows", type=int, default=500, help="DB sample rows per ordinary table")
    parser.add_argument("--full-table-max-rows", type=int, default=100000, help="safety cap for complete CRM/reference exports; 0 means no cap")
    parser.add_argument("--database-url", help="Override DATABASE_URL; full value is retained in this unfiltered bundle")
    parser.add_argument("--verify-artifacts", action="store_true", help="Recompute hashes for preview artifacts")
    parser.add_argument("--skip-db", action="store_true", help="Skip automation-platform database diagnostic collection")
    parser.add_argument("--frappe-root", type=Path, help="Frappe Helpdesk install root; auto-detects ~/frappe-helpdesk-docker")
    parser.add_argument("--frappe-site", help="Frappe site name; auto-detected when omitted")
    parser.add_argument("--frappe-max-rows", type=int, default=10000, help="maximum rows per Frappe Helpdesk table; 0 means no cap")
    parser.add_argument("--frappe-timeout", type=int, default=900, help="Frappe Docker collector timeout in seconds")
    parser.add_argument("--skip-frappe", action="store_true", help="Skip Frappe Helpdesk Docker/context collection")
    parser.add_argument("--db-timeout", type=int, default=900, help="Database collector timeout in seconds")
    parser.add_argument("--include-sensitive", action=argparse.BooleanOptionalAction, default=True, help="Compatibility flag; unfiltered mode is enabled by default")
    parser.add_argument("--include-binaries", action="store_true", help="Include bounded binary files; off by default")
    parser.add_argument("--max-file-mb", type=int, default=25)
    parser.add_argument("--max-total-source-mb", type=int, default=750)
    parser.add_argument("--log-tail-kb", type=int, default=512)
    parser.add_argument("--max-log-files", type=int, default=200, help="Newest ignored/non-ignored logs to tail")
    parser.add_argument("--max-local-dbs", type=int, default=50, help="Newest local SQLite databases to inspect")
    parser.add_argument("--max-local-db-mb", type=int, default=4096, help="Maximum size of each local DB inspected")
    parser.add_argument("--run-checks", choices=("none", "quick", "full"), default="none")
    parser.add_argument("--check-command", action="append", default=[], help="Additional read-only/check command; repeatable")
    parser.add_argument("--check-timeout", type=int, default=1200)
    parser.add_argument("--context-note", action="append", default=[], help="Current issue/task context; repeatable")
    parser.add_argument("--context-file", type=Path, help="Text/Markdown file with additional context")
    parser.add_argument("--compression-level", type=int, choices=range(0, 10), default=6)
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("-y", "--yes", action="store_true")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    workspace = Path(args.workspace).expanduser().resolve()
    if not workspace.is_dir():
        raise SystemExit(f"Workspace is not a directory: {workspace}")
    platform_root = resolve_platform_root(workspace, args.platform_root)
    timestamp = dt.datetime.now().strftime("%Y%m%d-%H%M%S")
    output = (args.output or (workspace.parent / f"{workspace.name}-ai-context-{timestamp}.zip")).expanduser().resolve()
    if output.suffix.lower() != ".zip":
        output = output.with_suffix(".zip")

    projects = select_projects(workspace, args.project, set(args.skip_project))
    if not projects:
        raise SystemExit("No projects selected")

    print(f"Workspace     : {workspace}")
    print(f"Platform root : {platform_root}")
    print(f"Projects      : {len(projects)}")
    print(f"Output        : {output}")

    candidates, initial_skipped, warnings = scan_workspace(
        workspace, projects, output, args.max_file_mb * 1024 * 1024,
    )
    operational, operational_skipped = scan_operational_files(
        workspace, projects, args.max_log_files, args.max_local_dbs,
        args.max_local_db_mb * 1024 * 1024,
    )
    sensitive, sensitive_skipped = scan_sensitive_files(
        workspace, projects, args.max_file_mb * 1024 * 1024,
    )
    by_source = {candidate.source.resolve(): candidate for candidate in candidates}
    for candidate in [*operational, *sensitive]:
        by_source.setdefault(candidate.source.resolve(), candidate)
    candidates = list(by_source.values())
    initial_skipped.extend(operational_skipped)
    initial_skipped.extend(sensitive_skipped)
    category_counts = Counter(c.category for c in candidates)
    candidate_size = sum(c.size for c in candidates)
    print(f"Candidates    : {len(candidates):,} files / {human_size(candidate_size)}")
    print(f"Categories    : {dict(category_counts)}")
    if warnings:
        print(f"Scan warnings : {len(warnings)}")

    if args.dry_run:
        print("Dry run complete; no database collection, checks or ZIP creation performed.")
        return 0

    if not args.yes:
        answer = input("Create the AI context bundle now? [y/N] ").strip().lower()
        if answer not in {"y", "yes"}:
            print("Cancelled.")
            return 0

    with tempfile.TemporaryDirectory(prefix="deomee-ai-context-") as temp_dir:
        temp_root = Path(temp_dir)
        bundle = temp_root / "bundle"
        bundle.mkdir()

        manifest, copy_skipped, redactions, symbols, local_data = copy_candidates(
            candidates, bundle, args.include_sensitive, args.include_binaries,
            args.max_total_source_mb * 1024 * 1024,
            args.log_tail_kb * 1024,
            args.full_table_max_rows,
        )
        all_skipped = [*initial_skipped, *copy_skipped]

        repositories = find_git_repositories(workspace, projects)
        git_reports = collect_git(bundle, repositories, args.include_sensitive)
        runtime = collect_runtime(bundle, workspace, platform_root, args.include_sensitive)
        checks = run_checks(
            bundle, platform_root, args.run_checks, args.check_command,
            args.check_timeout, args.include_sensitive,
        )
        crm_source_focus = collect_crm_source_focus(bundle, manifest, symbols)
        db_result = extract_and_run_database_component(bundle, temp_root, platform_root, args)
        frappe_result = collect_frappe_helpdesk(
            bundle,
            args.frappe_root,
            args.frappe_site,
            args.frappe_max_rows,
            args.frappe_timeout,
            args.include_sensitive,
            args.skip_frappe,
        )

        write_json(bundle / "manifest" / "files.json", [asdict(record) for record in manifest])
        write_manifest_csv(bundle / "manifest" / "files.csv", manifest)
        write_json(bundle / "manifest" / "summary.json", {
            "tool_version": TOOL_VERSION,
            "generated_at": dt.datetime.now(dt.timezone.utc).isoformat(),
            "workspace": str(workspace), "platform_root": str(platform_root),
            "projects": [str(project) for project in projects],
            "file_count": len(manifest), "source_bytes": sum(r.size_bundled for r in manifest),
            "languages": dict(Counter(r.language or "Other" for r in manifest)),
            "categories": dict(Counter(r.category for r in manifest)),
            "local_data_reports": local_data,
            "crm_source_focus": {key: value for key, value in crm_source_focus.items() if key not in {"files", "symbols"}},
            "database_collector": db_result,
            "frappe_collector": frappe_result,
            "command_line": [sys.executable, *sys.argv],
        })
        write_text(bundle / "analysis" / "PROJECT_TREE.txt", build_tree(manifest))
        write_json(bundle / "analysis" / "code_symbols.json", symbols)
        write_text(bundle / "analysis" / "CODE_MAP.md", generate_code_map(symbols))

        duplicate_groups: dict[str, list[str]] = defaultdict(list)
        for record in manifest:
            duplicate_groups[record.sha256_bundled].append(record.archive_name)
        write_json(bundle / "analysis" / "duplicate_files.json", [
            {"sha256": digest, "files": files}
            for digest, files in duplicate_groups.items() if len(files) > 1
        ])
        write_json(bundle / "analysis" / "largest_files.json", [
            asdict(record) for record in sorted(manifest, key=lambda r: r.size_bundled, reverse=True)[:100]
        ])
        write_json(bundle / "security" / "skipped.json", all_skipped)
        write_json(bundle / "security" / "redactions.json", redactions)
        write_json(bundle / "security" / "warnings.json", warnings)
        write_text(bundle / "START_HERE_FOR_AI.md", generate_start_here(
            args, workspace, platform_root, projects, manifest, all_skipped,
            db_result, checks, git_reports, frappe_result,
        ))
        write_text(bundle / "NEW_CHAT_PROMPT.md", generate_new_chat_prompt(args))
        write_text(bundle / "BUNDLE_CONTENTS.txt", textwrap.dedent("""
            START_HERE_FOR_AI.md      Reading order and interpretation guidance
            NEW_CHAT_PROMPT.md        Ready-to-paste prompt for a new chat
            source/                   Sanitized source snapshot
            manifest/                 File hashes, sizes, languages and source paths
            analysis/                 Project tree, code map, symbols, duplicates, largest files
            git/                      Repository state, recent commits and diffs
            runtime/                  Host/tool/dependency/service snapshots
            checks/                   Optional test/build/check logs
            database/                 Sanitized DB schema/data plus complete CRM linked context
            frappe/                   Frappe Helpdesk Docker state and HD ticket/reply context
            local_databases/          Sanitized local SQLite diagnostics
            data_previews/            Bounded XLSX sheet previews
            logs/                     Sanitized recent log tails
            security/                 Exclusions, empty redaction report and scan warnings
            tooling/                  Embedded database collector used by this bundle
        """).strip() + "\n")

        # Hash every bundled file before creating the ZIP.
        bundle_hashes = []
        for path in sorted(bundle.rglob("*")):
            if path.is_file():
                bundle_hashes.append({"path": path.relative_to(bundle).as_posix(), "size": path.stat().st_size, "sha256": sha256_file(path)})
        write_json(bundle / "manifest" / "bundle_files.json", bundle_hashes)

        zip_bundle(bundle, output, args.compression_level)

    zip_hash = sha256_file(output)
    sidecar = output.with_suffix(output.suffix + ".sha256")
    sidecar.write_text(f"{zip_hash}  {output.name}\n", encoding="ascii")
    print("\nCompleted")
    print("=" * 72)
    print(f"Bundle : {output}")
    print(f"Size   : {human_size(output.stat().st_size)}")
    print(f"SHA256 : {zip_hash}")
    print(f"Sidecar: {sidecar}")
    print("WARNING: this compact ZIP contains unredacted credentials, tokens, personal data and selected auth state. Keep it private.")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except KeyboardInterrupt:
        eprint("\nCancelled.")
        raise SystemExit(130)
    except subprocess.CalledProcessError as exc:
        eprint(f"\nCommand failed with exit code {exc.returncode}: {exc}")
        raise SystemExit(exc.returncode or 1)
